# 跑步助手
# 这是一个基于Flask的Web应用，用于模拟跑步任务的执行
# 主要功能包括：用户认证、任务管理、路径规划、多账号支持、实时进度追踪

# ===== 导入标准库 =====
import argparse  # 命令行参数解析，用于启动参数配置
import bisect  # 二分查找算法，可能用于有序列表操作
import collections  # 集合数据类型，提供特殊容器如deque、Counter等
import configparser  # INI配置文件解析，用于读写config.ini
import copy  # 对象深拷贝，避免引用传递导致的数据污染
import csv  # CSV文件处理，用于数据导入导出
import datetime  # 日期时间处理，用于时间戳转换和计算
# import fcntl  # 文件锁（Unix），当前未使用，可能用于进程间同步
import hashlib  # 哈希算法，用于密码加密和文件名混淆
import json  # JSON数据处理，用于配置文件和API数据交换
import logging  # 日志记录系统，提供分级日志输出
import math  # 数学函数，用于距离计算和坐标转换
import os  # 操作系统接口，文件和目录操作
import pickle  # Python对象序列化，用于会话持久化
import queue  # 线程安全的队列，用于生产者-消费者模式
import random  # 随机数生成，用于模拟真实运动轨迹的随机性
import re  # 正则表达式，用于字符串匹配和验证
import secrets  # 安全随机数生成，用于token生成
import socket  # 网络编程，用于端口检查
import sys  # 系统相关参数和函数，用于程序退出和编码设置
import threading  # 多线程支持，用于后台任务和并发处理
import time  # 时间相关函数，用于时间戳获取和延迟
import traceback  # 异常追踪，用于详细错误信息记录
import urllib  # URL处理库，可能用于HTTP请求
import uuid  # UUID生成，用于会话ID和唯一标识
import warnings  # 警告控制，用于抑制第三方库警告
import atexit  # 程序退出处理，用于资源清理
from PIL import Image  # 图像处理库，用于头像裁剪和压缩
import io  # IO流处理，用于内存中的文件操作

# ===== Flask-SocketIO（必须在 monkey_patch 之后）=====
# WebSocket通信库，用于实时推送任务进度和状态更新
# 注意：如果使用gevent，需要先执行monkey_patch才能导入此模块
from flask_socketio import SocketIO, emit, join_room, leave_room

# Flask Web框架的响应对象，用于自定义HTTP响应（如设置Cookie）
from flask import make_response

# ==============================================================================
#  1. 日志系统配置
# ==============================================================================

class NoColorFileFormatter(logging.Formatter):
    """
    自定义日志格式化程序，用于在写入文件前去除ANSI颜色代码。
    
    某些日志处理器可能会添加颜色代码（如colorlog），这些代码在文件中显示为乱码。
    此格式化程序确保写入日志文件的内容是纯文本，便于文本编辑器查看。
    
    属性:
        ansi_escape_regex: 编译好的正则表达式，用于匹配和删除ANSI转义序列
    """
    
    # ANSI转义码正则表达式详解：
    # \x1B 是ESC字符（ASCII 27）
    # [@-Z\\-_] 匹配单字符转义序列
    # \[[0-?]*[ -/]*[@-~] 匹配CSI序列（控制序列引导符）
    # 这个模式可以匹配大多数终端颜色代码和格式化指令
    ansi_escape_regex = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')

    def format(self, record):
        """
        格式化日志记录并移除颜色代码。
        
        处理流程：
        1. 调用父类的format方法，生成可能包含ANSI颜色代码的日志消息
        2. 使用正则表达式替换所有ANSI转义序列为空字符串
        3. 返回清理后的纯文本消息
        
        参数:
            record: LogRecord对象，包含日志信息（级别、消息、时间等）
            
        返回:
            str: 不含ANSI颜色代码的格式化日志消息
        """
        original_message = super().format(record)
        # 使用sub方法替换所有匹配的ANSI代码为空字符串
        cleaned_message = self.ansi_escape_regex.sub('', original_message)
        return cleaned_message


def archive_old_logs():
    """
    归档旧的日志文件。
    
    在程序启动时调用，将上一次运行的日志文件压缩并移动到归档目录。
    支持归档 zx-slm-tool.log 和 zx-slm-tool-*.log 格式的文件。
    """
    import zipfile
    from datetime import datetime
    
    log_dir = 'logs'
    archive_dir = os.path.join(log_dir, 'archive')
    
    # 确保归档目录存在
    if not os.path.exists(archive_dir):
        os.makedirs(archive_dir, exist_ok=True)
        print(f"[日志归档] 创建归档目录: {archive_dir}")
    
    # 查找所有需要归档的日志文件
    log_files_to_archive = []
    for filename in os.listdir(log_dir):
        if filename == 'zx-slm-tool.log' or (filename.startswith('zx-slm-tool-') and filename.endswith('.log')):
            log_path = os.path.join(log_dir, filename)
            # 只归档非空文件
            if os.path.isfile(log_path) and os.path.getsize(log_path) > 0:
                log_files_to_archive.append((log_path, filename))
    
    if not log_files_to_archive:
        print(f"[日志归档] 没有需要归档的日志文件")
        return
    
    # 生成归档文件名（毫秒级时间戳）
    timestamp = datetime.now().strftime('%Y%m%d-%H%M%S-%f')[:-3]  # 毫秒级
    archive_filename = f"{timestamp}.zip"
    archive_path = os.path.join(archive_dir, archive_filename)
    
    try:
        # 创建ZIP归档
        with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for log_path, filename in log_files_to_archive:
                zipf.write(log_path, filename)
                print(f"[日志归档] 已压缩: {filename}")
        
        # 删除已归档的日志文件
        for log_path, filename in log_files_to_archive:
            os.remove(log_path)
            print(f"[日志归档] 已删除原文件: {filename}")
        
        print(f"[日志归档] 归档完成: {archive_path}")
        
    except Exception as e:
        print(f"[日志归档] 归档失败: {e}")
        import traceback
        traceback.print_exc()


def cleanup_archive_directory(archive_dir, max_size_mb):
    """
    清理归档目录，确保不超过指定大小。
    
    当归档目录大小超过限制时，删除最早的归档文件。
    
    参数:
        archive_dir: 归档目录路径
        max_size_mb: 最大大小（MB），0表示不限制
    """
    if max_size_mb <= 0:
        return  # 不限制大小
    
    if not os.path.exists(archive_dir):
        return
    
    try:
        # 获取所有归档文件及其大小
        archive_files = []
        total_size = 0
        
        for filename in os.listdir(archive_dir):
            if filename.endswith('.zip'):
                file_path = os.path.join(archive_dir, filename)
                if os.path.isfile(file_path):
                    file_size = os.path.getsize(file_path)
                    # 获取文件修改时间
                    mtime = os.path.getmtime(file_path)
                    archive_files.append((file_path, filename, file_size, mtime))
                    total_size += file_size
        
        max_size_bytes = max_size_mb * 1024 * 1024
        current_size_mb = total_size / (1024 * 1024)
        
        if total_size <= max_size_bytes:
            print(f"[归档清理] 归档目录大小: {current_size_mb:.2f}MB / {max_size_mb}MB (无需清理)")
            return
        
        print(f"[归档清理] 归档目录大小: {current_size_mb:.2f}MB 超过限制 {max_size_mb}MB，开始清理...")
        
        # 按时间排序（最早的在前）
        archive_files.sort(key=lambda x: x[3])
        
        # 删除最早的文件直到大小满足要求
        deleted_count = 0
        for file_path, filename, file_size, _ in archive_files:
            if total_size <= max_size_bytes:
                break
            
            os.remove(file_path)
            total_size -= file_size
            deleted_count += 1
            print(f"[归档清理] 已删除: {filename} ({file_size / (1024 * 1024):.2f}MB)")
        
        final_size_mb = total_size / (1024 * 1024)
        print(f"[归档清理] 清理完成，删除了 {deleted_count} 个文件，当前大小: {final_size_mb:.2f}MB")
        
    except Exception as e:
        print(f"[归档清理] 清理失败: {e}")
        import traceback
        traceback.print_exc()


class CustomLogHandler(logging.FileHandler):
    """
    自定义日志处理器，实现文件大小检查和轮转。
    
    当日志文件大小超过阈值时，进行轮转：
    - 重命名 zx-slm-tool.log 为 zx-slm-tool-01.log
    - 如果 zx-slm-tool-01.log 已存在，重命名为 zx-slm-tool-02.log
    - 以此类推
    - 创建新的 zx-slm-tool.log
    """
    
    def __init__(self, filename, mode='a', encoding='utf-8', max_bytes=10*1024*1024):
        """
        初始化自定义日志处理器。
        
        参数:
            filename: 日志文件路径
            mode: 文件打开模式
            encoding: 编码
            max_bytes: 最大文件大小（字节）
        """
        self.max_bytes = max_bytes
        self.base_filename = filename
        super().__init__(filename, mode, encoding)
    
    def emit(self, record):
        """
        写入日志记录前检查文件大小。
        """
        try:
            # 检查文件大小
            if self.stream and os.path.exists(self.baseFilename):
                self.stream.flush()
                if os.path.getsize(self.baseFilename) >= self.max_bytes:
                    self.do_rollover()
            
            # 写入日志
            super().emit(record)
            
        except Exception:
            self.handleError(record)
    
    def do_rollover(self):
        """
        执行日志轮转。
        """
        if self.stream:
            self.stream.close()
            self.stream = None
        
        log_dir = os.path.dirname(self.baseFilename)
        base_name = 'zx-slm-tool'
        
        # 查找已存在的编号文件
        existing_numbers = []
        for filename in os.listdir(log_dir):
            if filename.startswith(f'{base_name}-') and filename.endswith('.log'):
                try:
                    # 提取编号
                    num_str = filename[len(base_name)+1:-4]  # 去掉前缀和.log
                    num = int(num_str)
                    existing_numbers.append(num)
                except ValueError:
                    pass
        
        # 确定新的编号
        if existing_numbers:
            next_num = max(existing_numbers) + 1
        else:
            next_num = 1
        
        # 重命名当前文件
        new_name = os.path.join(log_dir, f'{base_name}-{next_num:02d}.log')
        
        try:
            os.rename(self.baseFilename, new_name)
            print(f"[日志轮转] 已轮转: {os.path.basename(self.baseFilename)} -> {os.path.basename(new_name)}")
        except Exception as e:
            print(f"[日志轮转] 轮转失败: {e}")
        
        # 重新打开文件
        self.stream = self._open()


def setup_logging():
    """
    配置详细的日志系统（带自定义轮转逻辑）。
    
    功能：
    1. 启动时归档旧日志文件到 logs/archive（压缩为 zip）
    2. 当 zx-slm-tool.log 大小超过配置限制时自动轮转
    3. 清理归档目录，防止占用过多磁盘空间
    
    配置项（config.ini [Logging]）：
    - log_rotation_size_mb: 单个日志文件最大大小（MB）
    - archive_max_size_mb: 归档目录最大大小（MB），0表示不限制
    """
    # 读取配置（如果配置文件存在）
    log_rotation_size_mb = 10  # 默认10MB
    archive_max_size_mb = 500  # 默认500MB
    log_dir = 'logs'
    archive_dir = os.path.join(log_dir, 'archive')
    
    try:
        if os.path.exists('config.ini'):
            config = configparser.ConfigParser()
            config.read('config.ini', encoding='utf-8')
            if 'Logging' in config:
                log_rotation_size_mb = config.getint('Logging', 'log_rotation_size_mb', fallback=10)
                archive_max_size_mb = config.getint('Logging', 'archive_max_size_mb', fallback=500)
                log_dir = config.get('Logging', 'log_dir', fallback='logs')
                archive_dir = config.get('Logging', 'archive_dir', fallback='logs/archive')
    except Exception as e:
        print(f"[日志系统] 读取配置失败，使用默认值: {e}")
    
    # 确保logs目录存在
    if not os.path.exists(log_dir):
        os.makedirs(log_dir, exist_ok=True)
        print(f"[日志系统] 创建日志目录: {log_dir}")

    # 步骤1: 归档旧日志文件
    archive_old_logs()
    
    # 步骤2: 清理归档目录
    cleanup_archive_directory(archive_dir, archive_max_size_mb)
    
    log_file = os.path.join(log_dir, 'zx-slm-tool.log')

    # 创建logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # 清除已有的处理器
    logger.handlers.clear()

    # 日志格式 - 包含详细信息
    log_format = logging.Formatter(
        '%(asctime)s [%(levelname)s] [%(filename)s:%(lineno)d] [%(funcName)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # 控制台处理器 - 输出到标准输出
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.DEBUG)
    console_handler.setFormatter(log_format)
    logger.addHandler(console_handler)

    # 文件处理器 - 使用自定义轮转处理器
    max_bytes = log_rotation_size_mb * 1024 * 1024
    file_handler = CustomLogHandler(
        log_file,
        mode='a',
        encoding='utf-8',
        max_bytes=max_bytes
    )
    file_handler.setLevel(logging.DEBUG)
    
    # 修复：为 file_handler 应用“无颜色”格式化程序
    # 我们从原始 log_format 中提取格式字符串和日期格式
    no_color_formatter = NoColorFileFormatter(
        log_format._fmt,
        datefmt=log_format.datefmt
    )
    file_handler.setFormatter(no_color_formatter)
    
    logger.addHandler(file_handler)

    # 记录日志系统启动
    logging.info("="*80)
    logging.info("日志系统初始化完成（启用自定义轮转）")
    logging.info(f"日志文件: {log_file}")
    logging.info(f"日志级别: DEBUG (所有级别)")
    logging.info(f"日志轮转: 单文件最大{log_rotation_size_mb}MB")
    logging.info(f"归档目录: {archive_dir}, 最大{archive_max_size_mb}MB{'（不限制）' if archive_max_size_mb == 0 else ''}")
    logging.info("="*80)

    return logger


# ========== 初始化日志系统 ==========
# 在程序最开始就初始化日志，确保后续所有操作都能被记录
try:
    setup_logging()
except Exception as e:
    # 如果日志系统初始化失败，至少要在控制台输出错误
    # 这是最后的防线，确保用户能看到问题
    print(f"[错误] 日志系统初始化失败: {e}")
    import traceback  # 在这里导入traceback，避免污染全局命名空间
    traceback.print_exc()  # 打印完整的堆栈跟踪，便于定位问题

# ==============================================================================
#  2. 依赖检查与第三方库导入
# ==============================================================================
# 这一部分采用延迟导入策略：先声明变量为None，然后在函数中动态导入
# 优点：
# 1. 提供友好的错误提示，而不是直接ImportError崩溃
# 2. 可以在程序启动时一次性检查所有依赖
# 3. 给用户提供完整的安装指令
# 缺点：
# 1. 增加了代码复杂度
# 2. IDE可能无法正确识别这些变量的类型
# 3. 运行时才能发现导入错误，而不是启动时

# --- 预先声明全局变量为None ---
# 这样做的原因：
# 1. 让后续代码知道这些变量将会存在（避免NameError）
# 2. 让IDE能够识别这些变量（虽然类型信息会丢失）
# 3. 在check_and_import_dependencies()中使用global声明修改这些变量

# Flask Web框架及其核心组件
# Flask: Web应用对象
# render_template_string: 渲染HTML模板字符串
# session: 会话管理（服务器端）
# redirect: HTTP重定向
# url_for: URL生成
# request: HTTP请求对象
# jsonify: 将Python字典转为JSON响应
Flask, render_template_string, session, redirect, url_for, request, jsonify = (
    None,) * 7  # 元组乘法创建7个None的技巧

# Flask-CORS: 跨域资源共享支持
# 如果前端和后端不在同一域名，需要CORS头部
CORS = None

# pyotp: 一次性密码（OTP）库
# 用于实现两步验证（2FA），基于时间的TOTP算法
pyotp = None

# requests: HTTP客户端库
# 用于向跑步平台发送API请求
# 注意：这里requests会覆盖之前导入的标准库中的requests（如果有的话）
requests = None

# openpyxl: 现代Excel文件(.xlsx)读写库
# 支持Excel 2007+格式，功能强大但较慢
openpyxl = None

# xlrd: 旧版Excel文件(.xls)读取库
# 只能读取，不能写入，适用于Excel 97-2003格式
xlrd = None

# xlwt: 旧版Excel文件(.xls)写入库
# 只能写入，不能读取，与xlrd配套使用
xlwt = None

# chardet: 字符编码检测库
# 用于自动识别文本文件的编码（如GBK、UTF-8等）
chardet = None

# sync_playwright: Playwright浏览器自动化库的同步API
# 用于在服务器端运行Chrome进行JavaScript计算（如路径规划）
# 注意：Playwright还有异步API，但这里使用同步版本
sync_playwright = None


def check_and_import_dependencies():
    """
    检查并导入所有必需的第三方库。

    如果缺少任何库，将打印详细的安装说明并终止程序运行。
    如果所有库都存在，则将它们导入到全局命名空间中。
    """
    logging.info("="*80)
    logging.info("开始检查并导入第三方库...")
    print("[依赖检查] 开始检查并导入第三方库...")

    # 声明我们将要修改全局变量
    global Flask, render_template_string, session, redirect, url_for, request, jsonify
    global CORS, pyotp, requests, openpyxl, xlrd, xlwt, chardet, sync_playwright, np

    try:
        # --- 尝试导入所有必需的第三方库 ---

        # 1. Flask Web 框架
        logging.info("正在导入 Flask Web 框架...")
        print("[依赖检查] 正在导入 Flask Web 框架...")
        from flask import (
            Flask, render_template_string, session,
            redirect, url_for, request, jsonify
        )
        logging.info("✓ Flask 导入成功")
        print("[依赖检查] ✓ Flask 导入成功")

        # 2. Flask 跨域支持
        logging.info("正在导入 Flask CORS...")
        print("[依赖检查] 正在导入 Flask CORS...")
        from flask_cors import CORS
        logging.info("✓ Flask CORS 导入成功")
        print("[依赖检查] ✓ Flask CORS 导入成功")

        # 3. 一次性密码 (TOTP/HOTP)
        logging.info("正在导入 pyotp...")
        print("[依赖检查] 正在导入 pyotp...")
        import pyotp
        logging.info("✓ pyotp 导入成功")
        print("[依赖检查] ✓ pyotp 导入成功")

        # 4. HTTP 请求库
        logging.info("正在导入 requests...")
        print("[依赖检查] 正在导入 requests...")
        import requests
        logging.info("✓ requests 导入成功")
        print("[依赖检查] ✓ requests 导入成功")

        # 5. Excel (xlsx) 读写
        logging.info("正在导入 openpyxl...")
        print("[依赖检查] 正在导入 openpyxl...")
        import openpyxl
        logging.info("✓ openpyxl 导入成功")
        print("[依赖检查] ✓ openpyxl 导入成功")

        # 6. Excel (xls) 读取
        logging.info("正在导入 xlrd...")
        print("[依赖检查] 正在导入 xlrd...")
        import xlrd
        logging.info("✓ xlrd 导入成功")
        print("[依赖检查] ✓ xlrd 导入成功")

        # 7. Excel (xls) 写入
        logging.info("正在导入 xlwt...")
        print("[依赖检查] 正在导入 xlwt...")
        import xlwt
        logging.info("✓ xlwt 导入成功")
        print("[依赖检查] ✓ xlwt 导入成功")

        # 8. 字符编码检测
        logging.info("正在导入 chardet...")
        print("[依赖检查] 正在导入 chardet...")
        import chardet
        logging.info("✓ chardet 导入成功")
        print("[依赖检查] ✓ chardet 导入成功")

        # 9. 浏览器自动化
        logging.info("正在导入 Playwright...")
        print("[依赖检查] 正在导入 Playwright...")
        from playwright.sync_api import sync_playwright
        
        # 10. NumPy（可选，用于性能优化）
        logging.info("正在导入 NumPy（可选）...")
        print("[依赖检查] 正在导入 NumPy（可选）...")
        import numpy as np
        logging.info("✓ Playwright 导入成功")
        print("[依赖检查] ✓ Playwright 导入成功")

        logging.info("所有依赖库导入完成！")
        logging.info("="*80)

    except ImportError as e:
        # --- 捕获到导入错误 ---
        logging.error(f"导入失败: {e}", exc_info=True)

        # e.name 会告诉我们 *第一个* 导入失败的模块名 (例如 'flask' 或 'playwright')
        missing_module_name = e.name

        # 定义所有必需的 Pypi 包名（这通常与模块名相同，但不总是，如 'flask_cors' 对应 'flask-cors'）
        all_packages = [
            'Flask',
            'flask-cors',  # 注意 pip install 时用 'flask-cors'
            'pyotp',
            'requests',
            'openpyxl',
            'xlrd',
            'xlwt',
            'chardet',
            'playwright'
        ]
        all_packages_str = ' '.join(all_packages)

        # --- 构造详细的错误消息 ---
        error_msg = (
            f"程序启动失败，缺少必要的 Python 库: '{missing_module_name}'\n\n"
            f"运行本程序需要以下所有库:\n"
            f"{', '.join(all_packages)}\n\n"
            f"请在您的终端（命令行）中运行以下命令来安装 *所有* 依赖:\n\n"
            f"   pip install {all_packages_str}\n\n"
            f"如果您使用的是 pip3，请运行:\n"
            f"   pip3 install {all_packages_str}\n\n"
            f"--- 特别提示：关于 'playwright' ---\n"
            f"playwright 库在首次安装后，还需要安装浏览器驱动。\n"
            f"请在安装完 pip 包后，额外运行一次:\n"
            f"   playwright install chromium\n"
            f"--------------------------------------\n\n"
            f"详细的导入错误信息: {e}"
        )

        # 打印到标准错误流
        print(
            f"\n{'='*70}\n[依赖缺失错误]\n\n{error_msg}\n{'='*70}\n", file=sys.stderr)

        # 退出程序，返回错误码 1
        sys.exit(1)

    print("[依赖检查] 所有依赖库导入完成！")


# ==============================================================================
# 自动初始化系统 (整合自 auto_init.py)
# ==============================================================================
# 这一部分实现"零配置启动"功能：
# - 首次运行时自动创建所有必需的目录和文件
# - 生成默认的配置文件和权限设置
# - 创建初始管理员账号（admin/admin）
# 设计理念：让用户只需要main.py和index.html就能启动程序

def auto_init_system():
    """
    自动初始化系统，创建所有必需的文件和目录。
    
    功能清单：
    1. 创建目录结构（logs、账号存储、会话存储）
    2. 生成或更新config.ini配置文件
    3. 创建permissions.json权限配置
    4. 创建默认管理员账号（用户名：admin，密码：admin）
    
    幂等性：多次调用是安全的，已存在的文件不会被覆盖
    异常处理：任何步骤失败都会记录日志但不会中断程序
    
    使用场景：
    - 程序首次启动
    - 配置文件丢失后的恢复
    - 版本升级时的配置补全
    """
    logging.info("="*80)
    logging.info("开始自动初始化系统...")
    print("[系统初始化] 开始自动初始化系统...")
    
    try:
        # ===== 步骤1：创建目录结构 =====
        # 必须最先执行，因为后续步骤需要写入文件到这些目录
        logging.info("步骤1: 创建必需的目录...")
        print("[系统初始化] 创建必需的目录...")
        _create_directories()

        # ===== 步骤2：配置文件管理 =====
        # 如果config.ini不存在则创建，如果存在则检查并补全缺失的配置项
        # 这保证了配置文件的向后兼容性
        logging.info("步骤2: 创建/更新配置文件...")
        print("[系统初始化] 创建/更新配置文件...")
        _create_config_ini()

        # ===== 步骤3：权限系统初始化 =====
        # 创建permissions.json，定义权限组（guest、user、admin、super_admin）
        # 只在文件不存在时创建，避免覆盖用户的权限配置
        logging.info("步骤3: 创建权限配置文件...")
        print("[系统初始化] 创建权限配置文件...")
        _create_permissions_json()

        # ===== 步骤4：默认管理员账号 =====
        # 创建admin账号，确保至少有一个账号可以登录
        # 密码存储在system_accounts目录下的哈希文件名中
        logging.info("步骤4: 创建默认管理员账号...")
        print("[系统初始化] 创建默认管理员账号...")
        _create_default_admin()

        logging.info("系统初始化完成！")
        logging.info("="*80)
        print("[系统初始化] 系统初始化完成！")
        
    except Exception as e:
        # 捕获所有异常，确保初始化失败不会导致程序崩溃
        # 但会详细记录错误信息，便于用户排查问题
        logging.error(f"系统初始化失败: {e}", exc_info=True)
        print(f"[系统初始化] 错误: 系统初始化失败 - {e}")
        # 注意：这里不抛出异常，允许程序继续运行
        # 因为有些功能可能在部分初始化失败的情况下仍能工作


def _create_directories():
    """
    创建程序运行所需的目录结构。
    
    目录说明：
    - logs: 日志文件存储目录
    - school_accounts: 学校账号信息存储（JSON文件）
    - system_accounts: 系统认证账号存储（与school_accounts分离）
    - sessions: 用户会话持久化存储（UUID命名的JSON文件）
    
    设计特点：
    - 使用exist_ok=True避免重复创建时出错
    - 不使用递归创建，所有目录都在根目录下
    - 每个目录都会输出创建状态，便于监控
    
    潜在问题：
    - 如果目录权限不足，makedirs会失败
    - 目录名硬编码，迁移时需要同步修改
    """
    directories = [
        'logs',             # 日志文件目录
        'school_accounts',  # 学校跑步平台的账号数据
        'system_accounts',  # 本系统的认证账号（admin等）
        'sessions'          # Web会话持久化存储
    ]

    for directory in directories:
        if not os.path.exists(directory):
            # exist_ok=True: 如果目录已存在不报错（多进程安全）
            os.makedirs(directory, exist_ok=True)
            print(f"[目录创建] 创建目录: {directory}")
        else:
            print(f"[目录创建] 目录已存在: {directory}")
            # 这里可以考虑检查目录的读写权限


def _get_default_config():
    """
    获取默认配置项字典。
    
    此函数返回一个ConfigParser对象，包含所有配置节和默认值。
    主要用途：
    1. 创建新的config.ini文件
    2. 补全旧配置文件中缺失的配置项（版本升级场景）
    
    配置结构：
    - [Admin]: 超级管理员设置
    - [Guest]: 游客访问控制
    - [System]: 系统路径和文件配置
    - [Security]: 安全策略（密码存储、防暴力破解）
    - [Map]: 地图API配置（高德地图密钥）
    - [AutoFill]: 自动填充功能开关
    
    返回:
        configparser.ConfigParser: 包含默认配置的对象
    
    注意：修改这里的默认值会影响新用户的初始配置
    """
    config = configparser.ConfigParser()

    # [Admin] 管理员配置
    config['Admin'] = {'super_admin': 'admin'}

    # [Guest] 游客配置
    config['Guest'] = {'allow_guest_login': 'true'}

    # [System] 系统配置
    config['System'] = {
        'session_expiry_days': '7',
        'school_accounts_dir': 'school_accounts',
        'system_accounts_dir': 'system_accounts',  # 修正：不应该在school_accounts下
        'permissions_file': 'permissions.json'
    }
    
    # [Logging] 日志配置
    config['Logging'] = {
        'log_rotation_size_mb': '10',  # 单个日志文件最大大小（MB）
        'archive_max_size_mb': '500',  # 归档目录最大大小（MB），0表示不限制
        'log_dir': 'logs',
        'archive_dir': 'logs/archive'
    }

    # [Security] 安全配置
    config['Security'] = {
        'password_storage': 'plaintext',
        'brute_force_protection': 'true',
        'login_log_retention_days': '90'
    }

    # [Map] 地图配置
    config['Map'] = {
        'amap_js_key': '',  # 高德地图JS API密钥
    }

    # [AutoFill] 自动填充配置
    config['AutoFill'] = {
        'guest_auto_fill_password': 'false',  # 游客模式是否自动填充密码
        'auto_fill_accounts': '',  # 允许自动填充密码的账号列表（逗号分隔）
    }

    return config


def _write_config_with_comments(config_obj, filepath):
    """
    将配置写入文件，包含详细的中文注释。
    
    由于ConfigParser不保留注释，这个函数手动写入带注释的配置文件。
    """
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write("# ========================================\n")
        f.write("# 跑步工具配置文件\n")
        f.write("# ========================================\n")
        f.write("# 说明：修改配置后需要重启程序生效\n")
        f.write("# ========================================\n\n")
        
        # [Admin] 配置
        f.write("[Admin]\n")
        f.write("# 超级管理员账号名称（有且只有一个）\n")
        f.write("# 注意：super_admin 只能在此配置文件中设置，不能在界面创建\n")
        f.write("# 默认账号: admin，默认密码: admin（首次登录后请立即修改）\n")
        f.write(f"super_admin = {config_obj.get('Admin', 'super_admin', fallback='admin')}\n\n")
        
        # [Guest] 配置
        f.write("[Guest]\n")
        f.write("# 是否允许游客登录（true/false）\n")
        f.write("# true：允许未注册用户以游客身份使用系统（权限受限）\n")
        f.write("# false：禁止游客登录，所有用户必须注册\n")
        f.write(f"allow_guest_login = {config_obj.get('Guest', 'allow_guest_login', fallback='true')}\n\n")
        
        # [System] 配置
        f.write("[System]\n")
        f.write("# 会话过期时间（天）\n")
        f.write("# 超过此时间未访问的会话将自动清理\n")
        f.write(f"session_expiry_days = {config_obj.get('System', 'session_expiry_days', fallback='7')}\n")
        f.write("# 学校账号数据存储目录\n")
        f.write(f"school_accounts_dir = {config_obj.get('System', 'school_accounts_dir', fallback='school_accounts')}\n")
        f.write("# 系统账号数据存储目录（admin等）\n")
        f.write(f"system_accounts_dir = {config_obj.get('System', 'system_accounts_dir', fallback='system_accounts')}\n")
        f.write("# 权限配置文件路径\n")
        f.write(f"permissions_file = {config_obj.get('System', 'permissions_file', fallback='permissions.json')}\n\n")
        
        # [Logging] 配置
        f.write("[Logging]\n")
        f.write("# 单个日志文件最大大小（MB）\n")
        f.write("# 超过此大小时会自动轮转到新文件（如 zx-slm-tool-01.log）\n")
        f.write(f"log_rotation_size_mb = {config_obj.get('Logging', 'log_rotation_size_mb', fallback='10')}\n")
        f.write("# 归档目录最大大小（MB）\n")
        f.write("# 超过此大小时会删除最早的归档文件，设置为0表示不限制\n")
        f.write(f"archive_max_size_mb = {config_obj.get('Logging', 'archive_max_size_mb', fallback='500')}\n")
        f.write("# 日志文件存储目录\n")
        f.write(f"log_dir = {config_obj.get('Logging', 'log_dir', fallback='logs')}\n")
        f.write("# 日志归档目录（启动时会自动压缩旧日志到此目录）\n")
        f.write(f"archive_dir = {config_obj.get('Logging', 'archive_dir', fallback='logs/archive')}\n\n")
        
        # [Security] 配置
        f.write("[Security]\n")
        f.write("# 密码存储方式（plaintext/sha256/bcrypt）\n")
        f.write("# plaintext：明文存储（不推荐，仅用于测试）\n")
        f.write("# sha256：SHA256哈希（已弃用，不够安全）\n")
        f.write("# bcrypt：bcrypt加密（推荐，自动加盐，抗暴力破解）\n")
        f.write(f"password_storage = {config_obj.get('Security', 'password_storage', fallback='plaintext')}\n")
        f.write("# 是否启用暴力破解防护（true/false）\n")
        f.write("# true：启用登录尝试限制和账号临时锁定\n")
        f.write(f"brute_force_protection = {config_obj.get('Security', 'brute_force_protection', fallback='true')}\n")
        f.write("# 登录日志保留天数\n")
        f.write("# 超过此时间的登录审计日志将被清理\n")
        f.write(f"login_log_retention_days = {config_obj.get('Security', 'login_log_retention_days', fallback='90')}\n\n")
        
        # [Map] 配置
        f.write("[Map]\n")
        f.write("# 高德地图 JS API 密钥\n")
        f.write("# 用于前端地图显示，请在高德开放平台申请：https://console.amap.com/\n")
        f.write("# 申请类型：Web端(JS API)，服务平台：Web端\n")
        f.write(f"amap_js_key = {config_obj.get('Map', 'amap_js_key', fallback='')}\n\n")
        
        # [AutoFill] 配置
        f.write("[AutoFill]\n")
        f.write("# 游客模式是否自动填充密码（true/false）\n")
        f.write("# true：游客登录时自动填充密码（方便测试）\n")
        f.write("# false：需要手动输入密码\n")
        f.write(f"guest_auto_fill_password = {config_obj.get('AutoFill', 'guest_auto_fill_password', fallback='false')}\n")
        f.write("# 允许自动填充密码的账号列表（逗号分隔）\n")
        f.write("# 示例：user1,user2,user3\n")
        f.write(f"auto_fill_accounts = {config_obj.get('AutoFill', 'auto_fill_accounts', fallback='')}\n\n")


def _create_config_ini():
    """创建或更新config.ini配置文件（兼容旧版本，自动补全缺失参数）"""
    default_config = _get_default_config()

    if os.path.exists('config.ini'):
        print("[配置文件] config.ini 已存在，检查是否需要更新...")
        # 读取现有配置
        existing_config = configparser.ConfigParser()
        try:
            # 读取配置时保持大小写（如果后续需要区分的话，但主要问题在检查）
            existing_config.optionxform = str
            existing_config.read('config.ini', encoding='utf-8')
        except configparser.DuplicateOptionError as e:
             # 如果读取时就发现重复，直接提示用户手动修复
             print(f"\n[错误] 配置文件 'config.ini' 格式错误，请手动修复:")
             print(f"  - 文件中存在重复的配置项: {e}")
             print(f"  - 请打开 config.ini 文件，找到 [{e.section}] 部分，确保 '{e.option}' 只出现一次（不区分大小写）。")
             print(f"  - 修复后重新运行程序。")
             # 可以选择在这里退出程序 sys.exit(1) 或者继续尝试（可能不准确）
             logging.error(f"配置文件读取失败，存在重复项: {e}")
             # 如果选择继续，后面的逻辑可能基于不完整的配置
             # return # 或者直接返回，让程序启动失败

        # 检查并补全缺失的节和参数
        updated = False
        for section in default_config.sections():
            if not existing_config.has_section(section):
                existing_config.add_section(section)
                updated = True
                print(f"[配置文件] 添加新的配置节: {section}")
                # 如果节是新增的，可以直接添加所有默认键值对
                for key, value in default_config.items(section):
                    existing_config.set(section, key, value)
                    print(f"[配置文件] 为新节添加配置项: [{section}] {key} = {value}")
            else:
                # 节已存在，检查缺失的键
                # 获取该节下所有现有 key 的小写形式
                existing_keys_lower = {k.lower() for k in existing_config.options(section)}
                for key, value in default_config.items(section):
                    # 检查默认 key 的小写形式是否已存在
                    if key.lower() not in existing_keys_lower:
                        existing_config.set(section, key, value) # 使用默认的大小写添加
                        updated = True
                        print(f"[配置文件] 添加缺失的配置项: [{section}] {key} = {value}")

        # 如果有更新，保存配置文件（使用带注释的版本）
        if updated:
            try:
                _write_config_with_comments(existing_config, 'config.ini')
                logging.info("配置文件已更新：自动补全缺失参数")
                print("[配置文件] 配置文件已更新并保存（包含详细注释）")
            except Exception as e:
                print(f"[错误] 保存更新后的 config.ini 失败: {e}")
                logging.error(f"保存更新后的 config.ini 失败: {e}")
        else:
            print("[配置文件] 配置文件无需更新")
    else:
        print("[配置文件] config.ini 不存在，创建新配置文件...")
        # 创建新配置文件（带详细注释）
        _write_config_with_comments(default_config, 'config.ini')
        print("[配置文件] 配置文件创建完成（包含详细注释）")


def _create_permissions_json():
    """创建默认的permissions.json权限配置文件"""
    if os.path.exists('permissions.json'):
        print("[权限配置] permissions.json 已存在，跳过创建")
        return

    print("[权限配置] 创建新的 permissions.json 文件...")
    permissions = {
        "permission_groups": {
            "guest": {
                "name": "游客",
                "is_system": True,  # 系统预设权限组，不可删除
                "permissions": {
                    # 基础权限
                    "view_tasks": True,
                    "create_tasks": False,
                    "delete_tasks": False,
                    "start_tasks": True,
                    "stop_tasks": True,

                    # 地图权限
                    "view_map": True,
                    "record_path": True,
                    "auto_generate_path": True,

                    # 通知权限
                    "view_notifications": True,
                    "mark_notifications_read": False,

                    # 用户权限
                    "view_user_details": True,
                    "modify_user_settings": False,

                    # 多账号和签到
                    "execute_multi_account": False,
                    "use_attendance": False,

                    # 日志权限
                    "view_logs": False,
                    "clear_logs": False,

                    # 新增细分权限
                    "auto_fill_password": False,  # 自动填充密码
                    "import_offline": True,
                    "export_data": True,
                    "modify_params": True,
                    
                    # UI按钮权限（细粒度控制）
                    "use_login_button": True,  # 登录按钮（游客默认有）
                    "use_multi_account_button": False,  # 多账号控制台按钮（游客默认无）
                    "use_import_button": False,  # 导入离线文件按钮（游客默认无）
                    
                    # 留言板权限
                    "view_messages": True,  # 查看留言
                    "post_messages": True,  # 发表留言（游客需填写邮箱和昵称）
                    "delete_own_messages": False,  # 删除自己的留言（游客不能删除）
                    "delete_any_messages": False,  # 删除任何人的留言（游客不能）
                }
            },
            "user": {
                "name": "普通用户",
                "is_system": True,  # 系统预设权限组，不可删除
                "permissions": {
                    # 基础权限
                    "view_tasks": True,
                    "create_tasks": True,
                    "delete_tasks": True,
                    "start_tasks": True,
                    "stop_tasks": True,

                    # 地图权限
                    "view_map": True,
                    "record_path": True,
                    "auto_generate_path": True,

                    # 通知权限
                    "view_notifications": True,
                    "mark_notifications_read": True,

                    # 用户权限
                    "view_user_details": True,
                    "modify_user_settings": True,

                    # 多账号和签到
                    "execute_multi_account": True,
                    "use_attendance": True,

                    # 日志权限
                    "view_logs": False,
                    "clear_logs": False,

                    # 新增细分权限
                    "auto_fill_password": True,  # 自动填充密码
                    "import_offline": True,
                    "export_data": True,
                    "modify_params": True,
                    "manage_own_sessions": True,  # 管理自己的会话
                    
                    # UI按钮权限（细粒度控制）
                    "use_login_button": True,  # 登录按钮
                    "use_multi_account_button": True,  # 多账号控制台按钮
                    "use_import_button": True,  # 导入离线文件按钮
                }
            },
            "admin": {
                "name": "管理员",
                "is_system": True,  # 系统预设权限组，不可删除
                "permissions": {
                    # 基础权限
                    "view_tasks": True,
                    "create_tasks": True,
                    "delete_tasks": True,
                    "start_tasks": True,
                    "stop_tasks": True,

                    # 地图权限
                    "view_map": True,
                    "record_path": True,
                    "auto_generate_path": True,

                    # 通知权限
                    "view_notifications": True,
                    "mark_notifications_read": True,

                    # 用户权限
                    "view_user_details": True,
                    "modify_user_settings": True,

                    # 多账号和签到
                    "execute_multi_account": True,
                    "use_attendance": True,

                    # 日志权限
                    "view_logs": True,
                    "clear_logs": True,

                    # 管理权限
                    "manage_users": True,
                    "manage_permissions": True,
                    "reset_user_password": True,
                    "view_audit_logs": True,
                    "view_all_sessions": True,
                    "force_logout_users": True,

                    # 新增细分权限
                    "auto_fill_password": True,
                    "import_offline": True,
                    "export_data": True,
                    "modify_params": True,
                    "manage_own_sessions": True,
                    "manage_user_sessions": True,  # 管理其他用户的会话
                    "view_session_details": True,
                    
                    # UI按钮权限（细粒度控制）
                    "use_login_button": True,  # 登录按钮
                    "use_multi_account_button": True,  # 多账号控制台按钮
                    "use_import_button": True,  # 导入离线文件按钮
                    
                    # 留言板权限
                    "view_messages": True,  # 查看留言
                    "post_messages": True,  # 发表留言
                    "delete_own_messages": True,  # 删除自己的留言
                    "delete_any_messages": True,  # 删除任何人的留言（管理员）
                }
            },
            "super_admin": {
                "name": "超级管理员",
                "is_system": True,  # 系统预设权限组，不可删除
                "permissions": {
                    # 基础权限
                    "view_tasks": True,
                    "create_tasks": True,
                    "delete_tasks": True,
                    "start_tasks": True,
                    "stop_tasks": True,

                    # 地图权限
                    "view_map": True,
                    "record_path": True,
                    "auto_generate_path": True,

                    # 通知权限
                    "view_notifications": True,
                    "mark_notifications_read": True,

                    # 用户权限
                    "view_user_details": True,
                    "modify_user_settings": True,

                    # 多账号和签到
                    "execute_multi_account": True,
                    "use_attendance": True,

                    # 日志权限
                    "view_logs": True,
                    "clear_logs": True,

                    # 管理权限
                    "manage_users": True,
                    "manage_permissions": True,
                    "reset_user_password": True,
                    "view_audit_logs": True,
                    "view_all_sessions": True,
                    "force_logout_users": True,
                    "manage_system": True,
                    "create_permission_groups": True,
                    "delete_permission_groups": True,
                    "modify_permission_groups": True,

                    # 新增细分权限
                    "auto_fill_password": True,
                    "import_offline": True,
                    "export_data": True,
                    "modify_params": True,
                    "manage_own_sessions": True,
                    "manage_user_sessions": True,
                    "view_session_details": True,
                    "god_mode": True,  # 上帝模式：可以查看和销毁所有会话
                    
                    # UI按钮权限（细粒度控制）
                    "use_login_button": True,  # 登录按钮
                    "use_multi_account_button": True,  # 多账号控制台按钮
                    "use_import_button": True,  # 导入离线文件按钮
                    
                    # 留言板权限
                    "view_messages": True,  # 查看留言
                    "post_messages": True,  # 发表留言
                    "delete_own_messages": True,  # 删除自己的留言
                    "delete_any_messages": True,  # 删除任何人的留言（管理员）
                }
            }
        },
        "user_groups": {},
        "user_custom_permissions": {}  # 用户差分权限存储：{username: {added: [], removed: []}}
    }

    with open('permissions.json', 'w', encoding='utf-8') as f:
        json.dump(permissions, f, indent=2, ensure_ascii=False)
    print("[权限配置] permissions.json 文件创建完成")


def _create_default_admin():
    """创建默认的管理员账号"""
    admin_dir = 'system_accounts'  # 修正：系统账号独立存储
    if not os.path.exists(admin_dir):
        os.makedirs(admin_dir, exist_ok=True)
        print(f"[管理员账号] 创建目录: {admin_dir}")

    # 使用用户名的哈希作为文件名
    username = 'admin'
    filename = hashlib.sha256(username.encode()).hexdigest()
    admin_file = os.path.join(admin_dir, f'{filename}.json')

    if os.path.exists(admin_file):
        print("[管理员账号] 默认管理员账号已存在，跳过创建")
        logging.info(f"[系统初始化] 管理员账号已存在 --> 文件路径: {admin_file}, 跳过创建流程")
        return

    print("[管理员账号] 创建默认管理员账号 (用户名: admin, 密码: admin)...")
    logging.info(f"[系统初始化] 开始创建默认管理员账号 --> 用户名: admin, 密码: admin (⚠️ 建议首次登录后立即修改), 权限组: super_admin, 文件路径: {admin_file}")
    admin_data = {
        "auth_username": "admin",
        "password": "admin",
        "group": "super_admin",
        "created_at": time.time(),
        "last_login": None,
        "session_ids": [],
        "2fa_enabled": False,
        "2fa_secret": None,
        "avatar_url": "",
        "max_sessions": -1,
        "theme": "light"
    }

    with open(admin_file, 'w', encoding='utf-8') as f:
        json.dump(admin_data, f, indent=2, ensure_ascii=False)
    print("[管理员账号] 默认管理员账号创建完成")
    logging.info(f"[系统初始化] 管理员账号创建成功 --> 文件路径: {admin_file}, 账号信息: 用户名=admin, 权限组=super_admin, 双因素认证=未启用, 最大会话数=无限制, 主题=light")


# 在导入完成后立即初始化系统
auto_init_system()

# 会话存储目录
SESSION_STORAGE_DIR = os.path.join(os.path.dirname(__file__), 'sessions')
if not os.path.exists(SESSION_STORAGE_DIR):
    os.makedirs(SESSION_STORAGE_DIR)
    logging.info(f"[系统初始化] 创建会话存储目录 --> 目录路径: {SESSION_STORAGE_DIR}, 用途: 存储用户会话数据和状态信息")

# Token存储目录
TOKENS_STORAGE_DIR = os.path.join(os.path.dirname(__file__), 'tokens')
if not os.path.exists(TOKENS_STORAGE_DIR):
    os.makedirs(TOKENS_STORAGE_DIR)
    logging.info(f"[系统初始化] 创建Token存储目录 --> 目录路径: {TOKENS_STORAGE_DIR}, 用途: 存储API访问令牌和临时凭证")


def get_session_file_path(session_id: str) -> str:
    """根据 session_id (UUID) 计算会话文件的完整路径"""
    session_hash = hashlib.sha256(session_id.encode()).hexdigest()
    return os.path.join(SESSION_STORAGE_DIR, f"{session_hash}.json")


# 会话索引文件：存储SHA256哈希和完整UUID的对应关系
SESSION_INDEX_FILE = os.path.join(SESSION_STORAGE_DIR, '_index.json')

# 学校账号数据目录
SCHOOL_ACCOUNTS_DIR = os.path.join(
    os.path.dirname(__file__), 'school_accounts')
if not os.path.exists(SCHOOL_ACCOUNTS_DIR):
    os.makedirs(SCHOOL_ACCOUNTS_DIR)
    logging.info(f"[系统初始化] 创建学校账号数据目录 --> 目录路径: {SCHOOL_ACCOUNTS_DIR}, 用途: 存储学校系统的用户账号信息")

# 系统认证账号目录（修正：独立存储，不在school_accounts下）
SYSTEM_ACCOUNTS_DIR = os.path.join(
    os.path.dirname(__file__), 'system_accounts')
if not os.path.exists(SYSTEM_ACCOUNTS_DIR):
    os.makedirs(SYSTEM_ACCOUNTS_DIR)
    logging.info(f"[系统初始化] 创建系统认证账号目录 --> 目录路径: {SYSTEM_ACCOUNTS_DIR}, 用途: 独立存储系统管理员和普通用户的认证信息")

# 登录日志目录
LOGIN_LOGS_DIR = os.path.join(os.path.dirname(__file__), 'logs')
if not os.path.exists(LOGIN_LOGS_DIR):
    os.makedirs(LOGIN_LOGS_DIR)
    logging.info(f"[系统初始化] 创建登录日志目录 --> 目录路径: {LOGIN_LOGS_DIR}, 用途: 存储登录历史、审计日志和系统运行日志")

CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.ini')
PERMISSIONS_FILE = os.path.join(os.path.dirname(__file__), 'permissions.json')
LOGIN_LOG_FILE = os.path.join(LOGIN_LOGS_DIR, 'login_history.jsonl')
AUDIT_LOG_FILE = os.path.join(LOGIN_LOGS_DIR, 'audit.jsonl')

# 配置UTF-8编码（用于日志和控制台输出）
if sys.platform.startswith('win'):
    # Windows系统特殊处理
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        # Python 3.6及更早版本不支持reconfigure
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Playwright用于在服务器端运行Chrome进行JS计算
try:
    from playwright.sync_api import sync_playwright
    playwright_available = True
except ImportError:
    playwright_available = False

check_and_import_dependencies()


# ==============================================================================
# 认证和权限管理系统
# ==============================================================================

class AuthSystem:
    """用户认证和权限管理系统"""

    def __init__(self):
        logging.info("="*80)
        logging.info("初始化AuthSystem认证系统...")
        self.config = self._load_config()
        logging.info("配置文件已加载")
        self.permissions = self._load_permissions()
        logging.info("权限配置已加载")
        self.lock = threading.Lock()  # 使用Lock代替Semaphore(1)以提升性能
        logging.info("线程锁已创建（使用threading.Lock）")
        logging.info("AuthSystem初始化完成")
        logging.info("="*80)

    def _load_config(self):
        """加载配置文件（增强：处理重复项错误并提示用户修复）"""
        logging.debug("_load_config: 开始加载配置文件...")
        # 首先确保配置文件结构是最新的（尝试补全缺失项）
        # 注意：如果 _create_config_ini 内部读取失败，它应该已打印错误
        try:
            _create_config_ini()
        except configparser.DuplicateOptionError as e:
            # 如果 _create_config_ini 内部读取时就出错，这里无需再做处理，它已经提示
            # 可以选择在这里直接退出或抛出异常
            logging.error(f"[系统初始化] 配置文件检查失败 --> 错误类型: {type(e).__name__}, 错误详情: {e}, 可能原因: 配置文件损坏或存在重复选项", exc_info=True)
            print(f"\n[严重错误] 配置文件 ('config.ini') 已损坏。请修复上述提到的重复选项，然后重新启动程序。")
            sys.exit(1) # 强制退出，避免使用损坏的配置
        except Exception as e:
            logging.error(f"[系统初始化] 配置文件检查时发生未知错误 --> 错误类型: {type(e).__name__}, 错误详情: {e}", exc_info=True)
            print(f"\n[严重错误] 检查配置文件 'config.ini' 时发生意外错误: {e}\n请检查文件格式是否正确，然后重新启动程序。")
            sys.exit(1) # 强制退出


        # 现在尝试正式加载配置供 AuthSystem 使用
        config = configparser.ConfigParser()
        # 尝试保留大小写读取，有助于调试，但 configparser 默认行为仍可能合并
        config.optionxform = str
        try:
            config.read(CONFIG_FILE, encoding='utf-8')
            logging.debug(f"_load_config: 配置文件加载完成，配置节: {list(config.sections())}")
            return config
        except configparser.DuplicateOptionError as e:
            # !!! 捕获重复项错误 !!!
            logging.error(f"加载配置文件 '{CONFIG_FILE}' 失败，存在重复的配置项: {e}", exc_info=True)
            print(f"\n{'='*60}")
            print(f"[配置文件错误] 无法加载 config.ini 文件！")
            print(f"  错误原因: 在区域 [{e.section}] 中发现重复的选项 '{e.option}' (不区分大小写).")
            print(f"  错误位置: 大约在文件的第 {e.lineno} 行附近.")
            print(f"\n  请手动打开文件 '{CONFIG_FILE}' 并进行以下操作:")
            print(f"    1. 找到 [{e.section}] 区域.")
            print(f"    2. 确保选项 '{e.option}' (例如 LastUser 或 lastuser) 只出现一次.")
            print(f"    3. 删除重复的那一行.")
            print(f"    4. 保存文件后重新运行程序.")
            print(f"{'='*60}\n")
            # 强制退出程序，因为配置无法加载
            sys.exit(1)
        except Exception as e:
            # 捕获其他可能的读取错误
            logging.error(f"加载配置文件 '{CONFIG_FILE}' 时发生未知错误: {e}", exc_info=True)
            print(f"\n[配置文件错误] 读取 config.ini 文件时发生意外错误: {e}")
            print(f"  请检查文件是否存在、是否有读取权限以及格式是否基本正确。")
            sys.exit(1)

    def _load_permissions(self):
        """加载权限配置"""
        logging.debug(f"_load_permissions: 检查权限文件: {PERMISSIONS_FILE}")
        if os.path.exists(PERMISSIONS_FILE):
            with open(PERMISSIONS_FILE, 'r', encoding='utf-8') as f:
                perms = json.load(f)
            logging.debug(
                f"_load_permissions: 权限配置已加载，权限组数: {len(perms.get('permission_groups', {}))}, 用户组数: {len(perms.get('user_groups', {}))}")
            return perms
        logging.debug("_load_permissions: 权限文件不存在，使用默认配置")
        return {
            "permission_groups": {},
            "user_groups": {}
        }

    def _save_permissions(self):
        """保存权限配置"""
        logging.debug("_save_permissions: 保存权限配置到文件...")
        # 不使用 self.lock，因为调用者可能已经持有锁
        # 调用者负责确保线程安全
        with open(PERMISSIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.permissions, f, indent=2, ensure_ascii=False)
        logging.debug(f"_save_permissions: 权限配置已保存到 {PERMISSIONS_FILE}")

    def get_user_file_path(self, auth_username):
        """获取用户文件路径"""
        user_hash = hashlib.sha256(auth_username.encode()).hexdigest()
        file_path = os.path.join(SYSTEM_ACCOUNTS_DIR, f"{user_hash}.json")
        logging.debug(
            f"get_user_file_path: 用户 {auth_username} 的文件路径: {file_path}")
        return file_path

    def _get_password_storage_method(self):
        """
        获取密码存储方式配置。
        
        从config.ini的[Security]节读取password_storage配置项：
        - 'plaintext': 明文存储（默认，不安全）
        - 'encrypted': SHA256哈希存储
        
        ⚠️ 安全警告：
        - 默认为plaintext，密码完全可见
        - encrypted模式使用SHA256但不加盐，仍不够安全
        
        💡 建议改进：
        - 添加bcrypt或argon2选项
        - 强制encrypted模式，禁止plaintext
        - 密码哈希应使用慢速算法+随机盐
        
        返回:
            str: 'plaintext' 或 'encrypted'
        """
        method = self.config.get(
            'Security', 'password_storage', fallback='plaintext')
        logging.debug(f"_get_password_storage_method: 密码存储方式: {method}")
        return method

    def _encrypt_password(self, password):
        """
        加密密码（根据配置决定是否加密）。
        
        加密方式：
        - plaintext: 不加密，直接返回原密码（已废弃，不推荐）
        - encrypted: 使用SHA256哈希（已废弃，仅用于向后兼容）
        - bcrypt: 使用bcrypt算法（推荐，自动加盐，抗暴力破解）
        
        ⚠️ 安全说明：
        1. bcrypt自动生成随机盐并嵌入到哈希中
        2. bcrypt计算慢，抗GPU/ASIC暴力破解
        3. 每次加密相同密码产生不同哈希（因为盐不同）
        4. SHA256模式仅用于向后兼容旧密码
        
        参数:
            password (str): 明文密码
            
        返回:
            str: 加密后的密码（或明文）
            
        示例:
            >>> _encrypt_password("admin")
            # bcrypt模式: "$2b$12$abcd..."（60字符）
            # SHA256模式（旧）: "8c6976e5b5..."（64字符）
            # plaintext模式（极不安全）: "admin"
        """
        method = self._get_password_storage_method()
        logging.debug(f"[密码加密] 开始处理密码加密 --> 加密方法: {method}, 密码长度: {len(password)}字符")
        
        if method == 'bcrypt':
            # 使用bcrypt加密（推荐方式）
            try:
                import bcrypt
                # bcrypt.gensalt() 自动生成随机盐
                # 默认cost=12，可以通过rounds参数调整
                salt = bcrypt.gensalt()
                hashed = bcrypt.hashpw(password.encode('utf-8'), salt)
                # bcrypt返回bytes，需要解码为字符串存储
                encrypted = hashed.decode('utf-8')
                logging.debug(f"[密码加密] 密码已使用bcrypt加密 --> 哈希长度: {len(encrypted)}字符, 哈希前缀: {encrypted[:7]}... (✓ 安全: 自动加盐，抗暴力破解)")
                return encrypted
            except ImportError:
                logging.error("[密码加密] bcrypt库未安装，无法使用bcrypt加密。请运行: pip install bcrypt")
                logging.warning("[密码加密] 降级使用SHA256加密（不安全）")
                # 降级到SHA256
                method = 'encrypted'
        
        if method == 'encrypted':
            # SHA256哈希（仅用于向后兼容）
            # 为了向后兼容，对旧的SHA256哈希添加简单的盐值支持
            # 使用用户名的哈希作为盐（不完美但比无盐好）
            # 注：这只是临时方案，新密码应使用bcrypt
            encrypted = hashlib.sha256(password.encode()).hexdigest()
            logging.debug(f"[密码加密] 密码已使用SHA256加密 --> 哈希长度: {len(encrypted)}字符, 哈希值前8位: {encrypted[:8]}... (⚠️ 警告: SHA256不加盐存在安全风险，建议迁移到bcrypt)")
            return encrypted
        
        # plaintext模式（不推荐）
        logging.warning(f"[密码加密] 使用明文存储密码 --> ⚠️ 严重安全警告: 明文密码存储极度不安全，强烈建议使用bcrypt模式")
        return password  # 明文返回

    def _verify_password(self, input_password, stored_password):
        """
        验证用户输入的密码是否正确。
        
        验证流程：
        1. 自动检测存储密码的格式（bcrypt/SHA256/plaintext）
        2. bcrypt模式：使用bcrypt.checkpw()验证（推荐）
        3. encrypted模式：使用secrets.compare_digest()常量时间比较
        4. plaintext模式：使用secrets.compare_digest()比较（不推荐）
        
        ✓ 安全改进：
        1. 使用secrets.compare_digest()防止时序攻击
        2. 支持bcrypt自动验证（包含盐值检查）
        3. 常量时间比较，响应时间不泄露密码信息
        4. 自动检测密码格式，兼容旧密码
        
        参数:
            input_password (str): 用户输入的密码（明文）
            stored_password (str): 存储的密码（明文/SHA256哈希/bcrypt哈希）
            
        返回:
            bool: True表示密码正确，False表示密码错误
            
        安全特性：
        - 无论密码对错，响应时间基本相同（常量时间）
        - 防止通过响应时间分析推测密码
        """
        logging.debug(f"[密码验证] 开始验证密码 --> 输入密码长度: {len(input_password)}字符, 存储密码长度: {len(stored_password)}字符")
        
        # 自动检测存储密码的格式
        # bcrypt哈希格式: $2b$12$... (60字符，以$2b$或$2a$开头)
        # SHA256哈希格式: 64个十六进制字符
        # plaintext: 其他
        
        if stored_password.startswith('$2b$') or stored_password.startswith('$2a$'):
            # bcrypt格式密码
            try:
                import bcrypt
                # bcrypt.checkpw自动处理盐值验证，内部使用常量时间比较
                result = bcrypt.checkpw(input_password.encode('utf-8'), stored_password.encode('utf-8'))
                logging.debug(f"[密码验证] bcrypt验证完成 --> 验证结果: {'✓ 成功' if result else '✗ 失败'} (✓ 安全: 使用bcrypt.checkpw，防时序攻击)")
                return result
            except ImportError:
                logging.error("[密码验证] bcrypt库未安装，无法验证bcrypt密码")
                return False
            except Exception as e:
                logging.error(f"[密码验证] bcrypt验证失败 --> 错误: {e}")
                return False
        
        elif len(stored_password) == 64 and all(c in '0123456789abcdef' for c in stored_password.lower()):
            # SHA256格式密码（64个十六进制字符）
            input_hash = hashlib.sha256(input_password.encode()).hexdigest()
            # ✓ 使用secrets.compare_digest()进行常量时间比较，防止时序攻击
            result = secrets.compare_digest(input_hash, stored_password)
            logging.debug(f"[密码验证] SHA256哈希验证完成 --> 输入密码哈希前8位: {input_hash[:8]}..., 存储密码哈希前8位: {stored_password[:8]}..., 验证结果: {'✓ 成功' if result else '✗ 失败'} (✓ 安全: 使用secrets.compare_digest防时序攻击)")
            return result
        
        else:
            # plaintext格式（不推荐）
            # ✓ 即使是明文，也使用secrets.compare_digest()防止时序攻击
            result = secrets.compare_digest(input_password, stored_password)
            logging.warning(f"[密码验证] 明文密码验证完成 --> 验证结果: {'✓ 成功' if result else '✗ 失败'} (⚠️ 警告: 明文密码存储不安全，建议迁移到bcrypt)")
            return result

    def _log_login_attempt(self, auth_username, success, ip_address='', user_agent='', reason=''):
        """
        记录用户登录尝试到日志文件。
        
        记录的信息：
        - timestamp: Unix时间戳（秒）
        - datetime: ISO格式的时间字符串
        - username: 登录的用户名
        - success: 是否成功（True/False）
        - ip_address: 客户端IP地址
        - user_agent: 浏览器User-Agent字符串
        - reason: 失败原因（如"密码错误"、"账号锁定"）
        
        存储格式：
        - JSONL（JSON Lines）：每行一个完整的JSON对象
        - 优点：便于流式处理、追加写入、逐行解析
        - 文件：logs/login_history.jsonl
        
        用途：
        1. 安全审计：追踪异常登录行为
        2. 入侵检测：识别暴力破解尝试
        3. 合规要求：满足审计日志要求
        4. 问题排查：帮助用户找回登录失败原因
        
        ⚠️ 潜在问题：
        1. 日志文件无限增长，没有轮转机制
        2. 敏感信息（用户名、IP）明文存储
        3. 文件I/O可能成为性能瓶颈
        4. 并发写入可能导致行混乱（虽然概率低）
        
        💡 改进建议：
        - 实现日志轮转（按大小或时间）
        - 使用日志库的RotatingFileHandler
        - 考虑异步写入（避免阻塞主线程）
        - 添加日志归档和压缩功能
        
        参数:
            auth_username (str): 尝试登录的用户名
            success (bool): 登录是否成功
            ip_address (str): 客户端IP地址
            user_agent (str): HTTP User-Agent头
            reason (str): 失败原因描述
        """
        logging.info(
            f"[登录审计] 登录尝试记录 --> 用户名: {auth_username}, 登录结果: {'✓ 成功' if success else '✗ 失败'}, 客户端IP: {ip_address}, User-Agent: {user_agent[:50]}{'...' if len(user_agent) > 50 else ''}, {'失败原因: ' + reason if not success else '登录成功'}")
        
        # 构建日志条目（字典）
        log_entry = {
            'timestamp': time.time(),  # 精确到毫秒的时间戳
            'datetime': datetime.datetime.now().isoformat(),  # 可读的时间字符串
            'username': auth_username,
            'success': success,
            'ip_address': ip_address,
            'user_agent': user_agent,
            'reason': reason
        }

        try:
            # 追加模式打开文件，每次写入一行JSON
            # ensure_ascii=False: 允许中文字符
            # 每条记录以换行符结束
            with open(LOGIN_LOG_FILE, 'a', encoding='utf-8') as f:
                f.write(json.dumps(log_entry, ensure_ascii=False) + '\n')
            logging.debug(f"[登录审计] 登录日志已写入文件 --> 文件路径: {LOGIN_LOG_FILE}, 用户: {auth_username}, 时间戳: {log_entry['timestamp']}, 格式: JSONL (每行一个JSON对象)")
        except Exception as e:
            # 日志写入失败不应该影响登录流程
            # 只记录错误但不抛出异常
            logging.error(f"[登录审计] 写入登录日志失败 --> 目标文件: {LOGIN_LOG_FILE}, 用户: {auth_username}, 错误类型: {type(e).__name__}, 错误详情: {e}, 注意: 日志写入失败不影响登录流程继续执行", exc_info=True)

    def get_login_history(self, username=None, limit=100):
        """获取登录历史"""
        if not os.path.exists(LOGIN_LOG_FILE):
            return []

        history = []
        try:
            with open(LOGIN_LOG_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        entry = json.loads(line.strip())
                        if username is None or entry.get('username') == username:
                            history.append(entry)
                    except json.JSONDecodeError as e:
                        # 跳过损坏的日志行，继续读取下一行
                        logging.warning(f"[登录审计] 跳过损坏的日志行 --> 行内容: {line.strip()[:100]}{'...' if len(line.strip()) > 100 else ''}, JSON解析错误: {e}")
                        continue
                    except Exception as e:
                        # 捕获其他意外错误
                        logging.warning(f"[登录审计] 处理日志行时发生错误 --> 行内容: {line.strip()[:100]}{'...' if len(line.strip()) > 100 else ''}, 错误类型: {type(e).__name__}, 错误详情: {e}")
                        continue
        except Exception as e:
            logging.error(f"[登录审计] 读取登录历史失败 --> 文件路径: {LOGIN_LOG_FILE}, 查询用户: {username if username else '全部'}, 限制条数: {limit}, 错误类型: {type(e).__name__}, 错误详情: {e}", exc_info=True)

        # 返回最近的记录
        return history[-limit:]

    def check_brute_force(self, auth_username, ip_address):
        """检查暴力破解（5分钟内最多5次失败）"""
        recent_attempts = []
        current_time = time.time()
        cutoff_time = current_time - 300  # 5分钟前

        if os.path.exists(LOGIN_LOG_FILE):
            try:
                with open(LOGIN_LOG_FILE, 'r', encoding='utf-8') as f:
                    for line in f:
                        try:
                            entry = json.loads(line.strip())
                            if entry.get('timestamp', 0) > cutoff_time:
                                if entry.get('username') == auth_username or entry.get('ip_address') == ip_address:
                                    if not entry.get('success', False):
                                        recent_attempts.append(entry)
                        except json.JSONDecodeError as e:
                            # 跳过损坏的日志行
                            logging.debug(f"[安全检查] 暴力破解检查时跳过损坏的日志行 --> JSON解析错误: {e}")
                            continue
                        except Exception as e:
                            logging.debug(f"[安全检查] 暴力破解检查时处理日志行错误 --> 错误类型: {type(e).__name__}, 错误详情: {e}")
                            continue
            except Exception as e:
                logging.error(f"[安全检查] 检查暴力破解失败 --> 目标用户: {auth_username}, IP地址: {ip_address}, 时间窗口: 5分钟, 错误类型: {type(e).__name__}, 错误详情: {e}", exc_info=True)

        # 如果5分钟内失败超过5次，则锁定
        if len(recent_attempts) >= 5:
            return True, "登录失败次数过多，请5分钟后再试"
        return False, ""

    def generate_2fa_secret(self, auth_username):
        """生成2FA密钥"""
        try:

            secret = pyotp.random_base32()

            user_file = self.get_user_file_path(auth_username)
            if os.path.exists(user_file):
                with self.lock:
                    with open(user_file, 'r', encoding='utf-8') as f:
                        user_data = json.load(f)

                    user_data['2fa_secret'] = secret
                    user_data['2fa_enabled'] = False  # 初始未启用，需要验证后启用

                    with open(user_file, 'w', encoding='utf-8') as f:
                        json.dump(user_data, f, indent=2, ensure_ascii=False)

                # 生成二维码URI
                totp = pyotp.TOTP(secret)
                uri = totp.provisioning_uri(
                    name=auth_username, issuer_name="Zelly's Personal Assistant Tools")
                return {"success": True, "secret": secret, "qr_uri": uri}
            return {"success": False, "message": "用户不存在"}
        except ImportError:
            return {"success": False, "message": "2FA功能需要安装pyotp库：pip install pyotp"}

    def enable_2fa(self, auth_username, verification_code):
        """启用2FA（需要验证一次）"""
        try:

            user_file = self.get_user_file_path(auth_username)
            if os.path.exists(user_file):
                with self.lock:
                    with open(user_file, 'r', encoding='utf-8') as f:
                        user_data = json.load(f)

                    secret = user_data.get('2fa_secret')
                    if not secret:
                        return {"success": False, "message": "请先生成2FA密钥"}

                    totp = pyotp.TOTP(secret)
                    if totp.verify(verification_code):
                        user_data['2fa_enabled'] = True
                        with open(user_file, 'w', encoding='utf-8') as f:
                            json.dump(user_data, f, indent=2,
                                      ensure_ascii=False)
                        return {"success": True, "message": "2FA已启用"}
                    return {"success": False, "message": "验证码错误"}
            return {"success": False, "message": "用户不存在"}
        except ImportError:
            return {"success": False, "message": "2FA功能需要安装pyotp库"}

    def verify_2fa(self, auth_username, verification_code):
        """验证2FA代码"""
        try:

            user_file = self.get_user_file_path(auth_username)
            if os.path.exists(user_file):
                with open(user_file, 'r', encoding='utf-8') as f:
                    user_data = json.load(f)

                if not user_data.get('2fa_enabled', False):
                    return True  # 2FA未启用，直接通过

                secret = user_data.get('2fa_secret')
                if not secret:
                    return False

                totp = pyotp.TOTP(secret)
                return totp.verify(verification_code)
            return False
        except ImportError:
            logging.warning("2FA验证失败：pyotp库未安装")
            return True  # 如果库未安装，允许通过

    def register_user(self, auth_username, auth_password, group='user'):
        """注册新用户"""
        logging.info(f"register_user: 开始注册新用户: {auth_username}, 权限组: {group}")
        print(f"[用户注册] 开始注册新用户: {auth_username}, 权限组: {group}")
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            logging.debug(f"register_user: 检查用户文件是否存在: {user_file}")
            if os.path.exists(user_file):
                logging.warning(f"register_user: 用户名已存在: {auth_username}")
                print(f"[用户注册] 用户名已存在: {auth_username}")
                return {"success": False, "message": "用户名已存在"}

            # 根据配置选择密码存储方式
            logging.debug(f"register_user: 加密密码...")
            print(f"[用户注册] 加密密码...")
            stored_password = self._encrypt_password(auth_password)

            user_data = {
                'auth_username': auth_username,
                'password': stored_password,
                'group': group,
                'created_at': time.time(),
                'last_login': None,
                'session_ids': [],  # 用于关联用户的会话ID
                '2fa_enabled': False,
                '2fa_secret': None,
                'avatar_url': '',  # 用户头像URL
                'max_sessions': 1,  # 允许的最大会话数量 (1=单会话, >1=多会话, -1=无限制)
                'theme': 'light'  # 主题偏好：light/dark
            }

            logging.debug(f"register_user: 保存用户数据到文件: {user_file}")
            print(f"[用户注册] 保存用户数据到文件...")
            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            # 添加到权限组
            logging.debug(f"register_user: 添加用户到权限组: {group}")
            self.permissions['user_groups'][auth_username] = group
            self._save_permissions()

            logging.info(f"新用户注册: {auth_username} (组: {group})")
            print(f"[用户注册] ✓ 用户注册成功: {auth_username} (组: {group})")
            return {"success": True, "message": "注册成功"}

    def authenticate(self, auth_username, auth_password, ip_address='', user_agent='', two_fa_code=''):
        """验证用户登录（支持2FA和暴力破解防护）"""
        logging.info(
            f"authenticate: 开始认证用户: {auth_username}, IP: {ip_address}")
        print(f"[用户认证] 开始认证用户: {auth_username}, IP: {ip_address}")
        with self.lock:
            # 检查是否为游客登录
            if auth_username == 'guest' and self.config.getboolean('Guest', 'allow_guest_login', fallback=True):
                logging.info(f"authenticate: 游客登录成功: {auth_username}")
                print(f"[用户认证] 游客登录成功: {auth_username}")
                self._log_login_attempt(
                    auth_username, True, ip_address, user_agent, 'guest_login')
                return {
                    "success": True,
                    "auth_username": "guest",
                    "group": "guest",
                    "is_guest": True
                }

            # 检查暴力破解
            logging.debug(f"authenticate: 检查暴力破解防护: {auth_username}")
            print(f"[用户认证] 检查暴力破解: {auth_username}")
            is_locked, lock_message = self.check_brute_force(
                auth_username, ip_address)
            if is_locked:
                logging.warning(
                    f"authenticate: 用户被锁定（暴力破解防护）: {auth_username}")
                print(f"[用户认证] 用户被锁定（暴力破解防护）: {auth_username}")
                self._log_login_attempt(
                    auth_username, False, ip_address, user_agent, 'brute_force_locked')
                return {"success": False, "message": lock_message}

            user_file = self.get_user_file_path(auth_username)
            logging.debug(f"authenticate: 检查用户文件: {user_file}")
            if not os.path.exists(user_file):
                logging.warning(f"authenticate: 用户不存在: {auth_username}")
                print(f"[用户认证] 用户不存在: {auth_username}")
                self._log_login_attempt(
                    auth_username, False, ip_address, user_agent, 'user_not_found')
                return {"success": False, "message": "用户不存在"}

            logging.debug(f"authenticate: 读取用户数据: {auth_username}")
            print(f"[用户认证] 读取用户数据: {auth_username}")
            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            # 检查用户是否被封禁
            if user_data.get('banned', False):
                logging.warning(f"authenticate: 用户已被封禁: {auth_username}")
                print(f"[用户认证] 用户已被封禁: {auth_username}")
                self._log_login_attempt(
                    auth_username, False, ip_address, user_agent, 'user_banned')
                return {"success": False, "message": "账号已被封禁，请联系管理员"}

            # 验证密码
            logging.debug(f"authenticate: 验证密码: {auth_username}")
            print(f"[用户认证] 验证密码: {auth_username}")
            if not self._verify_password(auth_password, user_data.get('password')):
                logging.warning(f"authenticate: 密码错误: {auth_username}")
                print(f"[用户认证] 密码错误: {auth_username}")
                self._log_login_attempt(
                    auth_username, False, ip_address, user_agent, 'wrong_password')
                return {"success": False, "message": "密码错误"}

            # 验证2FA（如果启用）
            if user_data.get('2fa_enabled', False):
                logging.debug(f"authenticate: 检查2FA验证: {auth_username}")
                print(f"[用户认证] 检查2FA验证: {auth_username}")
                if not two_fa_code:
                    logging.info(f"authenticate: 需要2FA验证码: {auth_username}")
                    print(f"[用户认证] 需要2FA验证码: {auth_username}")
                    return {"success": False, "message": "需要2FA验证码", "requires_2fa": True}

                if not self.verify_2fa(auth_username, two_fa_code):
                    logging.warning(f"authenticate: 2FA验证失败: {auth_username}")
                    print(f"[用户认证] 2FA验证失败: {auth_username}")
                    self._log_login_attempt(
                        auth_username, False, ip_address, user_agent, '2fa_failed')
                    return {"success": False, "message": "2FA验证码错误"}

            # 更新最后登录时间
            logging.debug(f"authenticate: 更新最后登录时间: {auth_username}")
            print(f"[用户认证] 更新最后登录时间: {auth_username}")
            user_data['last_login'] = time.time()
            if 'session_ids' not in user_data:
                user_data['session_ids'] = []

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            # 检查是否为超级管理员
            super_admin = self.config.get(
                'Admin', 'super_admin', fallback='admin')
            if auth_username == super_admin:
                group = 'super_admin'
            else:
                group = user_data.get('group', 'user')

            self._log_login_attempt(
                auth_username, True, ip_address, user_agent, 'success')
            logging.info(
                f"用户登录: {auth_username} (组: {group}) from {ip_address}")
            print(f"[用户认证] ✓ 用户登录成功: {auth_username} (组: {group})")
            logging.info(f"authenticate: ✓ 认证成功，返回用户信息")
            return {
                "success": True,
                "auth_username": auth_username,
                "group": group,
                "is_guest": False,
                "max_sessions": user_data.get('max_sessions', 1),
                "avatar_url": user_data.get('avatar_url', ''),
                "theme": user_data.get('theme', 'light'),
                "session_ids": user_data.get('session_ids', [])
            }

    def check_permission(self, auth_username, permission):
        """检查用户是否有特定权限（支持差分化权限）

        权限计算顺序：
        1. 获取用户所属权限组的基础权限
        2. 应用用户的自定义权限（added/removed）
        """
        # 获取用户组
        group = self.permissions['user_groups'].get(auth_username, 'guest')

        # 检查是否为超级管理员
        super_admin = self.config.get('Admin', 'super_admin', fallback='admin')
        if auth_username == super_admin:
            group = 'super_admin'

        # 获取组权限（基础权限）
        group_perms = self.permissions['permission_groups'].get(
            group, {}).get('permissions', {})
        has_permission = group_perms.get(permission, False)

        # 应用用户的差分化权限
        user_custom = self.permissions.get(
            'user_custom_permissions', {}).get(auth_username, {})
        added_perms = user_custom.get('added', [])
        removed_perms = user_custom.get('removed', [])

        # 如果权限在added列表中，则有权限
        if permission in added_perms:
            has_permission = True

        # 如果权限在removed列表中，则无权限
        if permission in removed_perms:
            has_permission = False

        return has_permission

    def get_user_permissions(self, auth_username):
        """获取用户的完整权限列表（包含差分权限）"""
        # 获取用户组
        group = self.get_user_group(auth_username)

        # 获取组权限
        group_perms = self.permissions['permission_groups'].get(
            group, {}).get('permissions', {})

        # 复制基础权限
        user_perms = dict(group_perms)

        # 应用差分化权限
        user_custom = self.permissions.get(
            'user_custom_permissions', {}).get(auth_username, {})
        added_perms = user_custom.get('added', [])
        removed_perms = user_custom.get('removed', [])

        # 添加额外权限
        for perm in added_perms:
            user_perms[perm] = True

        # 移除权限
        for perm in removed_perms:
            user_perms[perm] = False

        return user_perms

    def set_user_custom_permission(self, auth_username, permission, grant):
        """为用户设置自定义权限（差分化存储）

        Args:
            auth_username: 用户名
            permission: 权限名
            grant: True=授予权限, False=移除权限
        """
        with self.lock:
            # 初始化user_custom_permissions结构
            if 'user_custom_permissions' not in self.permissions:
                self.permissions['user_custom_permissions'] = {}

            if auth_username not in self.permissions['user_custom_permissions']:
                self.permissions['user_custom_permissions'][auth_username] = {
                    'added': [],
                    'removed': []
                }

            user_custom = self.permissions['user_custom_permissions'][auth_username]

            # 获取用户组的基础权限
            group = self.get_user_group(auth_username)
            group_perms = self.permissions['permission_groups'].get(
                group, {}).get('permissions', {})
            base_has_permission = group_perms.get(permission, False)

            # 根据授予/移除状态更新差分列表
            if grant:
                # 授予权限
                if not base_has_permission:
                    # 基础权限没有，添加到added列表
                    if permission not in user_custom['added']:
                        user_custom['added'].append(permission)
                    # 从removed列表移除（如果存在）
                    if permission in user_custom['removed']:
                        user_custom['removed'].remove(permission)
                else:
                    # 基础权限已有，从removed列表移除（如果存在）
                    if permission in user_custom['removed']:
                        user_custom['removed'].remove(permission)
            else:
                # 移除权限
                if base_has_permission:
                    # 基础权限有，添加到removed列表
                    if permission not in user_custom['removed']:
                        user_custom['removed'].append(permission)
                    # 从added列表移除（如果存在）
                    if permission in user_custom['added']:
                        user_custom['added'].remove(permission)
                else:
                    # 基础权限没有，从added列表移除（如果存在）
                    if permission in user_custom['added']:
                        user_custom['added'].remove(permission)

            self._save_permissions()

            logging.info(f"设置用户 {auth_username} 权限 {permission} = {grant}")
            return {"success": True, "message": "权限已更新"}

    def get_user_group(self, auth_username):
        """获取用户所属组"""
        super_admin = self.config.get('Admin', 'super_admin', fallback='admin')
        if auth_username == super_admin:
            return 'super_admin'
        return self.permissions['user_groups'].get(auth_username, 'guest')

    def update_user_group(self, auth_username, new_group):
        """更新用户组（需要管理员权限）"""
        with self.lock:
            if new_group not in self.permissions['permission_groups']:
                return {"success": False, "message": "权限组不存在"}

            self.permissions['user_groups'][auth_username] = new_group
            self._save_permissions()

            # 同时更新用户文件
            user_file = self.get_user_file_path(auth_username)
            if os.path.exists(user_file):
                with open(user_file, 'r', encoding='utf-8') as f:
                    user_data = json.load(f)
                user_data['group'] = new_group
                with open(user_file, 'w', encoding='utf-8') as f:
                    json.dump(user_data, f, indent=2, ensure_ascii=False)

            return {"success": True, "message": "权限组已更新"}

    def create_permission_group(self, group_name, permissions, display_name):
        """创建新的权限组（需要超级管理员权限）"""
        with self.lock:
            if group_name in self.permissions['permission_groups']:
                return {"success": False, "message": "权限组已存在"}

            self.permissions['permission_groups'][group_name] = {
                'name': display_name,
                'is_system': False,  # 用户创建的权限组可以删除
                'permissions': permissions
            }
            self._save_permissions()
            return {"success": True, "message": "权限组已创建"}

    def update_permission_group(self, group_name, permissions):
        """更新权限组（需要超级管理员权限）

        注意：系统预设权限组可以修改，但不能删除
        """
        with self.lock:
            if group_name not in self.permissions['permission_groups']:
                return {"success": False, "message": "权限组不存在"}

            self.permissions['permission_groups'][group_name]['permissions'] = permissions
            self._save_permissions()
            return {"success": True, "message": "权限组已更新"}

    def delete_permission_group(self, group_name):
        """删除权限组（需要超级管理员权限）

        注意：不允许删除系统预设权限组（guest、user、admin、super_admin）
        """
        with self.lock:
            if group_name not in self.permissions['permission_groups']:
                return {"success": False, "message": "权限组不存在"}

            # 检查是否为系统预设权限组
            group_info = self.permissions['permission_groups'][group_name]
            if group_info.get('is_system', False):
                return {"success": False, "message": "不允许删除系统预设权限组"}

            # 检查是否有用户正在使用此权限组
            users_count = sum(
                1 for u, g in self.permissions['user_groups'].items() if g == group_name)
            if users_count > 0:
                return {
                    "success": False,
                    "message": f"无法删除：有 {users_count} 个用户正在使用此权限组"
                }

            # 删除权限组
            del self.permissions['permission_groups'][group_name]
            self._save_permissions()

            logging.info(f"权限组已删除: {group_name}")
            return {"success": True, "message": "权限组已删除"}

    def list_users(self):
        """列出所有用户"""
        users = []
        for filename in os.listdir(SYSTEM_ACCOUNTS_DIR):
            if filename.endswith('.json'):
                user_file = os.path.join(SYSTEM_ACCOUNTS_DIR, filename)
                try:
                    with open(user_file, 'r', encoding='utf-8') as f:
                        user_data = json.load(f)
                    users.append({
                        'auth_username': user_data['auth_username'],
                        'group': user_data.get('group', 'user'),
                        'created_at': user_data.get('created_at'),
                        'last_login': user_data.get('last_login'),
                        '2fa_enabled': user_data.get('2fa_enabled', False),
                        'banned': user_data.get('banned', False),
                        'max_sessions': user_data.get('max_sessions', 1)
                    })
                except Exception as e:
                    logging.error(f"[用户管理] 读取用户文件失败 --> 文件名: {filename}, 文件路径: {filepath}, 错误类型: {type(e).__name__}, 错误详情: {e}, 可能原因: 文件损坏、JSON格式错误或权限不足", exc_info=True)
        return users

    def get_all_groups(self):
        """获取所有权限组"""
        return self.permissions['permission_groups']

    def link_session_to_user(self, auth_username, session_id):
        """关联会话ID到用户账号（用于状态恢复）"""
        if auth_username == 'guest':
            return  # 游客不关联会话

        # 跳过无效的session_id
        if not session_id or session_id == 'null' or session_id.strip() == '':
            logging.debug(f"[会话管理] 跳过关联无效会话 --> 会话ID: '{session_id}', 目标用户: {auth_username}, 原因: 会话ID为空、null或仅包含空白字符")
            return

        user_file = self.get_user_file_path(auth_username)
        if os.path.exists(user_file):
            with self.lock:
                with open(user_file, 'r', encoding='utf-8') as f:
                    user_data = json.load(f)

                if 'session_ids' not in user_data:
                    user_data['session_ids'] = []

                # 添加新会话（如果不存在）
                if session_id not in user_data['session_ids']:
                    user_data['session_ids'].append(session_id)

                    # 根据max_sessions配置限制会话数量
                    max_sessions = user_data.get('max_sessions', 1)
                    if max_sessions > 0:  # 有限制
                        user_data['session_ids'] = user_data['session_ids'][-max_sessions:]
                    # max_sessions == -1 时不限制，保留所有会话

                with open(user_file, 'w', encoding='utf-8') as f:
                    json.dump(user_data, f, indent=2, ensure_ascii=False)

    def get_user_sessions(self, auth_username):
        """获取用户关联的会话ID列表"""
        if auth_username == 'guest':
            return []

        user_file = self.get_user_file_path(auth_username)
        if os.path.exists(user_file):
            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)
            return user_data.get('session_ids', [])
        return []

    def unlink_session_from_user(self, auth_username, session_id):
        """取消会话与用户的关联"""
        if auth_username == 'guest':
            return

        user_file = self.get_user_file_path(auth_username)
        if os.path.exists(user_file):
            with self.lock:
                with open(user_file, 'r', encoding='utf-8') as f:
                    user_data = json.load(f)

                if 'session_ids' in user_data and session_id in user_data['session_ids']:
                    user_data['session_ids'].remove(session_id)

                    with open(user_file, 'w', encoding='utf-8') as f:
                        json.dump(user_data, f, indent=2, ensure_ascii=False)

    def reset_user_password(self, auth_username, new_password):
        """重置用户密码（管理员功能）"""
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            # 加密新密码
            user_data['password'] = self._encrypt_password(new_password)

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            logging.info(f"管理员重置密码: {auth_username}")
            return {"success": True, "message": "密码已重置"}

    def update_user_avatar(self, auth_username, avatar_url):
        """更新用户头像"""
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            user_data['avatar_url'] = avatar_url

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            return {"success": True, "message": "头像已更新"}

    def update_user_theme(self, auth_username, theme):
        """更新用户主题偏好"""
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            user_data['theme'] = theme

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            return {"success": True, "message": "主题已更新"}

    def update_max_sessions(self, auth_username, max_sessions):
        """更新用户最大会话数量

        Args:
            auth_username: 用户名
            max_sessions: 最大会话数 (1=单会话, >1=指定数量, -1=无限制)
        """
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            user_data['max_sessions'] = max_sessions

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            # 生成提示信息
            if max_sessions == 1:
                msg = f"已设置为单会话模式：用户每次只能保持1个活跃会话"
            elif max_sessions == -1:
                msg = f"已设置为无限会话模式：用户可以创建任意数量的会话"
            else:
                msg = f"已设置最大会话数量为：{max_sessions}个，超出时将自动清理最旧的会话"

            return {"success": True, "message": msg}

    def ban_user(self, auth_username):
        """封禁用户"""
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            user_data['banned'] = True
            user_data['banned_at'] = time.time()

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            logging.info(f"用户已封禁: {auth_username}")
            return {"success": True, "message": "用户已封禁"}

    def unban_user(self, auth_username):
        """解封用户"""
        with self.lock:
            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            user_data['banned'] = False
            user_data['unbanned_at'] = time.time()

            with open(user_file, 'w', encoding='utf-8') as f:
                json.dump(user_data, f, indent=2, ensure_ascii=False)

            logging.info(f"用户已解封: {auth_username}")
            return {"success": True, "message": "用户已解封"}

    def delete_user(self, auth_username):
        """删除用户（需要管理员权限）"""
        with self.lock:
            # 不允许删除超级管理员
            super_admin = self.config.get(
                'Admin', 'super_admin', fallback='admin')
            if auth_username == super_admin:
                return {"success": False, "message": "不允许删除超级管理员"}

            user_file = self.get_user_file_path(auth_username)
            if not os.path.exists(user_file):
                return {"success": False, "message": "用户不存在"}

            # 删除用户文件
            try:
                os.remove(user_file)
            except Exception as e:
                logging.error(f"删除用户文件失败: {e}")
                return {"success": False, "message": f"删除失败: {e}"}

            # 从权限组映射中移除
            if auth_username in self.permissions.get('user_groups', {}):
                del self.permissions['user_groups'][auth_username]

            # 从差分权限中移除
            if 'user_custom_permissions' in self.permissions and auth_username in self.permissions['user_custom_permissions']:
                del self.permissions['user_custom_permissions'][auth_username]

            self._save_permissions()

            logging.info(f"用户已删除: {auth_username}")
            return {"success": True, "message": "用户已删除"}

    def get_user_details(self, auth_username):
        """获取用户详细信息"""
        user_file = self.get_user_file_path(auth_username)
        if not os.path.exists(user_file):
            return None

        with open(user_file, 'r', encoding='utf-8') as f:
            user_data = json.load(f)

        # 返回用户信息（不包含密码）
        return {
            'auth_username': user_data['auth_username'],
            'group': user_data.get('group', 'user'),
            'created_at': user_data.get('created_at'),
            'last_login': user_data.get('last_login'),
            '2fa_enabled': user_data.get('2fa_enabled', False),
            'avatar_url': user_data.get('avatar_url', ''),
            'max_sessions': user_data.get('max_sessions', 1),
            'theme': user_data.get('theme', 'light'),
            'session_ids': user_data.get('session_ids', [])
        }

    def check_single_session_enforcement(self, auth_username, new_session_id):
        """检查并强制执行会话数量限制

        根据用户的max_sessions设置：
        - max_sessions = 1: 单会话模式，移除所有旧会话
        - max_sessions > 1: 多会话模式，超出时移除最旧的会话
        - max_sessions = -1: 无限制模式，不移除任何会话

        Returns:
            tuple: (需要清理的旧会话列表, 提示信息)
        """
        if auth_username == 'guest':
            return [], ""  # 游客不受限制

        user_file = self.get_user_file_path(auth_username)
        if not os.path.exists(user_file):
            return [], ""

        with self.lock:
            with open(user_file, 'r', encoding='utf-8') as f:
                user_data = json.load(f)

            max_sessions = user_data.get('max_sessions', 1)
            old_sessions = user_data.get('session_ids', [])

            # 无限制模式，不清理
            if max_sessions == -1:
                return [], ""

            # # 单会话模式：移除所有旧会话
            # if max_sessions == 1:
            #     user_data['session_ids'] = [new_session_id]

            #     with open(user_file, 'w', encoding='utf-8') as f:
            #         json.dump(user_data, f, indent=2, ensure_ascii=False)

            #     # 过滤掉无效的session_id
            #     valid_old_sessions = [s for s in old_sessions if s and s != 'null' and s.strip() != '']
            #     message = "单会话模式：已自动清理所有旧会话"
            #     return valid_old_sessions, message

            # 多会话模式：检查是否超出限制
            current_count = len(old_sessions)
            if current_count > max_sessions:
                # 超出限制，移除最旧的会话
                sessions_to_remove = old_sessions[:current_count -
                                                  max_sessions + 1]
                remaining_sessions = old_sessions[current_count -
                                                  max_sessions + 1:]
                user_data['session_ids'] = remaining_sessions + \
                    [new_session_id]

                with open(user_file, 'w', encoding='utf-8') as f:
                    json.dump(user_data, f, indent=2, ensure_ascii=False)

                # 过滤掉无效的session_id
                valid_sessions_to_remove = [
                    s for s in sessions_to_remove if s and s != 'null' and s.strip() != '']
                message = ""
                if len(valid_sessions_to_remove) > 0:
                    message = f"已达到最大会话数量限制({max_sessions}个)，已自动清理{len(valid_sessions_to_remove)}个最旧的会话"
                return valid_sessions_to_remove, message
            else:
                # 未超出限制，正常添加
                return [], ""

    def log_audit(self, auth_username, action, details='', ip_address='', session_id=''):
        """记录审计日志"""
        audit_entry = {
            'timestamp': time.time(),
            'datetime': datetime.datetime.now().isoformat(),
            'username': auth_username,
            'action': action,
            'details': details,
            'ip_address': ip_address,
            'session_id': session_id
        }

        try:
            with open(AUDIT_LOG_FILE, 'a', encoding='utf-8') as f:
                f.write(json.dumps(audit_entry, ensure_ascii=False) + '\n')
        except Exception as e:
            logging.error(f"记录审计日志失败: {e}")

    def get_audit_logs(self, username=None, action=None, limit=100):
        """获取审计日志"""
        if not os.path.exists(AUDIT_LOG_FILE):
            return []

        logs = []
        try:
            with open(AUDIT_LOG_FILE, 'r', encoding='utf-8') as f:
                for line in f:
                    try:
                        entry = json.loads(line.strip())
                        if (username is None or entry.get('username') == username) and \
                           (action is None or entry.get('action') == action):
                            logs.append(entry)
                    except (json.JSONDecodeError, KeyError, ValueError) as e:
                        logging.debug(f"[审计日志] 跳过无效日志行: {e}")
                        continue
        except Exception as e:
            logging.error(f"读取审计日志失败: {e}")

        # 返回最近的记录
        return logs[-limit:]


# 创建全局认证系统实例
auth_system = AuthSystem()


# ==============================================================================
# Token管理系统 (2048位令牌)
# ==============================================================================

class TokenManager:
    """管理用户登录令牌的系统

    功能:
    - 生成2048位(256字节)安全令牌
    - 持久化存储令牌到独立文件
    - 验证令牌有效性(1小时过期)
    - 检测多设备登录
    - 自动刷新活跃用户的令牌
    """

    def __init__(self, tokens_dir):
        logging.info("="*80)
        logging.info(f"TokenManager: 初始化令牌管理器，目录: {tokens_dir}")
        self.tokens_dir = tokens_dir
        self.lock = threading.Lock()
        if not os.path.exists(tokens_dir):
            os.makedirs(tokens_dir)
            logging.info(f"TokenManager: 创建令牌目录: {tokens_dir}")
        else:
            logging.debug(f"TokenManager: 令牌目录已存在: {tokens_dir}")
        logging.info("TokenManager: 初始化完成")
        logging.info("="*80)

    def _get_token_file_path(self, username):
        """获取用户的token文件路径"""
        # 使用用户名的哈希作为文件名
        username_hash = hashlib.sha256(username.encode()).hexdigest()
        file_path = os.path.join(
            self.tokens_dir, f"{username_hash}_tokens.json")
        logging.debug(
            f"_get_token_file_path: 用户 {username} 的令牌文件: {file_path}")
        return file_path

    def generate_token(self):
        """生成2048位(256字节)的安全令牌"""
        # secrets.token_hex(256) 生成 256字节 = 2048位 的随机令牌
        token = secrets.token_hex(256)
        logging.debug(f"generate_token: 生成新令牌，长度: {len(token)} 字符")
        return token

    def create_token(self, username, session_id):
        """为用户创建新令牌并存储

        Args:
            username: 用户名
            session_id: 会话UUID

        Returns:
            token: 生成的令牌字符串
        """
        session_preview = session_id[:16] if session_id and len(
            session_id) >= 16 else session_id
        logging.info(
            f"create_token: 为用户 {username} 创建令牌，会话: {session_preview}...")
        token = self.generate_token()
        created_at = time.time()
        expires_at = created_at + 3600  # 1小时后过期

        token_data = {
            'token': token,
            'session_id': session_id,
            'created_at': created_at,
            'expires_at': expires_at,
            'last_activity': created_at
        }

        with self.lock:
            token_file = self._get_token_file_path(username)

            # 读取现有tokens
            all_tokens = {}
            if os.path.exists(token_file):
                try:
                    with open(token_file, 'r', encoding='utf-8') as f:
                        all_tokens = json.load(f)
                except (FileNotFoundError, json.JSONDecodeError) as e:
                    logging.debug(f"[Token存储] 无法读取token文件，初始化为空: {e}")
                    all_tokens = {}

            # 添加新token
            all_tokens[session_id] = token_data

            # 保存到文件
            with open(token_file, 'w', encoding='utf-8') as f:
                json.dump(all_tokens, f, indent=2, ensure_ascii=False)

        logging.info(f"为用户 {username} 创建新令牌，会话: {session_preview}...")
        return token

    def verify_token(self, username, session_id, token):
        """验证令牌是否有效

        Args:
            username: 用户名
            session_id: 会话UUID
            token: 要验证的令牌

        Returns:
            tuple: (is_valid, reason)
        """
        token_file = self._get_token_file_path(username)

        if not os.path.exists(token_file):
            return False, "no_token_file"

        try:
            with open(token_file, 'r', encoding='utf-8') as f:
                all_tokens = json.load(f)

            if session_id not in all_tokens:
                return False, "session_not_found"

            token_data = all_tokens[session_id]

            # 检查令牌是否匹配
            if token_data['token'] != token:
                return False, "token_mismatch"

            # 检查是否过期
            if time.time() > token_data['expires_at']:
                return False, "token_expired"

            return True, "valid"

        except Exception as e:
            logging.error(f"验证令牌时出错: {e}")
            return False, "error"

    def get_valid_token_for_session(self, username, session_id):
        """
        获取指定用户和会话ID的有效Token（如果存在且未过期）。
        如果找到有效Token，会刷新其活动时间。

        Args:
            username: 用户名
            session_id: 会话UUID

        Returns:
            str | None: 如果找到有效Token则返回Token字符串，否则返回None。
        """
        token_file = self._get_token_file_path(username)

        if not os.path.exists(token_file):
            return None

        try:
            # 使用锁确保读取和可能的刷新操作是原子的
            with self.lock:
                with open(token_file, 'r', encoding='utf-8') as f:
                    all_tokens = json.load(f)

                if session_id not in all_tokens:
                    logging.debug(
                        f"get_valid_token_for_session: 会话 {session_id[:8]} 未找到Token记录")
                    return None

                token_data = all_tokens[session_id]
                current_time = time.time()

                # 检查是否过期
                if current_time > token_data['expires_at']:
                    logging.debug(
                        f"get_valid_token_for_session: 会话 {session_id[:8]} 的Token已过期")
                    # (可选) 在这里可以顺便清理掉这个过期的记录，如果需要的话
                    # del all_tokens[session_id]
                    # with open(token_file, 'w', encoding='utf-8') as wf:
                    #     json.dump(all_tokens, wf, indent=2, ensure_ascii=False)
                    return None

                # Token有效，刷新活动时间并延长有效期
                token_data['last_activity'] = current_time
                token_data['expires_at'] = current_time + 3600  # 延长1小时

                # 写回更新后的Token数据
                with open(token_file, 'w', encoding='utf-8') as f:
                    json.dump(all_tokens, f, indent=2, ensure_ascii=False)

                logging.info(
                    f"get_valid_token_for_session: 找到并刷新了会话 {session_id[:8]} 的有效Token")
                return token_data['token']  # 返回有效的Token字符串

        except Exception as e:
            logging.error(f"获取会话Token时出错 ({username}, {session_id[:8]}): {e}")
            return None

    def refresh_token(self, username, session_id):
        """刷新令牌的过期时间和最后活动时间

        Args:
            username: 用户名
            session_id: 会话UUID

        Returns:
            bool: 是否刷新成功
        """
        with self.lock:
            token_file = self._get_token_file_path(username)

            if not os.path.exists(token_file):
                return False

            try:
                with open(token_file, 'r', encoding='utf-8') as f:
                    all_tokens = json.load(f)

                if session_id not in all_tokens:
                    return False

                # 更新过期时间和最后活动时间
                current_time = time.time()
                # 延长1小时
                all_tokens[session_id]['expires_at'] = current_time + 3600
                all_tokens[session_id]['last_activity'] = current_time

                # 保存
                with open(token_file, 'w', encoding='utf-8') as f:
                    json.dump(all_tokens, f, indent=2, ensure_ascii=False)

                return True

            except Exception as e:
                logging.error(f"刷新令牌时出错: {e}")
                return False

    def invalidate_token(self, username, session_id):
        """使令牌失效（用于登出）

        Args:
            username: 用户名
            session_id: 会话UUID
        """
        with self.lock:
            token_file = self._get_token_file_path(username)

            if not os.path.exists(token_file):
                return

            try:
                with open(token_file, 'r', encoding='utf-8') as f:
                    all_tokens = json.load(f)

                if session_id in all_tokens:
                    del all_tokens[session_id]

                    # 保存
                    with open(token_file, 'w', encoding='utf-8') as f:
                        json.dump(all_tokens, f, indent=2, ensure_ascii=False)

                    logging.info(
                        f"令牌已失效: {username}, 会话: {session_id[:16]}...")
            except Exception as e:
                logging.error(f"使令牌失效时出错: {e}")

    def get_active_sessions(self, username):
        """获取用户所有有效的会话

        Args:
            username: 用户名

        Returns:
            list: 有效会话ID列表
        """
        token_file = self._get_token_file_path(username)

        if not os.path.exists(token_file):
            return []

        try:
            with open(token_file, 'r', encoding='utf-8') as f:
                all_tokens = json.load(f)

            current_time = time.time()
            active_sessions = []

            for session_id, token_data in all_tokens.items():
                if current_time <= token_data['expires_at']:
                    active_sessions.append(session_id)

            return active_sessions

        except Exception as e:
            logging.error(f"获取活跃会话时出错: {e}")
            return []

    def cleanup_expired_tokens(self, username):
        """清理过期的令牌

        Args:
            username: 用户名
        """
        with self.lock:
            token_file = self._get_token_file_path(username)

            if not os.path.exists(token_file):
                return

            try:
                with open(token_file, 'r', encoding='utf-8') as f:
                    all_tokens = json.load(f)

                current_time = time.time()
                valid_tokens = {}

                for session_id, token_data in all_tokens.items():
                    if current_time <= token_data['expires_at']:
                        valid_tokens[session_id] = token_data

                # 保存清理后的tokens
                with open(token_file, 'w', encoding='utf-8') as f:
                    json.dump(valid_tokens, f, indent=2, ensure_ascii=False)

                removed_count = len(all_tokens) - len(valid_tokens)
                if removed_count > 0:
                    logging.info(f"清理了 {removed_count} 个过期令牌: {username}")

            except Exception as e:
                logging.error(f"清理过期令牌时出错: {e}")

    def detect_multi_device_login(self, username, new_session_id):
        """检测多设备登录

        Args:
            username: 用户名
            new_session_id: 新的会话ID

        Returns:
            list: 需要踢出的旧会话ID列表
        """
        active_sessions = self.get_active_sessions(username)

        # 排除新会话本身
        old_sessions = [s for s in active_sessions if s != new_session_id]

        if old_sessions:
            logging.info(f"检测到多设备登录: {username}, 旧会话数: {len(old_sessions)}")

        return old_sessions


# 创建全局Token管理器实例
token_manager = TokenManager(TOKENS_STORAGE_DIR)


# ==============================================================================
# 数据结构定义
# ==============================================================================

class UserData:
    """存储用户相关信息的类"""

    def __init__(self):
        self.name: str = ""
        self.phone: str = ""
        self.student_id: str = ""
        self.id: str = ""
        self.registration_time: str = ""
        self.first_login_time: str = ""
        self.id_card: str = ""
        self.last_login_time: str = ""
        self.current_login_time: str = ""
        self.username: str = ""
        self.gender: str = ""
        self.school_name: str = ""
        self.attribute_type: str = ""
        self.avatar_url: str = ""


class RunData:
    """存储单个跑步任务相关数据的类"""

    def __init__(self):
        # 路径数据
        # (经度, 纬度, 是否关键点) - 用户绘制的草稿路径
        self.draft_coords: list[tuple[float, float, int]] = []
        # (经度, 纬度, 与上一点的时间间隔ms) - 处理后用于模拟的最终路径
        self.run_coords: list[tuple[float, float, int]] = []
        # (经度, 纬度) - 服务器返回的推荐路线, (0,0)用于分段
        self.recommended_coords: list[tuple[float, float]] = []
        self.target_points: list[tuple[float, float]
                                 ] = []          # (经度, 纬度) - 任务必须经过的打卡点

        # 任务基本信息
        self.target_point_names: str = ""   # 打卡点名称，用'|'分隔
        self.upload_time: str = ""          # 完成时间
        self.start_time: str = ""           # 任务开始时间
        self.end_time: str = ""             # 任务结束时间
        self.run_name: str = ""             # 任务名称
        self.errand_id: str = ""            # 任务ID
        self.errand_schedule: str = ""      # 任务计划ID
        self.status: int = 0                # 任务状态 (0: 未完成, 1: 已完成)

        # 运行时状态
        self.target_sequence: int = 0       # 当前目标打卡点序号（0-based，0表示第一个打卡点）
        self.is_in_target_zone: bool = False  # 是否在当前打卡点范围内
        self.trid: str = ""                 # 本次跑步的唯一轨迹ID
        self.details_fetched: bool = False  # 任务详情是否已加载
        self.total_run_time_s: float = 0.0  # 模拟总时长(秒)
        self.total_run_distance_m: float = 0.0  # 模拟总距离(米)
        self.distance_covered_m: float = 0.0  # 实时已跑距离(米)


class AccountSession:
    """封装单个账号的所有运行时数据、状态和操作"""

    def __init__(self, username, password, api_bridge):
        self.username: str = username
        self.password: str = password
        self.api_bridge = api_bridge  # Api 类的实例引用
        self.window = api_bridge.window

        # 每个账号拥有独立的ApiClient和requests.Session，确保Cookie隔离
        self.api_client = ApiClient(self)
        self.user_data = UserData()
        self.all_run_data: list[RunData] = []

        # 从全局参数深拷贝一份独立的参数配置
        self.params = copy.deepcopy(api_bridge.global_params)
        self.device_ua: str = ""  # 将在登录时生成或加载

        # --- 账号独立的签到半径缓存 ---
        self.server_attendance_radius_m = 0.0
        self.last_radius_fetch_time = 0

        self.status_text: str = "待命"  # UI上显示的状态
        self.summary = {"total": 0, "completed": 0, "pending": 0,
                        "executable": 0, "expired": 0, "not_started": 0}

        self.worker_thread: threading.Thread | None = None
        self.stop_event = threading.Event()

    def log(self, message: str):
        """为日志自动添加账号前缀"""
        # 使用主 Api 实例的 log 方法来确保日志发送到UI
        self.api_bridge.log(f"[{self.username}] {message}")


class ApiClient:
    """处理与后端服务器网络请求的类"""
    BASE_URL = "https://zslf.zsc.edu.cn"
    API_VERSION = 66

    def __init__(self, owner_instance):
        self.session = requests.Session()
        self.app = owner_instance
        logging.debug("ApiClient已初始化，创建了新的requests.Session会话实例")

    def _get_headers(self) -> dict:
        """构建请求头，包含认证信息和设备信息"""
        headers = {
            'Connection': 'keep-alive',
            'Accept': 'application/json, text/plain, */*',
            # 'Content-Type': 'application/x-www-form-urlencoded',
            'X-Requested-With': 'com.zx.slm',
            'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
            'User-Agent': self.app.device_ua
        }
        # 仅当存在登录后返回的 shiroCookie 时才携带 Authorization
        auth_token = self.session.cookies.get('shiroCookie')
        if auth_token:
            logging.debug(f"使用shiroCookie作为认证令牌进行Authorization请求头设置: {auth_token}")
            headers['Authorization'] = auth_token
        else:
            logging.debug(
                "未找到shiroCookie认证令牌，将不会设置Authorization请求头")
        return headers

    def _request(self, method: str, url: str, data: dict = None, params: dict = None, is_post_str=False, force_content_type: str = None) -> requests.Response | None:
        """统一的网络请求方法（增强：支持取消）"""

        # 在多账号模式下，app 可能是 AccountSession 实例，需要获取正确的日志函数和离线模式状态
        log_func = self.app.log if hasattr(
            self.app, 'log') else self.app.api_bridge.log
        is_offline = self.app.is_offline_mode if hasattr(
            self.app, 'is_offline_mode') else self.app.api_bridge.is_offline_mode

        # 全局/局部取消检查
        cancel_requested = False
        # try:
        #     # 单账号线程停止
        #     if hasattr(self.app, 'stop_event') and isinstance(self.app.stop_event, threading.Event):
        #         cancel_requested = cancel_requested or self.app.stop_event.is_set()
        #     # 多账号全局停止
        #     if hasattr(self.app, 'api_bridge') and hasattr(self.app.api_bridge, 'multi_run_stop_flag'):
        #         cancel_requested = cancel_requested or self.app.api_bridge.multi_run_stop_flag.is_set()
        #     # 单账号模式下的 stop_run_flag
        #     if hasattr(self.app, 'api_bridge') and hasattr(self.app.api_bridge, 'stop_run_flag'):
        #         cancel_requested = cancel_requested or self.app.api_bridge.stop_run_flag.is_set()
        # except Exception:
        #     pass

        if cancel_requested:
            log_func("操作已取消，跳过网络请求。")
            logging.debug(f"[网络请求] 请求已取消 --> 请求方法: {method.upper()}, 目标URL: {url}, 取消原因: 用户停止操作或系统取消标志已设置")
            return None

        if is_offline:
            log_func("离线模式：网络请求已被禁用。")
            logging.debug(f"[网络请求] 离线模式已启用，已阻止网络请求 --> 请求方法: {method.upper()}, 目标URL: {url}, 说明: 离线模式下所有网络通信将被禁用")
            return None

        retries = 3
        connect_timeout = 5  # 连接超时设置为 5 秒
        read_timeout = 10    # 读取超时设置为 10 秒

        log_data = data
        if is_post_str and isinstance(data, str) and len(data) > 500:
            log_data = data[:500] + '... (已截断，完整数据长度: ' + str(len(data)) + ' 字节)'

        logging.debug(f"[网络请求] 准备发起HTTP请求 --> 请求方法: {method.upper()}, 目标URL: {url}, 重试次数配置: {retries}次, 连接超时: {connect_timeout}秒, 读取超时: {read_timeout}秒\n[请求数据]: {log_data}")

        for attempt in range(retries):
            try:
                # 再次检查取消（避免刚发起时被取消）
                # try:
                #     if hasattr(self.app, 'stop_event') and self.app.stop_event.is_set():
                #         log_func("操作已取消。")
                #         return None
                #     if hasattr(self.app, 'api_bridge') and self.app.api_bridge.multi_run_stop_flag.is_set():
                #         log_func("操作已取消。")
                #         return None
                #     if hasattr(self.app, 'api_bridge') and self.app.api_bridge.stop_run_flag.is_set():
                #         log_func("操作已取消。")
                #         return None
                # except Exception:
                #     pass

                # 获取基础 headers
                headers = self._get_headers()

                if method.upper() == 'POST':
                    post_data_bytes = b""

                    # 1. 确定 Content-Type
                    if force_content_type:
                        headers['Content-Type'] = force_content_type
                    elif is_post_str:
                        # 默认为 form-urlencoded (用于 submit_run_track)
                        # 如果data是JSON字符串，调用者应使用 force_content_type
                        headers['Content-Type'] = 'application/x-www-form-urlencoded'
                    else:
                        # 默认: form-urlencoded (用于 login)
                        headers['Content-Type'] = 'application/x-www-form-urlencoded'

                    # 2. 准备 Body
                    if is_post_str:
                        post_data_bytes = (data or "").encode('utf-8')
                    else:
                        # 默认: form-urlencoded
                        post_data_bytes = urllib.parse.urlencode(
                            data or {}).encode('utf-8')
                        
                    # 修复 Bug A: 移除了在 post 之前的错误日志记录
                    resp = self.session.post(
                        url, data=post_data_bytes, params=params, headers=headers,
                        timeout=(connect_timeout, read_timeout) # 使用元组分别设置连接和读取超时
                    )
                else:
                    # GET 请求 (data 被用作 params)
                    resp = self.session.get(
                            url, params=data, headers=headers,
                            timeout=(connect_timeout, read_timeout) # 使用元组分别设置连接和读取超时
                        )
                
                # --- 成功响应处理 (Bug A 和 B 均在此修复) ---
                # 修复 Bug B: 这段代码现在位于 try 块内部，但在所有 except 块之前
                logging.debug(
                    f"[网络请求] 收到服务器响应 <-- 状态码: {resp.status_code} ({resp.reason}), 来源URL: {url}, 响应头: {dict(resp.headers)}, 响应内容长度: {len(resp.content)} 字节")
                resp.raise_for_status() # 如果状态码不是 2xx，会抛出 HTTPError
                return resp

            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as net_err:
                # --- 新增：专门捕获连接错误和超时 ---
                log_func(f"网络连接失败 (第{attempt+1}/{retries}次): {net_err}")
                logging.error(
                    f"[网络请求] 网络连接失败 --> 重试次数: 第{attempt+1}次/共{retries}次, 请求方法: {method.upper()}, 目标URL: {url}, 错误类型: {type(net_err).__name__}, 错误详情: {net_err}, 连接超时配置: {connect_timeout}秒, 读取超时配置: {read_timeout}秒", exc_info=False)
                if attempt + 1 == retries:
                    log_func(f"网络连接最终失败: 无法连接到服务器 {self.BASE_URL}")
                    logging.error(f"[网络请求] 网络连接最终失败 --> 已达到最大重试次数({retries}次), 目标服务器: {self.BASE_URL}, 无法建立连接")
                    return None # 连接失败，直接返回 None
                logging.info(f"[网络请求] 准备重试 --> 等待1.5秒后进行第{attempt+2}次请求尝试")
                time.sleep(1.5) # 重试前等待
                continue # 继续下一次重试

                # 修复 Bug B: 原本错误放置在这里的代码已被移到 try 块末尾

            except requests.exceptions.HTTPError as http_err:
                # --- 处理非 2xx 状态码的错误 (例如 404, 500) ---
                log_func(f"服务器返回错误 (第{attempt+1}次): {http_err.response.status_code}")
                logging.error(
                    f"[网络请求] HTTP错误 --> 重试次数: 第{attempt+1}次/共{retries}次, 请求方法: {method.upper()}, 目标URL: {url}, HTTP状态码: {http_err.response.status_code}, 状态描述: {http_err.response.reason}, 服务器响应内容: {http_err.response.text[:200]}{'...(已截断)' if len(http_err.response.text) > 200 else ''}", exc_info=False)
                if attempt + 1 == retries:
                    log_func(f"服务器错误: {http_err.response.status_code}")
                    logging.error(f"[网络请求] HTTP请求最终失败 --> 已达到最大重试次数({retries}次), HTTP状态码: {http_err.response.status_code}, 请求无法成功完成")
                    # 可以考虑返回包含错误信息的 response 对象，或者依然返回 None
                    # return http_err.response # 如果上层需要处理具体错误
                    return None # 保持返回 None
                logging.info(f"[网络请求] 准备重试 --> 等待1.5秒后进行第{attempt+2}次请求尝试")
                time.sleep(1.5) # 重试前等待
                continue # 继续下一次重试

            except requests.exceptions.RequestException as req_err:
                # --- 捕获其他所有 requests 相关的异常 (作为兜底) ---
                log_func(f"请求发生意外错误 (第{attempt+1}次): {req_err}")
                logging.error(
                    f"[网络请求] 意外的请求异常 --> 重试次数: 第{attempt+1}次/共{retries}次, 请求方法: {method.upper()}, 目标URL: {url}, 异常类型: {type(req_err).__name__}, 异常详情: {req_err}, 完整堆栈信息如下:", exc_info=True)
                if attempt + 1 == retries:
                    log_func(f"请求最终失败: {req_err}")
                    logging.error(f"[网络请求] 请求最终失败 --> 已达到最大重试次数({retries}次), 所有重试均失败, 异常信息: {req_err}")
                    return None # 最终失败返回 None
                logging.info(f"[网络请求] 准备重试 --> 等待1.5秒后进行第{attempt+2}次请求尝试")
                time.sleep(1.5) # 重试前等待
                continue # 继续下一次重试

        return None # 如果循环结束仍未成功


    def _json(self, resp: requests.Response | None) -> dict | None:
        """安全地将Response对象解析为JSON字典"""
        log_func = self.app.log if hasattr(
            self.app, 'log') else self.app.api_bridge.log
        if resp:
            try:
                json_data = resp.json()
                logging.debug(f"[JSON解析] 成功解析JSON响应 --> 响应状态码: {resp.status_code}, JSON数据字段: {list(json_data.keys()) if isinstance(json_data, dict) else type(json_data).__name__}")
                return json_data
            except json.JSONDecodeError as e:
                log_func("服务器响应解析失败。")
                logging.error(
                    f"[JSON解析] JSON解码失败 --> 响应状态码: {resp.status_code}, 响应内容类型: {resp.headers.get('Content-Type', '未知')}, 解码错误位置: 第{e.lineno}行第{e.colno}列, 响应文本内容(前500字符): {resp.text[:500]}{'...(已截断)' if len(resp.text) > 500 else ''}, 错误详情: {e}")
                return None
        logging.debug(f"[JSON解析] 响应对象为空，无法解析JSON")
        return None

    def login(self, username, password):
        return self._json(self._request('POST', f"{self.BASE_URL}/app/login", {"username": username, "password": password, "appVersion": self.API_VERSION}))

    def get_run_list(self, user_id, offset=0):
        return self._json(self._request('GET', f"{self.BASE_URL}:9097/run/errand/getErrandList", {"userId": user_id, "offset": offset, "limit": 10, "appVersion": self.API_VERSION}))

    def get_run_details(self, errand_id, user_id, errand_schedule_id):
        return self._json(self._request('GET', f"{self.BASE_URL}:9097/run/errand/getErrandDetail", {"errandId": errand_id, "userId": user_id, "errandScheduleId": errand_schedule_id, "appVersion": self.API_VERSION}))

    def get_run_history_list(self, user_id, errand_schedule_id):
        return self._json(self._request('GET', f"{self.BASE_URL}:9097/run/errand/getUserErrandTrackRecord", {"errandScheduleId": errand_schedule_id, "userId": user_id, "offset": 0, "limit": 20, "appVersion": self.API_VERSION}))

    def get_history_track_by_trid(self, trid):
        return self._json(self._request('GET', f"{self.BASE_URL}:9097/run/errand/getTrackByTrid", {"trid": trid, "appVersion": self.API_VERSION}))

    def submit_run_track(self, payload_str):
        return self._json(self._request('POST', f"{self.BASE_URL}:9097/run/errand/addErrandTrack", payload_str, is_post_str=True))

    def get_run_info_by_trid(self, trid):
        return self._json(self._request('GET', f"{self.BASE_URL}:9097/run/errand/getTrackRecordByTrid", {"trid": trid, "appVersion": self.API_VERSION}))

    def get_unread_notice_count(self):
        """获取未读通知数量 (POST, 空body, application/json)"""
        # 该API需要POST一个空 body 并且 Content-Type 为 application/json
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}/app/appNotice/unreadNumber",
            data="",  # 发送空字符串
            is_post_str=True,
            force_content_type='application/json;charset=UTF-8'  # 强制指定类型
        ))

    def get_notice_list(self, offset=0, limit=10, type_id=0):
        """获取通知列表 (POST, 空body, 带URL参数)"""
        params = {"offset": offset, "limit": limit, "typeId": type_id}
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}/app/appNotice/noticeListByType",
            data="",  # 发送空字符串
            params=params,
            is_post_str=True,
            force_content_type='application/json;charset=UTF-8'  # 强制指定类型
        ))

    def mark_notice_as_read(self, notice_id):
        """将单个通知设为已读 (POST, 空body, 带URL参数)"""
        params = {"noticeId": notice_id}
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}/app/appNotice/updateNoticeIsRead",
            data="",  # 发送空字符串
            params=params,
            is_post_str=True,
            force_content_type='application/json;charset=UTF-8'  # 强制指定类型
        ))

    @staticmethod
    def generate_random_ua():
        """生成一个随机的、模拟安卓设备的User-Agent字符串"""
        build_texts = ["TD1A.221105.001.A1", "TP1A.221005.003", "SQ3A.220705.004",
                       "SP2A.220505.008", "SQ1D.220205.004", "RP1A.201005.004"]
        phone_models = ["Xiaomi 12", "Xiaomi 13 Pro", "Redmi K60", "vivo X90", "iQOO 11",
                        "OPPO Find X6 Pro", "Realme GT Neo5", "HONOR Magic5 Pro", "OnePlus 11"]
        android_version_map = {'T': 13, 'S': 12, 'R': 11, 'Q': 10, 'P': 9}
        random_build = random.choice(build_texts)
        build_letter = random_build.split('.')[0][0]
        android_version = android_version_map.get(build_letter, 13)
        chrome_version = f"Chrome/{random.randint(100, 120)}.0.{random.randint(4000, 6000)}.{random.randint(100, 200)}"
        return f"Mozilla/5.0 (Linux; Android {android_version}; {random.choice(phone_models)} Build/{random_build}; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 {chrome_version} Mobile Safari/537.36"

    def get_roll_call_info(self, roll_call_id, user_id):
        """获取指定签到活动的信息"""
        params = {"id": roll_call_id, "userId": user_id,
                  "appVersion": self.API_VERSION}
        # 注意：原始请求是 POST 但参数在 URL 中，且 Body 为空
        # 这里模拟这种行为，将参数放在 params 里，data 设为空字符串
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}:9097/run/attendanceRecord/getAttendanceByRollCallId",
            data="",  # 发送空字符串
            params=params,
            is_post_str=True,  # 告诉 _request data 是字符串
            force_content_type='application/json;charset=UTF-8'  # 强制类型以发送空 body
        ))

    def submit_attendance(self, payload: dict):
        """提交签到记录"""
        # 假设签到提交使用 form-urlencoded 格式
        # payload 应包含: rollCallId, userId, coordinate, distance, status, signCode 等
        payload["appVersion"] = self.API_VERSION
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}:9097/run/attendanceRecord/addAttendance",
            data=payload,
            is_post_str=False  # 使用默认的 form-urlencoded
        ))

    def get_attendance_radius(self):
        """获取服务器设定的签到半径"""
        # POST, 空body, application/json, 带URL参数
        params = {"code": "attendanceRadius", "num": 1}
        logging.debug("正在从服务器请求签到有效半径配置参数...")
        return self._json(self._request(
            'POST',
            f"{self.BASE_URL}/app/appFind/getDictTips",
            data="",  # 发送空字符串
            params=params,
            is_post_str=True,
            force_content_type='application/json;charset=UTF-8'
        ))

# ==============================================================================
# 3. 后端主逻辑 (Backend API Bridge)
#    作为Python后端和WebView前端之间的桥梁，处理所有业务逻辑。
# ==============================================================================


class Api:
    """此类的方法会暴露给WebView前端的JavaScript调用"""

    def __init__(self, args):
        self.args = args
        self.window = None
        self.path_gen_callbacks = {}

        # --- 路径和配置部分保持不变 ---
        self.run_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
        self.user_dir = SCHOOL_ACCOUNTS_DIR  # 使用新的学校账号目录
        os.makedirs(self.user_dir, exist_ok=True)
        self.config_path = os.path.join(self.run_dir, "config.ini")
        self.user_config_path = self.config_path

        # --- 为单用户模式保留一个api_client实例 ---
        self.api_client = ApiClient(self)

        # --- 初始化状态变量 ---
        self._init_state_variables()

        # --- 全局“数据提交”串行队列：同一时间只允许提交一个数据包 ---
        # 说明：
        # - 任何 addErrandTrack 的提交都会被封装为任务入队；
        # - 后台仅有一个工作线程串行处理，保证“同时只提交一个数据包”；
        # - 调用方在入队后阻塞等待本次提交的结果（带超时）。
        # - 调用方在入队后阻塞等待本次提交的结果（带超时）。
        # 提交串行队列相关的共享状态
        self._submission_queue: queue.Queue | None = getattr(
            self, '_submission_queue', None)
        self._submission_worker_thread: threading.Thread | None = getattr(
            self, '_submission_worker_thread', None)
        self._submission_worker_stop: threading.Event | None = getattr(
            self, '_submission_worker_stop', None)
        if self._submission_queue is None:
            self._submission_queue = queue.Queue()
        if self._submission_worker_stop is None:
            self._submission_worker_stop = threading.Event()
        if (self._submission_worker_thread is None) or (not self._submission_worker_thread.is_alive()):
            self._submission_worker_thread = threading.Thread(
                target=self._submission_worker_loop,
                name="SubmissionWorker",
                daemon=True,
            )
            self._submission_worker_thread.start()

    def _init_state_variables(self):
        """初始化或重置应用的所有状态变量"""
        # --- 单用户模式变量 (大部分保留) ---
        self.device_ua = ""
        self.is_offline_mode = False
        self.user_data = UserData()
        self.all_run_data: list[RunData] = []
        self.current_run_idx = -1
        self.stop_run_flag = threading.Event()
        self.stop_run_flag.set()
        # 修复打卡点检测：设置合理的打卡半径（米）
        # 如果未来服务器返回此值，应从任务详情中更新
        self.target_range_m = 30.0  # 默认30米打卡范围

        # 修复会话持久化：初始化登录状态标志（不清除已存在的）
        if not hasattr(self, 'login_success'):
            self.login_success = False
        if not hasattr(self, 'user_info'):
            self.user_info = None
        # 全局登录并发控制（多账号模式下串行登录）
        # 固定并发为 1；若已有实例，则保持现状（防止模式切换时重复创建）
        if not hasattr(self, 'multi_login_lock') or self.multi_login_lock is None:
            self.multi_login_lock = threading.Semaphore(1)

        # --- 将原params作为全局参数的模板 ---
        self.global_params = {
            "interval_ms": 3000, "interval_random_ms": 500,
            "speed_mps": 1.5, "speed_random_mps": 0.5,
            "location_random_m": 1.5, "task_gap_min_s": 600,
            "task_gap_max_s": 3600, "api_fallback_line": False,
            "api_retries": 2, "api_retry_delay_s": 0.5,
            "ignore_task_time": True,  # 新增：忽略任务起止时间，仅对比日期
            "theme_base_color": "#7dd3fc",
            "theme_style": "default",  # 新增：主题风格参数，默认为 'default'
            # 多账号模式下的路径规划参数
            "min_time_m": 20, "max_time_m": 30, "min_dist_m": 2000,

            # --- 自动签到参数 ---
            "auto_attendance_enabled": False,  # 是否开启自动签到
            "auto_attendance_refresh_s": 30,  # 自动刷新间隔（秒）
            "attendance_user_radius_m": 40,   # 用户期望的随机半径（默认40m）


            # 高德地图JS API Key
            "amap_js_key": ""
        }

        # --- 签到半径的全局缓存 ---
        self.server_attendance_radius_m = 0.0  # 服务器返回的最大半径，0表示精确签到
        self.last_radius_fetch_time = 0  # 上次获取半径的时间

        # --- 单账号自动刷新线程 ---
        self.auto_refresh_thread: threading.Thread | None = None
        self.stop_auto_refresh = threading.Event()
        self.stop_auto_refresh.set()  # 默认停止

        # --- 多账号自动签到线程 ---
        self.multi_auto_refresh_thread: threading.Thread | None = None
        self.stop_multi_auto_refresh = threading.Event()
        self.stop_multi_auto_refresh.set()  # 默认停止

        # self.params 将在单用户模式下使用，是 global_params 的一个副本
        self.params = self.global_params.copy()

        self._first_center_done = False

        # --- 新增多账号模式变量 ---
        self.is_multi_account_mode = False
        # 停止并清理所有账号会话
        if hasattr(self, 'accounts'):
            for acc in self.accounts.values():
                if acc.worker_thread and acc.worker_thread.is_alive():
                    acc.stop_event.set()
        # {username: AccountSession}
        self.accounts: dict[str, AccountSession] = {}
        # 使用 threading（已被 eventlet monkey_patch 绿化）
        self.multi_run_stop_flag = threading.Event()
        self.multi_run_stop_flag.set()

        self._load_tasks_lock = threading.RLock()
        self._load_tasks_inflight = False
        self.multi_run_only_incomplete = True

    def set_multi_run_only_incomplete(self, flag: bool):
        """
        设置多账号模式下“仅执行未完成”的全局开关，并立即刷新全局按钮。
        """
        try:
            self.multi_run_only_incomplete = bool(flag)
            # 轻量刷新一次全局按钮状态（不阻塞）
            self._update_multi_global_buttons()
            return {"success": True}
        except Exception as e:
            return {"success": False, "message": str(e)}

    def set_window(self, window):
        """由主程序调用，设置WebView窗口对象的引用"""
        logging.info("Python后端: set_window方法已被调用，准备设置WebView窗口对象引用")
        self.window = window
        if self.args.autologin:
            user, passwd = self.args.autologin
            self.log("收到自动登录指令。")
            logging.debug(f"收到自动登录请求，目标用户: {user}")
            # 使用 threading.Timer（已被绿化）
            timer = threading.Timer(2.0, lambda: self.window.evaluate_js(
                f'autoLogin("{user}", "{passwd}")'))
            timer.start()

    def log(self, message):
        """将日志消息通过 WebSocket 发送到前端界面显示"""
        # 尝试获取当前 Api 实例关联的 session_id
        session_id = getattr(self, '_web_session_id', None)
        logging.info(message)
        if session_id and socketio:
            try:
                # 使用 session_id 作为房间名，确保消息只发给对应的浏览器标签页
                socketio.emit('log_message', {
                              'msg': str(message)}, room=session_id)
            except Exception as e:
                # 在后台记录发送失败的日志，避免程序崩溃
                logging.error(
                    f"WebSocket emit log failed for session {session_id[:8]}: {e}")
        else:
            # 如果没有 session_id 或 socketio 未初始化，则仅记录到后端日志
            logging.debug(
                f"[Log Emission Skipped] Session ID or SocketIO missing. Message: {message}")

    def js_log(self, level, message):
        """接收并记录来自JavaScript的日志"""
        level = level.upper()
        if level == 'INFO':
            logging.info(f"{message}")
        elif level == 'DEBUG':
            logging.debug(f"{message}")
        elif level == 'WARNING':
            logging.warning(f"{message}")
        elif level == 'ERROR':
            logging.error(f"{message}")
        else:
            logging.info(f"[JS-{level}] {message}")

    @staticmethod
    def _robust_decode(raw: bytes) -> str:
        """尽可能兼容所有编码的解码函数"""
        # 1. 自动检测
        if chardet:
            det = chardet.detect(raw) or {}
            enc = det.get("encoding")
            if enc:
                try:
                    return raw.decode(enc)
                except UnicodeDecodeError:
                    pass

        # 2. 常见编码回退
        for enc in ("utf-8", "gbk", "big5", "shift_jis", "latin-1"):
            try:
                return raw.decode(enc)
            except UnicodeDecodeError:
                continue

        # 3. 最终兜底：强制替换非法字符
        return raw.decode("utf-8", errors="replace")

    def normalize_chinese_config_to_english(self, path: str) -> None:
        """
        将可能包含中文分区/键名的 INI 配置文件，转换为标准英文分区与键名，并统一写回 UTF-8。
        - [增强] 只保留当前程序定义的有效参数，抛弃所有未知或已弃用的参数。
        - 若文件已是英文分区（Config/System），不做任何改写（避免清空原配置）。
        - 若不存在中文分区时，保留原文（以 UTF-8 重写），避免写入空配置。
        - 写回前创建 .bak 备份。
        """
        if not os.path.exists(path):
            return

        # 原始字节与文本
        with open(path, "rb") as f:
            raw = f.read()
        text = self._robust_decode(raw)

        # 先用标准解析器尝试读取（判断是否已是英文分区）
        cfg_existing = configparser.ConfigParser(
            delimiters=("=", ":"),
            comment_prefixes=("#"),
            strict=False,
            interpolation=None
        )
        try:
            cfg_existing.read_string(text)
        except Exception:
            # 原文有异常时，直接以 UTF-8 重写原文并返回
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            return

        # 如果已存在英文分区（Config 或 System），直接保留，不转换不写回
        if cfg_existing.has_section("Config") or cfg_existing.has_section("System"):
            # 已是英文结构，尊重现有内容
            return

        # 继续尝试中文分区读取
        cfg_cn = configparser.ConfigParser(
            delimiters=("=", ":"),
            comment_prefixes=("#"),
            strict=False,
            interpolation=None
        )
        try:
            cfg_cn.read_string(text)
        except Exception:
            # 解析失败则以 UTF-8 重写原文并返回（不清空）
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            return

        # 若没有中文分区（“配置”/“系统”均不存在），保留原文，不写空文件
        if not (cfg_cn.has_section("配置") or cfg_cn.has_section("系统")):
            with open(path, "w", encoding="utf-8") as f:
                f.write(text)
            return

        # 分区与键映射
        config_key_map = {
            "间隔时间": "interval_ms",
            "间隔时间随机": "interval_random_ms",
            "速度": "speed_mps",
            "速度随机": "speed_random_mps",
            "定位随机": "location_random_m",
            "任务间速度": "task_interval_speed",
            "用户名": "Username",
            "密码": "Password",
            "有效": "Enabled",
            "未完成次数": "PendingCount",
            "未完成次数上次检查时间": "LastPendingCheckTs",
            "下次最早开始时间": "NextEarliestStartTs",
        }
        system_key_map = {
            "cookie": "AuthorizationCookie",
            "UA": "UA",
        }

        # 构建英文配置
        cfg_en = configparser.ConfigParser(
            delimiters=("=", ":"),
            comment_prefixes=("#"),
            strict=False,
            interpolation=None
        )
        cfg_en.add_section("Config")
        cfg_en.add_section("System")

        
        # 1. 定义所有当前有效的Config参数键（白名单）
        valid_config_keys = set(self.global_params.keys())
        valid_config_keys.add("Username")
        valid_config_keys.add("Password")

        # 2. 遍历旧配置，只写入白名单中的键
        if cfg_cn.has_section("配置"):
            for k, v in cfg_cn.items("配置"):
                # 转换中文键为英文键
                k_en = config_key_map.get(k, k)
                # 只有当转换后的键是有效参数时，才将其保留
                if k_en in valid_config_keys:
                    cfg_en.set("Config", k_en, v)

        if cfg_cn.has_section("系统"):
            for k, v in cfg_cn.items("系统"):
                k_en = system_key_map.get(k, k)
                if k_en == "UA" and isinstance(v, str) and v.lower().startswith("user-agent:"):
                    v = v.split(":", 1)[1].strip()
                # 系统区的参数也做白名单校验
                if k_en in ["AuthorizationCookie", "UA"]:
                    cfg_en.set("System", k_en, v)

        # 写回前做备份
        try:
            backup_path = f"{path}.bak"
            with open(backup_path, "wb") as bf:
                bf.write(raw)
        except Exception:
            # 备份失败不阻止写回
            pass

        # 写回 UTF-8（仅当确有中文分区内容需要转换）
        with open(path, "w", encoding="utf-8") as f:
            cfg_en.write(f)

    def _save_config(self, username, password=None, ua=None):
        """保存指定用户的配置到 user/<username>.ini；当 password 为 None 时保留现有密码；当 ua 为 None 时保留现有 UA。同时更新主 config.ini 的 LastUser 和 amap_js_key。"""
        logging.debug(
            f"Saving config: username={username!r}, password provided: {password is not None}, ua provided: {ua is not None}")

        # --- 1. 处理用户独立的 .ini 文件 ---
        user_ini_path = os.path.join(self.user_dir, f"{username}.ini")

        # 将旧配置文件（可能是中文各种编码）规范化为英文UTF-8 (这一步保持不变)
        self.normalize_chinese_config_to_english(user_ini_path)

        # 读取旧密码（仅当需要保留时）
        # --- 1. 创建或读取现有配置 ---
        cfg_to_save = configparser.RawConfigParser()
        cfg_to_save.optionxform = str  # 保持键的大小写
        if os.path.exists(user_ini_path):
            try:
                cfg_to_save.read(user_ini_path, encoding='utf-8')
            except Exception as e:
                logging.warning(f"读取旧配置文件 {user_ini_path} 失败: {e}, 将创建新的。")

        # --- 2. 确保分区存在 ---
        if not cfg_to_save.has_section('Config'):
            cfg_to_save.add_section('Config')
        if not cfg_to_save.has_section('System'):
            cfg_to_save.add_section('System')

        # --- 3. 设置 Username ---
        cfg_to_save.set('Config', 'Username', username)

        # --- 4. 智能处理密码 ---
        # 仅当 *提供了新的* password (非 None) 时，才覆盖密码
        if password is not None:
            # 场景: 提供了新密码 (来自 login, multi_add_account)
            cfg_to_save.set('Config', 'Password', password)
        else:
            # 场景: 未提供新密码 (来自 update_param)
            
            # 修复：检查密码行是否*完全不存在*于 .ini 文件中
            if not cfg_to_save.has_option('Config', 'Password'):
                # 选项不存在，尝试从备份文件中恢复
                backup_path = f"{user_ini_path}.bak"
                if os.path.exists(backup_path):
                    try:
                        # 尝试从备份文件中读取密码，使用robust_decode确保兼容各种编码
                        with open(backup_path, "rb") as bf:
                            raw_backup = bf.read()
                        backup_text = self._robust_decode(raw_backup)
                        for line in backup_text.splitlines():
                            clean_line = line.strip()
                            # 规范化用于匹配
                            normalized_line = clean_line.lower().replace(" ", "")
                            if normalized_line.startswith("password=") or normalized_line.startswith("密码="):
                                # 从原始clean_line中分割（保留原始大小写和值）
                                parts = clean_line.split("=", 1)
                                if len(parts) == 2:
                                    recovered_password = parts[1].strip()
                                    if recovered_password:
                                        cfg_to_save.set(
                                            'Config', 'Password', recovered_password)
                                        logging.info(
                                            f"已从备份文件恢复用户 {username} 的密码")
                                        break
                    except Exception as e:
                        logging.warning(f"从备份文件恢复密码失败: {e}")
            else:
                # 选项存在 (has_option 为 True)，无论其值是 "mypass" 还是 "" (空字符串)，
                # 我们都“什么也不做”，以保留 cfg_to_save 对象中已读取的状态。
                # 这可以防止 configparser 在写回时丢弃该行。
                pass

        # --- 5. 智能处理 UA ---
        # 仅当 *提供了新的* ua (非 None) 时，才覆盖 UA
        # 修复：避免在切换账号时误将当前实例的 UA 覆盖到其他用户的配置文件
        if ua is not None:
            # 场景: 明确提供了新 UA (来自 login, generate_new_ua)
            cfg_to_save.set('System', 'UA', ua)
        # 场景: 未提供新 UA (来自 update_param 等)
        # 则 *不* 触碰 UA 键，从而保留 cfg_to_save 中已加载的旧 UA(或它的缺失状态)。

        # --- 6. 参数：从当前实例状态获取 ---
        params_to_save = self.params
        if self.is_multi_account_mode and username in self.accounts:
            params_to_save = self.accounts[username].params
        for k, v in params_to_save.items():
            # 仅保存当前全局参数模板中存在的键
            # 排除 amap_js_key，因为它应该只保存在主 config.ini 中，不应在用户 .ini 中
            if k in self.global_params and k != 'amap_js_key':
                cfg_to_save.set('Config', k, str(v))

        # --- 7. 安全写入用户 .ini 文件 ---
        try:
            with open(user_ini_path, 'w', encoding='utf-8') as f:
                cfg_to_save.write(f)
            logging.debug(
                f"Saved user config for {username} -> {user_ini_path}")
        except Exception as e:
            logging.error(f"写入用户配置文件 {user_ini_path} 失败: {e}", exc_info=True)
            # 可以选择在这里向上抛出异常或返回错误状态

        # --- 2. 处理主 config.ini 文件 ---
        main_cfg = configparser.RawConfigParser()
        main_cfg.optionxform = str  # 保持键的大小写
        if os.path.exists(self.config_path):
            try:
                main_cfg.read(self.config_path, encoding='utf-8')
            except Exception as e:
                logging.warning(f"读取主配置文件 {self.config_path} 失败: {e}, 将创建新的。")

        # 确保 [Config] 分区存在并更新 LastUser
        if not main_cfg.has_section('Config'):
            main_cfg.add_section('Config')
        main_cfg.set('Config', 'LastUser', username)

        # 确保 [Map] 分区存在并更新 amap_js_key（新版）
        if not main_cfg.has_section('Map'):
            main_cfg.add_section('Map')
        # 从内存中的全局参数获取最新的 Key
        amap_key_in_memory = self.global_params.get('amap_js_key', '')
        main_cfg.set('Map', 'amap_js_key', amap_key_in_memory)

        # 安全写入主 config.ini 文件
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                main_cfg.write(f)
            logging.debug(
                f"Updated main config {self.config_path} with LastUser and amap_js_key")
        except Exception as e:
            logging.error(f"写入主配置文件 {self.config_path} 失败: {e}", exc_info=True)

    def _load_global_config(self):
        """从主 config.ini 加载全局配置（优先读取新版 Map.amap_js_key，兼容旧版 System.AmapJsKey）"""
        if not os.path.exists(self.config_path):
            return
        cfg = configparser.RawConfigParser()
        cfg.optionxform = str  # 保持键的大小写
        try:
            cfg.read(self.config_path, encoding='utf-8')

            # 优先读取新版配置 [Map] amap_js_key
            amap_key = cfg.get('Map', 'amap_js_key', fallback="")

            # 如果新版配置为空，尝试读取旧版配置 [System] AmapJsKey（兼容性）
            if not amap_key:
                amap_key = cfg.get('System', 'AmapJsKey', fallback="")
                # 如果旧版有值，迁移到新版
                if amap_key:
                    if not cfg.has_section('Map'):
                        cfg.add_section('Map')
                    cfg.set('Map', 'amap_js_key', amap_key)
                    with open(self.config_path, 'w', encoding='utf-8') as f:
                        cfg.write(f)
                    logging.info("已将AmapJsKey从旧版[System]迁移到新版[Map]")

            self.global_params['amap_js_key'] = amap_key
            logging.info(
                f"Loaded global Amap JS Key: {amap_key if amap_key else '(empty)'}")
        except Exception as e:
            logging.error(
                f"加载全局配置文件 {self.config_path} 失败: {e}", exc_info=True)

    def _load_config(self, username):
        """从.ini文件加载指定用户的配置"""
        self.user_config_path = os.path.join(self.user_dir, f"{username}.ini")
        if not os.path.exists(self.user_config_path):
            return None

        # 先将旧配置文件（可能是中文+各种编码）规范化为英文+UTF-8
        self.normalize_chinese_config_to_english(self.user_config_path)

        cfg = configparser.ConfigParser()
        cfg.read(self.user_config_path, encoding='utf-8')
        password = cfg.get('Config', 'Password', fallback='')
        if not password:
            # 如果密码为空，尝试从原始文件中扫描 "Password=" 或 "密码="，以兼容规范化失败的情况
            try:
                with open(self.user_config_path, "r", encoding="utf-8", errors="ignore") as rf:
                    for line in rf:
                        clean_line = line.strip()
                        # 兼容英文键（不区分大小写）和中文键
                        temp_line_for_check = clean_line.lower().replace(" ", "")
                        if temp_line_for_check.startswith("password=") or temp_line_for_check.startswith("密码="):
                            # 从第一个 "=" 处分割，避免密码中包含 "=" 导致的问题
                            parts = clean_line.split("=", 1)
                            if len(parts) == 2:
                                password = parts[1].strip()
                                # 只要成功获取到密码，就立即跳出循环
                                if password:
                                    break
            except Exception:
                pass

        # 加载的配置应该应用到正确的对象上
        target_params = self.params
        if self.is_multi_account_mode and username in self.accounts:
            target_params = self.accounts[username].params

        ua = cfg.get('System', 'UA', fallback="")
        if self.is_multi_account_mode and username in self.accounts:
            self.accounts[username].device_ua = ua
        else:
            self.device_ua = ua

        for k in self.global_params:  # 迭代全局模板以确保所有键都存在
            if k in target_params:
                try:
                    if cfg.has_option('Config', k):
                        val_str = cfg.get('Config', k)
                        original_type = type(target_params[k])
                        if original_type is bool:
                            target_params[k] = val_str.lower() in (
                                'true', '1', 't', 'yes')
                        else:
                            target_params[k] = original_type(val_str)
                except ValueError:
                    logging.warning(
                        f"Could not parse config value for '{k}' for user {username}. Using default.")
                    pass
        logging.debug(f"已成功加载用户配置: {username}")
        return password

    def _get_full_user_info_dict(self):
        """获取当前用户所有信息的字典"""
        return {k: v for k, v in self.user_data.__dict__.items() if not k.startswith('_')}

    def get_initial_data(self):
        """应用启动时由前端调用，获取初始用户列表和最后登录用户"""
        
        try:
            logging.info("API调用: get_initial_data - 获取应用初始数据（用户列表和最后登录用户）")

            # 获取当前已有的用户配置文件列表
            users = sorted([os.path.splitext(f)[0]
                        for f in os.listdir(self.user_dir) if f.endswith(".ini")])

            # 读取全局配置
            cfg = configparser.RawConfigParser()
            cfg.optionxform = str  # 保持键的大小写
            cfg.read(self.config_path, encoding='utf-8')

            # 确保有 Config 分段
            if not cfg.has_section('Config'):
                cfg.add_section('Config')

            last_user = cfg.get('Config', 'LastUser', fallback="").strip()

            # 如果 last_user 不为空但对应的 .ini 不存在，则清空并写回
            if last_user and last_user not in users:
                logging.warning(f"LastUser '{last_user}' 不存在对应的 .ini，自动清空。")
                cfg.set('Config', 'LastUser', '')
                try:
                    with open(self.config_path, 'w', encoding='utf-8') as f:
                        cfg.write(f)
                except Exception as e:
                    logging.error(f"写回 config.ini 失败：{e}", exc_info=True)
                last_user = ""

            # 在返回数据前，先加载一次全局配置
            self._load_global_config()

            # 修复Issue 5: 检查是否已登录（会话持久化）
            is_logged_in = hasattr(self, 'login_success') and self.login_success
            user_info = None
            if is_logged_in and hasattr(self, 'user_info'):
                user_info = self.user_info

            logging.debug(
                f"Initial users={users}, last user={last_user}, logged_in={is_logged_in}")

            # 检查认证状态
            is_authenticated = hasattr(
                self, 'is_authenticated') and self.is_authenticated
            auth_username = getattr(self, 'auth_username', None)
            auth_group = getattr(self, 'auth_group', 'guest')
            is_guest = getattr(self, 'is_guest', False)

            # [代码片段 2.1：替换掉旧的 return 语句]
            # 构造返回字典
            response_data = {
                "success": True,
                "users": users,
                "lastUser": last_user,
                "amap_key": self.global_params.get('amap_js_key', ''), # 确保使用 amap_js_key
                "isLoggedIn": is_logged_in,
                "userInfo": user_info,
                "is_authenticated": is_authenticated,
                "auth_username": auth_username,
                "auth_group": auth_group,
                "is_guest": is_guest
            }

            # [BUG 2 修复] 如果是已登录状态（会话恢复），则从实例中提取 device_ua
            # (注意：self.device_ua 是在 restore_session_to_api_instance 中被恢复的)
            if is_logged_in and hasattr(self, 'device_ua'):
                response_data['device_ua'] = self.device_ua

            logging.debug(
                f"Initial data response prepared: users={len(users)}, last user={last_user}, logged_in={is_logged_in}, ua_sent={response_data.get('device_ua') is not None}")

            return response_data
        except Exception as e:
            self.is_offline_mode = True
            return {
                "success": False,
                "offline": True,
                "message": "后端无法连接服务器，已切换到离线模式"
            }

    def get_user_sessions(self):
        """获取当前认证用户的会话列表（供前端调用）"""
        logging.info("API调用: get_user_sessions - 获取当前用户的所有活动会话列表")

        # 检查当前会话是否已认证且非游客
        auth_username = getattr(self, 'auth_username', None)
        is_guest = getattr(self, 'is_guest', True)

        if not auth_username or is_guest:
            logging.warning("get_user_sessions: 用户未登录或为游客，返回空列表。")
            # 游客或未登录用户没有关联会话，返回成功但列表为空
            # 对于游客，特殊处理：只返回当前会话信息
            if is_guest and hasattr(self, '_web_session_id'):
                current_session_id = self._web_session_id
                session_file = get_session_file_path(current_session_id)
                created_at = 0
                last_activity = 0
                if os.path.exists(session_file):
                    try:
                        with open(session_file, 'r', encoding='utf-8') as f:
                            s_data = json.load(f)
                        created_at = s_data.get('created_at', 0)
                        last_activity = s_data.get('last_accessed', 0)
                    except Exception:
                        pass  # 忽略读取错误

                guest_session_info = [{
                    'session_id': current_session_id,
                    'session_hash': hashlib.sha256(current_session_id.encode()).hexdigest()[:16],
                    'created_at': created_at,
                    'last_activity': last_activity,
                    'is_current': True,
                    'login_success': False,  # 游客没有学校账号登录状态
                    'user_data': {"username": "guest"}
                }]
                return {"success": True, "sessions": guest_session_info}
            else:
                # 非游客但未认证，或游客但无法获取当前session_id
                return {"success": True, "sessions": []}

        try:
            # 调用 AuthSystem 获取该用户关联的所有会话 ID
            session_ids = auth_system.get_user_sessions(auth_username)
            logging.debug(
                f"Found {len(session_ids)} linked session IDs for user {auth_username}")

            sessions_info = []
            current_session_id = getattr(
                self, '_web_session_id', None)  # 获取当前请求的会话ID

            for sid in session_ids:
                # 检查会话是否还存在（文件是否存在）
                session_file = get_session_file_path(sid)
                if os.path.exists(session_file):
                    try:
                        with open(session_file, 'r', encoding='utf-8') as f:
                            session_data = json.load(f)

                        # 确保加载的数据属于当前用户 (虽然理论上 link_session_to_user 保证了这一点)
                        if session_data.get('auth_username') == auth_username:
                            sessions_info.append({
                                'session_id': sid,
                                # 添加 session_hash 以便前端显示缩略ID
                                'session_hash': hashlib.sha256(sid.encode()).hexdigest()[:16],
                                'created_at': session_data.get('created_at', 0),
                                # 使用 last_accessed
                                'last_activity': session_data.get('last_accessed', 0),
                                'is_current': sid == current_session_id,  # 标记哪个是当前会话
                                # 学校账号登录状态
                                'login_success': session_data.get('login_success', False),
                                # 包含学号等信息
                                'user_data': session_data.get('user_data', {})
                            })
                    except Exception as e:
                        logging.warning(
                            f"读取会话文件 {session_file} 失败: {e}, 跳过该会话。")
                        continue  # 跳过损坏或无法读取的文件

            logging.info(
                f"成功获取用户 {auth_username} 的 {len(sessions_info)} 个会话信息。")
            return {"success": True, "sessions": sessions_info}
        except Exception as e:
            logging.error(f"获取用户会话列表时发生错误: {e}", exc_info=True)
            return {"success": False, "message": f"服务器内部错误: {e}"}

    def save_amap_key(self, api_key: str):
        """由JS调用，保存高德地图API Key到主配置文件"""
        try:
            self.global_params['amap_js_key'] = api_key
            cfg = configparser.RawConfigParser()
            cfg.optionxform = str  # 保持键的大小写
            if os.path.exists(self.config_path):
                cfg.read(self.config_path, encoding='utf-8')

            # 确保 [Map] 分区存在并更新 amap_js_key（新版）
            if not cfg.has_section('Map'):
                cfg.add_section('Map')
            cfg.set('Map', 'amap_js_key', api_key)

            with open(self.config_path, 'w', encoding='utf-8') as f:
                cfg.write(f)
            self.log("高德地图API Key已保存。")
            logging.info("已成功保存新的高德地图JavaScript API密钥")
            return {"success": True}
        except Exception as e:
            self.log(f"保存高德地图API Key失败: {e}")
            logging.error(f"保存高德地图JavaScript API密钥失败: {e}")
            return {"success": False, "message": str(e)}
        except Exception as e:
            self.log(f"API Key保存失败: {e}")
            logging.error(f"保存高德地图JavaScript API密钥时发生异常: {e}", exc_info=True)
            return {"success": False, "message": str(e)}

    def on_user_selected(self, username):
        # return
        """当用户在登录界面选择一个已有用户时调用"""
        logging.info(f"API调用: on_user_selected - 用户选择事件触发，选中的用户名: '{username}'")
        if not username:
            return {"password": "", "ua": "", "params": self.params, "userInfo": {}}
        password = self._load_config(username)

        # 修复：如果配置文件不存在（password返回None），UA应该返回空字符串
        # 这样前端会显示"(新用户将在登录时自动生成)"
        if password is None:
            ua = ""
        else:
            ua = self.device_ua or ""
        logging.debug(
            f"on_user_selected: username={username}, ua={ua}, password={'***' if password else 'empty'}")
        info = {"name": self.user_data.name,
                "student_id": self.user_data.student_id}
        return {"password": password or "", "ua": ua, "params": self.params, "userInfo": info}

    def generate_new_ua(self):
        """生成一个新的UA并保存"""
        logging.info("API调用: generate_new_ua - 生成新的随机User-Agent字符串")
        self.device_ua = ApiClient.generate_random_ua()
        cfg = configparser.ConfigParser()
        if os.path.exists(self.user_config_path):
            cfg.read(self.user_config_path, encoding='utf-8')
            if not cfg.has_section('System'):
                cfg.add_section('System')
            cfg.set('System', 'UA', self.device_ua)
            with open(self.user_config_path, 'w', encoding='utf-8') as f:
                cfg.write(f)
        logging.info(f"已成功生成新的User-Agent字符串: {self.device_ua}")
        return self.device_ua

    def login(self, username, password):
        logging.info(f"API调用: login - 用户登录请求，用户名: '{username}'")
        if not username or not password:
            return {"success": False, "message": "用户名和密码不能为空！"}

        self.log("正在登录...")
        logging.debug(f"正在尝试为用户进行登录认证: 用户名={username}")
        self.user_data = UserData()
        # 记录“登录时输入的账号”，作为兜底
        input_username = username
        # UA 如未生成则生成
        if not self.device_ua:
            self.device_ua = ApiClient.generate_random_ua()

        resp = self.api_client.login(input_username, password)
        if not resp or not resp.get('success'):
            msg = resp.get('message', '未知错误') if resp else '网络连接失败'
            self.log(f"登录失败：{msg}")
            logging.warning(f"用户登录失败: {msg}")
            return {"success": False, "message": msg}

        # 先解析用户信息，再决定用哪个“主键”保存
        self.log("登录成功，正在解析用户信息...")
        data = resp.get('data', {})
        user_info = data.get('userInfo', {})
        dept_info = data.get('deptInfo', {})
        ud = self.user_data

        ud.name = user_info.get('name', '')
        ud.phone = user_info.get('phone', '')
        ud.student_id = user_info.get('account', '')  # 学号/账号（后端唯一标识）
        ud.id = user_info.get('id', '')
        # 这里把 username 统一设成“学号”，而不是 nickname
        # 若后端未返回学号，则回退到输入的 username
        ud.username = ud.student_id or input_username

        ud.registration_time = user_info.get('createtime', '')
        ud.first_login_time = user_info.get('firstlogin', '')
        ud.id_card = user_info.get('iDcard', '')
        ud.current_login_time = user_info.get('logintime', '')
        ud.last_login_time = dept_info.get('logintime', '')
        ud.gender = dept_info.get('sexValue', '')
        ud.school_name = dept_info.get('schoolName', '')
        ud.attribute_type = dept_info.get('typeValue', '')

        # 至此，ud.username 已经是“用于保存配置的主键”（学号优先）
        # 现在再保存配置（文件名与 LastUser 都用 ud.username）
        self._save_config(ud.username, password, self.device_ua)

        # --- 新增：登录成功后，立即获取并缓存签到半径 ---
        try:
            # None 表示这是在单账号模式下调用
            self._fetch_server_attendance_radius_if_needed(
                self.api_client, None)
        except Exception as e:
            self.log(f"获取签到半径失败: {e}")
            logging.warning(
                f"Failed to fetch attendance radius post-login: {e}")
        # --- 结束新增 ---

        self._first_center_done = False
        logging.info(
            f"Login successful: uid={ud.id}, name={ud.name}, sid={ud.student_id}")
        # 双保险：后端主动提示前端刷新任务列表
        if self.window:
            try:
                self.window.evaluate_js('refreshTasks()')
            except Exception:
                logging.debug(
                    "Attempt to trigger front-end refreshTasks failed (non-fatal).")
        # --- 启动自动刷新线程 ---
        try:
            self.stop_auto_refresh.clear()
            if self.auto_refresh_thread is None or not self.auto_refresh_thread.is_alive():
                self.auto_refresh_thread = threading.Thread(
                    target=self._auto_refresh_worker, daemon=True)
                self.auto_refresh_thread.start()
        except Exception as e:
            self.log(f"启动自动刷新线程失败: {e}")

        # 构造要返回给前端的 userInfo 字典
        user_info_dict = self._get_full_user_info_dict()
        # --- 新增：将获取到的半径附加到返回信息中 ---
        user_info_dict['server_attendance_radius_m'] = self.server_attendance_radius_m

        # 修复Issue 5: 设置登录状态标志用于会话持久化
        self.login_success = True
        self.user_info = user_info_dict
        logging.info(
            f"会话状态已保存: login_success={self.login_success}, user_id={ud.id}")

        # --- 新增: 在成功登录的返回结果中包含 auth_group ---
        auth_group = getattr(self, 'auth_group', 'guest') # 从 Api 实例获取认证时确定的组

        return {
            "success": True,
            "userInfo": user_info_dict,
            "ua": self.device_ua,
            "amap_key": self.global_params.get('amap_js_key', ''),
            "auth_group": auth_group
        }

    def logout(self):
        """处理注销逻辑"""
        logging.info("API调用: logout - 用户注销登出操作")
        self.log("已注销。")
        logging.info("用户已成功登出，正在清除会话和状态数据")
        # --- 停止自动刷新线程 ---
        try:
            self.stop_auto_refresh.set()
            if self.auto_refresh_thread and self.auto_refresh_thread.is_alive():
                self.auto_refresh_thread.join(timeout=1.0)
            self.auto_refresh_thread = None
        except Exception as e:
            logging.warning(f"停止自动刷新线程失败: {e}")

        # 修复Issue 5: 清除登录状态标志
        self.login_success = False
        self.user_info = None

        self._init_state_variables()
        self._load_global_config()
        self.api_client.session.cookies.clear()
        return {"success": True}

    def load_tasks(self):
        """加载任务列表（增强：稳健去重 + 并发保护 + 离线模式支持）"""
        logging.info("API调用: load_tasks - 加载用户任务列表")

        # 离线模式或无用户ID但有任务数据时，直接返回已加载的任务
        if not self.user_data.id:
            # 检查是否有已加载的任务（例如从会话恢复或导入的离线文件）
            if hasattr(self, 'all_run_data') and self.all_run_data:
                logging.info(f"load_tasks: 离线模式，返回已加载的 {len(self.all_run_data)} 个任务")
                tasks_for_js = []
                for run in self.all_run_data:
                    task_dict = run.__dict__.copy()
                    task_dict['info_text'] = self._get_task_info_text(run)
                    tasks_for_js.append(task_dict)
                return {"success": True, "tasks": tasks_for_js}
            else:
                return {"success": False, "message": "用户未登录且无离线任务"}

        # 并发保护：避免多次快速点击导致并发刷新交错
        if not hasattr(self, "_load_tasks_lock"):
            self._load_tasks_lock = threading.RLock()
        if not hasattr(self, "_load_tasks_inflight"):
            self._load_tasks_inflight = False

        with self._load_tasks_lock:
            if self._load_tasks_inflight:
                logging.debug(
                    "load_tasks skipped: another refresh is in-flight.")
                # 返回当前缓存（避免空白）——即使并发点击，也不会重复追加
                tasks_for_js = []
                for run in self.all_run_data:
                    task_dict = run.__dict__.copy()
                    task_dict['info_text'] = self._get_task_info_text(run)
                    tasks_for_js.append(task_dict)
                return {"success": True, "tasks": tasks_for_js}

            self._load_tasks_inflight = True
            try:
                self.log("正在获取任务列表...")
                logging.debug("正在从服务器获取任务运行列表数据")

                # 重置缓存
                self.all_run_data = []
                # 复合去重：优先 errandId；若为空或重复，则落到 schedule+时间 的复合 key
                seen_keys: set[str] = set()
                offset = 0
                dup_count = 0

                while True:
                    resp = self.api_client.get_run_list(
                        self.user_data.id, offset)
                    if not resp or not resp.get('success'):
                        self.log("获取任务列表失败。")
                        logging.warning("从服务器获取任务列表失败")
                        break

                    tasks = resp.get('data', {}).get('errandList', [])
                    if not tasks:
                        break

                    for td in tasks:
                        eid = td.get('errandId') or ""
                        es = td.get('errandSchedule') or ""
                        st = td.get('startTime') or ""
                        et = td.get('endTime') or ""

                        # 构造稳健唯一键（兼容 errandId 缺失 / 重复）
                        # 优先 errandId；若为空，则用 "SCHEDULE|START|END" 兜底；若 errandId 存在也加入 schedule 时间以提高稳健性
                        unique_key = f"{eid}|{es}|{st}|{et}"

                        if unique_key in seen_keys:
                            dup_count += 1
                            continue
                        seen_keys.add(unique_key)

                        run = RunData()
                        run.run_name = td.get('eName')

                        # 统一将 isExecute 规范为整型 0/1，防止字符串导致后续比较失败
                        try:
                            run.status = int(td.get('isExecute') or 0)
                        except (TypeError, ValueError):
                            run.status = 1 if str(
                                td.get('isExecute')).strip() == '1' else 0

                        run.errand_id = td.get('errandId')
                        run.errand_schedule = td.get('errandSchedule')
                        run.start_time = td.get('startTime')
                        run.end_time = td.get('endTime')
                        run.upload_time = td.get('updateTime')
                        self.all_run_data.append(run)

                    offset += len(tasks)
                    if len(tasks) < 10:
                        break

                # 日志与前端数据整理
                self.log(f"任务列表加载完毕，共 {len(self.all_run_data)} 项。")
                if dup_count > 0:
                    logging.info(
                        f"Task de-dup completed: {dup_count} duplicates skipped (robust key).")

                tasks_for_js = []
                for run in self.all_run_data:
                    task_dict = run.__dict__.copy()
                    task_dict['info_text'] = self._get_task_info_text(run)
                    tasks_for_js.append(task_dict)
                return {"success": True, "tasks": tasks_for_js}
            finally:
                self._load_tasks_inflight = False

    def _get_task_info_text(self, run: RunData) -> str:
        """根据任务状态生成一个简短的信息文本（增强：多格式时间解析、稳健回退）"""
        if run.status == 1:
            return "已完成"

        now = datetime.datetime.now()
        ignore_time = self.params.get("ignore_task_time", True)

        def try_parse_dt(s):
            """尝试将不同格式的时间字符串解析为 datetime，支持多种常见格式和值类型。失败返回 None。"""
            if not s:
                return None
            # 若为数字字符串或数字（可能是时间戳毫秒/秒）
            try:
                if isinstance(s, (int, float)):
                    # 判断是秒还是毫秒（如果大于 10^12 认为是毫秒）
                    ts = int(s)
                    if ts > 1e12:
                        return datetime.datetime.fromtimestamp(ts / 1000.0)
                    if ts > 1e9:
                        return datetime.datetime.fromtimestamp(ts / 1000.0)
                    return datetime.datetime.fromtimestamp(ts)
                if s.isdigit():
                    ts = int(s)
                    if ts > 1e12:
                        return datetime.datetime.fromtimestamp(ts / 1000.0)
                    if ts > 1e9:
                        return datetime.datetime.fromtimestamp(ts / 1000.0)
                    return datetime.datetime.fromtimestamp(ts)
            except Exception:
                pass

            # 尝试多种常见格式
            fmts = [
                "%Y-%m-%d %H:%M:%S",
                "%Y/%m/%d %H:%M:%S",
                "%Y-%m-%d",
                "%Y/%m/%d",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%dT%H:%M:%S.%f",
                "%Y-%m-%d %H:%M"
            ]
            for f in fmts:
                try:
                    return datetime.datetime.strptime(s, f)
                except Exception:
                    continue
            # 尝试 ISO 解析的简易回退（去掉时区 Z）
            try:
                txt = s.rstrip('Z').split('+')[0]
                return datetime.datetime.fromisoformat(txt)
            except Exception:
                pass
            return None

        # 检查是否已过期
        if run.end_time:
            end_dt = try_parse_dt(run.end_time)
            if end_dt:
                if (ignore_time and end_dt.date() < now.date()) or (not ignore_time and end_dt < now):
                    return "已过期"

        # 检查是否未开始
        if run.start_time:
            start_dt = try_parse_dt(run.start_time)
            if start_dt:
                if (ignore_time and now.date() < start_dt.date()) or (not ignore_time and now < start_dt):
                    # 返回更简洁的未开始标记，便于前端直接判断显示“未开始”
                    return f"开始于: {start_dt.strftime('%Y-%m-%d')}"
            else:
                # 若解析失败但字符串明显为仅有日期（例如 '2025-09-30'）且长度合适，尝试按日期对比
                try:
                    s = str(run.start_time).strip()
                    if len(s) == 10 and s[4] == '-' and s[7] == '-':
                        sd = datetime.datetime.strptime(s, "%Y-%m-%d")
                        if (ignore_time and now.date() < sd.date()) or (not ignore_time and now < sd):
                            return f"开始于: {sd.strftime('%Y-%m-%d')}"
                except Exception:
                    pass

        # 进行中或带截止日期的显示
        if run.end_time:
            try:
                # 优先以可解析的 end_time 的日期部分显示
                end_dt = try_parse_dt(run.end_time)
                if end_dt:
                    return f"截止: {end_dt.strftime('%Y-%m-%d')}"
            except Exception:
                pass

        return "进行中"

    def get_task_details(self, index):
        """获取指定任务的详细信息"""
        logging.info(f"API调用: get_task_details - 获取任务详细信息，任务索引: {index}")
        if not (0 <= index < len(self.all_run_data)):
            return {"success": False, "message": "无效的任务索引"}

        self.current_run_idx = index
        run_data = self.all_run_data[index]

        if run_data.details_fetched:
            task_dict = run_data.__dict__.copy()
            task_dict['target_range_m'] = self.target_range_m
            return {"success": True, "details": task_dict}

        self.log(f"正在加载任务详情...")
        logging.debug(
            f"正在获取任务详细信息: 任务索引={index}, 任务名称={run_data.run_name}")
        resp = self.api_client.get_run_details(
            run_data.errand_id, self.user_data.id, run_data.errand_schedule)

        if resp and resp.get('success'):
            details = resp.get('data', {}).get('errandDetail', {})
            run_data.target_points = [(float(p['lon']), float(p['lat'])) for p in details.get(
                'geoCoorList', []) if p.get('lon') is not None and p.get('lat') is not None]
            run_data.target_point_names = "|".join(
                [p.get('name', '') for p in details.get('geoCoorList', [])])

            temp_coords = []
            walk_paths = details.get('walkPaths', [])
            if walk_paths:
                for i, seg in enumerate(walk_paths):
                    for pt in seg:
                        if isinstance(pt, list) and len(pt) == 2 and pt[0] is not None and pt[1] is not None:
                            try:
                                temp_coords.append(
                                    (float(pt[0]), float(pt[1])))
                            except (TypeError, ValueError):
                                logging.warning(
                                    f"Invalid coordinate in recommended path: {pt}")
                    if i < len(walk_paths) - 1:
                        temp_coords.append((0.0, 0.0))
            run_data.recommended_coords = temp_coords
            run_data.details_fetched = True
            self.log("任务详情加载成功。")
            logging.debug(
                f"任务详情获取成功: 目标点数量={len(run_data.target_points)}, 推荐路径点数量={len(run_data.recommended_coords)}")
            task_dict = run_data.__dict__.copy()
            task_dict['target_range_m'] = self.target_range_m
            return {"success": True, "details": task_dict}
        else:
            self.log("获取任务详情失败。")
            logging.warning("从服务器获取任务详情失败")
            return {"success": False, "message": "获取任务详情失败"}

    def set_draft_path(self, coords):
        """接收前端手动绘制的草稿路径"""
        logging.info(f"API调用: set_draft_path - 设置草稿路径，点数: {len(coords)}")
        if self.current_run_idx == -1:
            return {"success": False, "message": "未选择任务"}
        run = self.all_run_data[self.current_run_idx]
        run.draft_coords = [(c['lng'], c['lat'], c.get('isKey', 0))
                            for c in coords]
        logging.debug(f"已成功设置草稿路径，包含 {len(coords)} 个坐标点")
        return {"success": True}

    def _calculate_distance_m(self, lon1, lat1, lon2, lat2):
        """
        使用Haversine公式精确计算两个GPS坐标点之间的距离（米）。
        
        算法说明：
        - 使用Haversine公式考虑地球曲率
        - 适用于任意距离和任意纬度
        - 精度高，适合GPS应用
        
        Haversine公式：
        a = sin²(Δlat/2) + cos(lat1) * cos(lat2) * sin²(Δlon/2)
        c = 2 * atan2(√a, √(1-a))
        distance = R * c
        
        其中 R = 6371000 米（地球平均半径）
        
        ✓ 改进：
        1. 考虑地球曲率，精度更高
        2. 适用于任意纬度（赤道、极地均可）
        3. 适用于任意距离（短距离和长距离）
        4. 标准的GPS距离计算方法
        
        参数:
            lon1 (float): 起点经度（度）
            lat1 (float): 起点纬度（度）
            lon2 (float): 终点经度（度）
            lat2 (float): 终点纬度（度）
            
        返回:
            float: 距离（米）
            
        示例:
            >>> _calculate_distance_m(120.0, 30.0, 120.01, 30.01)
            1536.8  # 约1.5公里（更精确）
        """
        # 地球平均半径（米）
        R = 6371000
        
        # 转换为弧度
        lat1_rad = math.radians(lat1)
        lat2_rad = math.radians(lat2)
        delta_lat = math.radians(lat2 - lat1)
        delta_lon = math.radians(lon2 - lon1)
        
        # Haversine公式
        a = (math.sin(delta_lat / 2) ** 2 +
             math.cos(lat1_rad) * math.cos(lat2_rad) * 
             math.sin(delta_lon / 2) ** 2)
        c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
        distance = R * c
        
        return distance

    def _calculate_distances_vectorized(self, coords):
        """
        ✓ 使用NumPy向量化计算多个坐标点之间的距离（性能优化，问题#11）
        
        参数:
            coords: [(lon, lat, interval_ms), ...] 坐标列表
            
        返回:
            float: 总距离（米）
            
        说明:
            当NumPy可用时，使用向量化操作一次性计算所有相邻点的距离，
            比逐对计算快10-100倍（取决于点数）。如果NumPy不可用，回退到循环计算。
        """
        if len(coords) < 2:
            return 0.0
        
        # 尝试使用numpy向量化计算（如果可用）
        if np is not None:
            try:
                # 提取经纬度到numpy数组
                lons = np.array([c[0] for c in coords])
                lats = np.array([c[1] for c in coords])
                
                # 计算相邻点的差值
                delta_lons = np.diff(lons)
                delta_lats = np.diff(lats)
                
                # 转换为弧度（向量化）
                lat1_rad = np.radians(lats[:-1])
                lat2_rad = np.radians(lats[1:])
                delta_lat_rad = np.radians(delta_lats)
                delta_lon_rad = np.radians(delta_lons)
                
                # Haversine公式（向量化）
                R = 6371000
                a = (np.sin(delta_lat_rad / 2) ** 2 +
                     np.cos(lat1_rad) * np.cos(lat2_rad) * 
                     np.sin(delta_lon_rad / 2) ** 2)
                c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
                distances = R * c
                
                return np.sum(distances)
            except Exception as e:
                logging.debug(f"[性能优化] NumPy向量化计算失败，回退到循环: {e}")
        
        # 回退到循环计算（当NumPy不可用或失败时）
        total_dist = 0.0
        for i in range(len(coords) - 1):
            total_dist += self._calculate_distance_m(
                coords[i][0], coords[i][1], coords[i+1][0], coords[i+1][1])
        return total_dist

    def _gps_random_offset(self, lon, lat, params):
        """
        对GPS坐标添加随机偏移，模拟真实GPS的漂移误差。
        
        功能说明：
        真实GPS定位存在误差（通常5-10米），为了让模拟轨迹更真实，
        需要在关键点之间的插值点上添加随机偏移，避免完美的直线。
        
        偏移计算：
        1. 从params中读取偏移范围（location_random_m，单位：米）
        2. 在[-m, m]范围内随机选择偏移量
        3. 将米转换为经纬度偏移：
           - 经度偏移 = 米 / 102834.74
           - 纬度偏移 = 米 / 111712.69
        
        使用场景：
        - 关键点（打卡点）：不添加偏移，保持精确坐标
        - 插值点：添加偏移，模拟GPS漂移
        
        转换系数说明：
        - 同_calculate_distance_m，基于纬度30度的近似值
        - 经度转换随纬度变化，但对小偏移影响不大
        
        ⚠️ 注意：
        1. 偏移是独立的，经度和纬度分别随机
        2. 实际GPS误差可能有方向性（如受建筑遮挡）
        3. 偏移量过大会导致轨迹看起来不自然
        
        参数:
            lon (float): 原始经度
            lat (float): 原始纬度
            params (dict): 参数字典，必须包含'location_random_m'键
            
        返回:
            tuple: (偏移后的经度, 偏移后的纬度)
            
        示例:
            >>> params = {'location_random_m': 5}  # 5米偏移
            >>> _gps_random_offset(120.0, 30.0, params)
            (120.00003, 30.00004)  # 偏移约5米
        """
        m = params['location_random_m']  # 偏移范围（米）
        # 在[-m, m]范围内随机偏移，然后转换为经纬度
        return lon + random.uniform(-m, m) / 102834.74, lat + random.uniform(-m, m) / 111712.69

    def process_path(self):
        """
        处理草稿路径，生成带有时间戳的模拟运动轨迹。
        
        功能概述：
        将用户在地图上绘制的草稿路径（draft_coords）转换为详细的
        模拟跑步轨迹（run_coords），包含坐标、时间间隔、速度变化。
        
        算法流程：
        1. 读取草稿路径（关键点序列）
        2. 计算每个时间间隔的移动距离（速度×时间）
        3. 沿着草稿路径插值生成密集的坐标点
        4. 为非关键点添加GPS随机偏移，模拟真实GPS漂移
        5. 记录每个点的时间戳（累积时间）
        
        数据结构：
        - draft_coords: [(lon, lat, is_key_point), ...] 
          - is_key_point=1表示打卡点，必须精确经过
          - is_key_point=0表示普通路径点，可以添加偏移
        
        - run_coords: [(lon, lat, interval_ms), ...]
          - interval_ms: 距离上一个点的时间间隔（毫秒）
          - 第一个点的interval_ms为0
        
        模拟参数（从self.params读取）：
        - interval_ms: 上报间隔（毫秒），如5000 = 5秒上报一次
        - interval_random_ms: 间隔随机变化范围
        - speed_mps: 平均速度（米/秒），如3 m/s = 10.8 km/h
        - speed_random_mps: 速度随机变化范围
        - location_random_m: GPS随机偏移范围（米）
        
        算法特点：
        1. 真实性：速度、间隔、位置都添加随机性
        2. 精确性：保证经过所有打卡点
        3. 连续性：轨迹在草稿路径上连续移动
        4. 可控性：通过参数控制运动特征
        
        ⚠️ 潜在问题：
        1. 循环嵌套较深，大量坐标点时性能较差
        2. 浮点数累积误差可能导致总距离略有偏差
        3. 速度随机性可能导致局部速度异常（过快或过慢）
        
        💡 优化建议：
        1. 对超长路径进行分段处理
        2. 缓存距离计算结果
        3. 添加速度平滑算法（避免突变）
        
        返回:
            dict: {
                "success": True,
                "run_coords": [(lon, lat, interval_ms), ...],
                "total_dist": 总距离（米）,
                "total_time": 总时长（秒）
            }
        """
        logging.info("API调用: process_path - 处理路径，生成模拟运动轨迹")
        
        # 检查是否选择了任务
        if self.current_run_idx == -1:
            return {"success": False, "message": "未选择任务"}
        
        run = self.all_run_data[self.current_run_idx]
        
        # 检查草稿路径是否有效（至少需要起点和终点）
        if not run.draft_coords or len(run.draft_coords) < 2:
            return {"success": False, "message": "没有可处理的路径"}

        self.log("正在处理路径...")
        logging.debug(
            f"正在处理路径: 将 {len(run.draft_coords)} 个草稿点转换为运动坐标序列")

        draft = run.draft_coords  # 草稿路径：用户绘制的关键点
        run.run_coords = []  # 清空之前的模拟路径

        # ===== 步骤1：处理起点 =====
        start_lon, start_lat = draft[0][0], draft[0][1]
        # 如果起点是关键点（打卡点），保持精确坐标；否则添加随机偏移
        lon, lat = (start_lon, start_lat) if draft[0][2] == 1 else self._gps_random_offset(
            start_lon, start_lat, self.params)
        # 起点的时间间隔为0（初始位置）
        run.run_coords.append((lon, lat, 0))

        # 初始化统计变量
        total_dist, total_time = 0.0, 0.0
        current_gps_pos, draft_idx = (draft[0][0], draft[0][1]), 0  # 当前位置和草稿索引
        p = self.params  # 简化参数引用

        # ===== 步骤2：主循环 - 沿草稿路径插值生成坐标点 =====
        # 算法思想：
        # - 每次迭代代表一个时间间隔（如5秒）
        # - 根据速度和时间计算这个间隔内应该移动的距离
        # - 沿着草稿路径向前移动这段距离，找到新位置
        # - 如果这段距离跨越多个草稿段，需要累积移动
        
        # ✓ 速度平滑：使用移动平均避免速度突变（问题#12修复）
        speed_history = []  # 记录最近的速度值
        speed_window = 3    # 移动平均窗口大小
        
        while draft_idx < len(draft) - 1:
            interval_t = max(0.2, random.uniform(
                p['interval_ms'] - p['interval_random_ms'], p['interval_ms'] + p['interval_random_ms']) / 1000.0)
            
            # 生成随机速度
            raw_speed = max(0.2, random.uniform(
                p['speed_mps'] - p['speed_random_mps'], p['speed_mps'] + p['speed_random_mps']))
            
            # ✓ 应用速度平滑算法
            speed_history.append(raw_speed)
            if len(speed_history) > speed_window:
                speed_history.pop(0)  # 保持窗口大小
            # 使用移动平均作为实际速度
            speed = sum(speed_history) / len(speed_history)
            
            dist_to_go = speed * interval_t

            final_pos, temp_draft_idx = current_gps_pos, draft_idx
            # 内层循环：沿着草稿路径移动dist_to_go距离
            # 如果一条段不够长，就继续下一段，直到移动完成或到达路径终点
            while dist_to_go > 0 and temp_draft_idx < len(draft) - 1:
                # 当前段：从final_pos到下一个草稿点
                seg_start_gps, seg_end_gps = final_pos, (
                    draft[temp_draft_idx + 1][0], draft[temp_draft_idx + 1][1])
                # 计算这一段的长度
                seg_dist = self._calculate_distance_m(
                    seg_start_gps[0], seg_start_gps[1], seg_end_gps[0], seg_end_gps[1])

                # 情况1：这一段足够长，可以在段内插值
                if seg_dist >= dist_to_go:
                    # 线性插值：在段的起点和终点之间按比例找到位置
                    # ratio = 0 在起点，ratio = 1 在终点
                    # 🐛 潜在bug：seg_dist为0时会导致除零，已用三元运算符防护
                    ratio = dist_to_go / seg_dist if seg_dist > 0 else 0
                    # 计算插值位置：起点 + ratio × (终点 - 起点)
                    final_pos = (seg_start_gps[0] + ratio * (seg_end_gps[0] - seg_start_gps[0]),
                                 seg_start_gps[1] + ratio * (seg_end_gps[1] - seg_start_gps[1]))
                    # 已移动完成，更新索引
                    dist_to_go, draft_idx = 0, temp_draft_idx
                # 情况2：这一段不够长，移动到段终点后继续下一段
                else:
                    # 减去这一段的距离，还需要继续移动
                    dist_to_go -= seg_dist
                    # 移动到段终点
                    final_pos = seg_end_gps
                    # 前进到下一个草稿段
                    temp_draft_idx += 1
                    # 如果下一个点是关键点（打卡点），必须在此停止
                    # 因为关键点不能跳过，必须精确到达
                    if draft[temp_draft_idx][2] == 1:
                        dist_to_go = 0

            # 边界情况：如果所有草稿段都用完了还没移动够dist_to_go
            # 说明草稿路径太短，强制移动到终点
            # 🐛 这种情况理论上不应该发生，可能说明参数配置有问题
            if dist_to_go > 0:
                final_pos = (draft[-1][0], draft[-1][1])
            # 更新当前位置和草稿索引，为下一次迭代做准备
            draft_idx, current_gps_pos = temp_draft_idx, final_pos

            # 检查final_pos是否是关键点（打卡点）
            # 如果是关键点，保持精确坐标；如果不是，添加GPS随机偏移
            # ✓ 使用epsilon容差比较，避免浮点数精度问题
            epsilon = 1e-9  # 容差值，约0.0001米
            is_key_point = any(
                abs(d[0] - final_pos[0]) < epsilon and 
                abs(d[1] - final_pos[1]) < epsilon and 
                d[2] == 1 
                for d in draft
            )
            # 根据是否关键点决定是否添加偏移
            lon, lat = (final_pos[0], final_pos[1]) if is_key_point else self._gps_random_offset(
                final_pos[0], final_pos[1], self.params)
            
            # ✓ 优化：在生成点的同时计算距离，避免二次遍历
            if len(run.run_coords) > 0:
                prev_coord = run.run_coords[-1]
                segment_dist = self._calculate_distance_m(
                    prev_coord[0], prev_coord[1], lon, lat
                )
                total_dist += segment_dist
            
            # 添加到run_coords：(经度, 纬度, 距上一点的时间间隔毫秒)
            run.run_coords.append((lon, lat, int(interval_t * 1000)))
            # 累计总时间
            total_time += interval_t

        # ===== 步骤3：计算总距离 ===== (已优化，在生成点时同时计算)

        # 保存结果到run_data对象
        run.total_run_time_s, run.total_run_distance_m = total_time, total_dist
        self.log(f"处理完成。")
        logging.info(
            f"路径处理完成: 生成坐标点数={len(run.run_coords)}, 总距离={total_dist:.1f}米, 总时长={total_time:.1f}秒")
        # 返回成功结果，包含生成的坐标序列和统计信息
        return {"success": True, "run_coords": run.run_coords, "total_dist": total_dist, "total_time": total_time}

    def check_target_reached_during_run(self, run_data: RunData, current_lon: float, current_lat: float):
        """
        在模拟运行时，检查当前位置是否到达了打卡点。
        
        功能说明：
        跑步任务通常有多个打卡点（如起点、中间点、终点），必须依次经过。
        此函数在每次位置更新时调用，检测是否进入打卡点范围内。
        
        打卡点状态机：
        1. 未进入区域 (is_in_target_zone=False) → 进入区域后触发打卡
        2. 已在区域内 (is_in_target_zone=True) → 避免重复打卡
        3. 离开区域 → 重置状态，准备检测下一个打卡点
        
        关键逻辑：
        - 打卡点必须按顺序(target_sequence)依次到达
        - 每个打卡点只能打卡一次（通过is_in_target_zone标志防重）
        - 打卡成功后自动移动到下一个打卡点
        - 如果当前位置同时在下一个打卡点范围内，自动处理
        
        ⚠️ 潜在问题：
        1. 如果两个打卡点非常近（距离 < 2×range），可能跳过中间点
        
        参数:
            run_data (RunData): 任务数据对象，包含打卡点列表和状态
            current_lon (float): 当前经度
            current_lat (float): 当前纬度
        
        修改状态:
            run_data.target_sequence: 当前应到达的打卡点序号（✓ 0-based，0表示第一个打卡点）
            run_data.is_in_target_zone: 是否在打卡点范围内
        """
        # 边界检查：确保target_sequence有效
        # ✓ target_sequence从0开始（0-based索引），最大值为打卡点数量-1
        if not (0 <= run_data.target_sequence < len(run_data.target_points)):
            logging.debug(
                f"打卡点检查跳过: target_sequence={run_data.target_sequence}, total_points={len(run_data.target_points)}")
            return

        # 获取当前应到达的打卡点坐标
        # ✓ target_sequence直接作为数组索引使用（0-based）
        tar_lon, tar_lat = run_data.target_points[run_data.target_sequence]
        
        # 计算当前位置与打卡点的距离
        dist = self._calculate_distance_m(
            current_lon, current_lat, tar_lon, tar_lat)
        
        # 判断是否在打卡范围内
        # target_range_m通常为50-100米，可配置
        is_in_zone = (dist < self.target_range_m)

        # 详细的调试日志，记录检查过程（显示为1-based给用户看）
        logging.debug(f"打卡点检查: 当前位置=({current_lon:.6f}, {current_lat:.6f}), "
                      f"目标点{run_data.target_sequence+1}(索引{run_data.target_sequence})=({tar_lon:.6f}, {tar_lat:.6f}), "
                      f"距离={dist:.2f}米, 范围={self.target_range_m:.2f}米, "
                      f"在范围内={is_in_zone}, 已在区域内={run_data.is_in_target_zone}")

        # === 状态转换逻辑 ===
        # 情况1：进入打卡范围，且之前不在范围内（首次进入，触发打卡）
        if is_in_zone and not run_data.is_in_target_zone:
            # 标记为已在区域内，防止重复打卡
            run_data.is_in_target_zone = True
            # 日志显示为1-based（用户视角）
            logging.info(
                f"✓ 到达打卡点 {run_data.target_sequence+1}/{len(run_data.target_points)}")
            
            # 如果还有下一个打卡点，移动到下一个
            if run_data.target_sequence + 1 < len(run_data.target_points):
                # 递增打卡点序号
                run_data.target_sequence += 1
                # 获取下一个打卡点坐标（直接使用0-based索引）
                next_lon, next_lat = run_data.target_points[run_data.target_sequence]
                # 检查当前位置是否也在下一个打卡点范围内
                # 如果不在，重置is_in_target_zone，准备检测下一个
                # 💡 这个逻辑处理了两个打卡点很近的情况
                if self._calculate_distance_m(current_lon, current_lat, next_lon, next_lat) >= self.target_range_m:
                    run_data.is_in_target_zone = False
                    logging.debug(
                        f"移动到下一个打卡点 {run_data.target_sequence+1}（索引{run_data.target_sequence}），已离开区域")
                # else: 当前位置已在下一个打卡点范围内，保持is_in_target_zone=True
                # 下次调用此函数时会立即触发下一个打卡点的打卡
        
        # 情况2：不在打卡范围内，重置状态
        # 这确保了离开区域后可以重新检测（虽然正常流程不会重新检测同一个点）
        elif not is_in_zone:
            run_data.is_in_target_zone = False

    def start_single_run(self):
        """开始执行单个任务"""
        logging.info("API调用: start_single_run - 开始执行单个任务")
        if not self.stop_run_flag.is_set():
            return {"success": False, "message": "已有任务在运行"}
        if self.current_run_idx == -1 or not self.all_run_data[self.current_run_idx].run_coords:
            return {"success": False, "message": "请选择任务并生成路线"}

        self.stop_run_flag.clear()
        run_data = self.all_run_data[self.current_run_idx]
        run_data.target_sequence = 0  # ✓ 从0开始（0-based索引）
        run_data.is_in_target_zone = False
        self._first_center_done = False

        logging.info(f"正在启动单任务执行: 任务名称={run_data.run_name}")
        # 修正: 为单用户模式传递 self.api_client
        threading.Thread(target=self._run_submission_thread, args=(
            run_data, self.current_run_idx, self.api_client, False), daemon=True).start()
        return {"success": True}

    def stop_run(self):
        """停止当前所有正在执行的任务"""
        logging.info("API调用: stop_run - 停止当前所有正在执行的任务")
        self.log("正在停止任务...")
        logging.info("已收到停止运行任务的信号")
        self.stop_run_flag.set()
        return {"success": True}

    def get_run_status(self):
        """获取当前运行状态（用于前端轮询）"""
        # 检查是否有任务在运行
        is_running = not self.stop_run_flag.is_set()

        if not is_running or self.current_run_idx == -1:
            return {"running": False}

        run_data = self.all_run_data[self.current_run_idx]
        total_points = len(run_data.run_coords) if run_data.run_coords else 0

        # 修复Issue 1: 使用current_point_index追踪实际进度
        processed_points = getattr(run_data, 'current_point_index', 0)
        processed_points = min(processed_points, total_points)

        # 获取当前位置（如果有的话）
        current_position = None
        if processed_points > 0 and processed_points <= total_points:
            coord = run_data.run_coords[processed_points - 1]
            current_position = {
                "lon": coord[0],
                "lat": coord[1]
            }

        return {
            "running": True,
            "processed_points": processed_points,
            "total_points": total_points,
            "distance_covered": run_data.distance_covered_m,
            "target_sequence": run_data.target_sequence,
            "duration": sum(p[2] for p in run_data.run_coords[:processed_points]) if processed_points > 0 else 0,
            "current_position": current_position
        }

    # 修正: 添加 acc_session 或 api_client 参数
    def _submit_chunk(self, run_data: RunData, chunk, start_time, is_finish, chunk_start_index, client: ApiClient, user: UserData):
        """将一小块轨迹数据提交到服务器"""
        log_func = client.app.log if hasattr(
            client.app, 'log') else client.app.api_bridge.log

        # 修复离线测试模式：检查是否为离线模式
        if self.is_offline_mode:
            log_func(f"[离线测试模式] 模拟提交 {len(chunk)} 个GPS点...")
            logging.info(
                f"[离线测试模式] 模拟提交 chunk: start_index={chunk_start_index}, size={len(chunk)}, is_finish={is_finish}")
            time.sleep(0.1)  # 模拟网络延迟
            return True  # 离线模式总是返回成功

        log_func(f"正在提交数据...")
        logging.debug(
            f"Submitting chunk: start_index={chunk_start_index}, size={len(chunk)}, is_finish={is_finish}")

        last_point_gps = chunk[0] if chunk_start_index == 0 else run_data.run_coords[chunk_start_index - 1]
        time_elapsed_before_chunk_ms = sum(
            p[2] for p in run_data.run_coords[:chunk_start_index])
        coords_list, chunk_total_dist, chunk_total_dur = [], 0.0, 0

        for lon, lat, dur_ms in chunk:
            distance = self._calculate_distance_m(
                last_point_gps[0], last_point_gps[1], lon, lat)
            time_elapsed_before_chunk_ms += dur_ms
            coords_list.append({
                "location": f"{lon},{lat}",
                "locatetime": str(int(time.time() * 1000)),
                "dis": f"{distance:.1f}",
                "count": str(int(time_elapsed_before_chunk_ms / 1000))
            })
            last_point_gps = (lon, lat, dur_ms)
            chunk_total_dist += distance
            chunk_total_dur += dur_ms

        payload = {
            "scheduleId": run_data.errand_schedule, "userId": user.id,
            "userName": user.name or "", "runLength": str(int(chunk_total_dist)),
            "runTime": str(chunk_total_dur), "startPoint": "", "endPoint": "",
            "startTime": start_time,
            # endTime 将在下面根据条件添加
            "trid": run_data.trid, "sid": "", "tid": "",
            "speed": f"{(run_data.total_run_distance_m / run_data.total_run_time_s):.2f}" if run_data.total_run_time_s > 0 else "",
            "finishType": "1" if is_finish else "0",
            "coordinate": json.dumps(coords_list, separators=(',', ':')),
            "appVersion": ApiClient.API_VERSION
        }

        # 仅当这是最后一块(is_finish=True)时才添加 endTime
        if is_finish:
            payload["endTime"] = str(int(time.time() * 1000))
            
        payload_str = urllib.parse.urlencode(payload)

        # 通过全局串行队列进行提交，保证同一时间只提交一个数据包
        resp = self._enqueue_submission(client, payload_str, wait_timeout=60.0)
        success = bool(resp and resp.get('success'))
        log_func(f"数据提交{'成功' if success else '失败'}。")
        logging.debug(
            f"Chunk submission result: success={success}, msg={resp.get('message') if resp else 'network error'}")
        return success

    # ===================== 提交队列：串行化所有数据包提交 =====================
    def _submission_worker_loop(self):
        """后台工作线程：从队列取出提交任务并串行执行。"""
        while not getattr(self, '_submission_worker_stop', threading.Event()).is_set():
            try:
                task = self._submission_queue.get(timeout=0.5)
            except Exception:
                continue
            try:
                client: ApiClient = task.get('client')
                payload_str: str = task.get('payload')
                # 实际网络提交（内部已带重试）
                resp = client.submit_run_track(payload_str)
                task['response'] = resp
            except Exception as e:
                logging.error(f"提交任务执行异常: {e}", exc_info=True)
                task['response'] = None
            finally:
                try:
                    task['event'].set()
                except Exception:
                    pass
                try:
                    self._submission_queue.task_done()
                except Exception:
                    pass

    def _enqueue_submission(self, client: ApiClient, payload_str: str, wait_timeout: float = 30.0):
        """将一次提交加入全局队列，并等待结果返回。
        返回值：与 ApiClient.submit_run_track 一致的响应字典，或 None（失败/超时）。
        """
        # 入队任务包
        task = {
            'client': client,
            'payload': payload_str,
            'event': threading.Event(),  # 使用绿化后的 Event
            'response': None
        }
        # 简要日志：提示入队长度，帮助定位排队情况
        try:
            qsize = self._submission_queue.qsize()
            logging.debug(
                f"Enqueue submit task. Queue size before push: {qsize}")
        except Exception:
            pass
        self._submission_queue.put(task)
        # 同步等待结果
        signaled = task['event'].wait(timeout=wait_timeout)
        if not signaled:
            logging.warning("提交队列等待超时。")
            return None
        return task.get('response')

    # 修正: 添加 client 参数
    def _finalize_run(self, run_data: RunData, task_index: int, client: ApiClient):
        """在所有数据提交后，查询服务器确认任务是否已标记为完成"""
        log_func = client.app.log if hasattr(
            client.app, 'log') else client.app.api_bridge.log
        log_func("正在确认任务状态...")
        logging.debug(f"正在确认任务完成状态，任务追踪ID: trid={run_data.trid}")
        for _ in range(3):
            resp = client.get_run_info_by_trid(run_data.trid)
            if resp and resp.get('success'):
                record_map = resp.get('data', {}).get('recordMap', {})
                if record_map and record_map.get('status') == 1:
                    run_data.status = 1
                    log_func("任务已确认完成。")
                    logging.info(
                        f"任务已成功完成: 任务名称={run_data.run_name}")
                    # 修正：使用 SocketIO 发送 task_completed 事件
                    session_id = getattr(self, '_web_session_id', None)
                    if socketio and session_id and task_index != -1:
                        try:
                            socketio.emit('task_completed', {
                                'task_index': task_index
                            }, room=session_id)
                        except Exception as e:
                            logging.error(f"SocketIO发送'task_completed'事件失败: {e}")
                    return
            time.sleep(1)
        log_func("暂未确认完成，请稍后刷新。")
        logging.warning(
            f"任务完成状态确认失败: 任务名称={run_data.run_name}")

    # 修正: 添加 client 参数
    def _run_submission_thread(self, run_data: RunData, task_index: int, client: ApiClient, is_all: bool, finished_event: threading.Event | None = None):
        """模拟跑步和提交数据的主线程函数"""
        log_func = client.app.log if hasattr(
            client.app, 'log') else client.app.api_bridge.log
        user_data = client.app.user_data if hasattr(
            client.app, 'user_data') else self.user_data
        stop_flag = client.app.stop_event if hasattr(
            client.app, 'stop_event') else self.stop_run_flag

        # 获取当前会话ID（用于Web模式的自动保存）
        session_id = getattr(self, '_web_session_id', None)
        last_auto_save_time = time.time()

        try:
            log_func("开始执行任务。")
            logging.info(
                f"任务提交线程已启动: 任务名称={run_data.run_name}")

            run_data.trid = f"{user_data.student_id}{int(time.time() * 1000)}"
            start_time_ms = str(int(time.time() * 1000))
            run_data.distance_covered_m = 0.0
            last_point_gps = run_data.run_coords[0]
            submission_successful = True

            point_index = 0  # 修复Issue 1: 追踪当前处理的点索引
            for i in range(0, len(run_data.run_coords), 40):
                if stop_flag.is_set():
                    log_func("任务已中止。")
                    logging.info("检测到停止标志，正在中止任务运行")
                    break

                chunk = run_data.run_coords[i:i + 40]

                for lon, lat, dur_ms in chunk:
                    if stop_flag.wait(timeout=dur_ms / 1000.0):
                        logging.debug(
                            "等待下一个坐标点时被停止信号中断")
                        break

                    run_data.distance_covered_m += self._calculate_distance_m(
                        last_point_gps[0], last_point_gps[1], lon, lat)
                    last_point_gps = (lon, lat, dur_ms)
                    point_index += 1  # 修复Issue 1: 更新点索引
                    run_data.current_point_index = point_index  # 修复Issue 1: 保存到run_data
                    self.check_target_reached_during_run(run_data, lon, lat)

                    # 关键增强：每30秒自动保存会话状态（在Web模式下）
                    if session_id and (time.time() - last_auto_save_time >= 30):
                        try:
                            with web_sessions_lock:
                                if session_id in web_sessions:
                                    save_session_state(
                                        session_id, web_sessions[session_id])
                                    logging.debug(
                                        f"任务执行中自动保存会话状态 (进度: {point_index}/{len(run_data.run_coords)})")
                            last_auto_save_time = time.time()
                        except Exception as e:
                            logging.error(f"任务执行中自动保存会话失败: {e}")

                    # 修正：使用 SocketIO 发送 runner_position_update 事件
                    session_id = getattr(self, '_web_session_id', None)
                    if socketio and session_id and self.current_run_idx == task_index:
                        center_now = False
                        if not self._first_center_done:
                            center_now = True
                            self._first_center_done = True
                        try:
                            socketio.emit('runner_position_update', {
                                'lon': lon, 'lat': lat, 'distance': run_data.distance_covered_m,
                                'target_sequence': run_data.target_sequence, 'duration': dur_ms,
                                'center_now': center_now
                            }, room=session_id)
                        except Exception as e:
                            logging.error(f"SocketIO发送'runner_position_update'位置更新事件失败: {e}")

                if stop_flag.is_set():
                    break

                is_final_chunk = (i + 40 >= len(run_data.run_coords))
                # 尝试提交，失败时最多重试 3 次再放弃
                max_attempts = 3
                attempt = 1
                chunk_submitted = False
                while attempt <= max_attempts:
                    if self._submit_chunk(run_data, chunk, start_time_ms, is_final_chunk, i, client, user_data):
                        chunk_submitted = True
                        break
                    # 提交失败
                    submission_successful = False
                    if self.is_offline_mode:
                        logging.error(
                            f"[离线测试模式] 模拟提交失败（不应该发生），尝试 {attempt}/{max_attempts}")
                        # 离线模式下不做额外等待
                    else:
                        logging.warning(f"数据提交失败，重试 {attempt}/{max_attempts}")
                        # 短暂等待后重试
                        time.sleep(1)
                    attempt += 1

                if not chunk_submitted:
                    logging.error(f"数据提交在 {max_attempts} 次尝试后仍然失败，任务中止")
                    break

            if not stop_flag.is_set() and submission_successful:
                log_func("任务执行完毕，等待确认...")
                logging.info("任务运行执行完毕，等待最终确认")
                time.sleep(3)
                self._finalize_run(run_data, task_index, client)

                # 关键增强：任务完成后立即保存会话状态
                if session_id:
                    try:
                        with web_sessions_lock:
                            if session_id in web_sessions:
                                save_session_state(
                                    session_id, web_sessions[session_id], force_save=True)
                                logging.info(f"任务完成，已保存会话状态")
                    except Exception as e:
                        logging.error(f"任务完成后保存会话失败: {e}")

        finally:
            # 修复Issue: 只有在停止标志已设置（用户手动停止）或出现错误时才设置停止标志
            # 正常完成时不应该设置停止标志，避免中断后续操作
            if not is_all:
                # 仅在任务被手动停止或失败时才标记停止
                # 正常完成的情况下，stop_flag.is_set()应该为False
                if not submission_successful or stop_flag.is_set():
                    self.stop_run_flag.set()
                    logging.info(f"任务停止或失败，设置停止标志")
                else:
                    # 正常完成，设置停止标志以允许新任务开始
                    self.stop_run_flag.set()
                    logging.info(f"任务正常完成，重置停止标志")
                # 修正：使用 SocketIO 发送 run_stopped 事件
                session_id = getattr(self, '_web_session_id', None)
                if socketio and session_id:
                    try:
                        socketio.emit('run_stopped', {}, room=session_id)
                    except Exception as e:
                        logging.error(f"SocketIO发送'run_stopped'运行停止事件失败: {e}")
            if finished_event:
                finished_event.set()
            logging.info(
                f"Submission thread finished for task: {run_data.run_name}")

    def _get_path_for_distance(self, path, cumulative_distances, target_dist):
        """如果路径总长不足，则通过来回走的方式凑足目标距离"""
        total_len = cumulative_distances[-1]
        final_path = list(path)
        if 0 < total_len < target_dist:
            rem = target_dist - total_len
            rev = path[::-1]
            acc = 0.0
            for i in range(len(rev) - 1):
                seg = self._calculate_distance_m(
                    rev[i][0], rev[i][1], rev[i + 1][0], rev[i + 1][1])
                if acc + seg < rem:
                    final_path.append(rev[i + 1])
                    acc += seg
                else:
                    ratio = (rem - acc) / seg if seg > 0 else 0
                    s, e = rev[i], rev[i + 1]
                    final_path.append(
                        (s[0] + (e[0] - s[0]) * ratio, s[1] + (e[1] - s[1]) * ratio))
                    break
        return final_path

    def _get_point_at_distance(self, path, cumulative_distances, dist):
        """在一条路径上，根据距离找到精确的坐标点"""
        idx = bisect.bisect_left(cumulative_distances, dist)
        if idx == 0:
            return path[0]
        if idx >= len(cumulative_distances):
            return path[-1]

        d0, d1 = cumulative_distances[idx - 1], cumulative_distances[idx]
        seg_len = d1 - d0
        ratio = (dist - d0) / seg_len if seg_len > 0 else 0
        s, e = path[idx - 1], path[idx]
        return (s[0] + (e[0] - s[0]) * ratio, s[1] + (e[1] - s[1]) * ratio)

    def auto_generate_path_with_api(self, api_path_coords, min_t_m, max_t_m, min_d_m):
        """接收由前端JS API规划好的路径点，并生成模拟数据"""
        logging.info(
            f"API CALL: auto_generate_path_with_api with {len(api_path_coords)} points")
        if self.current_run_idx == -1:
            return {"success": False, "message": "请先选择任务"}
        run = self.all_run_data[self.current_run_idx]

        self.log("收到JS API路径，正在生成模拟数据...")
        logging.info(
            f"Auto-generating path from {len(api_path_coords)} Amap API points.")
        if not api_path_coords or len(api_path_coords) < 2:
            return {"success": False, "message": "高德API未能返回有效路径"}

        final_path_dedup = []
        last_coord = None
        for p in api_path_coords:
            # 修正BUG：同时兼容 'lng' 和 'lon' 两种经度键名
            longitude = p.get('lng', p.get('lon'))
            if longitude is None:
                continue  # 如果经度不存在，则跳过这个点
            coord = (longitude, p['lat'])
            if coord != last_coord:
                final_path_dedup.append(coord)
                last_coord = coord

        target_time_s = random.uniform(min_t_m * 60, max_t_m * 60)
        target_dist_m = random.uniform(min_d_m, min_d_m * 1.15)

        cumulative = [0.0]
        for i in range(len(final_path_dedup) - 1):
            cumulative.append(cumulative[-1] + self._calculate_distance_m(final_path_dedup[i][0],
                              final_path_dedup[i][1], final_path_dedup[i + 1][0], final_path_dedup[i + 1][1]))

        final_geo_path = self._get_path_for_distance(
            final_path_dedup, cumulative, target_dist_m)

        final_cumulative = [0.0]
        for i in range(len(final_geo_path) - 1):
            final_cumulative.append(final_cumulative[-1] + self._calculate_distance_m(
                final_geo_path[i][0], final_geo_path[i][1], final_geo_path[i + 1][0], final_geo_path[i + 1][1]))

        actual_total_dist = final_cumulative[-1] if final_cumulative else 0.0
        if actual_total_dist == 0:
            return {"success": False, "message": "路径计算距离为0"}

        avg_speed = actual_total_dist / target_time_s
        run_coords = []
        if not final_geo_path:
            return {"success": False, "message": "无法生成地理路径"}

        start = final_geo_path[0]
        run_coords.append((self._gps_random_offset(
            start[0], start[1], self.params) + (0,)))
        t_elapsed, d_covered = 0.0, 0.0

        while t_elapsed < target_time_s:
            interval = min(random.uniform(
                self.params['interval_ms'] * .9, self.params['interval_ms'] * 1.1) / 1000.0, target_time_s - t_elapsed)
            if interval <= 0.1:
                break

            d_covered = min(d_covered + random.uniform(avg_speed * .9,
                            avg_speed * 1.1) * interval, actual_total_dist)
            lon, lat = self._get_point_at_distance(
                final_geo_path, final_cumulative, d_covered)
            lon_o, lat_o = self._gps_random_offset(lon, lat, self.params)
            run_coords.append((lon_o, lat_o, int(interval * 1000)))
            t_elapsed += interval
            if d_covered >= actual_total_dist:
                break

        run.run_coords = run_coords
        run.total_run_time_s = t_elapsed
        run.total_run_distance_m = d_covered

        self.log("自动生成完成。")
        logging.info(
            f"Auto-generated path: points={len(run.run_coords)}, dist={d_covered:.1f}, time={t_elapsed:.1f}")
        return {"success": True, "run_coords": run.run_coords, "total_dist": d_covered, "total_time": t_elapsed}

    def start_all_runs(self, ignore_completed, auto_generate):
        """开始执行所有符合条件的任务"""
        logging.info(
            f"API CALL: start_all_runs (ignore_completed={ignore_completed}, auto_generate={auto_generate})")
        if not self.stop_run_flag.is_set():
            return {"success": False, "message": "已有任务在运行"}

        tasks_to_run = []
        for i, d in enumerate(self.all_run_data):
            # 修正BUG：如果任务已完成，并且没有勾选“忽略已完成”，则跳过
            if d.status == 1 and not ignore_completed:
                continue

            is_expired = False
            is_not_started = False
            now = datetime.datetime.now()
            ignore_time = self.params.get("ignore_task_time", True)

            try:
                if d.end_time:
                    end_dt = datetime.datetime.strptime(
                        d.end_time, '%Y-%m-%d %H:%M:%S')
                    if ignore_time:
                        is_expired = end_dt.date() < now.date()
                    else:
                        is_expired = end_dt < now
            except (ValueError, TypeError):
                is_expired = False

            try:
                if d.start_time:
                    start_dt = datetime.datetime.strptime(
                        d.start_time, '%Y-%m-%d %H:%M:%S')
                    if ignore_time:
                        is_not_started = now.date() < start_dt.date()
                    else:
                        is_not_started = now < start_dt
            except (ValueError, TypeError):
                is_not_started = False

            if is_expired or is_not_started:
                continue

            # 只有当需要自动生成，或者任务本身已有路径时，才加入队列
            if auto_generate or d.run_coords:
                tasks_to_run.append(i)

        if not tasks_to_run:
            msg = "没有符合条件的可执行任务。"
            if not auto_generate:
                msg += " (提示: 勾选'全部自动生成'可以为没有路线的任务自动规划)"
            return {"success": False, "message": msg}

        queue = collections.deque(tasks_to_run)
        self.stop_run_flag.clear()
        logging.info(
            f"Starting 'run all' process. Queue={list(queue)}, auto-generate={auto_generate}")
        threading.Thread(target=self._run_all_tasks_manager, args=(
            queue, auto_generate), daemon=True).start()
        return {"success": True}

    def _run_all_tasks_manager(self, queue: collections.deque, auto_gen_enabled: bool):
        """管理“执行所有”任务队列的线程函数"""
        self.log("开始执行所有任务。")
        if self.window:
            self.window.evaluate_js('onAllRunsToggled(true)')
        is_first_task = True

        while queue:
            if self.stop_run_flag.is_set():
                break

            idx = queue.popleft()
            run_data = self.all_run_data[idx]
            self.current_run_idx = idx

            if self.window:
                self.window.evaluate_js(f'selectTaskFromBackend({idx})')
            time.sleep(1.2)

            if not run_data.run_coords and auto_gen_enabled:
                self.log(f"正在为任务 '{run_data.run_name}' 自动生成路线...")
                logging.info(
                    f"Auto-generating path for task in 'run all' mode: {run_data.run_name}")

                # 1) 确保已加载任务详情，拿到打卡点
                if not run_data.details_fetched:
                    details_resp = self.get_task_details(idx)
                    if not details_resp.get("success"):
                        self.log("获取详情失败，跳过。")
                        logging.warning(
                            f"Skipping task {run_data.run_name}: failed to get details.")
                        continue
                    run_data = self.all_run_data[idx]

                if not run_data.target_points:
                    self.log(f"任务 '{run_data.run_name}' 无打卡点，无法自动生成，跳过。")
                    logging.warning(
                        f"Skipping task {run_data.run_name}: no target points for auto-generation.")
                    continue

                # 2) 触发前端JS路径规划，使用与多账号相同的回调机制
                try:
                    self.log("调用高德API进行路径规划...")
                    # 准备回调通道
                    callback_key = f"single_{idx}_{int(time.time() * 1000)}"
                    path_result: dict = {}
                    completion_event = threading.Event()  # 使用绿化后的 Event
                    self.path_gen_callbacks[callback_key] = (
                        path_result, completion_event)

                    # 将打卡点传给前端；JS会调用 getWalkingPath 并通过 multi_path_generation_callback 回传
                    # 形如 [(lon, lat), ...]，AMap.Walking.search 支持 [lng, lat] 数组
                    waypoints = run_data.target_points
                    if self.window:
                        self.window.evaluate_js(
                            f'triggerPathGenerationForPy("{callback_key}", {json.dumps(waypoints)})'
                        )

                    # 等待JS完成路径规划
                    path_received = completion_event.wait(timeout=120)
                    if 'path' not in path_result:
                        error_msg = path_result.get('error', '超时或未知错误')
                        self.log(f"路径规划失败或超时：{error_msg}，跳过此任务。")
                        logging.warning(
                            f"Path planning failed for task {run_data.run_name}: {error_msg}")
                        if callback_key in self.path_gen_callbacks:
                            self.path_gen_callbacks.pop(callback_key, None)
                        continue

                    # 形如 [{lng,lat}, ...]
                    api_path_coords = path_result['path']
                    self.log(f"路径规划成功，共 {len(api_path_coords)} 个点，正在生成模拟数据...")

                    # 3) 使用后端生成 run_coords（复用已有逻辑与参数）
                    p = self.params
                    gen_resp = self.auto_generate_path_with_api(
                        api_path_coords,
                        p.get("min_time_m", 20),
                        p.get("max_time_m", 30),
                        p.get("min_dist_m", 2000)
                    )
                    if not gen_resp.get("success"):
                        self.log(
                            f"自动生成失败：{gen_resp.get('message', '未知错误')}，跳过。")
                        logging.warning(
                            f"Auto-generation failed for task {run_data.run_name}: {gen_resp}")
                        continue

                    # 将生成结果回填到当前任务
                    run_data.run_coords = gen_resp["run_coords"]
                    run_data.total_run_distance_m = gen_resp["total_dist"]
                    run_data.total_run_time_s = gen_resp["total_time"]

                except Exception as e:
                    self.log(f"自动生成失败，跳过：{e}")
                    logging.error(
                        f"Auto-generation failed for {run_data.run_name}: {e}", exc_info=True)
                    continue

                # 路径生成成功后，通知前端刷新UI以显示新路径和数据
                self.log("路径已生成，正在更新界面...")
                if self.window:
                    self.window.evaluate_js(f'forceRefreshTaskUI({idx})')
                time.sleep(1.0)  # 短暂延时，让用户能看到UI更新

            if not run_data.run_coords:
                self.log(f"任务 '{run_data.run_name}' 无可用路线，跳过。")
                logging.warning(
                    f"Skipping task {run_data.run_name}: no route available.")
                continue

            if not is_first_task:
                wait_time = random.uniform(
                    self.params['task_gap_min_s'], self.params['task_gap_max_s'])
                self.log(f"任务间等待中...")
                logging.info(f"等待 {wait_time:.1f}秒 后开始执行下一个任务")
                if self.stop_run_flag.wait(timeout=wait_time):
                    break
            is_first_task = False

            run_data.target_sequence, run_data.is_in_target_zone = 0, False  # ✓ 从0开始
            self._first_center_done = False
            task_finished_event = threading.Event()  # 使用绿化后的 Event
            self._run_submission_thread(
                run_data, idx, self.api_client, True, task_finished_event)
            task_finished_event.wait()

        self.log("所有任务执行结束。")
        self.stop_run_flag.set()
        if self.window:
            self.window.evaluate_js('onAllRunsToggled(false)')

    def get_task_history(self, index):
        """获取任务的历史跑步记录"""
        logging.info(f"API调用: get_task_history - 获取任务历史跑步记录，任务索引: {index}")
        if not (0 <= index < len(self.all_run_data)):
            return {"success": False, "message": "任务索引无效"}
        run_data = self.all_run_data[index]
        self.log("正在获取历史记录...")
        logging.debug(f"正在获取任务的历史记录: 任务名称={run_data.run_name}")
        resp = self.api_client.get_run_history_list(
            self.user_data.id, run_data.errand_schedule)

        if resp and resp.get('success'):
            records = resp.get('data', {}).get('userErrandTrackRecord', [])
            history_list = []
            for rec in records:
                run_time_s = rec.get('runTime', 0) / 1000
                length_m = rec.get('runLength', 0)
                km = length_m / 1000 if length_m else 0
                speed_s_per_km = (run_time_s / km) if km > 0 else 0
                s_min, s_sec = divmod(speed_s_per_km, 60)
                total_seconds_int = int(run_time_s)
                run_minutes, run_seconds = divmod(total_seconds_int, 60)
                formatted_used_time = f"{run_minutes:02d}:{run_seconds:02d}"
                history_list.append({
                    "time": rec.get('createTime', 'NULL'),
                    "used_time": formatted_used_time,
                    "len": f"{length_m}m",
                    "speed": f"{int(s_min)}'{int(s_sec):02d}\"" if km > 0 else "NULL",
                    "trid": rec.get('trid')
                })
            self.log(f"历史记录已加载。")
            logging.debug(f"已成功加载 {len(records)} 条历史记录")
            return {"success": True, "history": history_list}
        return {"success": False, "message": "获取历史记录失败"}

    def get_historical_track(self, trid):
        """根据轨迹ID获取历史轨迹坐标点"""
        logging.info(f"API调用: get_historical_track - 根据轨迹ID获取历史轨迹，trid: {trid}")
        self.log("正在加载历史轨迹...")
        logging.debug(f"正在加载历史运动轨迹数据，轨迹ID: trid={trid}")
        resp = self.api_client.get_history_track_by_trid(trid)
        if resp and resp.get('success'):
            coords = []
            for track_point_list in resp.get('data', {}).get('trackPointList', []):
                try:
                    coords_list = json.loads(
                        track_point_list.get('coordinate', '[]'))
                    for item in coords_list:
                        lon, lat = map(float, item.get(
                            'location', '0,0').split(','))
                        coords.append((lon, lat))
                except (json.JSONDecodeError, ValueError):
                    continue
            self.log("历史轨迹加载成功。")
            logging.debug(
                f"历史轨迹数据加载成功，包含 {len(coords)} 个坐标点")
            return {"success": True, "coords": coords}
        self.log("加载历史轨迹失败。")
        logging.warning("加载历史运动轨迹数据失败")
        return {"success": False, "message": "加载历史轨迹失败"}

    def open_file_dialog(self, dialog_type, options):
        """打开系统文件对话框（Web模式不支持，返回错误）"""
        logging.info(f"API调用: open_file_dialog - 打开文件对话框，类型: {dialog_type}")
        # Web模式下无法使用文件对话框
        logging.error("文件对话框在Web模式下不可用")
        return None

    def show_confirm_dialog(self, title, message):
        """(已修复) 由JS调用，显示一个基于HTML的确认对话框(是/否)"""
        logging.info(f"API调用: show_confirm_dialog - 显示确认对话框，标题: {title}")

        if not self.window:
            logging.error(
                "窗口对象未设置，无法显示确认对话框")
            return False  # 无法显示，默认返回 "否"

        try:
            # 调用JS函数 (jsShowConfirm)，并等待其返回的 Promise
            # JS中的 jsShowConfirm 必须返回一个 Promise<boolean>
            js_code = f'jsShowConfirm({json.dumps(title)}, {json.dumps(message)})'

            # evaluate_js 会阻塞Python线程，直到JS的Promise解析
            result = self.window.evaluate_js(js_code)

            # 确保返回的是布尔值
            return bool(result)

        except Exception as e:
            # 如果JS函数不存在、JS执行出错或返回了非预期值
            logging.error(
                f"从JavaScript显示/获取确认对话框时出错: {e}", exc_info=True)
            # Web模式下无法使用tkinter回退
            return False

    def update_param(self, key, value):
        """更新并保存单个参数"""
        logging.info(f"API调用: update_param - 更新参数，键: {key}, 值: {value}")

        # 决定要更新哪个参数字典
        username_to_update = None
        if self.is_multi_account_mode:
            # 在多账号模式下，这被视为更新全局参数
            target_params = self.global_params
        else:
            target_params = self.params
            username_to_update = self.user_data.username

        if key in target_params:
            try:
                original_type = type(target_params[key])
                if original_type is bool:
                    target_params[key] = bool(value) if isinstance(
                        value, bool) else str(value).lower() in ('true', '1', 't', 'yes')
                else:
                    target_params[key] = original_type(value)

                # 如果是多账号模式, 更新所有已加载账号的对应参数
                if self.is_multi_account_mode:
                    for acc in self.accounts.values():
                        if key in acc.params:
                            acc.params[key] = target_params[key]
                            # 仅当该账号已有 .ini 时才写回，避免无用文件生成
                            ini_path = os.path.join(
                                self.user_dir, f"{acc.username}.ini")
                            if os.path.exists(ini_path):
                                self._save_config(acc.username)

                # 获取最后登录的用户名以保存配置
                if not username_to_update:
                    cfg = configparser.ConfigParser()
                    cfg.read(self.config_path, encoding='utf-8')
                    username_to_update = cfg.get(
                        'Config', 'LastUser', fallback=None)

                if username_to_update and not self.is_multi_account_mode:
                    self._save_config(username_to_update)

                logging.debug(f"参数已更新: 参数名={key}, 新值={target_params[key]}")

                # --- 响应自动签到参数变化 ---
                if key in ("auto_attendance_enabled", "auto_attendance_refresh_s") and not self.is_multi_account_mode:
                    self.stop_auto_refresh.set()  # 停止旧的
                    if self.auto_refresh_thread and self.auto_refresh_thread.is_alive():
                        self.auto_refresh_thread.join(timeout=1.0)

                    if self.params.get("auto_attendance_enabled", False):
                        # 如果是启用，则重启
                        self.stop_auto_refresh.clear()
                        self.auto_refresh_thread = threading.Thread(
                            target=self._auto_refresh_worker, daemon=True)
                        self.auto_refresh_thread.start()
                        self.log("自动刷新设置已更新并重启。")

                return {"success": True}
            except (ValueError, TypeError) as e:
                return {"success": False, "message": str(e)}
        return {"success": False, "message": "Unknown parameter"}

    def export_task_data(self):
        """导出当前任务数据为JSON文件（Web模式：返回JSON数据让前端下载）"""
        logging.info("API调用: export_task_data - 导出当前任务数据为JSON格式")
        logging.info("导出任务数据...")
        if self.current_run_idx == -1:
            logging.warning("未选择任务，无法导出")
            return {"success": False, "message": "请先选择一个任务。"}
        run_data = self.all_run_data[self.current_run_idx]
        if not run_data.draft_coords and not run_data.run_coords and not run_data.recommended_coords:
            logging.warning("任务无路径数据，无法导出")
            return {"success": False, "message": "当前任务没有可导出的路径数据。"}

        export_data = {
            "task_name": run_data.run_name, "errand_id": run_data.errand_id,
            "errand_schedule": run_data.errand_schedule, "target_points": run_data.target_points,
            "target_point_names": run_data.target_point_names, "recommended_coords": run_data.recommended_coords,
            "draft_coords (gps)": run_data.draft_coords, "run_coords (gps)": run_data.run_coords
        }

        # Web模式：返回数据让前端处理下载
        try:
            logging.info(f"导出任务数据成功: {run_data.run_name}")
            return {
                "success": True,
                "data": export_data,
                "filename": f"task_{run_data.errand_schedule or 'debug'}_{int(time.time())}.json",
                "message": "任务数据已准备完成"
            }
        except Exception as e:
            logging.error(f"导出失败: {e}", exc_info=True)
            return {"success": False, "message": f"导出失败: {e}"}

    def import_task_data(self, json_data=None):
        """导入JSON任务数据，进入离线调试模式（UA=Null，保留用户信息）"""
        logging.info("API调用: import_task_data - 导入JSON任务数据（离线调试模式）")
        logging.info("开始导入任务数据...")

        if not json_data:
            logging.error("导入失败：未提供有效的JSON数据")
            return {"success": False, "message": "未提供导入数据"}

        try:
            # 如果json_data是字符串，解析它
            if isinstance(json_data, str):
                logging.info("解析JSON字符串...")
                data = json.loads(json_data)
            else:
                data = json_data

            logging.info(f"JSON数据解析成功，任务名称: {data.get('task_name', '未知')}")

            # --- 软重置：保留用户信息，但 UA 强制置空 ---
            prev_user = copy.deepcopy(getattr(self, 'user_data', UserData()))

            self.is_offline_mode = True
            logging.info("切换到离线模式")
            # 停止任何运行中的单账号任务
            try:
                if hasattr(self, 'stop_run_flag') and isinstance(self.stop_run_flag, threading.Event):
                    self.stop_run_flag.set()
                    logging.info("停止运行中的任务")
            except Exception:
                pass
            # 清空运行态
            self.all_run_data = []
            self.current_run_idx = -1
            self._first_center_done = False

            # 恢复用户信息
            self.user_data = prev_user if prev_user else UserData()
            if not (self.user_data.name or "").strip():
                self.user_data.name = "离线调试"
            if not (self.user_data.student_id or "").strip():
                self.user_data.student_id = "NULL"

            logging.info(
                f"用户信息: {self.user_data.name} ({self.user_data.student_id})")

            # 离线模式下 UA 必须为 NULL
            self.device_ua = None

            # 组装离线任务
            debug_run = RunData()
            debug_run.run_name = data.get("task_name", "调试任务 (离线)")
            debug_run.errand_id = data.get("errand_id", "debug_id")
            debug_run.errand_schedule = data.get(
                "errand_schedule", "debug_schedule")
            debug_run.target_points = data.get("target_points", [])
            if "target_point_names" in data and data["target_point_names"]:
                debug_run.target_point_names = data["target_point_names"]
            else:
                num_points = len(debug_run.target_points)
                if num_points > 0:
                    debug_run.target_point_names = "|".join(
                        [f"打卡点 {i + 1}" for i in range(num_points)])

            debug_run.recommended_coords = data.get("recommended_coords", [])
            debug_run.draft_coords = data.get("draft_coords (gps)", [])
            debug_run.run_coords = data.get("run_coords (gps)", [])

            if debug_run.run_coords and len(debug_run.run_coords) > 1:
                total_dist_m = 0.0
                for i in range(len(debug_run.run_coords) - 1):
                    p1 = debug_run.run_coords[i]
                    p2 = debug_run.run_coords[i + 1]
                    total_dist_m += self._calculate_distance_m(
                        p1[0], p1[1], p2[0], p2[1])
                total_time_s = sum(p[2] for p in debug_run.run_coords) / 1000.0
                debug_run.total_run_distance_m = total_dist_m
                debug_run.total_run_time_s = total_time_s
                logging.info(
                    f"导入路径统计完成: 距离={total_dist_m:.1f}米, 时间={total_time_s:.1f}秒")

            debug_run.details_fetched = True
            self.all_run_data = [debug_run]
            self.current_run_idx = 0

            self.log("离线数据已导入。")
            logging.info(f"离线数据导入成功: {debug_run.run_name}")

            tasks_for_js = [r.__dict__.copy() for r in self.all_run_data]
            tasks_for_js[0]['info_text'] = "离线"
            tasks_for_js[0]['target_range_m'] = self.target_range_m

            return {
                "success": True,
                "tasks": tasks_for_js,
                "userInfo": self._get_full_user_info_dict()
            }
        except Exception as e:
            self.log("导入失败。")
            logging.error(f"导入失败: {e}", exc_info=True)
            return {"success": False, "message": f"导入失败: {e}"}

    def clear_current_task_draft(self):
        """清除当前任务的草稿路径和已生成路径"""
        logging.info("API调用: clear_current_task_draft - 清除当前任务的草稿路径和生成路径")
        if self.current_run_idx == -1 or not (0 <= self.current_run_idx < len(self.all_run_data)):
            return {"success": False, "message": "未选择任务"}
        run = self.all_run_data[self.current_run_idx]
        run.draft_coords = []
        run.run_coords = []
        run.total_run_distance_m = 0
        run.total_run_time_s = 0
        logging.info(f"已清除任务的草稿路径和运行数据: 任务名称={run.run_name}")
        return {"success": True}

    def enter_multi_account_mode(self):
        """切换到多账号模式（增强：先打断单账号运行）"""
        # 先停止单账号任何运行中的任务/线程
        try:
            if hasattr(self, 'stop_run_flag') and isinstance(self.stop_run_flag, threading.Event):
                self.stop_run_flag.set()
            # 清空路径规划回调，避免误回传
            for key, (path_result, completion_event) in list(self.path_gen_callbacks.items()):
                path_result['error'] = '模式切换（进入多账号）已取消'
                try:
                    completion_event.set()
                except Exception:
                    pass
            self.path_gen_callbacks.clear()
            # 通知前端停止单账号UI（可选）
            if self.window:
                try:
                    self.window.evaluate_js('onRunStopped()')
                except Exception:
                    pass
        except Exception:
            logging.debug(
                "enter_multi_account_mode pre-stop single failed (non-fatal).")

        self.log("进入多账号模式。")
        self.is_multi_account_mode = True
        self._update_multi_global_buttons()

        try:
            self.stop_multi_auto_refresh.clear()
            if self.multi_auto_refresh_thread is None or not self.multi_auto_refresh_thread.is_alive():
                self.multi_auto_refresh_thread = threading.Thread(
                    target=self._multi_auto_attendance_worker,
                    daemon=True
                )
                self.multi_auto_refresh_thread.start()
        except Exception as e:
            self.log(f"启动多账号自动刷新线程失败: {e}")

        return {"success": True, "params": self.global_params}

    def exit_multi_account_mode(self):
        """退出多账号模式，返回单用户登录页（增强：彻底停止所有执行）"""
        # 发送全局停止并终止所有账号线程
        try:
            self.multi_stop_all_accounts()
            # 等待线程尽量退出
            for acc in list(self.accounts.values()):
                try:
                    if acc.worker_thread and acc.worker_thread.is_alive():
                        acc.stop_event.set()
                        acc.worker_thread.join(timeout=1.0)
                except Exception:
                    pass
        except Exception:
            logging.debug(
                "exit_multi_account_mode: stop all accounts failed (non-fatal).")

        # 清空路径规划回调，避免延迟回传
        for key, (path_result, completion_event) in list(self.path_gen_callbacks.items()):
            path_result['error'] = '模式切换（退出多账号）已取消'
            try:

                completion_event.set()
            except Exception:
                pass
        self.path_gen_callbacks.clear()

        try:
            self.stop_multi_auto_refresh.set()
            if self.multi_auto_refresh_thread and self.multi_auto_refresh_thread.is_alive():
                self.multi_auto_refresh_thread.join(timeout=1.0)
            self.multi_auto_refresh_thread = None
        except Exception as e:
            logging.warning(f"停止多账号自动刷新线程失败: {e}")

        # 重置所有状态（包含 accounts 清理）
        self._init_state_variables()
        self._load_global_config()
        self.log("已退出多账号模式。")
        return {"success": True}

    def enter_single_account_mode(self):
        """进入单账号模式（增强：先停止多账号运行）"""
        # 如果有多账号线程，先全停
        try:
            self.multi_stop_all_accounts()
            # 等待线程尽量退出
            for acc in list(getattr(self, 'accounts', {}).values()):
                try:
                    if acc.worker_thread and acc.worker_thread.is_alive():
                        acc.stop_event.set()
                        acc.worker_thread.join(timeout=1.0)
                except Exception:
                    pass
        except Exception:
            logging.debug(
                "enter_single_account_mode: stop multi failed (non-fatal).")

        # 清空路径规划回调
        for key, (path_result, completion_event) in list(self.path_gen_callbacks.items()):
            path_result['error'] = '模式切换（进入单账号）已取消'
            try:
                completion_event.set()
            except Exception:
                pass
        self.path_gen_callbacks.clear()

        self.is_multi_account_mode = False
        self.log("进入单账号模式。")
        return {"success": True}

    def get_session_mode_info(self):
        """获取会话模式信息（单账号/多账号），用于页面刷新时恢复状态"""
        mode_info = {
            "success": True,
            "is_multi_account_mode": getattr(self, 'is_multi_account_mode', False),
            "school_account_logged_in": getattr(self, 'login_success', False),
            "is_offline_mode": getattr(self, 'is_offline_mode', False),
        }
        
        # 如果是多账号模式，返回更多信息
        if getattr(self, 'is_multi_account_mode', False):
            mode_info["multi_account_count"] = len(getattr(self, 'accounts', {}))
            mode_info["multi_account_usernames"] = list(getattr(self, 'accounts', {}).keys())
            mode_info["global_params"] = getattr(self, 'global_params', {})
        else:
            # 单账号模式信息
            mode_info["has_tasks"] = len(getattr(self, 'all_run_data', [])) > 0
            mode_info["task_count"] = len(getattr(self, 'all_run_data', []))
            mode_info["selected_task_index"] = getattr(self, 'current_run_idx', -1)
            
            # 保存用户数据（用于离线模式恢复）
            if hasattr(self, 'user_data') and self.user_data:
                user_data = self.user_data
                mode_info["user_data"] = {
                    'name': getattr(user_data, 'name', ''),
                    'phone': getattr(user_data, 'phone', ''),
                    'student_id': getattr(user_data, 'student_id', ''),
                    'id': getattr(user_data, 'id', ''),
                    'username': getattr(user_data, 'username', ''),
                    'gender': getattr(user_data, 'gender', ''),
                    'school_name': getattr(user_data, 'school_name', '')
                }
            
        return mode_info


    def multi_get_all_config_users(self):
        """获取所有存在配置文件的用户列表，用于前端便捷添加"""
        users = sorted([os.path.splitext(f)[0]
                       for f in os.listdir(self.user_dir) if f.endswith(".ini")])
        return {"users": users}

    def multi_load_accounts_from_config(self):
        """模式一：从所有.ini配置文件加载账号"""
        self.log("正在从配置文件加载账号列表...")
        users = sorted([os.path.splitext(f)[0]
                       for f in os.listdir(self.user_dir) if f.endswith(".ini")])
        loaded_count = 0
        for username in users:
            if username not in self.accounts:
                self.multi_add_account(username, "")
                loaded_count += 1
        self.log(f"已加载 {loaded_count} 个账号。")
        self._update_multi_global_buttons()

        # # 批量加载后立即触发一次“刷新全部”，便于自动拉取状态与汇总
        # try:
        #     self.multi_refresh_all_statuses()
        # except Exception:
        #     logging.error(f"触发刷新全部失败: {traceback.format_exc()}")

        return self.multi_get_all_accounts_status()

    def multi_add_account(self, username, password):
        """模式二：手动或选择性添加账号"""
        if username in self.accounts:
            # 如果账号已存在，智能处理密码更新或从文件刷新
            acc = self.accounts[username]

            # 场景1: 提供了新的、非空的密码 (来自导入或手动输入)。
            # 则更新内存和文件中的密码。
            if password and acc.password != password:
                acc.password = password
                self._save_config(username, password)
                self.log(f"已更新账号 [{username}] 的密码。")
            # 场景2: 未提供密码 (来自下拉框重复添加)。
            # 此时从.ini文件实时重新加载密码，确保使用的是最新的（例如刚刚导入的）。
            else:
                self.log(f"账号 [{username}] 已存在，正在从配置文件刷新...")
                reloaded_password = self._load_config(username)
                if reloaded_password:
                    acc.password = reloaded_password

            # 统一执行后续的强制刷新逻辑，以新密码或从文件加载的最新密码来重新验证
            try:
                # UI反馈为“刷新中…”
                self._update_account_status_js(acc, status_text="刷新中...")
                # 清理会话与登录态，强制走登录分支
                try:
                    acc.api_client.session.cookies.clear()
                except Exception:
                    pass
                acc.user_data.id = ""  # 标记为未登录，从而在刷新线程中执行登录流程
                # 启动刷新线程
                threading.Thread(
                    target=self._multi_refresh_worker,
                    args=(acc,),
                    daemon=True
                ).start()
            except Exception:
                logging.error(f"更新账号后触发刷新失败: {traceback.format_exc()}")
            # 刷新全局按钮状态并返回最新账号状态列表
            self._update_multi_global_buttons()
            return self.multi_get_all_accounts_status()

        # 创建账号会话（注意：若稍后判定密码为空，将撤销）
        self.accounts[username] = AccountSession(username, password, self)

        # 尝试从 .ini 加载 UA/参数/密码（如果存在）
        loaded_password = self._load_config(username)

        # 如果加载不到UA，则生成一个新的
        if not self.accounts[username].device_ua:
            self.accounts[username].device_ua = ApiClient.generate_random_ua()

        # 最终确定密码：优先使用传入的新密码；否则使用 .ini 中的密码；否则空
        final_password = password or (loaded_password or "")
        self.accounts[username].password = final_password

        # 若最终密码仍为空，则弹出前端模态框让用户补充密码，并撤销本次添加
        if not final_password:
            # 撤销创建的会话，避免半添加状态
            try:
                del self.accounts[username]
            except Exception:
                pass
            self.log(f"账号 {username} 缺少密码，已弹出输入框以完善密码。")
            if self.window:
                # 复用已有的“新用户”模态框，预填用户名
                js = (
                    f'openNewUserModal();'
                    f'document.getElementById("newUsername").value = {json.dumps(username)};'
                    f'document.getElementById("newPassword").value = "";'
                )
                try:
                    self.window.evaluate_js(js)
                except Exception:
                    logging.debug("打开新用户模态框失败（非致命错误）")
            # 返回当前状态列表（不新增该账号）
            self._update_multi_global_buttons()
            return self.multi_get_all_accounts_status()

        # 若密码已确定，保存到 .ini（避免刷新后丢失）
        ini_path = os.path.join(self.user_dir, f"{username}.ini")
        try:
            # 创建 .ini 或更新现有 .ini，保存 UA 与参数与密码
            # 修复：在多账号模式添加账号时，应该保存该账号的 UA
            self._save_config(
                username, self.accounts[username].password, self.accounts[username].device_ua)
        except Exception:
            logging.warning(
                f"保存 {ini_path} 失败（将继续运行）：{traceback.format_exc()}")

        self.log(f"已添加账号: {username}")
        # 添加账号后立即刷新“全部开始/全部停止”按钮状态
        self._update_multi_global_buttons()
        
        # 修正：使用 SocketIO 向特定会话发送“列表已更新”事件
        session_id = getattr(self, '_web_session_id', None)
        if session_id and socketio:
            try:
                current_accounts = self.multi_get_all_accounts_status().get("accounts", [])
                # 发送一个自定义事件 'accounts_updated'，前端JS需要监听这个事件
                socketio.emit('accounts_updated', {
                    'accounts': current_accounts
                }, room=session_id)
            except Exception as e:
                logging.error(f"SocketIO emit 'accounts_updated' failed: {e}")

        # 为“刚添加的账号”立即启动刷新线程（自动登录 + 拉摘要）
        try:
            threading.Thread(
                target=self._multi_refresh_worker,
                args=(self.accounts[username],),
                daemon=True
            ).start()
        except Exception:
            logging.error(f"启动刷新线程失败: {traceback.format_exc()}")

        return self.multi_get_all_accounts_status()

    def multi_remove_account(self, username):
        """移除一个账号"""
        if username in self.accounts:
            # 如果该账号正在运行，先停止它
            if self.accounts[username].worker_thread and self.accounts[username].worker_thread.is_alive():
                self.accounts[username].stop_event.set()
            del self.accounts[username]
            self.log(f"已移除账号: {username}")
        # 统一刷新多账号全局按钮状态
        self._update_multi_global_buttons()
        return self.multi_get_all_accounts_status()

    def multi_refresh_all_statuses(self):
        """(多线程)刷新所有账号的任务状态和统计信息"""
        if not self.accounts:
            self.log("账号列表为空，无需刷新。")
            return {"success": True}

        self.log("正在刷新所有账号状态...")
        for acc in self.accounts.values():
            is_running = bool(
                acc.worker_thread and acc.worker_thread.is_alive())
            # 仅当账号未在运行时才提示“刷新中...”
            if not is_running:
                self._update_account_status_js(acc, status_text="刷新中...")
            # 运行中的账号走保留状态模式
            threading.Thread(target=self._multi_refresh_worker, args=(
                acc, True if is_running else False), daemon=True).start()

        self._update_multi_global_buttons()
        return {"success": True}

    def multi_refresh_single_status(self, username: str):
        """刷新指定账号的状态与摘要（单账号）"""
        if username not in self.accounts:
            return {"success": False, "message": "账号不存在"}
        acc = self.accounts[username]

        # 若账号正在运行任务，仍允许运行刷新线程，只要刷新线程自身不调用执行逻辑
        try:
            # 立刻标记UI状态
            self._update_account_status_js(acc, status_text="刷新中...")
            # 启动单独刷新线程（复用已有的 _multi_refresh_worker）
            threading.Thread(
                target=self._multi_refresh_worker,
                args=(acc,),
                daemon=True
            ).start()
            return {"success": True}
        except Exception as e:
            logging.error(f"启动单账号刷新失败: {e}", exc_info=True)
            self._update_account_status_js(acc, status_text="刷新出错")
            return {"success": False, "message": "启动刷新失败"}

    def _multi_refresh_worker(self, acc: AccountSession, preserve_status: bool = False):
        """用于刷新的单个账号工作线程
        - preserve_status=True 时：若账号正在运行，不改写 status_text，只更新 name 与 summary
        """
        try:
            # 动态判断是否需要保留状态（仅当账号当前正在运行且调用方要求保留时）
            preserve_now = preserve_status and bool(
                acc.worker_thread and acc.worker_thread.is_alive())

            # 1) 登录（或确保已登录）
            if not acc.user_data.id:
                if not acc.device_ua:
                    acc.device_ua = ApiClient.generate_random_ua()
                login_resp = self._queued_login(acc, respect_global_stop=False)
                if not login_resp or not login_resp.get('success'):
                    # 登录失败：仅在不保留状态时才更新状态文案
                    if not preserve_now:
                        self._update_account_status_js(
                            acc, status_text="刷新失败(登录错误)")
                    return

                data = login_resp.get('data', {})
                user_info = data.get('userInfo', {})
                acc.user_data.name = user_info.get('name', '')
                acc.user_data.id = user_info.get('id', '')
                acc.user_data.student_id = user_info.get('account', '')

                # 写回 UA/参数（如有）
                ini_path = os.path.join(self.user_dir, f"{acc.username}.ini")
                if os.path.exists(ini_path):
                    self._save_config(acc.username)

                # 更新名字（允许）
                self._update_account_status_js(acc, name=acc.user_data.name)

            # 2) 拉取任务并汇总
            self._multi_fetch_and_summarize_tasks(acc)

            # 2.5) 拉取签到统计
            if not preserve_now:  # 仅在非保留状态（即主动刷新）时才获取签到
                self._multi_fetch_attendance_stats(acc)

            # 3) 决定最终状态文本（仅当不保留状态时才改 status_text）
            if not preserve_now:
                exe_cnt = acc.summary.get("executable", 0)
                not_started_cnt = acc.summary.get("not_started", 0)

                if exe_cnt > 0:
                    final_status = "待命"
                    acc.has_pending_tasks = True
                elif not_started_cnt > 0:
                    final_status = "无任务可执行"
                    acc.has_pending_tasks = False
                else:
                    final_status = "无任务可执行"
                    acc.has_pending_tasks = False

                # 如果当前是错误态，则保留，不覆盖
                if self._should_preserve_status(acc.status_text, new_status=final_status):
                    self._update_account_status_js(acc, summary=acc.summary)
                else:
                    self._update_account_status_js(
                        acc, status_text=final_status, summary=acc.summary)

            acc.log("状态刷新完成。")

        except Exception as e:
            logging.error(
                f"Error refreshing {acc.username}: {traceback.format_exc()}")
            # 仅当不保留状态时才显示出错文案
            if not preserve_status:
                self._update_account_status_js(acc, status_text="刷新出错")

    def multi_remove_selected_accounts(self, usernames: list[str]):
        """根据用户名列表移除多个账号"""
        if not usernames:
            return self.multi_get_all_accounts_status()

        removed_count = 0
        for username in usernames:
            if username in self.accounts:
                # 如果账号正在运行，先停止它
                if self.accounts[username].worker_thread and self.accounts[username].worker_thread.is_alive():
                    self.accounts[username].stop_event.set()
                del self.accounts[username]
                removed_count += 1

        self.log(f"移除了 {removed_count} 个选定账号。")
        self._update_multi_global_buttons()
        return self.multi_get_all_accounts_status()

    def multi_remove_all_accounts(self):
        """一键移除所有账号"""
        if not self.accounts:
            return self.multi_get_all_accounts_status()

        # 停止所有正在运行的线程
        self.multi_stop_all_accounts()

        count = len(self.accounts)
        self.accounts.clear()
        self.log(f"已移除全部 {count} 个账号。")
        self._update_multi_global_buttons()
        return self.multi_get_all_accounts_status()

    def multi_get_all_accounts_status(self):
        """获取所有账号的当前状态，用于刷新前端UI"""
        status_list = []
        for acc in self.accounts.values():
            status_list.append({
                "username": acc.username,
                "name": acc.user_data.name or "---",
                "status_text": acc.status_text,
                "summary": acc.summary,
            })
        return {"accounts": status_list}

    def multi_download_import_template(self):
        """下载导入模板（账号、密码），支持 xlsx/xls/csv"""
        # 选择保存位置与格式
        filepath = self.open_file_dialog('save', {
            'initialfile': f"账号导入模板_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            'filetypes': [
                ('Excel 模板 (*.xlsx)', '*.xlsx'),
                ('Excel 97-2003 模板 (*.xls)', '*.xls'),
                ('CSV 模板 (*.csv)', '*.csv'),
                ('所有文件 (*.*)', '*.*')
            ],
            'defaultextension': ".xlsx"
        })
        if not filepath:
            return {"success": False, "message": "用户取消操作"}

        try:
            ext = os.path.splitext(filepath)[1].lower()
            headers = ["账号", "密码"]

            if ext == ".xlsx":
                wb = openpyxl.Workbook()
                sh = wb.active
                sh.title = "模板"
                sh.append(headers)

                wb.save(filepath)
            elif ext == ".xls":
                wb = xlwt.Workbook()
                sh = wb.add_sheet("模板")
                for col, val in enumerate(headers):
                    sh.write(0, col, val)

                wb.save(filepath)
            elif ext == ".csv":

                with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)

            else:
                return {"success": False, "message": f"不支持的模板格式: {ext}"}

            self.log(f"模板已生成：{os.path.basename(filepath)}")
            return {"success": True}
        except Exception as e:
            self.log(f"模板生成失败：{e}")
            logging.error(
                f"Template generation failed: {traceback.format_exc()}")
            return {"success": False, "message": f"生成失败: {e}"}

    def safe_load_workbook(self, filepath, **kwargs):
        if not filepath:
            raise ValueError("未提供有效的文件路径，可能是用户取消了文件选择。")
        if not isinstance(filepath, (str, bytes, os.PathLike)):
            raise TypeError(f"文件路径类型错误: 期望 str/Path，实际 {type(filepath)}")
        try:
            wb = openpyxl.load_workbook(
                filepath,
                data_only=True,
                read_only=True,
                **kwargs
            )
            return wb
        except Exception as e:
            # 增强提示：可能是办公软件兼容性问题
            warnings.warn(
                f"Excel 打开失败，已尝试只读模式: {e}\n"
                f"提示: 请确认已安装兼容的办公软件 (如 Microsoft Office 或新版 WPS)，"
                f"并确保文件未被其他程序占用。"
            )
            return openpyxl.load_workbook(
                filepath,
                data_only=True,
                read_only=True
            )

    def multi_import_accounts(self):
        """从文件导入账号密码，支持 .xlsx/.xls/.csv"""
        filepath = self.open_file_dialog('open', {
            'filetypes': [
                ('Excel 文件 (*.xlsx;*.xls)', '*.xlsx;*.xls'),
                ('CSV 文件 (*.csv)', '*.csv'),
                ('所有文件 (*.*)', '*.*')
            ]
        })
        if not filepath:
            return {"success": False, "message": "用户取消操作"}

        try:
            imported = 0
            ext = os.path.splitext(filepath)[1].lower()
            seen_usernames: set[str] = set()  # 本次导入会话内去重，避免同一文件重复账号多次导入

            if ext == ".xlsx":
                # 使用只读模式安全加载工作簿
                try:
                    wb = self.safe_load_workbook(filepath, keep_links=False)
                except TypeError as e:
                    # 针对个别环境样式解析异常的兜底
                    warnings.warn(f"样式解析失败，已忽略样式: {e}")
                    wb = self.safe_load_workbook(
                        filepath, keep_links=False, keep_vba=False)

                sh = wb.active
                skipped_no_password = []  # 收集无密码且 .ini 无密码的账号
                # 从第2行开始，默认第1行为表头，避免重复处理与表头导入
                for row in sh.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 1:
                        continue
                    username = str(row[0] or '').strip()
                    password = str(row[1] or '').strip() if len(
                        row) > 1 else ''
                    if not username:
                        continue
                    # 本次导入内去重
                    if username in seen_usernames:
                        continue
                    seen_usernames.add(username)

                    # 1. 尝试从 .ini 兜底密码
                    loaded_password = self._load_config(username)
                    # 2. 确定最终密码
                    final_password = password or (loaded_password or "")

                    if not final_password:
                        # 3. 只有当导入文件和.ini都没有密码时，才跳过
                        skipped_no_password.append(username)
                        continue  # 跳过本账号，留给用户手动补密码

                    # 4. 使用 final_password 添加
                    self.multi_add_account(username, final_password)
                    imported += 1

            elif ext == ".xls":
                book = xlrd.open_workbook(filepath)
                sh = book.sheet_by_index(0)
                skipped_no_password = []
                for r in range(1, sh.nrows):
                    username = str(sh.cell_value(r, 0) or '').strip()
                    password = ""
                    if sh.ncols >= 2:
                        password = str(sh.cell_value(r, 1) or '').strip()
                    if username:
                        if username in seen_usernames:
                            continue
                        seen_usernames.add(username)

                        loaded_password = self._load_config(username)
                        final_password = password or (loaded_password or "")
                        if not final_password:
                            skipped_no_password.append(username)
                            continue

                        self.multi_add_account(username, final_password)
                        imported += 1

            elif ext == ".csv":
                skipped_no_password = []
                with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
                    reader = csv.reader(f)
                    first_row = True
                    for row in reader:
                        if not row or len(row) < 1:
                            continue
                        # 跳过可能的表头
                        if first_row and str(row[0]).strip() in ("账号", "用户名", "username", "user"):
                            first_row = False
                            continue
                        first_row = False
                        username = (row[0] or '').strip()
                        password = (row[1] or '').strip() if len(
                            row) > 1 else ''
                        if username:
                            if username in seen_usernames:
                                continue
                            seen_usernames.add(username)

                            loaded_password = self._load_config(username)
                            final_password = password or (
                                loaded_password or "")
                            if not final_password:
                                skipped_no_password.append(username)
                                continue

                            self.multi_add_account(username, final_password)
                            imported += 1

            else:
                return {"success": False, "message": f"不支持的导入格式: {ext}"}

            # 导入汇总日志
            self.log(f"成功导入 {imported} 个账号。")
            if skipped_no_password:
                # 前端弹窗提醒需要手动补密码
                msg = "以下账号缺少密码，已跳过导入。请在多账号控制台手动添加并输入密码：\n" + \
                    "\n".join(skipped_no_password[:20])
                # 提示过长时仅显示前 20 条，避免 UI 过载
                try:
                    if self.window:
                        self.window.evaluate_js(f'alert({json.dumps(msg)})')
                except Exception:
                    logging.debug(
                        "Alert skipped_no_password failed (non-fatal).")
                self.log(
                    f"缺少密码的账号（共 {len(skipped_no_password)}）：{', '.join(skipped_no_password)}")

            return self.multi_get_all_accounts_status()

        except Exception as e:
            self.log(f"导入失败: {e}")
            logging.error(f"导入多账号配置失败，详细错误信息: {traceback.format_exc()}")
            return {"success": False, "message": f"导入失败: {e}"}

    def multi_export_accounts_summary(self):
        """导出多账号汇总，支持 .xlsx/.xls/.csv"""
        filepath = self.open_file_dialog('save', {
            'initialfile': f"跑步任务汇总_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
            'filetypes': [
                ('Excel 文件 (*.xlsx)', '*.xlsx'),
                ('Excel 97-2003 文件 (*.xls)', '*.xls'),
                ('CSV 文件 (*.csv)', '*.csv'),
                ('所有文件 (*.*)', '*.*')
            ],
            'defaultextension': ".xlsx"
        })
        if not filepath:
            return {"success": False, "message": "用户取消操作"}

        try:

            headers = ["账号", "姓名", "状态", "总任务数",
                       "已完成完成", "未开始任务数", "可执行任务数", "已过期任务数"]

            rows = []
            for acc in sorted(self.accounts.values(), key=lambda x: x.username):
                s = acc.summary
                rows.append([
                    acc.username,
                    acc.user_data.name or "---",
                    acc.status_text,
                    s.get('total', 0),
                    s.get('completed', 0),
                    s.get('not_started', 0),  # <- 使用 not_started
                    s.get('executable', 0),
                    s.get('expired', 0)
                ])

            ext = os.path.splitext(filepath)[1].lower()

            if ext == ".xlsx":
                wb = openpyxl.Workbook()
                sh = wb.active
                sh.title = "任务汇总"
                sh.append(headers)
                for r in rows:
                    sh.append(r)
                # 自动列宽
                for col in sh.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        v = '' if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    sh.column_dimensions[col_letter].width = max_len + 2
                wb.save(filepath)

            elif ext == ".xls":
                wb = xlwt.Workbook()
                sh = wb.add_sheet("任务汇总")
                # 写表头
                for c, v in enumerate(headers):
                    sh.write(0, c, v)
                # 写数据
                for r_idx, r in enumerate(rows, start=1):
                    for c_idx, v in enumerate(r):
                        sh.write(r_idx, c_idx, v)
                wb.save(filepath)

            elif ext == ".csv":

                with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    writer.writerows(rows)

            else:
                return {"success": False, "message": f"不支持的导出格式: {ext}"}

            self.log(f"汇总信息已导出到 {os.path.basename(filepath)}")
            return {"success": True}
        except Exception as e:
            self.log(f"导出失败: {e}")
            logging.error(
                f"Export accounts summary failed: {traceback.format_exc()}")
            return {"success": False, "message": f"导出失败: {e}"}

    def multi_get_account_params(self, username):
        """
        获取指定账号的参数配置。
        
        功能说明：
        在多账号模式下，获取单个账号的完整参数配置字典。
        用于Web界面显示账号配置或进行参数修改前的查询。
        
        参数:
            username (str): 账号用户名
            
        返回:
            dict: 包含以下字段：
                - success (bool): 操作是否成功
                - params (dict): 账号参数字典（成功时）
                - message (str): 错误信息（失败时）
        
        使用示例:
            result = api.multi_get_account_params("user123")
            if result["success"]:
                print(result["params"])
        """
        if username in self.accounts:
            return {"success": True, "params": self.accounts[username].params}
        return {"success": False, "message": "账号不存在"}

    def multi_update_account_param(self, username, key, value):
        """
        更新指定账号的单个参数值。
        
        功能说明：
        动态修改账号的配置参数，支持类型自动转换和配置持久化。
        修改后立即保存到配置文件，确保重启后配置不丢失。
        
        类型转换规则：
        - bool类型：支持多种表示方式（'true', '1', 't', 'yes'等）
        - 其他类型：根据原始参数类型自动转换（int, float, str等）
        
        参数:
            username (str): 账号用户名
            key (str): 参数键名（必须是已存在的参数）
            value: 新的参数值（将自动转换为正确类型）
            
        返回:
            dict: 包含以下字段：
                - success (bool): 操作是否成功
                - message (str): 错误信息（失败时）
        
        异常处理:
            - ValueError: 类型转换失败
            - TypeError: 类型不兼容
        
        ⚠️ 注意事项：
        1. 只能修改已存在的参数，不能添加新参数
        2. 类型必须与原始参数类型兼容
        3. 修改会立即保存到磁盘
        
        使用示例:
            # 修改布尔参数
            api.multi_update_account_param("user1", "auto_run", "true")
            # 修改数值参数
            api.multi_update_account_param("user1", "speed", 1.5)
        """
        if username not in self.accounts:
            return {"success": False, "message": "账号不存在"}

        acc = self.accounts[username]
        target_params = acc.params
        if key in target_params:
            try:
                # 获取原始参数类型，用于类型转换
                original_type = type(target_params[key])
                # 布尔类型特殊处理：支持多种表示方式
                if original_type is bool:
                    target_params[key] = bool(value) if isinstance(
                        value, bool) else str(value).lower() in ('true', '1', 't', 'yes')
                else:
                    # 其他类型：强制转换为原始类型
                    target_params[key] = original_type(value)

                # 保存配置到文件（持久化）
                self._save_config(
                    username, self.accounts[username].password)
                self.log(f"已更新账号 [{username}] 的参数 {key}。")
                return {"success": True}
            except (ValueError, TypeError) as e:
                # 类型转换失败，返回错误
                return {"success": False, "message": str(e)}
        # 参数不存在
        return {"success": False, "message": "Unknown parameter"}

    def multi_start_single_account(self, username, run_only_incomplete: bool = True):
        """
        启动指定账号的任务执行线程。
        
        功能说明：
        在多账号模式下启动单个账号的自动任务执行。
        创建独立的工作线程，不阻塞主线程和其他账号。
        
        执行流程：
        1. 验证账号存在性
        2. 检查是否已在运行（避免重复启动）
        3. 清除停止标志，准备运行
        4. 创建daemon工作线程
        5. 更新UI状态
        
        参数:
            username (str): 要启动的账号用户名
            run_only_incomplete (bool): 是否只执行未完成的任务
                - True: 跳过已完成的任务，只执行新任务或失败任务
                - False: 执行所有任务，包括已完成的（重新执行）
                
        返回:
            dict: 包含以下字段：
                - success (bool): 操作是否成功
                - message (str): 错误信息（失败时）
        
        线程安全：
        - 使用stop_event控制线程生命周期
        - daemon线程确保主程序退出时自动清理
        - multi_run_stop_flag全局停止控制
        
        ⚠️ 注意事项：
        1. 重复启动会被拒绝（返回失败）
        2. 线程异常会被_multi_account_worker内部捕获
        3. 启动后无法修改run_only_incomplete参数
        """
        if username not in self.accounts:
            return {"success": False, "message": "账号不存在"}
        acc = self.accounts[username]
        if acc.worker_thread and acc.worker_thread.is_alive():
            self.log(f"账号 {username} 已在运行中。")
            return {"success": False, "message": "该账号已在运行"}

        acc.stop_event.clear()
        # 进入运行态
        if self.multi_run_stop_flag.is_set():
            self.multi_run_stop_flag.clear()
        # 使用传入的 run_only_incomplete 复选框状态
        acc.worker_thread = threading.Thread(
            target=self._multi_account_worker,
            args=(acc, 0, bool(run_only_incomplete)),
            daemon=True
        )
        acc.worker_thread.start()
        self.log(f"已启动账号: {username}")
        self._update_account_status_js(acc, status_text="已启动")
        self._update_multi_global_buttons()
        return {"success": True}

    def multi_stop_single_account(self, username):
        """
        停止指定账号的任务执行线程。
        
        功能说明：
        向账号的工作线程发送停止信号，优雅地终止任务执行。
        使用异步监视器等待线程完全退出后清理资源。
        
        停止流程：
        1. 验证账号存在性
        2. 检查线程是否在运行
        3. 设置stop_event信号
        4. 启动监视线程等待工作线程退出
        5. 线程退出后更新UI状态为"待命"
        
        设计特点：
        - 非阻塞：使用独立监视线程，不阻塞主线程
        - 优雅停止：通过Event信号，让线程自己退出
        - 资源清理：线程退出后自动清理worker_thread句柄
        - 状态同步：实时更新UI状态显示
        
        监视器(_watch_stop)机制：
        - 每200ms检查一次线程状态
        - 线程退出后立即清理和更新UI
        - daemon线程确保不会阻止程序退出
        - 捕获所有异常，避免监视器崩溃
        
        ⚠️ 注意事项：
        1. 停止是异步的，可能需要几秒钟完成
        2. 如果线程未在运行，返回失败但不报错
        3. 强制停止可能导致任务状态不完整
        """
        acc = self.accounts.get(username)
        if not acc:
            return {"success": False, "message": f"账号 {username} 不存在"}

        if acc.worker_thread and acc.worker_thread.is_alive():
            acc.stop_event.set()
            self.log(f"已向账号 {username} 发送停止信号。")
            self._update_account_status_js(acc, status_text="正在停止...")
            self._update_multi_global_buttons()

            # 监视该账号的工作线程退出，退出后立即刷新UI与按钮状态
            def _watch_stop():
                try:
                    while acc.worker_thread and acc.worker_thread.is_alive():
                        time.sleep(0.2)
                    # 线程已退出，清理句柄并刷新状态
                    acc.worker_thread = None
                    # 若用户主动停止，统一回到“待命”
                    self._update_account_status_js(acc, status_text="待命")
                    self._update_multi_global_buttons()
                except Exception:
                    logging.debug(
                        "Stop watcher encountered a non-fatal error.", exc_info=True)
            threading.Thread(target=_watch_stop, daemon=True).start()

            return {"success": True}
        else:
            self._update_account_status_js(acc, status_text="待命")
            self._update_multi_global_buttons()
            return {"success": False, "message": "该账号未在运行"}

    def multi_start_all_accounts(self, min_delay, max_delay, use_delay, run_only_incomplete):
        """一键启动所有账号的任务"""
        # 账号为空：直接失败
        if not self.accounts:
            self.log("账号列表为空，无法开始。请先添加账号。")
            return {"success": False, "message": "账号列表为空，无法开始。请先添加账号。"}

        # 仅当“所有账号都在运行中”才阻止
        total_accounts = len(self.accounts)
        running_count = sum(
            1 for acc in self.accounts.values()
            if acc.worker_thread and acc.worker_thread.is_alive()
        )
        if total_accounts > 0 and running_count == total_accounts:
            return {"success": False, "message": "任务已在运行中"}

        # 允许为未运行账号继续启动（不再用全局停止标志作为唯一门槛）
        self.log("开始执行所有账号...")
        self._update_multi_global_buttons()
        # 若之前已在运行（multi_run_stop_flag 已清除），无需重复清除；否则清除以表明进入“运行态”
        if self.multi_run_stop_flag.is_set():
            self.multi_run_stop_flag.clear()

        account_list = list(self.accounts.values())
        num_accounts = len(account_list)
        delays = [0] * num_accounts

        if use_delay and num_accounts > 0 and max_delay >= min_delay:
            try:
                delays = [random.uniform(min_delay, max_delay)
                          for _ in range(num_accounts)]
            except ValueError:
                delays = [random.uniform(min_delay, max_delay)
                          for _ in range(num_accounts)]

        random.shuffle(account_list)

        started_threads = 0
        for i, acc in enumerate(account_list):
            # 已在运行的账号跳过，仅启动未运行账号
            if acc.worker_thread and acc.worker_thread.is_alive():
                acc.log("已在运行，本次'全部开始'将跳过此账号。")
                continue

            delay = delays[i] if use_delay else 0
            acc.stop_event.clear()
            acc.worker_thread = threading.Thread(
                target=self._multi_account_worker,
                args=(acc, delay, run_only_incomplete),
                daemon=True
            )
            acc.worker_thread.start()
            started_threads += 1

        # 如果这次没有任何账号被启动（所有账号已在运行或无可执行任务）
        if started_threads == 0:
            # 保持标志位为“运行中”或“停止”的合理状态：
            # - 若此前有运行中的账号，则 multi_run_stop_flag 可能已清除，无需更改；
            # - 若此前标志为停止，则维持停止态，给出用户提示。
            self.log("当前没有可启动的账号任务。")
            return {"success": False, "message": "当前没有可启动的账号任务。"}

        self._update_multi_global_buttons()
        return {"success": True}

    def multi_stop_all_accounts(self):
        """停止所有账号的运行"""
        logging.info("API调用: multi_stop_all_accounts - 停止所有多账号任务的运行")
        # 账号为空的情况，直接提示，无需切状态
        if not self.accounts:
            self.log("当前无账号，无需停止。")
            # 保持一致性：确保全局停止标志为“已停止”
            self.multi_run_stop_flag.set()
            return {"success": False, "message": "当前无账号，无需停止。"}

        if self.multi_run_stop_flag.is_set():
            return {"success": True}

        self.log("正在发送停止信号...")
        self.multi_run_stop_flag.set()
        for acc in self.accounts.values():
            acc.stop_event.set()
        self._update_multi_global_buttons()
        self.log("所有任务将在当前步骤完成后停止。")
        self._update_multi_global_buttons()

        time.sleep(1)

        self._update_multi_global_buttons()
        return {"success": True}

    def _update_account_status_js(self, acc: AccountSession, status_text: str = None, summary: dict = None, name: str = None,
                                      progress_pct: int | None = None, progress_text: str | None = None, progress_extra: str | None = None):
            """一个辅助函数，用于向前端发送状态更新"""
            # 修正：使用 SocketIO 向特定会话发送更新，而不是 self.window
            session_id = getattr(self, '_web_session_id', None)
            if not session_id or not socketio:
                logging.debug(f"跳过账号状态更新（_update_account_status_js），账号: {acc.username}，原因: 缺少会话ID或socketio未初始化")
                return  # 如果没有会话ID或socketio未初始化，则无法发送

            update_data = {}
            if status_text is not None:
                acc.status_text = status_text
                update_data['status_text'] = status_text
            if summary is not None:
                acc.summary = summary
                update_data['summary'] = summary
            if name is not None:
                update_data['name'] = name

            # 进度条更新
            if progress_pct is not None:
                update_data['progress_pct'] = int(progress_pct)
            if progress_text is not None:
                update_data['progress_text'] = progress_text
            if progress_extra is not None:
                update_data['progress_extra'] = progress_extra

            if update_data:
                try:
                    # 发送一个自定义事件到该 session_id 对应的房间
                    socketio.emit('multi_status_update', {
                        'username': acc.username,
                        'data': update_data
                    }, room=session_id)
                except Exception as e:
                    logging.error(f"SocketIO emit 'multi_status_update' failed: {e}")
            
            # _update_multi_global_buttons 也需要用 socketio
            self._update_multi_global_buttons()

    def _update_multi_global_buttons(self):
            """根据当前多账号状态刷新“全部开始/全部停止/返回登录页”按钮的可用性"""
            # 修正：使用 SocketIO 向特定会话发送更新
            session_id = getattr(self, '_web_session_id', None)
            if not session_id or not socketio:
                logging.debug("跳过多账号全局按钮更新（_update_multi_global_buttons），原因: 缺少会话ID或socketio未初始化")
                return

            # 仅统计仍有任务可执行的账号（active）
            active_accounts = []
            for acc in self.accounts.values():
                total = acc.summary.get("total", 0)
                expired = acc.summary.get("expired", 0)
                not_started = acc.summary.get("not_started", 0)
                executable = acc.summary.get("executable", 0)
                # 候选任务：已开始且未截止（无论是否已完成）
                candidates = max(0, total - expired - not_started)
                has_tasks = (executable > 0) if self.multi_run_only_incomplete else (
                    candidates > 0)
                if has_tasks:
                    active_accounts.append(acc)

            total_active = len(active_accounts)
            running_count = sum(
                1 for acc in active_accounts
                if acc.worker_thread and acc.worker_thread.is_alive()
            )

            # 四种状态规则：
            # 1) 无账号 或 无active账号：start=禁用, stop=禁用, exit=启用
            # 2) 有active且不是全部在运行：start=启用, stop=启用, exit=禁用
            # 3) 全部active都在运行：start=禁用, stop=启用, exit=禁用
            if total_active == 0:
                # 无账号或无可执行任务
                start_disabled = True
                stop_disabled = True
                exit_disabled = False
            elif running_count == 0:
                # 有账号但没有任何账号在执行任务
                start_disabled = False
                stop_disabled = True
                exit_disabled = False
            elif running_count == total_active:
                # 全部 active 账号都在运行
                start_disabled = True
                stop_disabled = False
                exit_disabled = True
            else:
                # 部分账号在运行
                start_disabled = False
                stop_disabled = False
                exit_disabled = True

            # 推送到前端
            try:
                socketio.emit('multi_global_buttons_update', {
                    'start_disabled': start_disabled,
                    'stop_disabled': stop_disabled,
                    'exit_disabled': exit_disabled
                }, room=session_id)
            except Exception as e:
                logging.error(f"SocketIO emit 'multi_global_buttons_update' failed: {e}")

    def _queued_login(self, acc: AccountSession, respect_global_stop: bool = True) -> dict | None:
        """
        多账号模式下的“排队登录”（参数化是否尊重全局停止标志）：
        - 使用全局信号量 multi_login_lock 将登录并发固定为 1
        - 在排队等待期间，若收到停止信号（可选尊重 multi_run_stop_flag），立即退出
        - 进入队列前给出UI提示“排队登录...”
        """
        try:
            self._update_account_status_js(acc, status_text="排队登录...")
        except Exception:
            pass

        # 循环尝试获取登录槽位，期间尊重停止事件
        while True:
            # 刷新线程登录不受全局停止影响；任务执行阶段登录继续尊重全局停止
            if acc.stop_event.is_set() or (respect_global_stop and self.multi_run_stop_flag.is_set()):
                return None
            # 尝试以短超时获取锁，便于及时响应停止
            acquired = self.multi_login_lock.acquire(timeout=0.5)
            if acquired:
                break
        try:
            # 二次检查停止（同样按参数决定是否尊重全局停止）
            if acc.stop_event.is_set() or (respect_global_stop and self.multi_run_stop_flag.is_set()):
                return None
            # 执行实际登录请求
            return acc.api_client.login(acc.username, acc.password)
        finally:
            try:
                self.multi_login_lock.release()
            except Exception:
                pass

    def _multi_fetch_and_summarize_tasks(self, acc: AccountSession):
        """(辅助函数) 为单个账号获取任务列表并计算统计信息（按任务ID去重）"""
        acc.all_run_data = []
        seen_keys: set[str] = set()  # 记录已见过的 errandId（与其他处保持一致）
        offset = 0
        dup_count = 0  # 可选：统计重复条数

        while True:
            if acc.stop_event.is_set():
                return
            resp = acc.api_client.get_run_list(acc.user_data.id, offset)
            if not resp or not resp.get('success'):
                acc.log("获取任务列表失败。")
                break

            tasks = resp.get('data', {}).get('errandList', [])
            if not tasks:
                break

            for td in tasks:
                # 构造稳健唯一键（兼容 errandId 缺失 / 同一 errandId 不同 schedule 或时间）
                eid = td.get('errandId') or ""
                es = td.get('errandSchedule') or ""
                st = td.get('startTime') or ""
                et = td.get('endTime') or ""
                unique_key = f"{eid}|{es}|{st}|{et}"

                if unique_key in seen_keys:
                    dup_count += 1
                    continue
                seen_keys.add(unique_key)

                run = RunData()
                run.run_name = td.get('eName')

                # 规范化 status（防止字符串/布尔等造成后续比较错误）
                try:
                    run.status = int(td.get('isExecute') or 0)
                except (TypeError, ValueError):
                    run.status = 1 if str(td.get('isExecute')).strip(
                    ).lower() in ('1', 'true', 'yes') else 0

                run.errand_id = eid
                run.errand_schedule = es
                run.start_time = st
                run.end_time = et
                acc.all_run_data.append(run)

                # 临时调试日志（上线后可删除）
                logging.info(
                    f"[{acc.username}] Parsed task {eid!r}: raw_isExecute={td.get('isExecute')!r} -> status={run.status}")

            offset += len(tasks)
            if len(tasks) < 10:
                break

        if dup_count > 0:
            logging.info(
                f"[{acc.username}] Task de-dup: {dup_count} duplicates skipped (by errandId).")

        # --- 任务分类统计重构 ---
        total = len(acc.all_run_data)
        completed = 0
        expired = 0
        executable = 0
        not_started = 0

        now = datetime.datetime.now()
        ignore_time = acc.params.get("ignore_task_time", True)

        for r in acc.all_run_data:
            # 1. 首先处理已完成的任务
            if r.status == 1:
                completed += 1
                continue  # 已完成的任务无需再进行后续分类

            # 2. 对未完成的任务 (status != 1) 进行分类
            start_dt, end_dt = None, None
            try:
                if r.start_time:
                    start_dt = datetime.datetime.strptime(
                        r.start_time, '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                start_dt = None
            try:
                if r.end_time:
                    end_dt = datetime.datetime.strptime(
                        r.end_time, '%Y-%m-%d %H:%M:%S')
            except (ValueError, TypeError):
                end_dt = None

            # 检查是否过期
            if end_dt:
                is_expired = end_dt.date() < now.date() if ignore_time else end_dt < now
                if is_expired:
                    expired += 1
                    continue

            # 检查是否未开始
            if start_dt:
                is_not_started = now.date() < start_dt.date() if ignore_time else now < start_dt
                if is_not_started:
                    not_started += 1
                    continue

            # 剩余的未完成任务即为可执行
            executable += 1

        acc.summary.update({
            "total": total,
            "completed": completed,
            "pending": total - completed,
            "not_started": not_started,
            "executable": executable,
            "expired": expired
        })

        # 未开始任务仅在摘要中展示，不应影响“全部开始/停止”按钮或地图标记逻辑
        acc.has_pending_tasks = (executable > 0)

    def _multi_fetch_attendance_stats(self, acc: AccountSession):
        """(多账号) 获取单个账号的签到统计"""
        if not acc.user_data.id:
            return  # 尚未登录

        try:
            list_resp = acc.api_client.get_notice_list(
                offset=0, limit=20, type_id=0)
            if not (list_resp and list_resp.get('success')):
                acc.log("获取通知列表失败 (用于签到统计)")
                return

            notices = list_resp.get('data', {}).get('noticeList', [])
            att_pending = 0
            att_completed = 0
            att_expired = 0

            for notice in notices:
                is_attendance = notice.get(
                    'image') == 'attendance' or '签到' in notice.get('title', '')
                if is_attendance and notice.get('id'):
                    roll_call_id = notice['id']
                    info_resp = acc.api_client.get_roll_call_info(
                        roll_call_id, acc.user_data.id)

                    status = -2
                    finished = 0
                    if info_resp and info_resp.get('success'):
                        data = info_resp.get('data', {})
                        roll_call_info = data.get('rollCallInfo', {})
                        status = roll_call_info.get('status')
                        finished = data.get('attendFinish')

                    # --- 应用新逻辑 ---
                    if status == -1:
                        att_expired += 1
                    elif status != -1 and (finished == 1 or finished is True):
                        att_completed += 1
                    else:  # status != -1 and finished == 0
                        att_pending += 1

            # 更新到账号的 summary
            acc.summary.update({
                "att_pending": att_pending,
                "att_completed": att_completed,
                "att_expired": att_expired
            })
            logging.debug(
                f"[{acc.username}] 签到统计: 待签{att_pending}, 完成{att_completed}, 过期{att_expired}")

        except Exception as e:
            acc.log(f"刷新签到统计时出错: {e}")
            logging.error(
                f"[{acc.username}] Failed to fetch attendance stats: {e}", exc_info=True)

    def multi_path_generation_callback(self, username: str, success: bool, data: any):
        """接收来自JS的路径规划结果并唤醒对应线程"""
        if username in self.path_gen_callbacks:
            path_result, completion_event = self.path_gen_callbacks.pop(
                username)
            if success:
                path_result['path'] = data
            else:
                path_result['error'] = data
            completion_event.set()

    def _multi_account_worker(self, acc: AccountSession, delay: float, run_only_incomplete: bool):
        try:
            if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                self._update_account_status_js(acc, status_text="已取消")
                return

            self._update_account_status_js(acc, status_text="登录中...")
            if not acc.device_ua:
                acc.device_ua = ApiClient.generate_random_ua()

          # 登录前再次检查是否已被停止
            if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                self._update_account_status_js(acc, status_text="已中止")
                self._update_multi_global_buttons()
                return

            login_resp = self._queued_login(acc)

            # 登录返回后立即检查是否已被停止（避免继续向下流程）
            if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                self._update_account_status_js(acc, status_text="已中止")
                self._update_multi_global_buttons()
                return

            if not login_resp or not login_resp.get('success'):
                msg = login_resp.get(
                    'message', '未知错误') if login_resp else '网络错误'
                self._update_account_status_js(acc, status_text=f"登录失败: {msg}")
                return

            data = login_resp.get('data', {})
            user_info = data.get('userInfo', {})
            acc.user_data.name = user_info.get('name', '')
            acc.user_data.id = user_info.get('id', '')
            acc.user_data.student_id = user_info.get('account', '')
            acc.log("登录成功。")
            self._update_account_status_js(
                acc, status_text="分析任务", name=acc.user_data.name)

            if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                self._update_account_status_js(acc, status_text="已中止")
                return

            self._multi_fetch_and_summarize_tasks(acc)
            self._update_account_status_js(acc, summary=acc.summary)

            tasks_to_run_candidates = []
            now = datetime.datetime.now()
            ignore_time = acc.params.get("ignore_task_time", True)
            for r in acc.all_run_data:
                # 解析时间
                start_dt = None
                end_dt = None
                try:
                    if r.start_time:
                        start_dt = datetime.datetime.strptime(
                            r.start_time, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    start_dt = None
                try:
                    if r.end_time:
                        end_dt = datetime.datetime.strptime(
                            r.end_time, '%Y-%m-%d %H:%M:%S')
                except (ValueError, TypeError):
                    end_dt = None

                # 排除过期
                if end_dt:
                    is_expired = end_dt.date() < now.date() if ignore_time else end_dt < now
                    if is_expired:
                        continue

                # 排除未开始
                if start_dt:
                    is_not_started = now.date() < start_dt.date() if ignore_time else now < start_dt
                    if is_not_started:
                        continue

                tasks_to_run_candidates.append(r)

            tasks_to_run = [t for t in tasks_to_run_candidates if t.status ==
                            0] if run_only_incomplete else tasks_to_run_candidates

            if not tasks_to_run:
                self._update_account_status_js(acc, status_text="无任务可执行")
                self._update_multi_global_buttons()
                return
            else:
                acc.has_pending_tasks = True  # 有任务可执行

            if delay > 0:
                end_time = time.time() + delay
                while time.time() < end_time:
                    if acc.stop_event.is_set() or self.multi_run_stop_flag.is_set():
                        self._update_account_status_js(acc, status_text="已中止")
                        self._update_multi_global_buttons()
                        return
                    remaining = end_time - time.time()
                    self._update_account_status_js(
                        acc, status_text=f"延迟 {remaining:.0f}s")
                    time.sleep(1)

            for i, run_data in enumerate(tasks_to_run):
                if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                    self._update_account_status_js(acc, status_text="已中止")
                    self._update_multi_global_buttons()
                    break

                task_name_short = run_data.run_name[:10] + '...' if len(
                    run_data.run_name) > 10 else run_data.run_name

                # 任务开始：显示初始进度 0%
                self._update_account_status_js(
                    acc,
                    status_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short}",
                    progress_pct=0,
                    progress_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short} · 0%",
                    progress_extra=""
                )

                acc.log(f"开始执行任务: {run_data.run_name}")

                # 获取详情前检查停止
                if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                    self._update_account_status_js(acc, status_text="已中止")
                    break

                details_resp = acc.api_client.get_run_details(
                    run_data.errand_id, acc.user_data.id, run_data.errand_schedule)
                if not (details_resp and details_resp.get('success')):
                    acc.log(f"获取任务详情失败，跳过。")
                    continue

                details = details_resp.get('data', {}).get('errandDetail', {})
                waypoints = [(float(p['lon']), float(p['lat'])) for p in details.get(
                    'geoCoorList', []) if p.get('lon') is not None]
                if not waypoints:
                    acc.log(f"任务无打卡点，无法自动规划，跳过。")
                    continue
                run_data.target_points = waypoints

                path_result = {}
                completion_event = threading.Event()  # 使用绿化后的 Event
                self.path_gen_callbacks[acc.username] = (
                    path_result, completion_event)
                self.window.evaluate_js(
                    f'triggerPathGenerationForPy("{acc.username}", {json.dumps(waypoints)})')

                acc.log("等待前端JS规划路径...")
                # path_received = completion_event.wait(timeout=120)

                # 等待期间若被停止，立即返回
                while not completion_event.wait(timeout=120):
                    if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                        # 回调通道清理
                        if acc.username in self.path_gen_callbacks:
                            self.path_gen_callbacks.pop(acc.username, None)
                        self._update_account_status_js(acc, status_text="已中止")
                        return

                if 'path' not in path_result:
                    error_msg = path_result.get('error', '超时')
                    acc.log(f"路径规划失败或超时: {error_msg}")
                    if acc.username in self.path_gen_callbacks:
                        del self.path_gen_callbacks[acc.username]
                    continue

                api_path_coords = path_result['path']
                acc.log(f"路径规划成功，共 {len(api_path_coords)} 点。")

                min_t_m, max_t_m, min_d_m = acc.params.get("min_time_m", 20), acc.params.get(
                    "max_time_m", 30), acc.params.get("min_dist_m", 2000)

                final_path_dedup = []
                last_coord = None
                for p in api_path_coords:
                    # 修正BUG：同时兼容 'lng' 和 'lon' 两种经度键名
                    longitude = p.get('lng', p.get('lon'))
                    if longitude is None:
                        continue  # 如果经度不存在，则跳过这个点
                    coord = (longitude, p['lat'])
                    if coord != last_coord:
                        final_path_dedup.append(coord)
                        last_coord = coord

                if not final_path_dedup:
                    acc.log("路径处理失败：无有效坐标点。")
                    continue

                target_time_s = random.uniform(min_t_m * 60, max_t_m * 60)
                target_dist_m = random.uniform(min_d_m, min_d_m * 1.15)
                cumulative = [0.0]
                for idx_c in range(len(final_path_dedup) - 1):
                    cumulative.append(cumulative[-1] + self._calculate_distance_m(final_path_dedup[idx_c][0],
                                      final_path_dedup[idx_c][1], final_path_dedup[idx_c + 1][0], final_path_dedup[idx_c + 1][1]))

                final_geo_path = self._get_path_for_distance(
                    final_path_dedup, cumulative, target_dist_m)
                final_cumulative = [0.0]
                for idx_c in range(len(final_geo_path) - 1):
                    final_cumulative.append(final_cumulative[-1] + self._calculate_distance_m(
                        final_geo_path[idx_c][0], final_geo_path[idx_c][1], final_geo_path[idx_c + 1][0], final_geo_path[idx_c + 1][1]))

                actual_total_dist = final_cumulative[-1] if final_cumulative else 0.0
                if actual_total_dist == 0:
                    acc.log("路径计算距离为0，跳过。")
                    continue

                avg_speed = actual_total_dist / target_time_s
                new_run_coords = []
                start = final_geo_path[0]
                new_run_coords.append(self._gps_random_offset(
                    start[0], start[1], acc.params) + (0,))
                t_elapsed, d_covered = 0.0, 0.0

                while t_elapsed < target_time_s:
                    interval = min(random.uniform(
                        acc.params['interval_ms'] * .9, acc.params['interval_ms'] * 1.1) / 1000.0, target_time_s - t_elapsed)
                    if interval <= 0.1:
                        break

                    d_covered = min(d_covered + random.uniform(avg_speed * .9,
                                    avg_speed * 1.1) * interval, actual_total_dist)
                    lon, lat = self._get_point_at_distance(
                        final_geo_path, final_cumulative, d_covered)
                    lon_o, lat_o = self._gps_random_offset(
                        lon, lat, acc.params)
                    new_run_coords.append((lon_o, lat_o, int(interval * 1000)))
                    t_elapsed += interval
                    if d_covered >= actual_total_dist:
                        break

                run_data.run_coords = new_run_coords
                run_data.total_run_distance_m = d_covered
                run_data.total_run_time_s = t_elapsed

                run_data.trid = f"{acc.user_data.student_id}{int(time.time() * 1000)}"
                start_time_ms = str(int(time.time() * 1000))
                submission_successful = True
                # 总点数用于进度计算（更高频率的进度更新）
                total_points = max(1, len(run_data.run_coords))

                for chunk_idx in range(0, len(run_data.run_coords), 40):
                    if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                        submission_successful = False
                        break

                    chunk = run_data.run_coords[chunk_idx: chunk_idx + 40]
                    processed_points = chunk_idx  # 从当前分块起始已处理的点数
                    for lon, lat, dur_ms in chunk:
                        if self.multi_run_stop_flag.is_set() or acc.stop_event.is_set():
                            submission_successful = False
                            break

                        # 位置与前端地图更新（保持原有逻辑）
                        if self.window:
                            self.window.evaluate_js(
                                f'multi_updateRunnerPosition("{acc.username}", {lon}, {lat}, "{acc.user_data.name}")'
                            )

                        # 等待当前点的“上报时长”（保持原有逻辑）
                        if acc.stop_event.wait(timeout=dur_ms / 1000.0):
                            submission_successful = False
                            break

                        # 每个点完成后实时推进进度（仅更新进度相关字段，避免频繁改写状态文本）
                        processed_points += 1
                        try:
                            pct = int(processed_points * 100 / total_points)
                            self._update_account_status_js(
                                acc,
                                progress_pct=pct,
                                progress_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short} · {pct}%",
                                progress_extra=f"{processed_points}/{total_points} 点"
                            )
                        except Exception:
                            pass

                    if not submission_successful:
                        break

                    # # 按已处理点数推送进度百分比
                    # try:
                    #     total_points = max(1, len(run_data.run_coords))
                    #     processed_points = min(total_points, chunk_idx + len(chunk))
                    #     pct = int(processed_points * 100 / total_points)
                    #     self._update_account_status_js(
                    #         acc,
                    #         status_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short}",
                    #         progress_pct=pct,
                    #         progress_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short} · {pct}%",
                    #         progress_extra=f"{processed_points}/{total_points} 点"
                    #     )
                    # except Exception:
                    #     pass

                    is_final_chunk = (chunk_idx + 40 >=
                                      len(run_data.run_coords))
                    if not self._submit_chunk(run_data, chunk, start_time_ms, is_final_chunk, chunk_idx, acc.api_client, acc.user_data):
                        submission_successful = False
                        break

                if submission_successful:
                    acc.log(f"任务 {run_data.run_name} 数据提交完毕，等待服务器确认...")
                    time.sleep(3)
                    self._finalize_run(run_data, -1, acc.api_client)
                    run_data.status = 1
                    self._multi_fetch_and_summarize_tasks(acc)
                    self._update_account_status_js(acc, summary=acc.summary)
                    acc.log(f"任务 {run_data.run_name} 执行流程完成。")
                    # 任务完成后置为 100%
                    self._update_account_status_js(
                        acc,
                        status_text="任务已完成",
                        progress_pct=100,
                        progress_text=f"运行 {i+1}/{len(tasks_to_run)}: {task_name_short} · 100%",
                        progress_extra=""
                    )
                else:
                    acc.log(f"任务 {run_data.run_name} 执行被中止。")
                    # 中止时维持当前进度文本（不强制 100）
                    self._update_account_status_js(
                        acc,
                        status_text="已中止"
                    )

                if i < len(tasks_to_run) - 1:
                    wait_time = random.uniform(
                        acc.params['task_gap_min_s'], acc.params['task_gap_max_s'])
                    should_break_worker = False
                    end_time = time.time() + wait_time
                    while time.time() < end_time:
                        if acc.stop_event.is_set() or self.multi_run_stop_flag.is_set():
                            should_break_worker = True
                            break
                        remaining = end_time - time.time()
                        self._update_account_status_js(
                            acc, status_text=f"等待 {remaining:.0f}s")
                        time.sleep(1)
                    if should_break_worker:
                        self._update_account_status_js(acc, status_text="已中止")
                        break

            if not acc.stop_event.is_set():
                self._update_account_status_js(acc, status_text="全部完成")

        except Exception:
            logging.error(
                f"Error in worker for {acc.username}: {traceback.format_exc()}")
            self._update_account_status_js(acc, status_text="执行出错")
        finally:
            try:
                acc.worker_thread = None
                if acc.stop_event.is_set():
                    self._update_account_status_js(acc, status_text="已停止")
                else:
                    # 保留错误/中止/网络错误等状态，不回落为“待命”
                    if not self._should_preserve_status(acc.status_text, new_status="待命"):
                        # 仅当当前状态是临时运行态或空状态时才回落为“待命”
                        if acc.status_text in ("登录中...", "排队登录...", "已启动", "分析任务", "运行", "等待") or not acc.status_text:
                            self._update_account_status_js(
                                acc, status_text="待命")

            except Exception:
                logging.debug(
                    "Finalize worker status update failed (non-fatal).", exc_info=True)
            finally:
                self._update_multi_global_buttons()

    def _get_device_sign_code(self, username):
        """生成或获取设备标识码 (signCode)"""
        # 方案1：使用固定的 UUID (每次运行可能不同，但同一用户会话内可能一致)
        # return str(uuid.uuid4())
        # 方案2：尝试更稳定的标识 (可能需要额外库或权限)
        # try:
        #     # mac 地址 (可能需要 pip install getmac)
        #     import getmac
        #     mac = getmac.get_mac_address()
        #     if mac: return mac
        # except ImportError:
        #     pass
        # 方案3：基于用户名和固定字符串生成 UUID (同一用户始终一致)
        # return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"zslf_runner_{username}"))
        # 方案4：简单模拟 trid 格式 (不保证唯一性)
        # return f"{username}_dev_{int(time.time() * 100)}"
        # 方案5：直接使用随机 UUID
        return str(uuid.uuid4())

    def _fetch_server_attendance_radius_if_needed(self, client: ApiClient, acc: AccountSession | None = None):
        """
        (辅助函数) 检查并获取服务器签到半径。
        - 单账号模式: 缓存到 self.server_attendance_radius_m
        - 多账号模式: 缓存到 acc.server_attendance_radius_m
        """
        log_func = acc.log if acc else self.log
        cache_duration_s = 3600  # 缓存1小时

        # 确定使用哪个缓存
        if acc:
            last_fetch_time = acc.last_radius_fetch_time
            current_radius = acc.server_attendance_radius_m
        else:
            last_fetch_time = self.last_radius_fetch_time
            current_radius = self.server_attendance_radius_m

        # 检查缓存是否有效
        if time.time() - last_fetch_time < cache_duration_s and current_radius > 0:
            logging.debug(f"使用缓存的服务器签到半径值: {current_radius}米")
            return current_radius

        log_func("正在获取服务器签到半径...")
        new_radius = 0.0
        try:
            resp = client.get_attendance_radius()
            if resp and resp.get('success'):
                info_list = resp.get('data', {}).get('info', [])
                if info_list and isinstance(info_list, list) and len(info_list) > 0:
                    radius_str = info_list[0].get('code')
                    new_radius = float(radius_str)
                    if new_radius <= 0:
                        log_func("服务器返回半径为0，启用精确签到。")
                        new_radius = 0.0
                    else:
                        log_func(f"服务器签到半径更新为: {new_radius} 米")
                else:
                    log_func("服务器未返回半径信息，默认为精确签到。")
            else:
                log_func("获取半径API失败，默认为精确签到。")
        except Exception as e:
            log_func(f"获取半径时出错: {e}，默认为精确签到。")
            logging.error(
                f"Failed to fetch attendance radius: {e}", exc_info=True)

        # 更新缓存
        if acc:
            acc.server_attendance_radius_m = new_radius
            acc.last_radius_fetch_time = time.time()
        else:
            self.server_attendance_radius_m = new_radius
            self.last_radius_fetch_time = time.time()

        return new_radius

    def trigger_attendance(self, roll_call_id: str, target_coords: tuple[float, float], location_choice: str = 'random', specific_coords: tuple[float, float] | None = None, is_makeup: bool = False, acc: AccountSession | None = None):
        """
        (已重构) 触发签到流程。
        - acc: (可选) 在多账号模式下传入 AccountSession。

        Args:
            roll_call_id: 签到活动的 ID.
            target_coords: 签到目标的坐标 (经度, 纬度).
            location_choice: 'random' (在目标点范围内随机) 或 'specific' (使用 specific_coords).
            specific_coords: 当 location_choice 为 'specific' 时使用的具体坐标 (经度, 纬度).
            is_makeup: 是否为补签。
            acc: (可选) 在多账号模式下传入 AccountSession。
        """

        # 确定上下文 (单账号 vs 多账号)
        if acc:
            client = acc.api_client
            log_func = acc.log
            user = acc.user_data
            params = acc.params
        elif not self.is_multi_account_mode:
            client = self.api_client
            log_func = self.log
            user = self.user_data
            params = self.params
        else:
            self.log("trigger_attendance 在多账号模式下被错误调用（未传入acc）")
            return {"success": False, "message": "多账号模式内部错误"}

        if not user.id:
            log_func("用户未登录，无法签到。")
            return {"success": False, "message": "用户未登录"}

        log_func(f"正在处理签到: {roll_call_id}")

        # 1. 获取服务器最大半径 (使用新函数，带缓存)
        server_radius_m = self._fetch_server_attendance_radius_if_needed(
            client, acc)
        is_precise_checkin = (server_radius_m <= 0)

        # 2. 检查当前签到状态 (使用你的新逻辑)
        try:
            info_resp = client.get_roll_call_info(roll_call_id, user.id)
            if info_resp and info_resp.get('success'):
                data = info_resp.get('data', {})
                roll_call_info = data.get('rollCallInfo', {})

                status = roll_call_info.get(
                    'status')      # 关键字段1: -1=过期, 0=进行中
                finished = data.get('attendFinish')  # 关键字段2: 1=已签, 0=未签

                # 逻辑: 只要 status == -1，就是已过期且未签到
            if status == -1:
                if not is_makeup:
                    # 如果不是补签，则按原逻辑阻止
                    log_func("此签到任务已过期（status=-1）。")
                    return {"success": False, "message": "任务已过期"}
                else:
                    # 如果是补签，则记录日志并继续
                    log_func(f"任务 {roll_call_id} 已过期，正在尝试[补签]...")

                # 逻辑: 只有 status != -1 且 finished == 1 才算完成
                if status != -1 and (finished == 1 or finished is True):
                    log_func("你已经签到过了 (status!=-1 and attendFinish=1)。")
                    return {"success": True, "message": "已签到"}

                # (其他情况，如 status=0 and finished=0，都视为 "待签到"，流程继续)
                log_func("任务状态：待签到。")

            else:
                log_func("获取签到信息失败，将继续尝试签到...")
        except Exception as e:
            log_func(f"检查签到状态时出错: {e}，将继续尝试...")
            logging.error(
                f"Error checking roll call status: {e}", exc_info=True)

        # 3. 确定签到坐标
        final_lon, final_lat = 0.0, 0.0
        target_lon, target_lat = target_coords

        if location_choice == 'specific' and specific_coords:
            final_lon, final_lat = specific_coords
            log_func(f"使用指定坐标签到: ({final_lon:.6f}, {final_lat:.6f})")

        elif location_choice == 'random':
            # 默认为精确签到
            radius_for_random = 0.0

            if not is_precise_checkin:
                # 仅在服务器允许范围签到时，才计算随机半径
                user_radius_m = params.get("attendance_user_radius_m", 40)
                # 随机半径不能超过服务器的半径
                radius_for_random = max(
                    0.0, min(user_radius_m, server_radius_m))

            if radius_for_random <= 0:
                # 精确签到
                final_lon, final_lat = target_lon, target_lat
                log_func(
                    f"精确签到模式：使用目标点坐标签到: ({final_lon:.6f}, {final_lat:.6f})")
            else:
                # 范围随机签到
                angle = random.uniform(0, 2 * math.pi)
                # 在 0 到 1 倍的 *有效半径* 内随机
                radius_ratio = random.uniform(0, 1.0)
                radius_m = radius_for_random * radius_ratio

                offset_lon = (radius_m * math.cos(angle)) / 102834.74
                offset_lat = (radius_m * math.sin(angle)) / 111712.69

                final_lon = target_lon + offset_lon
                final_lat = target_lat + offset_lat
                log_func(
                    f"在 {radius_for_random:.1f} 米范围内随机生成坐标签到: ({final_lon:.6f}, {final_lat:.6f})")

        else:
            log_func("无效的位置选择，无法签到。")
            return {"success": False, "message": "无效的位置选择"}

        # 4. 计算实际距离
        actual_distance = self._calculate_distance_m(
            final_lon, final_lat, target_lon, target_lat)
        log_func(f"签到点距离目标点 {actual_distance:.2f} 米。")

        # 5. 获取设备标识码
        sign_code = self._get_device_sign_code(user.username)

        # 6. 准备提交负载
        if is_makeup:
            # 补签负载
            payload = {
                "rollCallId": roll_call_id,
                "userId": user.id,
                "coordinate": f"{final_lon},{final_lat}",
                "distance": int(actual_distance),  # 转为整数
                "status": 2,  # 1=正常, 2=补签 (如果补签状态码不是2，请修改这里)
                "signCode": sign_code,
                "reason": "补签"  # 补签时填写原因
            }
            log_func("使用[补签]负载提交...")
        else:
            # 正常签到负载
            payload = {
                "rollCallId": roll_call_id,
                "userId": user.id,
                "coordinate": f"{final_lon},{final_lat}",
                "distance": int(actual_distance),  # 转为整数
                "status": 1,  # 1 表示成功签到
                "signCode": sign_code,
                "reason": ""  # 签到时为空
            }
            log_func("使用[正常签到]负载提交...")

        # 7. 提交签到
        try:
            log_func("正在提交签到...")
            submit_resp = client.submit_attendance(payload)
            if submit_resp and submit_resp.get('success'):
                log_func("签到成功！")
                logging.info(
                    f"Attendance submitted successfully for {roll_call_id}")
                return {"success": True, "message": "签到成功"}
            else:
                msg = submit_resp.get(
                    'message', '未知错误') if submit_resp else '网络错误'
                log_func(f"签到失败: {msg}")
                logging.warning(
                    f"Attendance submission failed for {roll_call_id}: {msg}")
                return {"success": False, "message": f"签到失败: {msg}"}
        except Exception as e:
            log_func(f"提交签到时出错: {e}")
            logging.error(f"提交签到请求时发生错误，异常信息: {e}", exc_info=True)
            return {"success": False, "message": f"提交签到时出错: {e}"}

    # --- 获取通知功能 ---

    def get_notifications(self, is_auto_refresh: bool = False):
        """
        (已重构) 获取未读通知数量和通知列表。
        - 附加基于新逻辑的签到状态。
        - (自动签到逻辑已移至 _check_and_trigger_auto_attendance)
        """
        logging.info("API调用: get_notifications - 获取用户通知消息列表")
        if not self.user_data.id or self.is_multi_account_mode:
            return {"success": False, "message": "仅单账号登录模式可用"}

        # 【修复 Requirement 13】离线模式下直接返回空白通知列表，不读取缓存
        if self.is_offline_mode:
            return {
                "success": True,
                "data": {
                    "unreadNumber": 0,
                    "notices": []
                }
            }

        try:
            # 1. 获取未读数量
            count_resp = self.api_client.get_unread_notice_count()
            unread_count = 0
            if count_resp and count_resp.get('success'):
                unread_count = count_resp.get(
                    'data', {}).get('unreadNumber', 0)

            # 2. 【重构】获取所有通知
            all_notices = []
            offset = 0
            limit = 10
            while True:
                list_resp = self.api_client.get_notice_list(
                    offset=offset, limit=limit, type_id=0)
                if list_resp and list_resp.get('success'):
                    current_notices = list_resp.get(
                        'data', {}).get('noticeList', [])
                    if not current_notices:
                        # 没有更多通知了，退出循环
                        break
                    all_notices.extend(current_notices)
                    offset += limit
                    # 如果返回的通知数小于请求的 limit，说明是最后一页
                    if len(current_notices) < limit:
                        break
                else:
                    self.log("获取通知列表时失败。")
                    # 发生错误时中断，避免无限循环
                    break

            notices = all_notices

            # 3. 【重构】遍历通知，仅为签到任务附加状态
            if notices:
                logging.debug(f"正在为 {len(notices)} 条通知附加签到状态...")
                for notice in notices:
                    try:
                        is_attendance = notice.get(
                            'image') == 'attendance' or '签到' in notice.get('title', '')
                        if is_attendance and notice.get('id'):
                            roll_call_id = notice['id']

                            info_resp = self.api_client.get_roll_call_info(
                                roll_call_id, self.user_data.id)

                            status = -2  # 默认为未知
                            finished = 0

                            if info_resp and info_resp.get('success'):
                                data = info_resp.get('data', {})
                                roll_call_info = data.get('rollCallInfo', {})
                                status = roll_call_info.get(
                                    'status')  # -1=过期, 0=进行中
                                finished = data.get(
                                    'attendFinish')  # 1=已签, 0=未签

                            # 附加新字段到 notice 对象
                            notice['attendance_finished'] = finished
                            notice['attendance_status_code'] = status

                            if status == -1:
                                notice['attendance_code'] = -1  # 已过期
                            elif status != -1 and (finished == 1 or finished is True):
                                notice['attendance_code'] = 1  # 已签到
                            else:
                                notice['attendance_code'] = 0  # 待签到

                    except Exception as e:
                        logging.warning(
                            f"附加签到状态失败 (ID: {notice.get('id')}): {e}")

            if not is_auto_refresh:
                self.log(f"获取到 {unread_count} 条未读通知，共有 {len(notices)} 条通知。")

            return {
                "success": True,
                "unreadCount": unread_count,
                "notices": notices
            }
        except Exception as e:
            self.log(f"获取通知失败: {e}")
            logging.error(f"get_notifications failed: {e}", exc_info=True)
            return {"success": False, "message": str(e)}

    def mark_notification_read(self, notice_id):
        """(单账号) 将指定ID的通知设为已读"""
        logging.info(f"API调用: mark_notification_read - 标记通知为已读，通知ID: {notice_id}")
        if not self.user_data.id or self.is_multi_account_mode:
            return {"success": False, "message": "仅单账号登录模式可用"}

        resp = self.api_client.mark_notice_as_read(notice_id)
        if resp and resp.get('success'):
            return {"success": True}
        else:
            return {"success": False, "message": resp.get('message', '标记已读失败')}

    def _auto_refresh_worker(self):
        """(单账号) 后台自动刷新通知和签到的线程 (已修复)"""
        while not self.stop_auto_refresh.is_set():
            try:
                # --- 修复：从循环顶部移除 is_enabled 检查 ---

                # 1. 检查是否满足运行条件（已登录 且 处于单账号模式）
                if (not self.user_data.id or self.is_multi_account_mode):
                    # 如果未登录或在多账号模式，则休眠
                    # (使用 wait 替代 sleep 以便快速响应停止)
                    if self.stop_auto_refresh.wait(timeout=5.0):
                        break  # 收到停止信号
                    continue  # 回去重新检查登录状态

                # --- 修复：移除 "Bug 2" 相关的300秒长休眠逻辑 ---
                # if not is_enabled: ... continue ...

                # 2. 线程应始终按 'refresh_interval_s' 频率运行
                
                # 读取刷新间隔，并设置合理的默认值和最小值
                refresh_interval_s = self.params.get(
                    "auto_attendance_refresh_s", 60) # 默认60秒
                # 确保间隔不小于一个最小值，例如15秒，防止过高频率
                refresh_interval_s = max(15, refresh_interval_s)

                # (修复：将等待日志移到 wait 之前)
                # self.log(f"自动刷新: 等待 {refresh_interval_s} 秒...")

                # 使用 wait() 替代 time.sleep()，以便在等待期间可以被 'stop_auto_refresh.set()' 立即中断
                if self.stop_auto_refresh.wait(timeout=refresh_interval_s):
                    break  # 如果在等待时收到停止信号，则退出循环

                # 再次检查登录状态 (可能在等待时退出了)
                if (self.is_multi_account_mode or not self.user_data.id):
                    continue

                # --- 关键修复：在执行操作前，才读取最新的 is_enabled 状态 ---
                is_enabled = self.params.get("auto_attendance_enabled", False)

                # 3. 只有在启用时才执行自动签到
                if is_enabled:
                    self.log("(后台) 自动签到已启用，正在检查...")
                    self._check_and_trigger_auto_attendance(self)
                else:
                    # 如果未启用，则不执行也不记录日志，保持安静
                    pass

                # 4. 获取通知并推送到UI (无论是否启用自动签到，都应该执行)
                self.log("正在自动刷新通知 (后台)...") # <-- 这是您看到的日志
                result = self.get_notifications(is_auto_refresh=True) # (修复：传入True以避免日志重复)

                # [BUG 修复]：替换 self.window 为 socketio.emit
                if result.get('success'):
                    # 修复：使用 SocketIO 向特定会话推送更新，而不是 self.window
                    session_id = getattr(self, '_web_session_id', None)
                    # 确保 socketio 变量在全局可用 (它在 start_web_server 中定义)
                    if session_id and 'socketio' in globals():
                        try:
                            # 发送一个自定义事件，前端JS需要监听这个事件
                            globals()['socketio'].emit('onNotificationsUpdated', result, room=session_id)
                            logging.debug(f"[_auto_refresh_worker] 已向会话 {session_id[:8]} 推送通知更新")
                        except Exception as e:
                            logging.error(f"[_auto_refresh_worker] SocketIO推送通知失败: {e}", exc_info=True)
                    elif not session_id:
                         logging.warning(f"[_auto_refresh_worker] 无法推送通知：未找到 _web_session_id")
                    else:
                         logging.warning(f"[_auto_refresh_worker] 无法推送通知：socketio 实例不可用")


            except Exception as e:
                self.log(f"自动刷新线程出错: {e}")
                logging.error(f"Auto-refresh worker error: {e}", exc_info=True)
                # 修复：使用 wait() 替换 time.sleep() 以便能被立即停止
                if self.stop_auto_refresh.wait(timeout=60):
                    break # 如果在等待时收到停止信号，则退出循环

        logging.info("Auto-refresh worker stopped.")


    def _check_and_trigger_auto_attendance(self, context: 'Api | AccountSession'):
        """
        (辅助函数) 检查并执行单个上下文(Api或AccountSession)的自动签到。
        """
        # 1. 确定上下文
        if isinstance(context, AccountSession):
            client = context.api_client
            log_func = context.log
            user = context.user_data
            params = context.params
            # 多账号模式下，使用账号自己的参数来判断是否启用
            if not params.get("auto_attendance_enabled", False):
                return
        else:  # context is Api (self)
            client = self.api_client
            log_func = self.log
            user = self.user_data
            params = self.params
            # 单账号模式下，也使用自己的参数判断
            if not params.get("auto_attendance_enabled", False):
                return

        if not user.id:
            log_func("用户未登录，跳过自动签到。")
            return

        log_func("(后台) 正在检查自动签到任务...")

        try:
            list_resp = client.get_notice_list(offset=0, limit=20, type_id=0)
            if not (list_resp and list_resp.get('success')):
                log_func("获取通知列表失败，跳过自动签到。")
                return

            notices = list_resp.get('data', {}).get('noticeList', [])
            if not notices:
                log_func("(后台) 通知列表为空。")
                return

            triggered_count = 0
            for notice in notices:
                is_attendance = notice.get(
                    'image') == 'attendance' or '签到' in notice.get('title', '')
                if not (is_attendance and notice.get('id')):
                    continue

                roll_call_id = notice['id']
                info_resp = client.get_roll_call_info(roll_call_id, user.id)

                status = -2
                finished = 0
                if info_resp and info_resp.get('success'):
                    data = info_resp.get('data', {})
                    roll_call_info = data.get('rollCallInfo', {})
                    status = roll_call_info.get('status')  # -1=过期, 0=进行中
                    finished = data.get('attendFinish')  # 1=已签, 0=未签

                # 检查是否为 "待签到" (进行中 且 未完成)
                if status != -1 and not (finished == 1 or finished is True):
                    log_func(f"检测到待签到任务 '{notice.get('title')}'，正在自动签到...")
                    coords_str = notice.get('updateBy', '').split(',')
                    if len(coords_str) == 2:
                        try:
                            target_lat, target_lon = float(
                                coords_str[0]), float(coords_str[1])
                            target_coords = (target_lon, target_lat)

                            # 关键：调用 trigger_attendance
                            auto_result = self.trigger_attendance(
                                roll_call_id,
                                target_coords,
                                'random',

                                specific_coords=None,
                                is_makeup=False,  # 自动签到总是尝试正常签到
                                acc=context if isinstance(
                                    context, AccountSession) else None
                            )

                            if auto_result.get('success'):
                                log_func(f"自动签到 '{notice.get('title')}' 成功。")
                                triggered_count += 1
                            else:
                                log_func(
                                    f"自动签到 '{notice.get('title')}' 失败: {auto_result.get('message', '')}")
                        except Exception as e:
                            log_func(f"签到坐标解析或执行失败: {e}")
                    else:
                        log_func("签到通知坐标格式错误，跳过。")

            if triggered_count == 0:
                log_func("(后台) 未发现待处理的签到任务。")

        except Exception as e:
            log_func(f"自动签到检查时出错: {e}")
            logging.error(
                f"[_check_and_trigger_auto_attendance] Error: {e}", exc_info=True)

    def _multi_auto_attendance_worker(self):
        """(多账号) 后台自动刷新和签到所有账号的线程"""
        while not self.stop_multi_auto_refresh.wait(timeout=1.0):
            try:
                if (not self.is_multi_account_mode or
                    not self.global_params.get("auto_attendance_enabled", False) or
                        not self.accounts):
                    # 如果未启用、或不在多账号模式、或没账号，则休眠
                    time.sleep(5)
                    continue

                # 读取刷新间隔
                refresh_interval_s = self.global_params.get(
                    "auto_attendance_refresh_s")

                self.log(f"(多账号) 自动签到: 等待 {refresh_interval_s} 秒...")
                if self.stop_multi_auto_refresh.wait(timeout=refresh_interval_s):
                    break  # 收到停止信号

                if (not self.is_multi_account_mode or
                        not self.global_params.get("auto_attendance_enabled", False)):
                    continue

                self.log("(多账号) 正在为所有账号执行后台签到检查...")

                # 迭代当前账号列表的快照
                accounts_to_check = list(self.accounts.values())

                for acc in accounts_to_check:
                    if self.stop_multi_auto_refresh.is_set():
                        break  # 检查过程中被停止

                    # 检查该账号是否也启用了自动签到 (从它自己的参数里读)
                    if acc.params.get("auto_attendance_enabled", False):
                        if acc.user_data.id:
                            self._check_and_trigger_auto_attendance(acc)
                            # 刷新该账号的签到统计 (这会更新UI)
                            self._multi_fetch_attendance_stats(acc)
                            self._update_account_status_js(
                                acc, summary=acc.summary)
                            # 简单延迟，避免API风暴
                            time.sleep(random.uniform(1.0, 3.0))
                        else:

                            acc.log("(后台) 尚未登录，跳过自动签到检查。")
                    else:

                        acc.log("(后台) 自动签到未在此账号上启用，跳过。")

            except Exception as e:
                self.log(f"(多账号) 自动签到线程出错: {e}")
                logging.error(
                    f"Multi-auto-attendance worker error: {e}", exc_info=True)
                time.sleep(60)

        logging.info("Multi-auto-attendance worker stopped.")

    def _normalize_status_flag(self, v) -> int:
        """将后端返回的各种 isExecute 表达式规范为 0 或 1"""
        try:
            if v is None:
                return 0
            if isinstance(v, bool):
                return 1 if v else 0
            s = str(v).strip()
            if s == '':
                return 0
            # 纯数字字符串
            if s.isdigit():
                return 1 if int(s) == 1 else 0
            # 常见布尔文本
            if s.lower() in ('true', 'yes', 'y', '1'):
                return 1
            return 0
        except Exception:
            return 0

    def _should_preserve_status(self, status: str, new_status: str = None) -> bool:
        """
        判断当前状态是否应保留。
        - 如果当前是错误/中止态，则保留，除非新状态明确表示成功或完成。
        """
        if not status:
            return False
        s = str(status).strip()
        error_keywords = ("登录失败", "刷新失败", "执行出错", "网络错误", "已中止", "已停止")

        # 如果当前是错误态
        if any(k in s for k in error_keywords):
            # 只有当新状态是“登录成功”“全部完成”“无任务可执行”等明确的成功/完成态时，才允许覆盖
            if new_status and new_status in ("登录成功", "全部完成", "无任务可执行", "待命"):
                return False
            return True
        return False


# ==============================================================================
# 4. 前端界面 (HTML/CSS/JS)
# ==============================================================================

def resource_path(relative_path):
    """获取资源文件的绝对路径，兼容 PyInstaller 打包和开发环境"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


html_content = ""

try:
    html_path = resource_path("index.html")
    with open(html_path, 'r', encoding='utf-8') as file:
        html_content = file.read()
    logging.info("成功读取 index.html 文件！")

except FileNotFoundError:
    logging.error("错误: 未在资源路径中找到 'index.html' 文件。")
    sys.exit(1)
except Exception as e:
    logging.error(f"读取文件时发生错误: {e}")
    sys.exit(1)


def check_port_available(host, port):
    """检查端口是否可用"""

    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind((host, port))
            return True
    except OSError:
        return False


# ==============================================================================
# 5. Web服务器模式
#    支持通过浏览器访问，使用UUID进行会话管理
#    使用Playwright在服务器端运行Chrome进行JS计算
# ==============================================================================
# 全局会话存储：{session_id: Api实例}
web_sessions = {}
web_sessions_lock = threading.Lock()
# 会话文件锁，用于并发保护
session_file_locks = {}  # {session_hash: threading.Lock}
session_file_locks_lock = threading.Lock()
# 会话活动追踪：{session_id: last_activity_time}
session_activity = {}
session_activity_lock = threading.Lock()


def update_session_activity(session_id):
    """更新会话活动时间"""
    with session_activity_lock:
        session_activity[session_id] = time.time()


def cleanup_session(session_id, reason="manual"):
    """清理指定会话（支持指定原因）"""
    # 跳过无效的session_id
    if not session_id or session_id == 'null' or session_id.strip() == '':
        logging.debug(f"跳过清理无效会话ID: '{session_id}'")
        return

    logging.info(f"清理会话: {session_id[:32]}... (原因: {reason})")
    cleanup_inactive_session(session_id)


def cleanup_inactive_session(session_id):
    """清理不活跃的会话"""
    # 跳过无效的session_id
    if not session_id or session_id == 'null' or session_id.strip() == '':
        logging.debug(f"跳过清理无效会话ID: '{session_id}'")
        return

    try:
        logging.info(f"清理不活跃会话: {session_id[:32]}...")

        # 从内存中移除会话
        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]

                # 停止运行中的任务
                if hasattr(api_instance, 'stop_run_flag'):
                    api_instance.stop_run_flag.set()

                # 如果是注册用户，取消会话关联并使token失效
                if hasattr(api_instance, 'auth_username') and not getattr(api_instance, 'is_guest', True):
                    username = api_instance.auth_username
                    auth_system.unlink_session_from_user(username, session_id)
                    # 使token失效
                    token_manager.invalidate_token(username, session_id)
                    logging.info(
                        f"已使用户 {username} 的会话 {session_id[:16]}... 的token失效")

                del web_sessions[session_id]

        # 删除会话文件

        session_hash = hashlib.sha256(session_id.encode()).hexdigest()
        session_file = os.path.join(
            SESSION_STORAGE_DIR, f"{session_hash}.json")
        if os.path.exists(session_file):
            os.remove(session_file)
            logging.info(f"已删除会话文件: {session_hash[:16]}...")

        # 从活动追踪中移除
        with session_activity_lock:
            if session_id in session_activity:
                del session_activity[session_id]

        # 从索引中移除
        index = _load_session_index()
        if session_id in index:
            del index[session_id]
            _save_session_index(index)

        logging.info(f"会话清理完成: {session_id[:32]}...")
    except Exception as e:
        logging.error(f"清理会话失败 {session_id[:32]}...: {e}")


def monitor_session_inactivity():
    """监控会话不活跃状态并清理（5分钟无活动）"""
    while True:
        try:
            time.sleep(60)  # 每分钟检查一次

            current_time = time.time()
            inactive_sessions = []

            with session_activity_lock:
                for session_id, last_activity in list(session_activity.items()):
                    # 检查是否在登录页面且5分钟无活动
                    if current_time - last_activity > 300:  # 5分钟 = 300秒
                        with web_sessions_lock:
                            if session_id in web_sessions:
                                api_instance = web_sessions[session_id]
                                # 只清理已认证但未登录应用的会话
                                is_authenticated = getattr(
                                    api_instance, 'is_authenticated', False)
                                is_logged_in = getattr(
                                    api_instance, 'login_success', False)

                                # 检查是否有任务正在执行
                                has_active_task = False

                                # 单账号模式：检查是否有跑步任务正在执行
                                if hasattr(api_instance, 'stop_run_flag'):
                                    has_active_task = has_active_task or not api_instance.stop_run_flag.is_set()

                                # 单账号模式：检查自动签到是否正在运行
                                if hasattr(api_instance, 'auto_refresh_thread'):
                                    thread = api_instance.auto_refresh_thread
                                    has_active_task = has_active_task or (
                                        thread is not None and thread.is_alive())

                                # 多账号模式：检查多账号自动签到是否正在运行
                                if hasattr(api_instance, 'multi_auto_refresh_thread'):
                                    thread = api_instance.multi_auto_refresh_thread
                                    has_active_task = has_active_task or (
                                        thread is not None and thread.is_alive())

                                # 多账号模式：检查是否有账号的任务线程正在运行
                                if hasattr(api_instance, 'multi_accounts'):
                                    for acc in api_instance.multi_accounts:
                                        if hasattr(acc, 'worker_thread'):
                                            thread = acc.worker_thread
                                            has_active_task = has_active_task or (
                                                thread is not None and thread.is_alive())

                                # 只清理：已认证但未登录应用，并且没有任务正在执行的会话
                                if is_authenticated and not is_logged_in and not has_active_task:
                                    inactive_sessions.append(session_id)
                                    logging.debug(
                                        f"会话 {session_id[:32]}... 标记为不活跃（无任务执行）")
                                elif has_active_task:
                                    logging.debug(
                                        f"会话 {session_id[:32]}... 有活跃任务，跳过清理")

            # 清理不活跃会话
            for session_id in inactive_sessions:
                cleanup_inactive_session(session_id)

        except Exception as e:
            logging.error(f"会话监控线程错误: {e}")

# 启动会话监控线程


def start_session_monitor():
    """启动会话不活跃监控"""
    monitor_thread = threading.Thread(
        target=monitor_session_inactivity, daemon=True)
    monitor_thread.start()
    logging.info("会话监控线程已启动")

# 会话持久化函数


def _load_session_index():
    """加载会话索引文件"""
    try:
        if os.path.exists(SESSION_INDEX_FILE):
            with open(SESSION_INDEX_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        logging.warning(f"加载会话索引失败: {e}")
    return {}


def _save_session_index(index):
    """保存会话索引文件"""
    try:
        with open(SESSION_INDEX_FILE, 'w', encoding='utf-8') as f:
            json.dump(index, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logging.error(f"保存会话索引失败: {e}")


def save_session_state(session_id, api_instance, force_save=False):
    """
    将会话状态保存到文件（完整版：保存所有应用状态包括离线任务数据）

    Args:
        session_id: 会话UUID
        api_instance: Api实例
        force_save: 强制保存，即使距离上次保存时间很短
    """
    if not session_id or session_id == "null":
        logging.warning(f"拒绝保存会话：无效的 session_id: '{session_id}'")
        # 可以在这里根据需要决定是否要抛出异常或返回特定值
        return  # 直接返回，不执行后续保存逻辑
    try:
        # 修复Windows路径长度限制：使用SHA256哈希作为文件名
        # 2048位UUID(512字符)会导致Windows文件名过长错误
        session_hash = hashlib.sha256(session_id.encode()).hexdigest()
        session_file = os.path.join(
            SESSION_STORAGE_DIR, f"{session_hash}.json")

        # 线程安全：获取或创建此会话文件的锁
        with session_file_locks_lock:
            if session_hash not in session_file_locks:
                session_file_locks[session_hash] = threading.Lock()
            file_lock = session_file_locks[session_hash]

        # 使用文件锁保护并发写入
        with file_lock:
            # 防抖：避免过于频繁的保存（除非强制保存）
            if not force_save:
                last_save_time = getattr(
                    api_instance, '_last_session_save_time', 0)
                if time.time() - last_save_time < 2.0:  # 最少间隔2秒
                    return

            # 记录本次保存时间
            api_instance._last_session_save_time = time.time()

            # 基础状态
            state = {
                'session_id': session_id,  # 在文件内容中保存完整的UUID
                'school_account_logged_in': getattr(api_instance, 'login_success', False),  # 更明确的命名：学校账号是否已登录
                'login_success': getattr(api_instance, 'login_success', False),  # 保留兼容性
                'user_info': getattr(api_instance, 'user_info', None),
                'created_at': getattr(api_instance, '_session_created_at', time.time()),
                'last_accessed': time.time(),
                'last_saved': time.time()  # 最后保存时间
            }

            # 增强：保存认证信息
            if hasattr(api_instance, 'auth_username'):
                state['auth_username'] = api_instance.auth_username
                state['auth_group'] = getattr(
                    api_instance, 'auth_group', 'guest')
                state['is_guest'] = getattr(api_instance, 'is_guest', False)
                state['is_authenticated'] = getattr(
                    api_instance, 'is_authenticated', False)

            # 增强：保存用户配置参数
            if hasattr(api_instance, 'params'):
                state['params'] = api_instance.params

            # 增强：保存User-Agent
            if hasattr(api_instance, 'device_ua'):
                state['device_ua'] = api_instance.device_ua

            # 增强：保存用户数据
            if hasattr(api_instance, 'user_data') and api_instance.user_data:
                user_data = api_instance.user_data
                state['user_data'] = {
                    'name': getattr(user_data, 'name', ''),
                    'phone': getattr(user_data, 'phone', ''),
                    'student_id': getattr(user_data, 'student_id', ''),
                    'id': getattr(user_data, 'id', ''),
                    'username': getattr(user_data, 'username', ''),
                    'gender': getattr(user_data, 'gender', ''),
                    'school_name': getattr(user_data, 'school_name', '')
                }

            # 增强：保存任务选择状态
            if hasattr(api_instance, 'current_run_idx'):
                state['current_run_idx'] = api_instance.current_run_idx

            # 保存完整的任务列表（包括所有离线任务字段）
            if hasattr(api_instance, 'all_run_data') and api_instance.all_run_data:
                loaded_tasks = []
                for run_data in api_instance.all_run_data:
                    task_dict = {
                        # 任务基本信息
                        'run_name': getattr(run_data, 'run_name', ''),
                        'errand_id': getattr(run_data, 'errand_id', ''),
                        'errand_schedule': getattr(run_data, 'errand_schedule', ''),
                        'status': getattr(run_data, 'status', 0),
                        'start_time': getattr(run_data, 'start_time', ''),
                        'end_time': getattr(run_data, 'end_time', ''),
                        'upload_time': getattr(run_data, 'upload_time', ''),
                        'total_run_time_s': getattr(run_data, 'total_run_time_s', 0.0),
                        'total_run_distance_m': getattr(run_data, 'total_run_distance_m', 0.0),

                        # 离线任务完整数据（核心）
                        'target_points': getattr(run_data, 'target_points', []),
                        'target_point_names': getattr(run_data, 'target_point_names', ''),
                        'recommended_coords': getattr(run_data, 'recommended_coords', []),
                        'draft_coords': getattr(run_data, 'draft_coords', []),
                        'run_coords': getattr(run_data, 'run_coords', []),

                        # 运行时状态
                        'target_sequence': getattr(run_data, 'target_sequence', 0),
                        'is_in_target_zone': getattr(run_data, 'is_in_target_zone', False),
                        'trid': getattr(run_data, 'trid', ''),
                        'details_fetched': getattr(run_data, 'details_fetched', False),
                        'distance_covered_m': getattr(run_data, 'distance_covered_m', 0.0)
                    }
                    loaded_tasks.append(task_dict)
                state['loaded_tasks'] = loaded_tasks

            # 保存运行状态
            state['is_offline_mode'] = getattr(
                api_instance, 'is_offline_mode', False)
            
            if hasattr(api_instance, 'api_client') and api_instance.api_client.session.cookies:
                try:
                    # 将 CookieJar 转换为可序列化的字典
                    state['api_cookies'] = requests.utils.dict_from_cookiejar(api_instance.api_client.session.cookies)
                    logging.debug(f"会话保存: 正在保存 {len(state['api_cookies'])} 个 API Cookies...")
                except Exception as e:
                    logging.warning(f"会话保存: 保存 API Cookies 失败: {e}")
            
            # 保存多账号模式状态
            state['is_multi_account_mode'] = getattr(api_instance, 'is_multi_account_mode', False)
            
            # 如果是多账号模式，保存多账号相关信息
            if getattr(api_instance, 'is_multi_account_mode', False):
                # 保存已加载的账号列表（用户名列表）
                state['multi_account_usernames'] = list(getattr(api_instance, 'accounts', {}).keys())
                
                # 保存全局参数
                state['multi_global_params'] = getattr(api_instance, 'global_params', {})
                
                # 保存每个账号的状态
                multi_account_states = {}
                accounts = getattr(api_instance, 'accounts', {})
                for username, account_session in accounts.items():
                    try:
                        account_state = {
                            'username': username,
                            'is_running': getattr(account_session, 'is_running', False),
                            'status_text': getattr(account_session, 'status_text', ''),
                            'progress': getattr(account_session, 'progress', 0),
                            'school_account_logged_in': getattr(account_session, 'login_success', False),
                            'has_tasks': len(getattr(account_session, 'all_run_data', [])) > 0,
                            'task_count': len(getattr(account_session, 'all_run_data', [])),
                            'completed_tasks': sum(1 for task in getattr(account_session, 'all_run_data', []) 
                                                 if getattr(task, 'status', 0) == 1),
                        }
                        # 保存账号特定参数（如果有）
                        if hasattr(account_session, 'params'):
                            account_state['params'] = account_session.params
                        multi_account_states[username] = account_state
                    except Exception as e:
                        logging.warning(f"保存账号 {username} 状态时出错: {e}")
                        continue
                        
                state['multi_account_states'] = multi_account_states
                
                # 保存多账号控制面板信息（仪表盘信息）
                state['multi_dashboard_info'] = {
                    'total_accounts': len(accounts),
                    'running_accounts': sum(1 for acc in accounts.values() if getattr(acc, 'is_running', False)),
                    'logged_in_accounts': sum(1 for acc in accounts.values() if getattr(acc, 'login_success', False)),
                    'total_tasks': sum(len(getattr(acc, 'all_run_data', [])) for acc in accounts.values()),
                    'completed_tasks': sum(
                        sum(1 for task in getattr(acc, 'all_run_data', []) if getattr(task, 'status', 0) == 1)
                        for acc in accounts.values()
                    )
                }
            
            # 单账号模式的仪表盘信息
            if not getattr(api_instance, 'is_multi_account_mode', False):
                # 统计任务信息
                all_tasks = getattr(api_instance, 'all_run_data', [])
                state['dashboard_info'] = {
                    'total_tasks': len(all_tasks),
                    'completed_tasks': sum(1 for task in all_tasks if getattr(task, 'status', 0) == 1),
                    'pending_tasks': sum(1 for task in all_tasks if getattr(task, 'status', 0) == 0),
                    'selected_task_index': getattr(api_instance, 'current_run_idx', -1),
                    'is_offline_mode': getattr(api_instance, 'is_offline_mode', False),
                    'school_account_logged_in': getattr(api_instance, 'login_success', False)
                }
                
                # 添加当前选中任务的信息（如果有）
                if hasattr(api_instance, 'current_run_idx') and api_instance.current_run_idx >= 0:
                    if hasattr(api_instance, 'all_run_data') and api_instance.current_run_idx < len(api_instance.all_run_data):
                        current_task = api_instance.all_run_data[api_instance.current_run_idx]
                        state['dashboard_info']['current_task'] = {
                            'run_name': getattr(current_task, 'run_name', ''),
                            'status': getattr(current_task, 'status', 0),
                            'total_distance': getattr(current_task, 'total_run_distance_m', 0.0),
                            'total_time': getattr(current_task, 'total_run_time_s', 0.0),
                            'has_path': len(getattr(current_task, 'run_coords', [])) > 0
                        }


            # 增强：保存停止标志状态
            if hasattr(api_instance, 'stop_run_flag'):
                state['stop_run_flag_set'] = api_instance.stop_run_flag.is_set()

            # 增强：保存UI状态（如果存在）
            if hasattr(api_instance, 'ui_state'):
                state['ui_state'] = api_instance.ui_state

            # 增强：保存用户设置（如果存在）
            if hasattr(api_instance, 'user_settings'):
                state['user_settings'] = api_instance.user_settings

            # 写入文件
            with open(session_file, 'w', encoding='utf-8') as f:
                json.dump(state, f, indent=2, ensure_ascii=False)

            # 更新索引文件：存储完整UUID到SHA256哈希的映射
            index = _load_session_index()
            index[session_id] = session_hash
            _save_session_index(index)

            tasks_count = len(state.get('loaded_tasks', []))
            logging.debug(
                f"会话状态已保存: {session_id[:32]}... (任务数:{tasks_count}, 选中索引:{state.get('current_run_idx', -1)})")
    except Exception as e:
        logging.error(f"保存会话状态失败: {e}", exc_info=True)


def load_session_state(session_id):
    """从文件加载会话状态"""
    try:
        # 优化：首先从索引文件查找哈希，避免每次都计算
        index = _load_session_index()

        # 如果索引中存在，直接使用索引中的哈希值
        if session_id in index:
            session_hash = index[session_id]
            logging.debug(
                f"从索引找到会话哈希: {session_id[:32]}... -> {session_hash[:16]}...")
        else:
            # 索引中不存在，计算哈希值（兼容旧文件或索引损坏情况）
            session_hash = hashlib.sha256(session_id.encode()).hexdigest()
            logging.debug(
                f"索引中未找到，计算会话哈希: {session_id[:32]}... -> {session_hash[:16]}...")

        session_file = os.path.join(
            SESSION_STORAGE_DIR, f"{session_hash}.json")

        if os.path.exists(session_file):
            # 线程安全：获取或创建此会话文件的锁
            with session_file_locks_lock:
                if session_hash not in session_file_locks:
                    session_file_locks[session_hash] = threading.Lock()
                file_lock = session_file_locks[session_hash]

            # 使用文件锁保护并发读取
            with file_lock:
                with open(session_file, 'r', encoding='utf-8') as f:
                    state = json.load(f)

            # 验证加载的UUID是否匹配
            if state.get('session_id') == session_id:
                # 检查会话是否过期（7天未访问）
                last_accessed = state.get('last_accessed', 0)
                session_age_days = (time.time() - last_accessed) / 86400  # 转换为天数
                max_age_days = 7  # 会话最大保留7天
                
                if session_age_days > max_age_days:
                    logging.warning(f"[会话管理] 会话已过期 --> 会话ID: {session_id[:32]}..., 最后访问: {session_age_days:.1f}天前, 最大保留期限: {max_age_days}天, 将被自动清理")
                    # 删除过期的会话文件
                    try:
                        os.remove(session_file)
                        logging.info(f"[会话管理] 已删除过期会话文件: {session_file}")
                    except Exception as remove_err:
                        logging.error(f"[会话管理] 删除过期会话文件失败: {remove_err}")
                    return None  # 返回None，表示会话不存在或已过期
                
                tasks_count = len(state.get('loaded_tasks', []))
                logging.info(
                    f"[会话管理] 从文件加载会话 --> 会话ID: {session_id[:32]}..., 登录状态: {state.get('login_success')}, 任务数: {tasks_count}, 最后访问: {session_age_days:.1f}天前")
                return state
            else:
                logging.warning(f"[会话管理] 会话文件UUID不匹配，忽略")
    except Exception as e:
        logging.error(f"[会话管理] 加载会话状态失败 --> 错误: {e}", exc_info=True)
    return None


def cleanup_expired_sessions():
    """
    清理过期的会话文件（7天未访问）
    
    此函数遍历所有会话文件，删除超过7天未访问的会话。
    建议在应用启动时调用一次，以及定期后台调用。
    """
    try:
        if not os.path.exists(SESSION_STORAGE_DIR):
            return
        
        max_age_days = 7
        max_age_seconds = max_age_days * 86400
        current_time = time.time()
        cleaned_count = 0
        error_count = 0
        
        logging.info(f"[会话清理] 开始清理过期会话 --> 最大保留期限: {max_age_days}天")
        
        for filename in os.listdir(SESSION_STORAGE_DIR):
            if not filename.endswith('.json') or filename == '_index.json':
                continue
            
            session_file = os.path.join(SESSION_STORAGE_DIR, filename)
            
            try:
                with open(session_file, 'r', encoding='utf-8') as f:
                    state = json.load(f)
                
                last_accessed = state.get('last_accessed', 0)
                if current_time - last_accessed > max_age_seconds:
                    session_id = state.get('session_id', 'unknown')[:32]
                    age_days = (current_time - last_accessed) / 86400
                    
                    os.remove(session_file)
                    cleaned_count += 1
                    logging.info(f"[会话清理] 已删除过期会话 --> 会话ID: {session_id}..., 文件: {filename}, 年龄: {age_days:.1f}天")
                    
            except Exception as e:
                error_count += 1
                logging.debug(f"[会话清理] 处理会话文件失败 --> 文件: {filename}, 错误: {e}")
        
        logging.info(f"[会话清理] 清理完成 --> 已删除: {cleaned_count}个过期会话, 错误: {error_count}个")
        
    except Exception as e:
        logging.error(f"[会话清理] 清理过期会话失败 --> 错误: {e}", exc_info=True)


def restore_session_to_api_instance(api_instance, state):
    """
    将保存的会话状态恢复到Api实例

    Args:
        api_instance: Api实例
        state: 从文件加载的状态字典
    """
    try:
        # 恢复认证信息
        if 'auth_username' in state:
            api_instance.auth_username = state['auth_username']
            api_instance.auth_group = state.get('auth_group', 'guest')
            api_instance.is_guest = state.get('is_guest', False)
            api_instance.is_authenticated = state.get(
                'is_authenticated', False)

        # 恢复登录状态和用户信息
        if 'login_success' in state:
            api_instance.login_success = state['login_success']
        if 'user_info' in state:
            api_instance.user_info = state['user_info']

        # 恢复用户配置参数
        if 'params' in state:
            api_instance.params = state['params']

        # 恢复User-Agent
        if 'device_ua' in state:
            api_instance.device_ua = state['device_ua']

        # 恢复用户数据
        if 'user_data' in state:
            user_data_dict = state['user_data']
            api_instance.user_data.name = user_data_dict.get('name', '')
            api_instance.user_data.phone = user_data_dict.get('phone', '')
            api_instance.user_data.student_id = user_data_dict.get(
                'student_id', '')
            api_instance.user_data.id = user_data_dict.get('id', '')
            api_instance.user_data.username = user_data_dict.get(
                'username', '')
            api_instance.user_data.gender = user_data_dict.get('gender', '')
            api_instance.user_data.school_name = user_data_dict.get(
                'school_name', '')

        # 恢复任务列表
        if 'loaded_tasks' in state:
            api_instance.all_run_data = []
            for task_dict in state['loaded_tasks']:
                run_data = RunData()
                # 任务基本信息
                run_data.run_name = task_dict.get('run_name', '')
                run_data.errand_id = task_dict.get('errand_id', '')
                run_data.errand_schedule = task_dict.get('errand_schedule', '')
                run_data.status = task_dict.get('status', 0)
                run_data.start_time = task_dict.get('start_time', '')
                run_data.end_time = task_dict.get('end_time', '')
                run_data.upload_time = task_dict.get('upload_time', '')
                run_data.total_run_time_s = task_dict.get(
                    'total_run_time_s', 0.0)
                run_data.total_run_distance_m = task_dict.get(
                    'total_run_distance_m', 0.0)

                # 离线任务完整数据
                run_data.target_points = [
                    tuple(p) for p in task_dict.get('target_points', [])]
                run_data.target_point_names = task_dict.get(
                    'target_point_names', '')
                run_data.recommended_coords = [
                    tuple(p) for p in task_dict.get('recommended_coords', [])]
                run_data.draft_coords = [
                    tuple(p) for p in task_dict.get('draft_coords', [])]
                run_data.run_coords = [
                    tuple(p) for p in task_dict.get('run_coords', [])]

                # 运行时状态
                run_data.target_sequence = task_dict.get('target_sequence', 0)
                run_data.is_in_target_zone = task_dict.get(
                    'is_in_target_zone', False)
                run_data.trid = task_dict.get('trid', '')
                run_data.details_fetched = task_dict.get(
                    'details_fetched', False)
                run_data.distance_covered_m = task_dict.get(
                    'distance_covered_m', 0.0)

                api_instance.all_run_data.append(run_data)

        # 恢复任务选择状态
        if 'current_run_idx' in state:
            api_instance.current_run_idx = state['current_run_idx']

        # 恢复离线模式标志
        if 'is_offline_mode' in state:
            api_instance.is_offline_mode = state['is_offline_mode']
        
        # 恢复多账号模式状态
        if 'is_multi_account_mode' in state:
            api_instance.is_multi_account_mode = state['is_multi_account_mode']
            
            # 如果是多账号模式，恢复多账号相关数据
            if state['is_multi_account_mode']:
                # 恢复全局参数
                if 'multi_global_params' in state:
                    api_instance.global_params = state['multi_global_params']
                
                # 注意：这里不会完全恢复每个账号的完整状态
                # 因为账号状态包含复杂的运行时对象（如线程等）
                # 仅标记模式，前端会重新加载账号列表
                logging.info(f"会话恢复：检测到多账号模式，账号数: {len(state.get('multi_account_usernames', []))}")

        # 恢复停止标志状态
        if 'stop_run_flag_set' in state:
            if state['stop_run_flag_set']:
                api_instance.stop_run_flag.set()
            else:
                api_instance.stop_run_flag.clear()

        # 恢复UI状态
        if 'ui_state' in state:
            api_instance.ui_state = state['ui_state']

        # 恢复用户设置
        if 'user_settings' in state:
            api_instance.user_settings = state['user_settings']

        # 恢复 API Client Cookies (用于拉取通知等)
        if 'api_cookies' in state and state['api_cookies']:
            try:
                cookies_dict = state['api_cookies']
                # 将 cookie 字典加载回 session
                api_instance.api_client.session.cookies = requests.utils.cookiejar_from_dict(cookies_dict)
                logging.info(f"会话恢复: 成功恢复 {len(cookies_dict)} 个 API Cookies (shiroCookie等)。")
            except Exception as e:
                logging.warning(f"会话恢复: 恢复 API Cookies 失败: {e}")

        # 恢复自动签到后台线程 (如果启用)
        # 单账号模式
        if not api_instance.is_multi_account_mode and api_instance.params.get("auto_attendance_enabled", False):
            api_instance.stop_auto_refresh.clear()
            api_instance.auto_refresh_thread = threading.Thread(
                target=api_instance._auto_refresh_worker, daemon=True)
            api_instance.auto_refresh_thread.start()
            logging.info(f"会话恢复: 已重启单账号自动签到后台线程")
        
        # 多账号模式
        if api_instance.is_multi_account_mode and api_instance.global_params.get("auto_attendance_enabled", False):
            api_instance.stop_multi_auto_refresh.clear()
            api_instance.multi_auto_refresh_thread = threading.Thread(
                target=api_instance._multi_auto_attendance_worker, daemon=True)
            api_instance.multi_auto_refresh_thread.start()
            logging.info(f"会话恢复: 已重启多账号自动签到后台线程")

        logging.info(
            f"会话状态恢复完成: 任务数={len(api_instance.all_run_data)}, 选中索引={api_instance.current_run_idx}, 多账号模式={api_instance.is_multi_account_mode}")

    except Exception as e:
        logging.error(f"恢复会话状态失败: {e}", exc_info=True)


def load_all_sessions(args):
    """启动时加载所有持久化会话"""
    if not os.path.exists(SESSION_STORAGE_DIR):
        return

    # 加载或重建索引文件
    index = _load_session_index()
    new_index = {}

    successful_sessions = {}

    loaded_count = 0
    for filename in os.listdir(SESSION_STORAGE_DIR):
        if filename == '_index.json':
            continue

        if filename.endswith('.json'):
            session_file = os.path.join(SESSION_STORAGE_DIR, filename)
            session_id = None  # 先声明 session_id
            session_hash = filename[:-5]
            try:
                # 尝试读取文件基本信息
                with open(session_file, 'r', encoding='utf-8') as f:
                    state = json.load(f)

                session_id = state.get('session_id')
                if not session_id:
                    raise ValueError(f"文件 {filename} 缺少 session_id")

                # 检查会话是否过期
                last_accessed = state.get('last_accessed', 0)
                if time.time() - last_accessed > 7 * 24 * 3600:
                    logging.info(
                        f"清理过期会话: {session_id[:8]}... (文件: {filename})")
                    try:
                        os.remove(session_file)
                    except Exception as remove_err:
                        logging.error(f"删除过期会话文件 {filename} 失败: {remove_err}")
                    continue  # 过期则跳过，不尝试恢复

                # --- 尝试恢复完整状态 ---
                api_instance = Api(args)
                api_instance._session_created_at = state.get(
                    'created_at', time.time())
                # restore_session_to_api_instance 现在包含恢复 login_success 等所有状态
                restore_session_to_api_instance(api_instance, state)

                # --- 如果成功恢复 ---
                logging.info(
                    f"成功恢复会话: {session_id[:8]}... (用户: {api_instance.auth_username if hasattr(api_instance, 'auth_username') else 'Unknown'}, 文件: {filename})")
                web_sessions[session_id] = api_instance  # 加入内存
                session_activity[session_id] = last_accessed  # 恢复活动时间
                successful_sessions[session_id] = session_hash  # 记录成功加载的会话及其哈希
                loaded_count += 1

            except (json.JSONDecodeError, ValueError, KeyError, TypeError, AttributeError) as e:
                # 加载或恢复过程中发生任何错误
                # 可以关闭 exc_info 避免过多日志
                logging.error(f"加载或恢复会话文件 {filename} 失败: {e}", exc_info=False)
                logging.warning(f"将删除损坏的/无法恢复的会话文件: {filename}")
                try:
                    os.remove(session_file)  # 删除损坏的文件
                except Exception as remove_err:
                    logging.error(f"删除损坏的会话文件 {filename} 失败: {remove_err}")
                # 不将失败的会话加入 web_sessions 或 successful_sessions
                continue  # 继续处理下一个文件
            except Exception as e:  # 捕获其他意外错误
                logging.error(f"处理会话文件 {filename} 时发生未知错误: {e}", exc_info=True)
                try:
                    os.remove(session_file)  # 也尝试删除
                except Exception as remove_err:
                    logging.error(f"删除未知错误的会话文件 {filename} 失败: {remove_err}")
                continue

    # 保存重建的索引（只包含成功加载的会话）
    if successful_sessions:
        # 使用 successful_sessions 作为新的索引内容
        _save_session_index(successful_sessions)
        logging.info(f"会话索引已更新，包含 {len(successful_sessions)} 个有效会话")
    elif not any(f.endswith('.json') and f != '_index.json' for f in os.listdir(SESSION_STORAGE_DIR)):
        # 如果 sessions 目录为空 (除了 index)，则清空 index 文件
        _save_session_index({})
        logging.info("会话目录为空，已清空会话索引。")

    # 保存重建的索引（清理过期条目）
    if new_index:
        _save_session_index(new_index)
        logging.debug(f"会话索引已更新，包含 {len(new_index)} 个有效会话")

    if loaded_count > 0:
        logging.info(f"共加载 {loaded_count} 个持久化会话")

# Playwright浏览器池管理


class BackgroundTaskManager:
    """管理服务器端后台任务执行"""
    
    def __init__(self):
        self.tasks = {}  # {session_id: task_info}
        self.lock = threading.Lock()
        self.task_storage_dir = os.path.join(os.path.dirname(__file__), 'background_tasks')
        if not os.path.exists(self.task_storage_dir):
            os.makedirs(self.task_storage_dir)
        logging.info("BackgroundTaskManager initialized")
        
    def _get_task_file_path(self, session_id):
        """获取任务状态文件路径"""
        task_hash = hashlib.sha256(session_id.encode()).hexdigest()
        return os.path.join(self.task_storage_dir, f"{task_hash}.json")
    
    def save_task_state(self, session_id, task_state):
        """保存任务状态到文件"""
        task_file = self._get_task_file_path(session_id)
        try:
            with open(task_file, 'w', encoding='utf-8') as f:
                json.dump(task_state, f, indent=2, ensure_ascii=False)
            logging.debug(f"后台任务状态已保存，会话ID前缀: {session_id[:8]}")
        except Exception as e:
            logging.error(f"保存后台任务状态失败: {e}")
    
    def load_task_state(self, session_id):
        """从文件加载任务状态"""
        task_file = self._get_task_file_path(session_id)
        if not os.path.exists(task_file):
            return None
        try:
            with open(task_file, 'r', encoding='utf-8') as f:
                task_state = json.load(f)
            logging.debug(f"后台任务状态已加载，会话ID前缀: {session_id[:8]}")
            return task_state
        except Exception as e:
            logging.error(f"加载后台任务状态失败: {e}")
            return None
    
    def start_background_task(self, session_id, api_instance, task_indices, auto_generate=False):
        """启动后台任务执行"""
        with self.lock:
            # 初始化任务状态
            task_state = {
                'session_id': session_id,
                'total_tasks': len(task_indices),
                'completed_tasks': 0,
                'current_task_index': 0,
                'task_indices': task_indices,
                'auto_generate': auto_generate,
                'status': 'running',
                'start_time': time.time(),
                'last_update': time.time(),
                'progress_percent': 0,
                'current_task_progress': 0,
                'singleProcessedPoints': 0,  # 已处理的GPS点数
                'singleTotalPoints': 0  # 总GPS点数
            }
            
            self.tasks[session_id] = task_state
            self.save_task_state(session_id, task_state)
            
            # 启动后台线程执行任务
            thread = threading.Thread(
                target=self._execute_tasks_background,
                args=(session_id, api_instance, task_indices, auto_generate),
                daemon=True
            )
            thread.start()
            
            logging.info(f"后台任务已启动，会话ID前缀: {session_id[:8]}, 总任务数: {len(task_indices)}")
            return {"success": True, "message": f"已启动后台任务，共{len(task_indices)}个任务"}
    
    def _execute_tasks_background(self, session_id, api_instance, task_indices, auto_generate):
        """后台执行任务的线程函数"""
        try:
            tasks_executed = 0  # 追踪实际执行的任务数量
            
            for i, task_idx in enumerate(task_indices):
                # 检查是否需要停止
                with self.lock:
                    if session_id not in self.tasks:
                        logging.info(f"后台任务已取消，会话ID前缀: {session_id[:8]}")
                        return
                    
                    task_state = self.tasks[session_id]
                    if task_state.get('status') == 'stopped':
                        logging.info(f"后台任务已停止，会话ID前缀: {session_id[:8]}")
                        return
                
                # 更新当前任务
                with self.lock:
                    task_state['current_task_index'] = i
                    task_state['last_update'] = time.time()
                    self.save_task_state(session_id, task_state)
                
                # 执行单个任务
                logging.info(f"正在执行后台任务 {i+1}/{len(task_indices)}，会话ID前缀: {session_id[:8]}")
                run_data = api_instance.all_run_data[task_idx]
                
                # 如果需要自动生成路径
                if auto_generate and not run_data.run_coords:
                    logging.info(f"正在为任务自动生成路径: 任务名称={run_data.run_name}")
                    try:
                        # 确保已加载任务详情，拿到打卡点
                        if not run_data.details_fetched:
                            logging.info(f"正在获取任务详细信息: {run_data.run_name}...")
                            details_resp = api_instance.get_task_details(task_idx)
                            if not details_resp.get("success"):
                                logging.error(f"获取任务详情失败，任务名称: {run_data.run_name}，错误信息: {details_resp.get('message', '未知错误')}")
                                continue
                            run_data = api_instance.all_run_data[task_idx]
                            logging.info(f"任务详情获取成功: {run_data.run_name}")
                        
                        if not run_data.target_points:
                            logging.error(f"任务缺少目标打卡点，无法自动生成路径: {run_data.run_name}")
                            continue
                        
                        logging.info(f"任务包含 {len(run_data.target_points)} 个目标打卡点: {run_data.run_name}")
                        
                        # 使用Chrome池进行服务器端路径规划
                        waypoints = run_data.target_points
                        logging.info(f"正在规划路径，包含 {len(waypoints)} 个路点，任务名称: {run_data.run_name}")
                        logging.info(f"Waypoints: {waypoints[:3]}..." if len(waypoints) > 3 else f"Waypoints: {waypoints}")
                        
                        
                        # 必须在调用Chrome池之前获取 Key
                        amap_key = api_instance.global_params.get('amap_js_key', '')
                        if not amap_key:
                            logging.error(f"无法为 {run_data.run_name} 自动规划路径：缺少高德地图 API Key")
                            # 标记任务状态为错误并跳过
                            with self.lock:
                                task_state['status'] = 'error'
                                task_state['error'] = '缺少高德地图API Key'
                                self.save_task_state(session_id, task_state)
                            continue # 跳过此任务
                        

                        # 调用Chrome池执行路径规划
                        global chrome_pool
                        if not chrome_pool:
                            logging.error("Chrome浏览器池不可用，无法进行路径规划！")
                            continue
                            
                        if chrome_pool:
                            try:
                                logging.info(f"正在获取Chrome浏览器上下文，会话ID前缀: {session_id[:8]}...")
                                # 获取页面并确保加载了AMap
                                ctx = chrome_pool.get_context(session_id)
                                page = ctx['page']
                                logging.info("Chrome浏览器上下文获取成功")
                                
                                # 首先加载包含AMap的页面
                                logging.info("正在向Chrome页面加载高德地图SDK...")
                                page.goto("about:blank")
                                page.set_content("""
                                <!DOCTYPE html>
                                <html>
                                <head>
                                    <meta charset="utf-8">
                                    <script type="text/javascript" src="https://webapi.amap.com/loader.js"></script>
                                </head>
                                <body></body>
                                </html>
                                """)
                                
                                # 等待AMap加载 (修复：等待 AMapLoader 加载完成)
                                logging.info("等待高德地图加载器(AMapLoader)加载完成...")
                                page.wait_for_function("typeof AMapLoader !== 'undefined'", timeout=10000)
                                logging.info("高德地图加载器在Chrome上下文中加载成功")
                                
                                # 使用Chrome池执行AMap路径规划JavaScript
                                logging.info("正在Chrome浏览器中执行路径规划JavaScript代码...")
                                path_coords = chrome_pool.execute_js(
                                    session_id,
                                    """
                                    async function planPath(waypoints, apiKey) {
                                        // 1. 确保 AMapLoader (来自 loader.js) 存在
                                        if (typeof AMapLoader === 'undefined') {
                                            return {error: 'AMapLoader not loaded'};
                                        }

                                        // 2. (修复BUG) 调用 AMapLoader.load 并传入 key
                                        try {
                                            // 确保 AMap 和 AMap.Walking 插件被加载
                                            await AMapLoader.load({
                                                "key": apiKey,
                                                "version": "2.0",
                                                "plugins": ["AMap.Walking"]
                                            });
                                        } catch (e) {
                                            // 返回详细的加载错误
                                            return {error: 'AMapLoader.load failed: ' + (e ? e.message : 'Unknown error')};
                                        }

                                        // 3. (修复BUG) 检查 AMap.Walking 插件是否真的加载成功
                                        if (typeof AMap.Walking === 'undefined') {
                                            return {error: 'AMap.Walking plugin failed to load'};
                                        }

                                        // 4. (原逻辑) 执行路径规划
                                        return new Promise((resolve) => {
                                            const walking = new AMap.Walking({
                                                map: null  // 不需要地图显示
                                            });
                                            
                                            // 将waypoints转换为AMap.LngLat格式
                                            const points = waypoints.map(p => new AMap.LngLat(p[0], p[1]));
                                            
                                            walking.search(points, function(status, result) {
                                                if (status === 'complete' && result.routes && result.routes.length > 0) {
                                                    const route = result.routes[0];
                                                    const path = [];
                                                    route.steps.forEach(step => {
                                                        if (step.path) {
                                                            step.path.forEach(p => {
                                                                path.push({lng: p.lng, lat: p.lat});
                                                            });
                                                        }
                                                    });
                                                    resolve({path: path});
                                                } else {
                                                    // 返回更详细的错误信息
                                                    const errorInfo = result ? result.info : status;
                                                    resolve({error: 'Path planning failed: ' + errorInfo});
                                                }
                                            });
                                        });
                                    }
                                    
                                    // 修复BUG：将 apiKey 作为 arguments[1] 传入
                                    return planPath(arguments[0], arguments[1]);
                                    """,
                                    waypoints,
                                    amap_key  # 修复BUG：传入从Python获取的 amap_key
                                )
                                
                                logging.info(f"路径规划JavaScript返回结果: 类型={type(path_coords)}, 包含'path'键={'是' if (path_coords and 'path' in path_coords) else '否'}")
                                
                                if path_coords and 'path' in path_coords:
                                    api_path_coords = path_coords['path']
                                    logging.info(f"路径规划成功，包含 {len(api_path_coords)} 个坐标点")
                                    
                                    # 使用后端生成 run_coords
                                    p = api_instance.params
                                    logging.info(f"正在生成运动模拟数据，参数: 最小时长={p.get('min_time_m', 20)}分钟, 最大时长={p.get('max_time_m', 30)}分钟, 最小距离={p.get('min_dist_m', 2000)}米")
                                    gen_resp = api_instance.auto_generate_path_with_api(
                                        api_path_coords,
                                        p.get("min_time_m", 20),
                                        p.get("max_time_m", 30),
                                        p.get("min_dist_m", 2000)
                                    )
                                    
                                    logging.info(f"auto_generate_path_with_api函数返回: 成功={gen_resp.get('success')}")
                                    
                                    if gen_resp.get("success"):
                                        # 将生成结果回填到当前任务
                                        run_data.run_coords = gen_resp["run_coords"]
                                        run_data.total_run_distance_m = gen_resp["total_dist"]
                                        run_data.total_run_time_s = gen_resp["total_time"]
                                        logging.info(f"路径自动生成成功，任务: {run_data.run_name}，坐标点数: {len(gen_resp['run_coords'])}, 总距离: {gen_resp['total_dist']}米, 总时长: {gen_resp['total_time']}秒")
                                    else:
                                        logging.error(f"生成运动坐标序列失败: {gen_resp.get('message')}")
                                        continue
                                else:
                                    error_msg = path_coords.get('error', 'Unknown error') if path_coords else 'No response from path planning'
                                    logging.error(f"任务路径规划失败，任务名称: {run_data.run_name}，错误信息: {error_msg}")
                                    continue
                            except Exception as e:
                                logging.error(f"Chrome浏览器池路径规划失败，任务名称: {run_data.run_name}，异常信息: {e}", exc_info=True)
                                continue
                        else:
                            logging.error("Chrome浏览器池不可用，无法进行路径规划")
                            continue
                            
                    except Exception as e:
                        logging.error(f"自动生成路径失败，异常信息: {e}", exc_info=True)
                        continue
                
                # 检查任务是否有路径
                if not run_data.run_coords:
                    logging.warning(f"任务没有可用路径，跳过执行: {run_data.run_name}")
                    # 跳过此任务，继续执行下一个任务
                    continue
                
                # 设置当前任务
                api_instance.current_run_idx = task_idx
                run_data.target_sequence = 0  # ✓ 从0开始（0-based索引）
                run_data.is_in_target_zone = False
                api_instance._first_center_done = False
                api_instance.stop_run_flag.clear()
                
                # 创建完成事件
                finished_event = threading.Event()
                
                # 执行任务（使用实际的提交线程）
                try:
                    # 调用实际的执行逻辑
                    thread = threading.Thread(
                        target=api_instance._run_submission_thread,
                        args=(run_data, task_idx, api_instance.api_client, False, finished_event),
                        daemon=True
                    )
                    thread.start()
                    
                    # 标记正在执行此任务
                    tasks_executed += 1
                    
                    # 等待任务完成或超时（最多等待任务预计时间的2倍）
                    total_time_s = sum(p[2] for p in run_data.run_coords) / 1000.0
                    timeout = max(total_time_s * 2, 300)  # 至少5分钟
                    
                    # 监控任务进度并更新状态
                    start_wait = time.time()
                    while not finished_event.is_set():
                        if time.time() - start_wait > timeout:
                            logging.warning(f"任务执行超时: {run_data.run_name}")
                            api_instance.stop_run_flag.set()
                            break
                        
                        # 更新当前任务进度
                        with self.lock:
                            if hasattr(run_data, 'current_point_index'):
                                total_points = len(run_data.run_coords)
                                current_progress = int(run_data.current_point_index / total_points * 100)
                                task_state['current_task_progress'] = current_progress
                                task_state['last_update'] = time.time()
                                
                                # 添加点数进度信息（供前端直接使用）
                                current_idx = run_data.current_point_index
                                task_state['singleProcessedPoints'] = current_idx
                                task_state['singleTotalPoints'] = total_points
                                
                                # 添加详细的任务数据（供前端完整恢复任务状态）
                                task_state['target_points'] = run_data.target_points if hasattr(run_data, 'target_points') else []
                                task_state['target_point_names'] = run_data.target_point_names if hasattr(run_data, 'target_point_names') else ''
                                task_state['recommended_coords'] = run_data.recommended_coords if hasattr(run_data, 'recommended_coords') else []
                                task_state['run_coords'] = run_data.run_coords if hasattr(run_data, 'run_coords') else []
                                
                                # 添加打卡点进度信息
                                task_state['checked_targets_count'] = getattr(run_data, 'target_sequence', 0)  # 已打卡的打卡点数量
                                task_state['total_targets_count'] = len(run_data.target_points) if hasattr(run_data, 'target_points') else 0
                                
                                # 添加时间和距离信息
                                task_state['elapsed_time_s'] = time.time() - start_wait  # 当前已用时间（秒）
                                task_state['current_distance_m'] = getattr(run_data, 'distance_covered_m', 0)  # 当前已跑距离（米）
                                task_state['estimated_total_time_s'] = getattr(run_data, 'total_run_time_s', 0)  # 预计总时间（秒）
                                task_state['estimated_total_distance_m'] = getattr(run_data, 'total_run_distance_m', 0)  # 预计总距离（米）
                                
                                # 添加实时位置数据
                                if current_idx > 0 and current_idx <= total_points:
                                    coord = run_data.run_coords[current_idx - 1]
                                    task_state['current_position'] = {
                                        'lon': coord[0],
                                        'lat': coord[1],
                                        'distance': getattr(run_data, 'distance_covered_m', 0),
                                        'target_sequence': getattr(run_data, 'target_sequence', 0),
                                        'point_index': current_idx  # 添加点索引用于前端进度计算
                                    }
                        
                        # 每5秒保存一次状态
                        if int(time.time() - start_wait) % 5 == 0:
                            with self.lock:
                                self.save_task_state(session_id, task_state)
                        
                        time.sleep(1)
                    
                    # 等待线程结束
                    thread.join(timeout=10)
                    
                except Exception as e:
                    logging.error(f"任务执行失败，异常信息: {e}", exc_info=True)
                
                # 更新完成状态
                with self.lock:
                    task_state['completed_tasks'] = i + 1
                    task_state['progress_percent'] = int((i + 1) / len(task_indices) * 100)
                    task_state['current_task_progress'] = 100
                    task_state['last_update'] = time.time()
                    self.save_task_state(session_id, task_state)
                
                logging.info(f"任务 {i+1}/{len(task_indices)} 已完成，会话ID前缀: {session_id[:8]}")
            
            # 所有任务完成 - 仅当至少执行了一个任务时
            if tasks_executed > 0:
                with self.lock:
                    task_state['status'] = 'completed'
                    task_state['last_update'] = time.time()
                    self.save_task_state(session_id, task_state)
                
                logging.info(f"所有后台任务已完成，会话ID前缀: {session_id[:8]}")
            else:
                # 没有任务被执行
                with self.lock:
                    if task_state.get('status') != 'error':  # 不覆盖错误状态
                        task_state['status'] = 'error'
                        # 根据是否启用自动生成，设置不同的错误消息
                        if auto_generate:
                            task_state['error'] = '自动生成路径失败，请查看服务器日志了解详情或手动生成路径'
                        else:
                            task_state['error'] = '所有任务都没有路径，请先生成路径'
                        task_state['last_update'] = time.time()
                        self.save_task_state(session_id, task_state)
                if auto_generate:
                    logging.error(f"没有任务被执行，会话ID前缀: {session_id[:8]} - 所有自动生成路径尝试均失败")
                else:
                    logging.warning(f"没有任务被执行，会话ID前缀: {session_id[:8]} - 没有可用的路径")
            
        except Exception as e:
            logging.error(f"后台任务执行失败，异常信息: {e}", exc_info=True)
            with self.lock:
                if session_id in self.tasks:
                    self.tasks[session_id]['status'] = 'error'
                    self.tasks[session_id]['error'] = str(e)
                    self.save_task_state(session_id, self.tasks[session_id])
    
    def get_task_status(self, session_id):
        """获取任务状态"""
        with self.lock:
            # 先从内存中获取
            if session_id in self.tasks:
                return self.tasks[session_id]
        
        # 如果内存中没有，从文件加载
        task_state = self.load_task_state(session_id)
        if task_state:
            with self.lock:
                self.tasks[session_id] = task_state
            return task_state
        
        return None
    
    def stop_task(self, session_id):
        """停止后台任务"""
        with self.lock:
            if session_id in self.tasks:
                self.tasks[session_id]['status'] = 'stopped'
                self.save_task_state(session_id, self.tasks[session_id])
                logging.info(f"后台任务已停止，会话ID前缀: {session_id[:8]}")
                return {"success": True, "message": "后台任务已停止"}
            return {"success": False, "message": "未找到运行中的后台任务"}
    
    def cleanup_old_tasks(self, max_age_hours=24):
        """清理旧的任务状态文件"""
        try:
            current_time = time.time()
            max_age_seconds = max_age_hours * 3600
            
            for filename in os.listdir(self.task_storage_dir):
                if not filename.endswith('.json'):
                    continue
                
                filepath = os.path.join(self.task_storage_dir, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        task_state = json.load(f)
                    
                    last_update = task_state.get('last_update', 0)
                    if current_time - last_update > max_age_seconds:
                        os.remove(filepath)
                        logging.info(f"已删除旧的任务状态文件: {filename}")
                except Exception as e:
                    logging.warning(f"处理任务文件失败，文件名: {filename}，错误: {e}")
        except Exception as e:
            logging.error(f"清理旧任务文件失败，异常信息: {e}")


class ChromeBrowserPool:
    """管理服务器端Chrome浏览器实例，用于执行JS计算"""

    def __init__(self, headless=True, max_instances=5):
        self.headless = headless
        self.max_instances = max_instances
        self.playwright = None
        self.browser = None
        self.contexts = {}  # {session_id: browser_context}
        self.lock = threading.Lock()
        self._initialize()

    def _initialize(self):
        """初始化Playwright和浏览器"""
        try:
            self.playwright = sync_playwright().start()
            self.browser = self.playwright.chromium.launch(
                headless=self.headless,
                args=['--no-sandbox', '--disable-setuid-sandbox']
            )
            logging.info(f"Chrome浏览器已启动 (headless={self.headless})")
        except Exception as e:
            logging.error(f"启动Chrome失败: {e}")
            raise

    def get_context(self, session_id):
        """获取或创建指定会话的浏览器上下文"""
        with self.lock:
            if session_id not in self.contexts:
                # 创建新的浏览器上下文
                context = self.browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                )
                page = context.new_page()
                self.contexts[session_id] = {'context': context, 'page': page}
                logging.info(f"为会话 {session_id} 创建Chrome上下文")
            return self.contexts[session_id]

    def execute_js(self, session_id, script, *args):
        """在指定会话的Chrome中执行JavaScript代码"""
        try:
            ctx = self.get_context(session_id)
            page = ctx['page']
            result = page.evaluate(script, *args)
            return result
        except Exception as e:
            logging.error(f"执行JS失败 (session={session_id}): {e}")
            return None

    def close_context(self, session_id):
        """关闭指定会话的浏览器上下文"""
        with self.lock:
            if session_id in self.contexts:
                try:
                    ctx = self.contexts[session_id]
                    ctx['context'].close()
                    del self.contexts[session_id]
                    logging.info(f"关闭会话 {session_id} 的Chrome上下文")
                except Exception as e:
                    logging.error(f"关闭上下文失败: {e}")

    def cleanup(self):
        """清理所有资源"""
        with self.lock:
            # 首先尝试关闭所有上下文
            for session_id in list(self.contexts.keys()):
                try:
                    # 增加日志，明确关闭哪个上下文
                    logging.debug(
                        f"正在关闭会话 {session_id[:8]}... 的 Chrome 上下文...")
                    self.contexts[session_id]['context'].close()
                    logging.debug(f"会话 {session_id[:8]}... 的上下文已关闭。")
                except Exception as e:
                    # 捕获关闭上下文时的错误，记录并继续
                    logging.warning(f"关闭会话 {session_id[:8]}... 上下文失败: {e}")
            self.contexts.clear()
            logging.debug("所有 Chrome 上下文已清理。")

            # 然后尝试关闭浏览器
            if self.browser:
                try:
                    logging.debug("正在关闭主 Chrome 浏览器实例...")
                    self.browser.close()  # <--- 在这里捕获异常
                    logging.debug("主 Chrome 浏览器实例已关闭。")
                except Exception as e:
                    # 捕获关闭浏览器时的错误，记录日志但允许程序继续退出
                    logging.warning(f"关闭 Chrome 浏览器实例时发生错误: {e}")
                    # 注意：这里我们只记录警告，因为程序正在退出

            # 最后停止 Playwright 实例
            if self.playwright:
                try:
                    logging.debug("正在停止 Playwright 实例...")
                    self.playwright.stop()
                    logging.debug("Playwright 实例已停止。")
                except Exception as e:
                    # 捕获停止 Playwright 时的错误
                    logging.warning(f"停止 Playwright 实例时发生错误: {e}")


# 全局Chrome浏览器池和后台任务管理器
chrome_pool = None
background_task_manager = None


def _cleanup_playwright():
    """
    在程序退出时清理Playwright资源。
    
    功能说明：
    - 这是一个注册在atexit的清理函数，确保程序退出时正确关闭所有浏览器实例
    - 防止遗留的浏览器进程占用系统资源
    
    清理流程：
    1. 检查chrome_pool全局实例是否存在
    2. 调用chrome_pool.cleanup()关闭所有浏览器
    3. 记录清理结果
    
    异常处理：
    - 使用try-except捕获清理过程中的所有异常
    - exc_info=False避免在退出时打印完整堆栈（可能引起混淆）
    - 只记录错误，不重新抛出，确保程序正常退出
    
    设计考虑：
    - 使用global访问chrome_pool而不是参数传递，因为atexit不支持参数
    - 判空检查避免未初始化时的错误
    
    使用场景：
    - 正常退出（Ctrl+C、关闭终端）
    - 异常退出（未捕获的异常）
    - 系统信号（SIGTERM等）
    
    注意：
    - 此函数由atexit.register()自动调用，不需手动调用
    - 执行时机在Python解释器关闭前
    """
    global chrome_pool
    if chrome_pool:
        logging.info("捕获到程序退出信号，正在清理 Playwright 资源...")
        try:
            chrome_pool.cleanup()
            logging.info("Playwright浏览器自动化框架资源清理完成")
        except Exception as e:
            # 在退出时尽量不抛出新异常，只记录错误
            logging.error(f"清理 Playwright 资源时发生错误: {e}", exc_info=False)
    else:
        logging.debug("Playwright 池未初始化，无需清理。")

def start_background_auto_attendance(args):
    """
    在服务器启动时扫描所有.ini配置文件，为启用自动签到的账号启动后台工作线程。
    
    功能说明：
    - 自动发现并加载所有启用了auto_attendance_enabled的账号配置
    - 根据账号数量智能选择单账号或多账号模式
    - 在后台守护线程中持续运行，无需用户干预
    
    工作模式：
    
    **单账号模式**（只有1个启用自动签到的账号）：
    - 为该账号创建独立的Api实例
    - 启动单独的_auto_refresh_worker线程
    - 资源占用更少，适合个人使用
    
    **多账号模式**（2个或更多账号）：
    - 所有账号共享一个Api实例
    - 启动_multi_auto_attendance_worker统一管理
    - 更高效的资源利用，适合批量管理
    
    执行流程：
    1. 扫描SCHOOL_ACCOUNTS_DIR目录中的所有.ini文件
    2. 读取每个文件的配置，检查auto_attendance_enabled参数
    3. 收集所有启用自动签到的账号信息
    4. 根据账号数量选择模式并启动相应的后台线程
    
    参数说明：
    - args: 命令行参数对象，包含--no-auto-start等配置
    
    异常处理：
    - 单个账号加载失败不影响其他账号
    - 登录失败会记录错误但不终止服务
    - 所有异常都会记录到日志中
    
    全局变量：
    - _background_service_api: 保持Api实例存活的引用（单账号模式）
    - chrome_pool: 共享的浏览器池（如果需要）
    
    设计考虑：
    - 使用daemon线程确保主程序退出时自动停止
    - 临时创建Api实例仅用于读取配置，避免资源浪费
    - 账号信息缓存在内存中，避免重复读取文件
    
    注意事项：
    - 此函数应在Flask服务器启动前调用
    - 需要确保SCHOOL_ACCOUNTS_DIR目录存在且有读权限
    - 账号的.ini文件必须包含密码和auto_attendance_enabled配置
    
    使用示例：
    ```python
    if __name__ == "__main__":
        args = parse_args()
        if not args.no_auto_start:
            start_background_auto_attendance(args)
        app.run()
    ```
    """
    try:
        logging.info("正在启动后台自动签到服务...")
        
        accounts_dir = SCHOOL_ACCOUNTS_DIR
        
        if not os.path.exists(accounts_dir):
            logging.warning(f"后台签到：未找到账号目录 {accounts_dir}，跳过。")
            return

        # 收集所有启用自动签到的账号
        enabled_accounts = []
        
        for filename in os.listdir(accounts_dir):
            if filename.endswith(".ini"):
                username = os.path.splitext(filename)[0]
                try:
                    # 临时创建一个Api实例来加载配置
                    temp_api = Api(args)
                    password = temp_api._load_config(username)
                    
                    if not password:
                        logging.debug(f"后台签到：跳过账号 {username}，因为未在 {filename} 中找到密码。")
                        continue
                    
                    # 检查是否启用了自动签到
                    if temp_api.params.get("auto_attendance_enabled", False):
                        enabled_accounts.append({
                            'username': username,
                            'password': password,
                            'params': dict(temp_api.params)
                        })
                        logging.info(f"后台签到：找到启用自动签到的账号: {username}")
                    else:
                        logging.debug(f"后台签到：账号 {username} 未启用自动签到，跳过。")
                        
                except Exception as e:
                    logging.error(f"后台签到：加载账号 {username} 失败: {e}", exc_info=True)

        if not enabled_accounts:
            logging.info("后台签到：未找到启用自动签到的账号。")
            return
        
        # 根据账号数量选择模式
        if len(enabled_accounts) == 1:
            # 单账号模式：为该账号创建独立的Api实例
            account = enabled_accounts[0]
            logging.info(f"后台签到：使用单账号模式，账号: {account['username']}")
            
            service_api = Api(args)
            service_api.is_multi_account_mode = False
            service_api.params = account['params']
            
            # 执行登录
            try:
                login_result = service_api.login(account['username'], account['password'])
                if login_result.get('success'):
                    logging.info(f"后台签到：账号 {account['username']} 登录成功")
                    
                    # 启动单账号自动刷新线程
                    service_api.stop_auto_refresh.clear()
                    service_api.auto_refresh_thread = threading.Thread(
                        target=service_api._auto_refresh_worker,
                        daemon=True,
                        name=f"BackgroundAttendance-{account['username']}"
                    )
                    service_api.auto_refresh_thread.start()
                    
                    # 保持实例存活
                    globals()['_background_service_api'] = service_api
                    logging.info(f"后台签到：单账号模式启动成功")
                else:
                    logging.error(f"后台签到：账号 {account['username']} 登录失败: {login_result.get('message')}")
            except Exception as e:
                logging.error(f"后台签到：账号 {account['username']} 登录时发生错误: {e}", exc_info=True)
                
        else:
            # 多账号模式：所有账号共享一个Api实例
            logging.info(f"后台签到：使用多账号模式，共 {len(enabled_accounts)} 个账号")
            
            service_api = Api(args)
            service_api.is_multi_account_mode = True
            service_api._load_global_config()
            service_api.global_params["auto_attendance_enabled"] = True
            
            # 为每个账号创建AccountSession
            for account in enabled_accounts:
                try:
                    acc_session = AccountSession(account['username'], account['password'], service_api)
                    acc_session.params = account['params']
                    service_api.accounts[account['username']] = acc_session
                except Exception as e:
                    logging.error(f"后台签到：创建账号会话失败 {account['username']}: {e}", exc_info=True)
            
            # 启动多账号自动刷新线程
            service_api.stop_multi_auto_refresh.clear()
            service_api.multi_auto_refresh_thread = threading.Thread(
                target=service_api._multi_auto_attendance_worker,
                daemon=True,
                name="BackgroundAttendanceWorker-Multi"
            )
            service_api.multi_auto_refresh_thread.start()
            
            # 保持实例存活
            globals()['_background_service_api'] = service_api
            logging.info(f"后台签到：多账号模式启动成功，已加载 {len(service_api.accounts)} 个账号")
            
    except Exception as e:
        logging.error(f"启动后台自动签到服务时发生严重错误: {e}", exc_info=True)


def start_web_server(args_param):
    """
    启动Flask Web服务器主函数，集成SocketIO实时通信和Chrome浏览器自动化。
    
    功能说明：
    - 初始化Flask应用和所有必需的全局组件
    - 配置跨域请求（CORS）和WebSocket实时通信（SocketIO）
    - 启动Chrome浏览器池用于服务端JS渲染
    - 初始化后台任务管理器
    - 设置会话管理和安全密钥
    
    主要组件：
    
    **1. Chrome浏览器池（ChromeBrowserPool）**
    - 管理可重用的浏览器实例
    - 支持headless模式（无界面运行）
    - 自动清理和资源回收
    
    **2. 后台任务管理器（BackgroundTaskManager）**
    - 管理长时间运行的任务（如批量任务执行）
    - 任务状态持久化到文件
    - 启动时清理历史任务记录
    
    **3. 会话管理系统**
    - web_sessions: 存储用户会话状态（登录信息、Api实例）
    - session_file_locks: 防止会话文件并发冲突
    - session_activity: 跟踪会话活跃时间
    - 所有会话数据都有对应的线程锁保护
    
    **4. SocketIO实时通信**
    - async_mode='threading': 使用线程模式处理异步请求
    - 支持后台任务进度实时推送
    - 双向通信（服务器可主动推送消息到客户端）
    
    参数说明：
    - args_param: 命令行参数对象，包含headless、port等配置
    
    初始化流程：
    1. 重置所有内存锁和会话状态（防止重启后的状态污染）
    2. 初始化Chrome浏览器池并注册退出清理函数
    3. 初始化后台任务管理器并清理历史任务
    4. 创建Flask应用并配置CORS、SocketIO
    5. 配置会话管理（SESSION_TYPE=filesystem，7天有效期）
    6. 注册所有Flask路由（在后续代码中）
    7. 启动Flask开发服务器
    
    全局变量：
    - chrome_pool: ChromeBrowserPool实例
    - background_task_manager: BackgroundTaskManager实例
    - web_sessions: 用户会话字典 {session_id: {'api': Api实例, ...}}
    - web_sessions_lock: 保护web_sessions的线程锁
    - session_file_locks: 会话文件锁字典 {username: Lock}
    - session_file_locks_lock: 保护session_file_locks的线程锁
    - session_activity: 会话活跃时间字典 {session_id: timestamp}
    - session_activity_lock: 保护session_activity的线程锁
    - socketio: SocketIO实例用于实时通信
    - args: 命令行参数（全局可访问）
    
    安全特性：
    - 使用secrets.token_hex(32)生成强随机密钥（256位）
    - 会话数据加密存储在文件系统中
    - 跨域请求受CORS控制
    
    错误处理：
    - Chrome池或任务管理器初始化失败会sys.exit(1)终止程序
    - 单个任务文件删除失败只记录错误，不影响整体启动
    
    注意事项：
    - 此函数会阻塞当前线程（Flask服务器运行在主线程）
    - 需要先调用check_install_dependencies()确保依赖已安装
    - 建议在后台自动签到服务启动后调用
    
    使用示例：
    ```python
    if __name__ == "__main__":
        args = parse_args()
        check_install_dependencies()
        initialize_dirs()
        if not args.no_auto_start:
            start_background_auto_attendance(args)
        start_web_server(args)  # 此函数会阻塞
    ```
    """
    global chrome_pool, background_task_manager, web_sessions, web_sessions_lock, session_file_locks, session_file_locks_lock, session_activity, session_activity_lock, args

    # Make args available globally for Flask routes
    # 将命令行参数存储为全局变量，供Flask路由函数访问
    args = args_param

    # --- 新增：显式初始化/重置内存锁状态 ---
    # 重置所有会话相关的全局变量，防止程序重启后出现状态污染
    web_sessions = {}
    web_sessions_lock = threading.Lock()
    session_file_locks = {}
    session_file_locks_lock = threading.Lock()
    session_activity = {}
    session_activity_lock = threading.Lock()
    logging.info("内存锁和会话状态已重置。")
    # --- 结束新增 ---

    # 初始化Chrome浏览器池
    try:
        chrome_pool = ChromeBrowserPool(
            headless=getattr(args, 'headless', True))
        logging.info("Chrome浏览器池初始化成功")
        atexit.register(_cleanup_playwright)
        logging.info("已注册 Playwright 退出清理函数。")
    except Exception as e:
        logging.error(f"无法初始化Chrome浏览器池: {e}")
        sys.exit(1)
    
    
    # 初始化后台任务管理器
    try:
        background_task_manager = BackgroundTaskManager()
        logging.info("后台任务管理器初始化成功")

        logging.info("程序启动：正在清理所有历史后台任务记录（内存和文件）...")
        cleaned_files_count = 0
        if background_task_manager and hasattr(background_task_manager, 'task_storage_dir'):
            task_dir = background_task_manager.task_storage_dir
            if os.path.exists(task_dir):
                # 清理内存
                with background_task_manager.lock:
                    background_task_manager.tasks.clear()
                # 清理文件
                for filename in os.listdir(task_dir):
                    if filename.endswith(".json"):
                        file_path = os.path.join(task_dir, filename)
                        try:
                            os.remove(file_path)
                            cleaned_files_count += 1
                        except Exception as e:
                            logging.error(f"无法删除后台任务文件 {filename}: {e}")
            logging.info(f"已清空后台任务管理器内存状态，并删除了 {cleaned_files_count} 个任务状态文件。")
        else:
            logging.warning("无法清理后台任务文件：BackgroundTaskManager 或其 task_storage_dir 未定义。")

    except Exception as e:
        logging.error(f"无法初始化后台任务管理器: {e}")
        sys.exit(1)

    app = Flask(__name__)
    app.secret_key = secrets.token_hex(32)  # 生成安全的密钥
    CORS(app)  # 允许跨域请求

    # 声明 socketio 为全局变量
    global socketio
    # 初始化 SocketIO
    socketio = SocketIO(app, async_mode='threading')

    # 会话配置
    app.config['SESSION_TYPE'] = 'filesystem'
    app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(
        days=7)  # 会话保持7天

    # ====================
    # 认证相关API路由
    # ====================

    @app.route('/auth/register', methods=['POST'])
    def auth_register():
        """
        用户注册API端点。
        
        请求方法：POST
        请求路径：/auth/register
        Content-Type：application/json
        
        请求体（JSON）：
        {
            "auth_username": "用户名",  # 必填，会被trim()处理
            "auth_password": "密码"     # 必填，会被trim()处理
        }
        
        响应体（JSON）：
        {
            "success": true/false,
            "message": "成功/失败信息"
        }
        
        处理流程：
        1. 解析JSON请求体（使用 or {} 防止None）
        2. 提取并trim用户名和密码
        3. 验证非空（两者都必填）
        4. 调用auth_system.register_user()执行注册逻辑
        5. 返回注册结果
        
        安全特性：
        - 密码由AuthSystem内部加密存储（SHA256或明文，取决于配置）
        - 用户名去除首尾空格，防止输入错误
        - 所有验证逻辑委托给AuthSystem处理
        
        错误情况：
        - 用户名或密码为空：返回 {"success": false, "message": "用户名和密码不能为空"}
        - 用户名已存在：由auth_system返回相应错误信息
        - 其他错误：由auth_system捕获并返回
        
        注意：
        - 此接口不需要认证（公开接口）
        - 注册成功后用户需要再调用 /auth/login 登录
        - 默认分组由AuthSystem配置决定
        """
        data = request.get_json() or {}
        auth_username = data.get('auth_username', '').strip()
        auth_password = data.get('auth_password', '').strip()

        if not auth_username or not auth_password:
            return jsonify({"success": False, "message": "用户名和密码不能为空"})

        result = auth_system.register_user(auth_username, auth_password)
        return jsonify(result)

    @app.route('/auth/login', methods=['POST'])
    def auth_login():
        """
        用户登录认证API端点。
        
        请求方法：POST
        请求路径：/auth/login
        Content-Type：application/json
        请求头：X-Session-ID: <客户端会话ID>（可选）
        
        请求体（JSON）：
        {
            "auth_username": "用户名",      # 必填
            "auth_password": "密码",        # 必填
            "two_fa_code": "双因素认证码"   # 可选，仅在启用2FA时需要
        }
        
        响应体（JSON）：
        {
            "success": true/false,
            "message": "成功/失败信息",
            "auth_username": "用户名",
            "group": "用户组",
            "is_guest": false
        }
        
        处理流程：
        1. 解析请求体和请求头（session_id、IP、User-Agent）
        2. 调用auth_system.authenticate()验证用户凭据
        3. 验证成功后创建或获取Api实例
        4. 将认证信息附加到web_sessions
        5. 执行单会话强制策略（可选，非游客用户）
        6. 返回认证结果
        
        会话管理：
        - 每个客户端通过X-Session-ID标识唯一会话
        - 如果session_id已存在，复用现有Api实例
        - 如果session_id不存在，创建新的Api实例
        - Api实例包含：
          * auth_username: 认证用户名
          * auth_group: 用户组
          * is_guest: 是否为游客
          * is_authenticated: 认证状态标志
          * _session_created_at: 会话创建时间戳
          * _web_session_id: 关联的会话ID
        
        单会话强制策略（仅注册用户）：
        - check_single_session_enforcement()检查同一用户的活跃会话数
        - 如果超过限制，返回需要清理的旧会话列表
        - 使用后台线程异步清理旧会话（不阻塞登录响应）
        - 清理过程：
          1. 从web_sessions中移除旧会话
          2. 尝试关闭旧会话的Api实例资源
          3. 记录清理结果到日志
        
        安全特性：
        - 密码由AuthSystem验证（支持明文或加密存储）
        - IP地址和User-Agent用于登录日志和安全审计
        - 双因素认证支持（如果用户启用）
        - 会话数量限制防止会话劫持
        - 所有会话操作都有web_sessions_lock保护（线程安全）
        
        双因素认证（2FA）：
        - 如果用户启用了2FA，必须提供正确的two_fa_code
        - AuthSystem会验证TOTP代码（Time-based One-Time Password）
        - 验证失败会拒绝登录
        
        错误情况：
        - 用户名或密码错误：由auth_system返回 {"success": false, "message": "..."}
        - 2FA代码错误：由auth_system返回相应错误
        - 内部错误：捕获并返回错误信息
        
        注意：
        - 此接口不需要预先认证（公开接口）
        - 注释掉的session_id检查代码表明曾经需要会话ID，现在已改为可选
        - 游客用户不受单会话限制
        - 旧会话清理是异步的，不影响新登录的响应速度
        """
        data = request.get_json() or {}
        auth_username = data.get('auth_username', '').strip()
        auth_password = data.get('auth_password', '').strip()
        two_fa_code = data.get('two_fa_code', '').strip()
        session_id = request.headers.get('X-Session-ID', '')

        # 获取IP和UA
        # 用于安全审计和登录日志记录
        ip_address = request.remote_addr or ''
        user_agent = request.headers.get('User-Agent', '')

        # if not session_id:
        #     return jsonify({"success": False, "message": "缺少会话ID"})

        # # 更新会话活动时间
        # update_session_activity(session_id)

        # 验证用户
        # authenticate()方法会：
        # 1. 验证密码（明文或加密比较）
        # 2. 验证2FA代码（如果启用）
        # 3. 检查账号状态（是否锁定等）
        # 4. 记录登录日志（IP、UA、时间、结果）
        auth_result = auth_system.authenticate(
            auth_username, auth_password, ip_address, user_agent, two_fa_code)
        if not auth_result['success']:
            return jsonify(auth_result)

        # 将认证信息附加到会话
        # 使用锁保护web_sessions字典的并发访问
        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]
            else:
                api_instance = Api(args)
                api_instance._session_created_at = time.time()
                api_instance._web_session_id = session_id
                web_sessions[session_id] = api_instance

            api_instance.auth_username = auth_result['auth_username']
            api_instance.auth_group = auth_result['group']
            api_instance.is_guest = auth_result.get('is_guest', False)
            api_instance.is_authenticated = True

            # 如果是注册用户（非游客），处理会话关联
            cleanup_message = ""
            if not auth_result.get('is_guest', False):
                try:
                    # 检查并强制执行会话数量限制
                    old_sessions, cleanup_message = auth_system.check_single_session_enforcement(
                        auth_username, session_id)

                    # 清理旧会话（如果超出限制）- 使用后台线程异步清理，不阻塞登录响应
                    if old_sessions:
                        def cleanup_old_sessions_async():
                            for old_sid in old_sessions:
                                try:
                                    cleanup_session(
                                        old_sid, "session_limit_exceeded")
                                except Exception as e:
                                    logging.error(
                                        f"后台清理旧会话失败 {old_sid[:16]}...: {e}")

                        cleanup_thread = threading.Thread(
                            target=cleanup_old_sessions_async, daemon=True)
                        cleanup_thread.start()

                    # 关联新会话到用户账号
                    auth_system.link_session_to_user(auth_username, session_id)

                    # 记录审计日志
                    audit_details = f'登录成功，会话ID: {session_id}'
                    if cleanup_message:
                        audit_details += f'; {cleanup_message}'

                    auth_system.log_audit(
                        auth_username,
                        'user_login',
                        audit_details,
                        ip_address,
                        session_id
                    )
                except Exception as e:
                    # 即使会话管理失败，也不应阻止登录
                    logging.error(f"会话管理过程出错，但继续登录流程: {e}")
                    cleanup_message = ""

            try:
                save_session_state(session_id, api_instance, force_save=True)
            except Exception as e:
                logging.error(f"保存会话状态失败: {e}")

        # 如果是注册用户，返回已保存的会话ID列表
        user_sessions = []
        max_sessions = 1
        if not auth_result.get('is_guest', False):
            try:
                user_sessions = auth_system.get_user_sessions(auth_username)
                user_details = auth_system.get_user_details(auth_username)
                if user_details:
                    max_sessions = user_details.get('max_sessions', 1)
            except Exception as e:
                logging.error(f"获取用户会话信息失败: {e}")
                user_sessions = []
                max_sessions = 1

        # 生成会话限制提示信息
        session_limit_info = ""
        if max_sessions == 1:
            session_limit_info = "您的账号为单会话模式，新登录将自动清理所有旧会话"
        elif max_sessions == -1:
            session_limit_info = "您的账号为无限会话模式，可以创建任意数量的会话"
        else:
            session_limit_info = f"您的账号最多可以同时保持{max_sessions}个活跃会话，超出时将自动清理最旧的会话"

        # ===== 生成和存储Token (仅非游客) =====
        token = None
        kicked_sessions = []
        if not auth_result.get('is_guest', False) and session_id:
            try:
                # 1. 生成2048位token
                token = token_manager.create_token(auth_username, session_id)

                # 2. 检测多设备登录
                kicked_sessions = token_manager.detect_multi_device_login(
                    auth_username, session_id)

                # 3. 清理过期token
                token_manager.cleanup_expired_tokens(auth_username)

                # 4. 踢出旧设备（使token失效并清理会话）
                if kicked_sessions:
                    for old_sid in kicked_sessions:
                        # 使token失效
                        token_manager.invalidate_token(auth_username, old_sid)
                        # 清理会话
                        # cleanup_session(old_sid, "logged_in_elsewhere")

                    logging.info(
                        f"用户 {auth_username} 从新设备登录，检测到 {len(kicked_sessions)} 个其他活跃会话。")
            except Exception as e:
                # 即使token管理失败，也不应阻止登录流程
                logging.error(f"Token管理过程出错，但继续登录流程: {e}")
                token = None
                kicked_sessions = []

        try:
            response_data = {
                "success": True,
                "session_id": session_id,  # 修正：返回session_id给前端
                "auth_username": auth_result['auth_username'],
                "group": auth_result['group'],
                "is_guest": auth_result.get('is_guest', False),
                "user_sessions": user_sessions,  # 用于状态恢复
                "max_sessions": max_sessions,
                "session_limit_info": session_limit_info,
                "avatar_url": auth_result.get('avatar_url', ''),
                "theme": auth_result.get('theme', 'light'),
                "token": token,  # 返回token给前端
                "kicked_sessions_count": len(kicked_sessions)  # 踢出的设备数量
            }

            # 添加清理提示（如果有）
            if cleanup_message:
                response_data['cleanup_message'] = cleanup_message

            # 添加多设备登录提示
            if kicked_sessions:
                response_data['multi_device_warning'] = f"检测到该账号在其他 {len(kicked_sessions)} 个设备上登录，已自动登出旧设备"

            # 创建响应并设置Cookie (仅非游客)
            response = jsonify(response_data)
            if token:
                # 设置1小时过期的httponly cookie
                response.set_cookie(
                    'auth_token',
                    value=token,
                    max_age=3600,  # 1小时
                    httponly=True,  # 防止JavaScript访问
                    secure=False,  # 开发环境设为False，生产环境应为True
                    samesite='Lax'
                )

            return response
        except Exception as e:
            # 最后的安全网：即使响应创建失败，也返回基本的成功响应
            logging.error(f"创建登录响应失败: {e}")
            return jsonify({
                "success": True,
                "session_id": session_id,
                "auth_username": auth_result.get('auth_username', auth_username),
                "group": auth_result.get('group', 'user'),
                "is_guest": False
            })

    @app.route('/auth/guest_login', methods=['POST'])
    def auth_guest_login():
        """
        游客登录API - 无需密码的快速访问入口。
        
        功能说明：
        - 为未注册用户提供受限的访问权限
        - 无需用户名和密码，只需会话ID
        - 适用于试用、演示或临时访问场景
        
        请求格式：
        - 方法：POST
        - 路径：/auth/guest_login
        - 请求头：
          * X-Session-ID: 会话标识符（必需）
        - 请求体：无需
        
        响应格式：
        {
          "success": true/false,
          "message": "错误信息（仅失败时）",
          "auth_username": "guest",
          "group": "guest",
          "is_guest": true
        }
        
        处理流程：
        1. 从请求头获取session_id
        2. 验证session_id存在性
        3. 更新会话活动时间（防止超时）
        4. 检查系统是否允许游客登录（从配置文件读取）
        5. 创建或获取游客的Api实例
        6. 设置游客属性（auth_username='guest', auth_group='guest', is_guest=True）
        7. 保存会话状态到磁盘
        8. 返回成功响应
        
        游客权限限制：
        - 无法使用需要认证的高级功能
        - 会话不支持多设备同步
        - 可能无法访问某些敏感API
        - 数据不会长期保存
        
        配置项：
        - auth.ini [Guest] allow_guest_login：控制是否允许游客登录（默认true）
        
        错误情况：
        - 缺少session_id：返回400错误
        - 系统不允许游客登录：返回失败消息
        
        安全考虑：
        - 游客会话不受单会话强制限制
        - 游客不生成token和cookie
        - 游客数据可能被定期清理
        
        注意：
        - 游客会话不会记录到登录日志
        - 游客切换到注册用户需要重新登录
        """
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id:
            return jsonify({"success": False, "message": "缺少会话ID"})

        # 更新会话活动时间
        update_session_activity(session_id)

        # 检查是否允许游客登录
        if not auth_system.config.getboolean('Guest', 'allow_guest_login', fallback=True):
            return jsonify({"success": False, "message": "系统不允许游客登录"})

        # 创建游客会话
        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]
            else:
                api_instance = Api(args)
                api_instance._session_created_at = time.time()
                api_instance._web_session_id = session_id
                web_sessions[session_id] = api_instance

            api_instance.auth_username = 'guest'
            api_instance.auth_group = 'guest'
            api_instance.is_guest = True
            api_instance.is_authenticated = True

            save_session_state(session_id, api_instance, force_save=True)

        return jsonify({
            "success": True,
            "auth_username": "guest",
            "group": "guest",
            "is_guest": True
        })

    @app.route('/auth/logout', methods=['POST'])
    def auth_logout():
        """
        用户登出API - 安全退出系统并清理所有会话数据。
        
        功能说明：
        - 终止用户的当前会话
        - 使token失效（防止重放攻击）
        - 清除浏览器cookie
        - 释放服务器端资源
        
        请求格式：
        - 方法：POST
        - 路径：/auth/logout
        - 请求头：
          * X-Session-ID: 要登出的会话标识符（必需）
        - 请求体：无需
        
        响应格式：
        {
          "success": true,
          "message": "登出成功"
        }
        
        响应头：
        - Set-Cookie: auth_token=; Max-Age=0（清除cookie）
        
        处理流程：
        1. 从请求头获取session_id
        2. 验证session_id存在性
        3. 获取用户信息（用户名、是否游客）
        4. 如果是注册用户，使其token失效：
           - 调用token_manager.invalidate_token()
           - 记录登出事件到日志
        5. 清理服务器端会话：
           - 从web_sessions移除会话
           - 关闭Api实例资源
           - 删除会话持久化文件
        6. 清除客户端cookie（auth_token）
        7. 返回成功响应
        
        安全特性：
        - Token失效机制：防止token被重复使用
        - Cookie清除：防止浏览器自动重新登录
        - 资源释放：避免会话泄漏
        - 审计日志：记录登出时间和原因
        
        游客处理：
        - 游客用户没有token，只清理会话即可
        - 不记录游客的登出事件
        
        错误情况：
        - 缺少session_id：返回400错误
        - 会话不存在：仍然返回成功（幂等性）
        
        注意：
        - 此接口幂等性强，重复调用不会出错
        - 登出后前端应导航到登录页
        - 多标签页的其他标签会话也会失效
        - 调用cleanup_session时原因设为"user_logout"
        
        使用场景：
        - 用户主动点击"退出"按钮
        - 会话超时后的强制登出
        - 安全策略要求的定期重新认证
        """
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id:
            return jsonify({"success": False, "message": "缺少会话ID"}), 400

        # 获取用户信息
        username = None
        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]
                if hasattr(api_instance, 'auth_username'):
                    username = api_instance.auth_username
                    is_guest = getattr(api_instance, 'is_guest', True)

                    # 如果是非游客用户，使token失效
                    if not is_guest and username:
                        token_manager.invalidate_token(username, session_id)
                        logging.info(
                            f"用户 {username} 登出，session: {session_id[:16]}...")

        # 清理会话
        cleanup_session(session_id, "user_logout")

        # 创建响应并清除cookies
        response = jsonify({"success": True, "message": "登出成功"})
        response.set_cookie('auth_token', '', max_age=0)

        return response

    @app.route('/auth/check_permission', methods=['POST'])
    def auth_check_permission():
        """
        权限检查API - 验证用户是否拥有特定权限。
        
        功能说明：
        - 基于RBAC（基于角色的访问控制）模型
        - 检查用户所在组是否有特定权限
        - 用于前端UI控制（显示/隐藏功能按钮）
        - 用于API调用前的权限验证
        
        请求格式：
        - 方法：POST
        - 路径：/auth/check_permission
        - 请求头：
          * X-Session-ID: 会话标识符（必需）
        - 请求体：
          {
            "permission": "权限名称"
          }
        
        响应格式：
        {
          "success": true/false,
          "has_permission": true/false
        }
        
        处理流程：
        1. 从请求头获取session_id
        2. 从请求体获取permission名称
        3. 验证session_id存在且有效
        4. 从web_sessions获取Api实例
        5. 验证用户已认证（has auth_username）
        6. 调用auth_system.check_permission()检查权限
        7. 返回检查结果
        
        权限系统：
        - 权限定义在auth.ini的[Permissions]段
        - 格式：permission_name = group1,group2,group3
        - 例如：admin_panel = admin
        - 支持多组共享同一权限
        
        常见权限示例：
        - "admin_panel"：访问管理面板
        - "multi_account"：多账号管理
        - "export_data"：导出数据
        - "view_logs"：查看日志
        - "manage_users"：用户管理
        
        错误情况：
        - session_id缺失或无效：返回{"success": false, "has_permission": false}
        - 用户未认证：返回{"success": false, "has_permission": false}
        - permission参数缺失：检查空字符串权限（通常返回false）
        
        安全考虑：
        - 不泄露权限系统的内部结构
        - 失败时统一返回has_permission=false
        - 不区分"用户不存在"和"无权限"
        
        注意：
        - 此接口只做检查，不执行实际操作
        - 前端应根据返回值动态调整UI
        - 后端API仍需在执行前再次验证权限
        - 游客用户通常无任何特殊权限
        
        使用场景：
        - 页面加载时检查功能可用性
        - 点击操作前的预先验证
        - 动态菜单生成
        """
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        permission = data.get('permission', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "has_permission": False})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "has_permission": False})

        has_permission = auth_system.check_permission(
            api_instance.auth_username, permission)
        return jsonify({"success": True, "has_permission": has_permission})

    @app.route('/auth/switch_session', methods=['POST'])
    def auth_switch_session():
        """
        会话切换API - 在多标签页间切换时更新认证token和cookie。
        
        功能说明：
        - 支持同一用户在多个浏览器标签页间无缝切换
        - 为目标会话生成新的token
        - 更新浏览器cookie以匹配目标会话
        - 预加载目标会话状态（如果存在）
        
        应用场景：
        - 用户在标签页A登录后，打开标签页B
        - 用户从会话列表中选择要切换到的会话
        - 多窗口协作（如同时查看不同任务）
        
        请求格式：
        - 方法：POST
        - 路径：/auth/switch_session
        - 请求头：
          * X-Session-ID: 当前活动窗口的会话ID（必需）
          * Cookie: auth_token=...（用于验证身份）
        - 请求体：
          {
            "target_session_id": "目标会话的UUID"
          }
        
        响应格式：
        {
          "success": true/false,
          "message": "成功/错误信息",
          "need_login": true（仅当token失效时）
        }
        
        响应头：
        - Set-Cookie: auth_token=新token; Max-Age=3600; HttpOnly; SameSite=Lax
        
        处理流程（5步）：
        1. 验证当前用户身份：
           - 从current_session_id获取用户名
           - 检查是否为游客（游客不支持切换）
           - 验证is_authenticated状态
        
        2. 验证当前token有效性：
           - 从cookie读取auth_token
           - 调用token_manager.verify_token()验证
           - 如果失效，返回401并要求重新登录
        
        3. 为目标会话生成新token：
           - 调用token_manager.create_token()
           - 新token关联到target_session_id
           - 记录切换事件到日志
        
        4. 更新cookie：
           - 设置新的auth_token cookie
           - 1小时过期时间
           - HttpOnly防止XSS攻击
           - SameSite=Lax防止CSRF
        
        5. 预加载目标会话状态（可选但推荐）：
           - 从文件加载目标会话的历史状态
           - 如果成功，创建Api实例并恢复
           - 如果失败，前端访问时会创建新会话
        
        安全特性：
        - **双重验证**：既验证session_id又验证token
        - **游客限制**：游客用户不允许切换会话
        - **Token刷新**：即使目标已有token，也生成新的
        - **自动清理**：无效token会自动清除cookie
        - **线程安全**：web_sessions_lock保护并发访问
        
        错误情况：
        - 缺少参数：返回400错误
        - 未登录或游客：返回401错误
        - Token失效：返回401并清除cookie，设置need_login=true
        - 目标会话不存在：仍然成功（前端访问时创建）
        
        设计考虑：
        - **为什么生成新token**：
          * 增强安全性，每次切换都是新的认证
          * 避免token复用导致的安全问题
          * 便于追踪用户的会话切换行为
        
        - **为什么预加载状态**：
          * 提升用户体验，减少加载时间
          * 但即使失败也不影响功能
          * 前端路由会处理会话创建
        
        注意：
        - 切换后前端应导航到 /uuid=<target_session_id>
        - 旧标签页的会话不会被销毁（用户可能返回）
        - 每个会话都有独立的状态和数据
        - Token有效期1小时，过期需重新登录
        
        与单会话强制的关系：
        - 此功能允许同一用户多会话并存
        - 单会话强制针对的是"并发登录设备数"
        - 会话切换是在已允许的设备内切换
        
        使用流程示例：
        1. 用户在标签A登录（会话ID: session-A）
        2. 用户打开标签B（自动分配会话ID: session-B）
        3. 用户在标签B点击"切换到会话A"
        4. 调用此API，target_session_id=session-A
        5. 服务器生成新token for session-A
        6. 标签B的cookie更新为新token
        7. 标签B导航到/uuid=session-A，成功接管会话A的状态
        """
        current_session_id = request.headers.get(
            'X-Session-ID', '')  # 当前活动窗口的 Session ID
        data = request.get_json() or {}
        target_session_id = data.get(
            'target_session_id', '')  # 用户想要切换到的 Session ID

        if not current_session_id or not target_session_id:
            return jsonify({"success": False, "message": "缺少会话ID参数"}), 400

        # 1. 验证当前用户身份 (基于 current_session_id 和 cookie)
        username = None
        is_guest = True
        auth_username = ''
        with web_sessions_lock:
            if current_session_id in web_sessions:
                api_instance = web_sessions[current_session_id]
                if getattr(api_instance, 'is_authenticated', False):
                    username = getattr(api_instance, 'auth_username', None)
                    is_guest = getattr(api_instance, 'is_guest', True)

        if not username or is_guest:
            return jsonify({"success": False, "message": "用户未登录或为游客，无法切换会话"}), 401

        # 2. 检查 token (这一步是必要的，确保当前操作是合法的)
        token_from_cookie = request.cookies.get('auth_token')
        if not token_from_cookie:
            return jsonify({"success": False, "message": "缺少认证令牌(cookie)"}), 401

        is_valid, reason = token_manager.verify_token(
            username, current_session_id, token_from_cookie)
        if not is_valid:
            # 如果当前 token 都无效了，直接要求重新登录
            logging.warning(
                f"切换会话失败：用户 {username} 的当前会话 {current_session_id[:8]} token 无效 ({reason})")
            response_data = {
                "success": False, "message": f"当前认证已失效({reason})，请重新登录", "need_login": True}
            response = make_response(jsonify(response_data), 401)
            response.set_cookie('auth_token', '', max_age=0)  # 清除无效 cookie
            return response

        # 3. 为目标会话生成新 Token
        # 注意：即使目标会话已存在 token，这里也生成一个新的，以确保安全性和一致性
        new_token_for_target = token_manager.create_token(
            username, target_session_id)
        logging.info(
            f"用户 {username} 切换会话：为目标会话 {target_session_id[:8]} 生成新 token")

        # 4. 创建响应，并设置新的 auth_token cookie
        response_data = {"success": True, "message": "Token已更新，可以跳转"}
        response = make_response(jsonify(response_data))
        response.set_cookie(
            'auth_token',
            value=new_token_for_target,
            max_age=3600,  # 1 小时
            httponly=True,
            secure=False,  # 开发环境 False，生产环境 True
            samesite='Lax'
        )

        # 5. （可选但推荐）确保目标会话状态存在于内存或文件中
        with web_sessions_lock:
            if target_session_id not in web_sessions:
                # 尝试从文件加载，如果文件不存在或加载失败，也没关系，
                # 因为 /uuid=<target_id> 路由会处理新会话的创建
                state = load_session_state(target_session_id)
                if state:
                    target_api_instance = Api(args)
                    target_api_instance._web_session_id = target_session_id
                    restore_session_to_api_instance(target_api_instance, state)
                    web_sessions[target_session_id] = target_api_instance
                    logging.info(f"切换会话时，预加载目标会话 {target_session_id[:8]} 状态成功")
                else:
                    logging.warning(
                        f"切换会话时，目标会话 {target_session_id[:8]} 状态文件不存在或加载失败，将在访问时创建")

        return response

    @app.route('/auth/admin/list_users', methods=['GET'])
    def auth_admin_list_users():
        """
        管理员API - 列出所有用户信息。
        
        功能说明：
        - 获取系统中所有注册用户的列表
        - 包含用户名、用户组、创建时间等信息
        - 用于管理员查看和管理用户
        
        权限要求：
        - 必须登录
        - 必须具有'manage_users'权限
        
        请求头：
        - X-Session-ID: 会话ID（必需）
        
        响应格式：
        成功：{"success": true, "users": [用户列表]}
        失败：{"success": false, "message": "错误信息"}
        
        用户对象结构：
        {
            "username": "用户名",
            "group": "用户组",
            "created_at": "创建时间",
            "last_login": "最后登录时间"
        }
        
        使用场景：
        - 管理员查看所有用户
        - 用户管理界面的数据源
        - 审计和统计分析
        """
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查权限
        if not auth_system.check_permission(api_instance.auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"})

        users = auth_system.list_users()
        return jsonify({"success": True, "users": users})

    @app.route('/auth/admin/update_user_group', methods=['POST'])
    def auth_admin_update_user_group():
        """
        管理员API - 修改用户所属的权限组。
        
        功能说明：
        - 将指定用户分配到新的权限组
        - 影响用户的权限范围
        - 立即生效，无需重新登录
        
        权限要求：
        - 必须登录
        - 必须具有'manage_users'权限
        
        请求头：
        - X-Session-ID: 会话ID（必需）
        
        请求体（JSON）：
        {
            "target_username": "目标用户名",
            "new_group": "新权限组名"
        }
        
        响应格式：
        成功：{"success": true, "message": "用户组已更新"}
        失败：{"success": false, "message": "错误信息"}
        
        注意事项：
        - 目标用户组必须存在
        - 不能修改自己的用户组（防止锁定）
        - 更改会立即影响用户权限
        
        使用场景：
        - 提升/降低用户权限
        - 调整用户角色
        - 权限管理界面
        """
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        target_username = data.get('target_username', '')
        new_group = data.get('new_group', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查权限
        if not auth_system.check_permission(api_instance.auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"})

        result = auth_system.update_user_group(target_username, new_group)
        return jsonify(result)

    @app.route('/auth/admin/list_groups', methods=['GET'])
    def auth_admin_list_groups():
        """
        管理员API - 列出所有权限组及其权限配置。
        
        功能说明：
        - 获取系统中定义的所有权限组
        - 包含每个组的权限列表
        - 用于权限管理和配置
        
        权限要求：
        - 必须登录
        - 必须具有'manage_permissions'权限
        
        请求头：
        - X-Session-ID: 会话ID（必需）
        
        响应格式：
        成功：{"success": true, "groups": {权限组配置}}
        失败：{"success": false, "message": "错误信息"}
        
        权限组结构示例：
        {
            "admin": ["manage_users", "manage_permissions", ...],
            "user": ["view_data", "edit_own_data"],
            "guest": ["view_public_data"]
        }
        
        使用场景：
        - 权限配置界面数据源
        - 显示可用的权限组列表
        - 权限分配参考
        """
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查权限
        if not auth_system.check_permission(api_instance.auth_username, 'manage_permissions'):
            return jsonify({"success": False, "message": "权限不足"})

        groups = auth_system.get_all_groups()
        return jsonify({"success": True, "groups": groups})

    @app.route('/auth/admin/create_group', methods=['POST'])
    def auth_admin_create_group():
        """超级管理员：创建权限组"""
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        group_name = data.get('group_name', '')
        display_name = data.get('display_name', '')
        permissions = data.get('permissions', {})

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查是否为超级管理员
        if api_instance.auth_group != 'super_admin':
            return jsonify({"success": False, "message": "仅超级管理员可创建权限组"})

        result = auth_system.create_permission_group(
            group_name, permissions, display_name)
        return jsonify(result)

    @app.route('/auth/admin/update_group', methods=['POST'])
    def auth_admin_update_group():
        """超级管理员：更新权限组"""
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        group_name = data.get('group_name', '')
        permissions = data.get('permissions', {})

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查是否为超级管理员
        if api_instance.auth_group != 'super_admin':
            return jsonify({"success": False, "message": "仅超级管理员可更新权限组"})

        result = auth_system.update_permission_group(group_name, permissions)
        return jsonify(result)

    @app.route('/auth/admin/delete_group', methods=['POST'])
    def auth_admin_delete_group():
        """超级管理员：删除权限组"""
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        group_name = data.get('group_name', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"})

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"})

        # 检查是否为超级管理员
        if api_instance.auth_group != 'super_admin':
            return jsonify({"success": False, "message": "仅超级管理员可删除权限组"})

        result = auth_system.delete_permission_group(group_name)
        return jsonify(result)

    @app.route('/auth/admin/get_user_permissions', methods=['POST'])
    def auth_admin_get_user_permissions():
        """管理员：获取用户的完整权限列表"""
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        target_username = data.get('username', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_permissions'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        # 获取用户的完整权限列表（包含差分化权限）
        all_permissions = auth_system.get_user_permissions(target_username)

        # 获取用户所属的权限组
        group = auth_system.get_user_group(target_username)

        # 获取权限组的基础权限
        group_permissions = auth_system.permissions['permission_groups'].get(
            group, {}).get('permissions', {})

        # 获取用户的差分权限信息
        user_custom = auth_system.permissions.get(
            'user_custom_permissions', {}).get(target_username, {})
        added_list = user_custom.get('added', [])
        removed_list = user_custom.get('removed', [])

        # 将差分权限数组转换为对象格式（前端需要）
        added_permissions = {perm: True for perm in added_list}
        removed_permissions = {perm: True for perm in removed_list}

        return jsonify({
            "success": True,
            "group": group,
            "all_permissions": all_permissions,
            "group_permissions": group_permissions,
            "added_permissions": added_permissions,
            "removed_permissions": removed_permissions
        })

    @app.route('/auth/admin/set_user_permission', methods=['POST'])
    def auth_admin_set_user_permission():
        """管理员：为用户设置自定义权限（差分化存储）

        支持两种模式：
        1. 单个权限设置：传递 permission 和 grant 参数
        2. 批量权限设置：传递 added_permissions 和 removed_permissions 对象
        """
        session_id = request.headers.get('X-Session-ID', '')
        data = request.get_json() or {}
        target_username = data.get('username', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_permissions'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        # 检查是批量更新还是单个更新
        added_permissions = data.get('added_permissions', {})
        removed_permissions = data.get('removed_permissions', {})

        if added_permissions or removed_permissions:
            # 批量更新模式
            try:
                # 清空用户的差分权限，然后设置新的
                if 'user_custom_permissions' not in auth_system.permissions:
                    auth_system.permissions['user_custom_permissions'] = {}

                auth_system.permissions['user_custom_permissions'][target_username] = {
                    'added': list(added_permissions.keys()),
                    'removed': list(removed_permissions.keys())
                }

                auth_system._save_permissions()

                # 记录审计日志
                ip_address = request.headers.get(
                    'X-Forwarded-For', request.remote_addr)
                auth_system.log_audit(
                    auth_username,
                    'set_user_permissions_batch',
                    f'批量更新用户 {target_username} 的差分权限: 添加{len(added_permissions)}个, 移除{len(removed_permissions)}个',
                    ip_address,
                    session_id
                )

                return jsonify({"success": True, "message": "权限已更新"})
            except Exception as e:
                logging.error(f"批量更新权限失败: {e}", exc_info=True)
                return jsonify({"success": False, "message": f"更新失败: {str(e)}"}), 500
        else:
            # 单个权限更新模式（兼容旧代码）
            permission = data.get('permission', '')
            grant = data.get('grant', False)

            


            if not permission:
                # return jsonify({"success": False, "message": "缺少permission参数"}), 400
                logging.info(f"管理员 {auth_username} 尝试为用户 {target_username} 设置权限时，缺少 permission 参数")
                try:
                    # 清空用户的差分权限，然后设置新的
                    if 'user_custom_permissions' not in auth_system.permissions:
                        auth_system.permissions['user_custom_permissions'] = {}
                        
                    auth_system.permissions['user_custom_permissions'][target_username] = {
                    'added':  [],
                    'removed': []
                    }

                    auth_system._save_permissions()

                    # 记录审计日志
                    ip_address = request.headers.get(
                        'X-Forwarded-For', request.remote_addr)
                    auth_system.log_audit(
                        auth_username,
                        'set_user_permissions_batch',
                        f'批量更新用户 {target_username} 的差分权限: 清空用户的差分权限',
                        ip_address,
                        session_id
                    )

                    return jsonify({"success": True, "message": "权限已更新"})
                except Exception as e:
                    logging.error(f"批量更新权限失败: {e}", exc_info=True)
                    return jsonify({"success": False, "message": f"更新失败: {str(e)}"}), 500
                

            result = auth_system.set_user_custom_permission(
                target_username, permission, grant)

            # 记录审计日志
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'set_user_permission',
                f'为用户 {target_username} {"授予" if grant else "移除"} 权限: {permission}',
                ip_address,
                session_id
            )

            return jsonify(result)

    @app.route('/auth/get_config', methods=['GET'])
    def auth_get_config():
        """获取认证配置（用于前端显示）"""
        return jsonify({
            "success": True,
            "allow_guest_login": auth_system.config.getboolean('Guest', 'allow_guest_login', fallback=True),
            "guest_auto_fill_password": auth_system.config.getboolean('AutoFill', 'guest_auto_fill_password', fallback=False),
            "amap_js_key": auth_system.config.get('Map', 'amap_js_key', fallback='')
        })

    @app.route('/auth/check_uuid_type', methods=['POST'])
    def auth_check_uuid_type():
        """
        检查UUID类型：游客UUID、系统账号UUID或未知UUID
        用于实现访问控制和会话验证
        """
        data = request.json
        check_uuid = data.get('uuid', '')

        if not check_uuid:
            return jsonify({"success": False, "message": "UUID参数缺失"}), 400

        # UUID格式验证 - 标准UUID v4格式

        uuid_pattern = re.compile(
            r'^[a-f0-9]{8}-[a-f0-9]{4}-4[a-f0-9]{3}-[89ab][a-f0-9]{3}-[a-f0-9]{12}$',
            re.IGNORECASE
        )
        if not uuid_pattern.match(check_uuid):
            return jsonify({
                "success": False,
                "message": "无效的UUID格式，请使用标准UUID v4格式"
            }), 400

        # 检查会话文件是否存在
        session_file = get_session_file_path(check_uuid)
        file_exists = os.path.exists(session_file)
        logging.debug(
            f"/auth/check_uuid_type: Checking UUID {check_uuid[:8]}..., File path: {session_file}, Exists: {file_exists}")  # 新增日志

        if not file_exists:
            return jsonify({
                "success": True,
                "uuid_type": "unknown",
                "message": "UUID不存在 (文件未找到)"  # 更明确的消息
            })

        # 读取会话文件判断类型 - 使用文件锁避免并发问题

        max_retries = 3
        retry_delay = 0.1

        import sys
        # 在循环开始前尝试导入 fcntl (仅非Windows)
        fcntl = None
        if sys.platform != 'win32':
            try:
                import fcntl
            except ImportError:
                logging.warning("fcntl 模块在当前平台不可用，文件锁将被跳过。")

        for attempt in range(max_retries):
            try:
                with open(session_file, 'r', encoding='utf-8') as f:
                    # 加共享锁（读锁），仅当 fcntl 可用时
                    if fcntl:  # <- 添加判断
                        fcntl.flock(f.fileno(), fcntl.LOCK_SH)
                    try:
                        session_data = json.load(f)
                    finally:
                        # 释放锁，仅当 fcntl 可用时
                        if fcntl:  # <- 添加判断
                            fcntl.flock(f.fileno(), fcntl.LOCK_UN)

                # 检查是否为游客会话
                is_guest = session_data.get('is_guest', False)
                auth_username = session_data.get('auth_username', '')

                if is_guest or auth_username == 'guest':
                    return jsonify({
                        "success": True,
                        "uuid_type": "guest",
                        "message": "游客UUID"
                    })
                elif auth_username:
                    return jsonify({
                        "success": True,
                        "uuid_type": "system_account",
                        "auth_username": auth_username,
                        "message": "系统账号UUID"
                    })
                else:
                    logging.warning(
                        f"/auth/check_uuid_type: 文件 {session_file} 存在但内容无法识别用户类型 (auth_username='{auth_username}', is_guest={is_guest})，返回 unknown")  # 新增日志
                    return jsonify({
                        "success": True,
                        "uuid_type": "unknown",
                        "message": "未知类型UUID (内容无法识别)"  # 更明确的消息
                    })
            except (IOError, OSError) as e:
                # 文件锁定或IO错误，重试
                if attempt < max_retries - 1:

                    time.sleep(retry_delay)
                    continue
                else:
                    logging.error(f"读取会话文件失败（已重试{max_retries}次）: {e}")
                    return jsonify({
                        "success": False,
                        "message": "读取会话失败，请稍后重试"
                    }), 500
            except json.JSONDecodeError as e:
                logging.error(f"会话文件JSON解析失败: {e}")
                return jsonify({
                    "success": False,
                    "message": "会话数据损坏"
                }), 500
            except Exception as e:
                logging.error(f"检查UUID类型失败: {e}")
                return jsonify({
                    "success": False,
                    "message": f"服务器错误: {str(e)}"
                }), 500

    @app.route('/auth/2fa/generate', methods=['POST'])
    def auth_2fa_generate():
        """生成2FA密钥"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        if not getattr(api_instance, 'is_authenticated', False):
            return jsonify({"success": False, "message": "未认证"}), 401

        auth_username = getattr(api_instance, 'auth_username', '')
        if auth_username == 'guest':
            return jsonify({"success": False, "message": "游客不支持2FA"}), 403

        result = auth_system.generate_2fa_secret(auth_username)
        return jsonify(result)

    @app.route('/auth/2fa/enable', methods=['POST'])
    def auth_2fa_enable():
        """启用2FA"""
        data = request.get_json() or {}
        verification_code = data.get('code', '').strip()
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        if not getattr(api_instance, 'is_authenticated', False):
            return jsonify({"success": False, "message": "未认证"}), 401

        auth_username = getattr(api_instance, 'auth_username', '')
        result = auth_system.enable_2fa(auth_username, verification_code)
        return jsonify(result)

    @app.route('/auth/2fa/verify', methods=['POST'])
    def auth_2fa_verify():
        """测试验证2FA代码（用于用户测试2FA是否工作正常）"""
        data = request.get_json() or {}
        verification_code = data.get('code', '').strip()
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        if not getattr(api_instance, 'is_authenticated', False):
            return jsonify({"success": False, "message": "未认证"}), 401

        auth_username = getattr(api_instance, 'auth_username', '')
        if auth_username == 'guest':
            return jsonify({"success": False, "message": "游客不支持2FA"}), 403

        # 验证2FA代码
        if auth_system.verify_2fa(auth_username, verification_code):
            return jsonify({"success": True, "message": "验证码正确"})
        else:
            return jsonify({"success": False, "message": "验证码错误"})

    @app.route('/auth/2fa/disable', methods=['POST'])
    def auth_2fa_disable():
        """关闭2FA"""
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        if not getattr(api_instance, 'is_authenticated', False):
            return jsonify({"success": False, "message": "未认证"}), 401

        auth_username = getattr(api_instance, 'auth_username', '')
        if auth_username == 'guest':
            return jsonify({"success": False, "message": "游客不支持2FA"}), 403

        # 禁用2FA
        user_file = auth_system.get_user_file_path(auth_username)
        if os.path.exists(user_file):
            try:
                with auth_system.lock:
                    with open(user_file, 'r', encoding='utf-8') as f:
                        user_data = json.load(f)

                    user_data['2fa_enabled'] = False
                    # 可选：也可以删除2fa_secret，但保留它允许用户稍后重新启用
                    # user_data['2fa_secret'] = None

                    with open(user_file, 'w', encoding='utf-8') as f:
                        json.dump(user_data, f, indent=2, ensure_ascii=False)

                logging.info(f"用户 {auth_username} 已关闭2FA")
                return jsonify({"success": True, "message": "2FA已关闭"})
            except Exception as e:
                logging.error(f"关闭2FA失败: {e}", exc_info=True)
                return jsonify({"success": False, "message": f"关闭失败: {str(e)}"}), 500
        else:
            return jsonify({"success": False, "message": "用户不存在"}), 404

    @app.route('/auth/2fa/verify_login', methods=['POST'])
    def auth_2fa_verify_login():
        """验证2FA代码并完成登录（用于登录流程中的2FA验证）"""
        data = request.get_json() or {}
        auth_username = data.get('auth_username', '').strip()
        verification_code = data.get('code', '').strip()

        if not auth_username:
            return jsonify({"success": False, "message": "缺少用户名"}), 400

        if not verification_code:
            return jsonify({"success": False, "message": "缺少验证码"}), 400

        # 验证2FA代码
        if not auth_system.verify_2fa(auth_username, verification_code):
            logging.warning(f"2FA登录验证失败: {auth_username}")
            return jsonify({"success": False, "message": "验证码错误"})

        # 2FA验证成功，创建会话
        session_id = str(uuid.uuid4())
        api_instance = Api(args)
        api_instance._session_created_at = time.time()
        api_instance._web_session_id = session_id
        api_instance.is_authenticated = True
        api_instance.auth_username = auth_username

        # 获取用户组信息
        user_group = auth_system.get_user_group(auth_username)
        api_instance.auth_group = user_group

        # 检查是否为游客
        is_guest = (auth_username == 'guest')
        api_instance.is_guest = is_guest

        # 保存会话
        web_sessions[session_id] = api_instance

        # 更新用户最后登录时间
        user_file = auth_system.get_user_file_path(auth_username)
        if os.path.exists(user_file):
            try:
                with auth_system.lock:
                    with open(user_file, 'r', encoding='utf-8') as f:
                        user_data = json.load(f)

                    user_data['last_login'] = time.time()
                    if 'session_ids' not in user_data:
                        user_data['session_ids'] = []
                    if session_id not in user_data['session_ids']:
                        user_data['session_ids'].append(session_id)

                    with open(user_file, 'w', encoding='utf-8') as f:
                        json.dump(user_data, f, indent=2, ensure_ascii=False)
            except Exception as e:
                logging.error(f"更新用户登录信息失败: {e}", exc_info=True)

        # 生成token（非游客）
        token = None
        if not is_guest:
            try:
                token = token_manager.create_token(auth_username, session_id)
                token_manager.cleanup_expired_tokens(auth_username)
            except Exception as e:
                logging.error(f"Token管理过程出错: {e}")
                token = None

        logging.info(f"用户 {auth_username} 通过2FA验证登录成功")

        response_data = {
            "success": True,
            "message": "2FA验证成功",
            "session_id": session_id,
            "is_guest": is_guest,
            "token": token
        }

        # 创建响应并设置Cookie（非游客）
        response = jsonify(response_data)
        if token:
            response.set_cookie(
                'auth_token',
                value=token,
                max_age=3600,  # 1小时
                httponly=True,
                secure=False,
                samesite='Lax'
            )

        return response

    @app.route('/auth/admin/create_user', methods=['POST'])
    def auth_admin_create_user():
        """管理员：创建新用户"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        new_username = data.get('username', '')
        password = data.get('password', '')
        group = data.get('group', 'user')

        if not new_username or not password:
            return jsonify({"success": False, "message": "用户名和密码不能为空"})

        result = auth_system.register_user(new_username, password, group)

        # 记录审计日志
        if result.get('success'):
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'create_user',
                f'创建新用户: {new_username} (组: {group})',
                ip_address,
                session_id
            )

        return jsonify(result)

    @app.route('/auth/admin/ban_user', methods=['POST'])
    def auth_admin_ban_user():
        """管理员：封禁用户"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        target_username = data.get('username', '')

        result = auth_system.ban_user(target_username)

        # 记录审计日志
        if result.get('success'):
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'ban_user',
                f'封禁用户: {target_username}',
                ip_address,
                session_id
            )

        return jsonify(result)

    @app.route('/auth/admin/unban_user', methods=['POST'])
    def auth_admin_unban_user():
        """管理员：解封用户"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        target_username = data.get('username', '')

        result = auth_system.unban_user(target_username)

        # 记录审计日志
        if result.get('success'):
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'unban_user',
                f'解封用户: {target_username}',
                ip_address,
                session_id
            )

        return jsonify(result)

    @app.route('/auth/admin/delete_user', methods=['POST'])
    def auth_admin_delete_user():
        """管理员：删除用户"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        target_username = data.get('username', '')

        result = auth_system.delete_user(target_username)

        # 记录审计日志
        if result.get('success'):
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'delete_user',
                f'删除用户: {target_username}',
                ip_address,
                session_id
            )

        return jsonify(result)

    @app.route('/auth/admin/login_logs', methods=['GET'])
    def auth_admin_login_logs():
        """获取登录日志（管理员）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_group = getattr(api_instance, 'auth_group', 'guest')

        # 只有管理员和超级管理员可以查看
        if auth_group not in ['admin', 'super_admin']:
            return jsonify({"success": False, "message": "权限不足"}), 403

        # 获取参数
        username = request.args.get('username', None)
        limit = int(request.args.get('limit', 100))

        logs = auth_system.get_login_history(username, limit)
        return jsonify({
            "success": True,
            "logs": logs
        })

    @app.route('/logs/view', methods=['GET'])
    def view_logs():
        """查看应用日志（管理员）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_group = getattr(api_instance, 'auth_group', 'guest')

        # 只有管理员和超级管理员可以查看
        if auth_group not in ['admin', 'super_admin']:
            return jsonify({"success": False, "message": "权限不足"}), 403

        # 读取最近的日志（最多1000行）
        lines = int(request.args.get('lines', 100))
        lines = min(lines, 1000)  # 限制最多1000行

        log_content = []
        # 尝试读取日志文件
        log_files = []

        # 查找日志目录中的日志文件
        if os.path.exists(LOGIN_LOGS_DIR):
            for f in os.listdir(LOGIN_LOGS_DIR):
                if f.endswith('.log') or f.endswith('.jsonl'):
                    log_files.append(os.path.join(LOGIN_LOGS_DIR, f))

        # 读取最近的日志
        for log_file in log_files[:5]:  # 最多读取5个日志文件
            try:
                with open(log_file, 'r', encoding='utf-8') as f:
                    content = f.readlines()
                    log_content.extend(content[-lines:])
            except (FileNotFoundError, PermissionError, UnicodeDecodeError) as e:
                logging.debug(f"[日志读取] 无法读取日志文件 {log_file}: {e}")
                continue

        return jsonify({
            "success": True,
            "logs": log_content[-lines:]  # 返回最近的N行
        })

    @app.route('/auth/admin/reset_password', methods=['POST'])
    def auth_admin_reset_password():
        """重置用户密码（管理员）或修改自己的密码"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')
        auth_group = getattr(api_instance, 'auth_group', 'guest')

        data = request.json
        target_username = data.get('username', '')
        new_password = data.get('new_password', '')
        old_password = data.get('old_password', '')  # 可选：用于用户自己修改密码时验证

        if not target_username or not new_password:
            return jsonify({"success": False, "message": "参数缺失"})

        # 判断是用户自己修改密码还是管理员重置他人密码
        is_self_change = (target_username == auth_username)

        if is_self_change:
            # 用户修改自己的密码：需要验证原密码
            if not old_password:
                return jsonify({"success": False, "message": "请提供当前密码"})

            # 验证原密码
            if not auth_system.verify_password(target_username, old_password):
                return jsonify({"success": False, "message": "当前密码错误"}), 401
        else:
            # 管理员重置他人密码：需要管理权限
            if not auth_system.check_permission(auth_username, 'reset_user_password'):
                return jsonify({"success": False, "message": "权限不足"}), 403

        # 执行密码重置
        result = auth_system.reset_user_password(target_username, new_password)
        return jsonify(result)

    @app.route('/auth/user/update_avatar', methods=['POST'])
    def auth_user_update_avatar():
        """更新用户头像（URL方式）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not auth_username or auth_username == 'guest':
            return jsonify({"success": False, "message": "游客无法设置头像"})

        data = request.json
        avatar_url = data.get('avatar_url', '')

        result = auth_system.update_user_avatar(auth_username, avatar_url)
        return jsonify(result)

    @app.route('/auth/user/upload_avatar', methods=['POST'])
    def auth_user_upload_avatar():
        """上传用户头像文件（multipart/form-data）"""
        import hashlib
        from PIL import Image
        import io

        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not auth_username or auth_username == 'guest':
            return jsonify({"success": False, "message": "游客无法上传头像"})

        # 检查是否有文件上传
        if 'avatar' not in request.files:
            return jsonify({"success": False, "message": "未找到头像文件"}), 400

        file = request.files['avatar']

        if file.filename == '':
            return jsonify({"success": False, "message": "未选择文件"}), 400

        # 检查文件类型
        allowed_extensions = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in allowed_extensions:
            return jsonify({"success": False, "message": "不支持的文件格式，请上传图片文件"}), 400

        # 读取文件内容
        file_content = file.read()

        # 检查文件大小（限制为5MB）
        max_size = 5 * 1024 * 1024  # 5MB
        if len(file_content) > max_size:
            return jsonify({"success": False, "message": "文件过大，请上传小于5MB的图片"}), 400

        try:
            # 使用PIL打开图片并转换为PNG格式
            img = Image.open(io.BytesIO(file_content))

            # 转换为RGB模式（PNG不支持CMYK等模式）
            if img.mode in ('RGBA', 'LA', 'P'):
                # 保留透明度
                pass
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            # 将图片转换为PNG格式的字节流
            png_buffer = io.BytesIO()
            img.save(png_buffer, format='PNG', optimize=True)
            png_content = png_buffer.getvalue()

            # 计算SHA256哈希
            sha256_hash = hashlib.sha256(png_content).hexdigest()

            # 创建存储目录
            images_dir = os.path.join('system_accounts', 'images')
            os.makedirs(images_dir, exist_ok=True)

            # 保存为PNG文件
            filename = f"{sha256_hash}.png"
            filepath = os.path.join(images_dir, filename)

            with open(filepath, 'wb') as f:
                f.write(png_content)

            # 更新索引文件，记录上传信息
            index_file = os.path.join(images_dir, '_index.json')
            try:
                # 读取现有索引
                if os.path.exists(index_file):
                    with open(index_file, 'r', encoding='utf-8') as f:
                        index_data = json.load(f)
                else:
                    index_data = {
                        'version': '1.0',
                        'description': '用户头像索引文件，记录每个文件的上传信息',
                        'files': {}
                    }

                # 获取客户端IP
                ip_address = request.headers.get(
                    'X-Forwarded-For', request.remote_addr)

                # 添加/更新文件信息
                index_data['files'][filename] = {
                    'username': auth_username,
                    'upload_time': time.time(),
                    'upload_time_str': datetime.datetime.now().isoformat(),
                    'ip_address': ip_address,
                    'original_filename': file.filename,
                    'file_size': len(png_content),
                    'sha256': sha256_hash
                }

                # 保存索引
                with open(index_file, 'w', encoding='utf-8') as f:
                    json.dump(index_data, f, indent=2, ensure_ascii=False)

                logging.info(
                    f"用户 {auth_username} 从 {ip_address} 上传头像: {filename}")
            except Exception as e:
                logging.error(f"更新头像索引失败: {e}", exc_info=True)
                # 不影响上传流程，继续执行

        except Exception as e:
            return jsonify({"success": False, "message": f"图片处理失败: {str(e)}"}), 500

        # 构建头像URL
        avatar_url = f"/api/avatar/{filename}"

        # 更新用户头像
        result = auth_system.update_user_avatar(auth_username, avatar_url)

        if result.get('success'):
            # 返回新的头像URL
            return jsonify({
                "success": True,
                "avatar_url": avatar_url,
                "message": "头像上传成功"
            })
        else:
            # 如果更新失败，删除已上传的文件和索引
            try:
                os.remove(filepath)
                # 从索引中移除
                if os.path.exists(index_file):
                    with open(index_file, 'r', encoding='utf-8') as f:
                        index_data = json.load(f)
                    if filename in index_data.get('files', {}):
                        del index_data['files'][filename]
                        with open(index_file, 'w', encoding='utf-8') as f:
                            json.dump(index_data, f, indent=2,
                                      ensure_ascii=False)
            except (OSError, PermissionError) as e:
                logging.warning(f"[会话索引] 更新索引文件失败: {e}")
            return jsonify(result)

    @app.route('/api/avatar/<filename>', methods=['GET'])
    def serve_avatar(filename):
        """提供头像图片服务（需要会话认证，管理员可访问）"""
        from flask import send_file

        # 验证会话 - 支持从header、cookie或query参数获取session_id
        session_id = request.headers.get('X-Session-ID', '') or \
                     request.cookies.get('session_id', '') or \
                     request.args.get('session_id', '')

        # 如果没有会话ID或会话无效，返回401
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权访问"}), 401

        # 获取当前用户信息，检查是否为管理员
        api_instance = web_sessions[session_id]
        current_username = getattr(api_instance, 'auth_username', '')
        is_admin = auth_system.check_permission(
            current_username, 'manage_users') if current_username else False

        # 管理员可以访问所有头像，无需进一步验证
        # 普通用户也可以访问（因为他们已通过会话验证）

        # 验证文件名格式（只允许PNG文件，且文件名为64字符的十六进制哈希值）
        # 64 chars hash + .png (4 chars)
        if not filename.endswith('.png') or len(filename) != 68:
            return jsonify({"success": False, "message": "无效的文件名"}), 400

        # 构建文件路径
        filepath = os.path.join('system_accounts', 'images', filename)

        # 检查文件是否存在
        if not os.path.exists(filepath):
            return jsonify({"success": False, "message": "头像不存在"}), 404

        # 返回图片文件
        try:
            return send_file(filepath, mimetype='image/png')
        except Exception as e:
            return jsonify({"success": False, "message": f"读取文件失败: {str(e)}"}), 500

    @app.route('/auth/admin/clear_user_avatar', methods=['POST'])
    def auth_admin_clear_user_avatar():
        """管理员：强制清除用户头像"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查管理权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        target_username = data.get('username', '')

        if not target_username:
            return jsonify({"success": False, "message": "缺少用户名参数"}), 400

        # 获取目标用户的头像URL
        user_details = auth_system.get_user_details(target_username)
        if not user_details:
            return jsonify({"success": False, "message": "用户不存在"}), 404

        old_avatar_url = user_details.get('avatar_url', '')

        # 清除用户头像URL
        result = auth_system.update_user_avatar(target_username, '')

        if result.get('success'):
            # 如果有旧头像，尝试删除文件和索引（但不强制，因为可能被其他用户共享）
            if old_avatar_url:
                try:
                    # 提取文件名
                    filename = old_avatar_url.split('/')[-1]
                    filepath = os.path.join(
                        'system_accounts', 'images', filename)
                    index_file = os.path.join(
                        'system_accounts', 'images', '_index.json')

                    # 检查索引中的文件信息
                    if os.path.exists(index_file):
                        with open(index_file, 'r', encoding='utf-8') as f:
                            index_data = json.load(f)

                        file_info = index_data.get(
                            'files', {}).get(filename, {})

                        # 只有当文件是该用户上传的才删除（避免误删共享头像）
                        if file_info.get('username') == target_username:
                            # 删除文件
                            if os.path.exists(filepath):
                                os.remove(filepath)
                                logging.info(
                                    f"管理员 {auth_username} 删除了用户 {target_username} 的头像文件: {filename}")

                            # 从索引中移除
                            if filename in index_data.get('files', {}):
                                del index_data['files'][filename]
                                with open(index_file, 'w', encoding='utf-8') as f:
                                    json.dump(index_data, f, indent=2,
                                              ensure_ascii=False)
                        else:
                            logging.info(
                                f"管理员 {auth_username} 清除了用户 {target_username} 的头像URL，但文件由其他用户上传，未删除文件")

                except Exception as e:
                    logging.error(f"清除头像文件时出错: {e}", exc_info=True)
                    # 不影响主要功能，继续

            # 记录审计日志
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            auth_system.log_audit(
                auth_username,
                'clear_user_avatar',
                f'强制清除用户 {target_username} 的头像',
                ip_address,
                session_id
            )

            return jsonify({
                "success": True,
                "message": f"已清除用户 {target_username} 的头像"
            })
        else:
            return jsonify(result)

    @app.route('/auth/user/update_theme', methods=['POST'])
    def auth_user_update_theme():
        """更新用户主题"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not auth_username or auth_username == 'guest':
            return jsonify({"success": False, "message": "游客无法设置主题"})

        data = request.json
        theme = data.get('theme', 'light')

        result = auth_system.update_user_theme(auth_username, theme)
        return jsonify(result)

    @app.route('/auth/user/details', methods=['GET'])
    def auth_user_details():
        """获取用户详细信息"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not auth_username or auth_username == 'guest':
            return jsonify({"success": False, "message": "游客无详细信息"})

        details = auth_system.get_user_details(auth_username)
        if details:
            return jsonify({"success": True, "user": details})
        return jsonify({"success": False, "message": "用户不存在"})

    @app.route('/auth/user/avatar', methods=['GET'])
    def auth_get_user_avatar():
        """根据用户名获取头像URL（管理员可查询所有用户）

        查询参数:
        - username: 用户名（可选，不提供则返回当前用户）
        """
        session_id = request.headers.get('X-Session-ID', '')
        target_username = request.args.get('username', '')

        # 验证会话
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未授权访问"}), 401

        api_instance = web_sessions[session_id]
        if not hasattr(api_instance, 'auth_username'):
            return jsonify({"success": False, "message": "未登录"}), 401

        current_username = getattr(api_instance, 'auth_username', '')
        is_admin = auth_system.check_permission(
            current_username, 'manage_users')

        # 如果没有指定用户名，返回当前用户的头像
        if not target_username:
            target_username = current_username

            if not target_username or target_username == 'guest':
                return jsonify({
                    "success": True,
                    "avatar_url": "",
                    "message": "游客无头像"
                })

        # 管理员可以查询任何用户的头像
        # 普通用户只能查询自己的头像
        if not is_admin and target_username != current_username:
            return jsonify({
                "success": False,
                "message": "权限不足，只能查询自己的头像"
            }), 403

        # 获取指定用户的头像
        details = auth_system.get_user_details(target_username)
        if details:
            return jsonify({
                "success": True,
                "avatar_url": details.get('avatar_url', ''),
                "username": target_username
            })

        return jsonify({
            "success": False,
            "message": "用户不存在",
            "avatar_url": ""
        })

    @app.route('/auth/admin/update_max_sessions', methods=['POST'])
    def auth_admin_update_max_sessions():
        """更新用户最大会话数量（管理员）

        请求参数:
        - username: 目标用户名
        - max_sessions: 最大会话数 (1=单会话, >1=指定数量, -1=无限制)
        """
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')
        auth_group = getattr(api_instance, 'auth_group', 'guest')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'manage_users'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        data = request.json
        target_username = data.get('username', '')
        max_sessions = data.get('max_sessions', 1)

        # 验证参数
        if not isinstance(max_sessions, int) or (max_sessions < 1 and max_sessions != -1):
            return jsonify({
                "success": False,
                "message": "无效的会话数量：必须为正整数或-1（无限制）"
            }), 400

        result = auth_system.update_max_sessions(target_username, max_sessions)

        # 记录审计日志
        ip_address = request.headers.get(
            'X-Forwarded-For', request.remote_addr)
        auth_system.log_audit(
            auth_username,
            'update_max_sessions',
            f'修改用户 {target_username} 的最大会话数为: {max_sessions}',
            ip_address,
            session_id
        )

        return jsonify(result)

    @app.route('/auth/user/sessions', methods=['GET'])
    def auth_user_sessions():
        """获取用户的所有会话"""
        session_id = request.headers.get('X-Session-ID', '')
        api_instance = None
        is_guest = False
        auth_username = ''

        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]
                is_guest = getattr(api_instance, 'is_guest', False)
                auth_username = getattr(api_instance, 'auth_username', '')
            else:
                # 尝试从文件加载，以应对内存中可能不存在的情况 (例如刚恢复)
                state = load_session_state(session_id)
                if state:
                    # 仅获取必要信息，不创建完整实例
                    is_guest = state.get('is_guest', False)
                    auth_username = state.get('auth_username', '')
                else:
                    # 如果内存和文件都没有，则视为未登录
                    return jsonify({"success": False, "message": "会话无效或未登录"}), 401

        sessions_info = []

        if is_guest:
            # 游客模式：只显示当前会话
            logging.debug(
                f"auth_user_sessions: Handling guest session {session_id[:8]}")
            session_file = get_session_file_path(session_id)
            created_at = 0
            last_activity = 0
            login_success_status = False  # 游客没有学校账号登录状态

            if os.path.exists(session_file):
                try:
                    with open(session_file, 'r', encoding='utf-8') as f:
                        session_data = json.load(f)
                    created_at = session_data.get('created_at', 0)
                    last_activity = session_data.get(
                        'last_accessed', 0)  # 使用 last_accessed
                except Exception as e:
                    logging.warning(
                        f"Failed to read guest session file {session_file}: {e}")

            sessions_info.append({
                'session_id': session_id,
                # 添加哈希用于显示
                'session_hash': hashlib.sha256(session_id.encode()).hexdigest()[:16],
                'created_at': created_at,
                'last_activity': last_activity,
                'is_current': True,  # 游客访问时，这个会话总是当前的
                'login_success': login_success_status,  # 游客的学校登录状态总是 False
                'user_data': {"username": "guest"}  # 简单标识
            })
            logging.debug(
                f"auth_user_sessions: Guest session info prepared: {sessions_info}")
            return jsonify({"success": True, "sessions": sessions_info, "max_sessions": -1})

        elif auth_username:
            # 注册用户模式：保持原有逻辑
            logging.debug(
                f"auth_user_sessions: Handling registered user {auth_username}")
            session_ids = auth_system.get_user_sessions(auth_username)
            logging.debug(
                f"auth_user_sessions: Found linked session IDs for {auth_username}: {session_ids}")

            # 获取用户的最大会话数设置
            user_details = auth_system.get_user_details(auth_username)
            max_sessions = user_details.get(
                'max_sessions', 1) if user_details else 1

            for sid in session_ids:
                session_file = get_session_file_path(sid)
                if os.path.exists(session_file):
                    try:
                        with open(session_file, 'r', encoding='utf-8') as f:
                            session_data = json.load(f)

                        # 确保加载的数据属于当前用户
                        if session_data.get('auth_username') == auth_username:
                            sessions_info.append({
                                'session_id': sid,
                                # 添加哈希
                                'session_hash': hashlib.sha256(sid.encode()).hexdigest()[:16],
                                'created_at': session_data.get('created_at', 0),
                                # 使用 last_accessed
                                'last_activity': session_data.get('last_accessed', 0),
                                'is_current': sid == session_id,
                                'login_success': session_data.get('login_success', False),
                                'user_data': session_data.get('user_data', {})
                            })
                    except Exception as e:
                        logging.warning(
                            f"Failed to read session file {session_file} for user {auth_username}: {e}")
                        continue  # 跳过损坏的文件
            logging.debug(
                f"auth_user_sessions: Registered user session info prepared: {len(sessions_info)} sessions")
            return jsonify({"success": True, "sessions": sessions_info, "max_sessions": max_sessions})
        else:
            # 既不是游客也不是有效注册用户（理论上不应发生，除非状态异常）
            logging.warning(
                f"auth_user_sessions: Invalid state for session {session_id[:8]} - neither guest nor valid user.")
            return jsonify({"success": False, "message": "会话状态异常"}), 500

    @app.route('/auth/user/delete_session', methods=['POST'])
    def auth_user_delete_session():
        """删除用户的一个会话"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not auth_username or auth_username == 'guest':
            return jsonify({"success": False, "message": "游客无会话管理"})

        data = request.json
        target_session_id = data.get('session_id', '')

        if not target_session_id:
            return jsonify({"success": False, "message": "会话ID缺失"})

        # 不能删除当前会话
        if target_session_id == session_id:
            return jsonify({"success": False, "message": "不能删除当前会话"})

        # 从用户数据中移除
        auth_system.unlink_session_from_user(auth_username, target_session_id)

        # 删除会话文件
        session_file = get_session_file_path(target_session_id)
        if os.path.exists(session_file):
            try:
                os.remove(session_file)
            except (FileNotFoundError, PermissionError) as e:
                logging.debug(f"[会话删除] 删除会话文件失败: {e}")

        # 从内存中移除
        with web_sessions_lock:
            if target_session_id in web_sessions:
                del web_sessions[target_session_id]

        return jsonify({"success": True, "message": "会话已删除"})

    @app.route('/auth/user/create_session_persistence', methods=['POST'])
    def auth_user_create_session_persistence():
        """创建会话持久化文件（登录状态下）"""
        session_id = request.headers.get('X-Session-ID', '')

        api_instance = None
        is_guest = True  # 默认假设是游客或未登录

        if session_id:
            with web_sessions_lock:
                if session_id in web_sessions:
                    api_instance = web_sessions[session_id]
                    is_guest = getattr(
                        api_instance, 'is_guest', True)  # 检查内存中的实例
                else:
                    # 尝试从文件加载判断是否游客
                    state = load_session_state(session_id)
                    if state:
                        is_guest = state.get('is_guest', True)
                    # else: # 文件不存在，保持 is_guest = True (视为未登录)

        # 如果是游客或根本没有有效的会话实例，则禁止创建
        if is_guest:
            logging.warning(
                f"Attempt by guest session {session_id[:8]} to create persistent session blocked.")
            return jsonify({"success": False, "message": "游客不允许创建额外的会话"}), 403

    # （确保 api_instance 在非游客情况下是有效的）
        if not api_instance:  # 双重检查，理论上非游客时 api_instance 应该已加载
            return jsonify({"success": False, "message": "会话无效或未登录"}), 401

        auth_username = getattr(api_instance, 'auth_username', '')

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 游客也可以创建会话持久化，但不关联到用户账号
        is_guest = getattr(api_instance, 'is_guest', False)

        # 获取新的会话ID（从请求中或生成新的）
        data = request.json or {}
        new_session_id = data.get('session_id', '')

        if not new_session_id:
            return jsonify({"success": False, "message": "缺少会话ID"}), 400

        # 验证UUID格式
        import re
        uuid_pattern = re.compile(
            r'^[a-f0-9]{8}-[a-f0-9]{4}-4[a-f0-9]{3}-[89ab][a-f0-9]{3}-[a-f0-9]{12}$',
            re.IGNORECASE
        )
        if not uuid_pattern.match(new_session_id):
            return jsonify({"success": False, "message": "无效的UUID格式"}), 400

        # 创建新的API实例，继承当前用户的认证信息
        new_api_instance = Api(args)
        new_api_instance._session_created_at = time.time()
        new_api_instance._web_session_id = new_session_id

        # 继承认证信息
        if hasattr(api_instance, 'auth_username'):
            new_api_instance.auth_username = api_instance.auth_username
            new_api_instance.auth_group = getattr(
                api_instance, 'auth_group', 'guest')
            new_api_instance.is_guest = is_guest
            new_api_instance.is_authenticated = True

        # 继承参数配置
        if hasattr(api_instance, 'params'):
            new_api_instance.params = copy.deepcopy(api_instance.params)

        # 继承User-Agent
        if hasattr(api_instance, 'device_ua'):
            new_api_instance.device_ua = api_instance.device_ua

        # 如果是注册用户（非游客），处理会话关联
        cleanup_message = ""
        if not is_guest and auth_username:
            # 检查并强制执行会话数量限制
            old_sessions, cleanup_message = auth_system.check_single_session_enforcement(
                auth_username, new_session_id)

            # 清理旧会话（如果超出限制）- 使用后台线程异步清理，不阻塞响应
            if old_sessions:
                def cleanup_old_sessions_async():
                    for old_sid in old_sessions:
                        try:
                            cleanup_session(old_sid, "session_limit_exceeded")
                        except Exception as e:
                            logging.error(f"后台清理旧会话失败 {old_sid[:16]}...: {e}")

                cleanup_thread = threading.Thread(
                    target=cleanup_old_sessions_async, daemon=True)
                cleanup_thread.start()

            # 关联新会话到用户账号
            auth_system.link_session_to_user(auth_username, new_session_id)

            # 记录审计日志
            ip_address = request.headers.get(
                'X-Forwarded-For', request.remote_addr)
            audit_details = f'创建新会话持久化文件，会话ID: {new_session_id}'
            if cleanup_message:
                audit_details += f'; {cleanup_message}'

            auth_system.log_audit(
                auth_username,
                'create_session_persistence',
                audit_details,
                ip_address,
                session_id
            )

        # 保存新会话状态到文件
        with web_sessions_lock:
            web_sessions[new_session_id] = new_api_instance
            save_session_state(
                new_session_id, new_api_instance, force_save=True)

        response_data = {
            "success": True,
            "message": "会话持久化文件已创建",
            "session_id": new_session_id
        }

        # 添加清理提示（如果有）
        if cleanup_message:
            response_data['cleanup_message'] = cleanup_message

        logging.info(f"用户 {auth_username} 创建新会话持久化: {new_session_id[:32]}...")
        # 创建 Flask 响应对象
        response = make_response(jsonify(response_data))

        # 如果是注册用户（非游客），为其生成新 Token 并设置 Cookie
        if not is_guest and auth_username:
            try:
                # 1. 为新会话生成新 Token
                new_token = token_manager.create_token(
                    auth_username, new_session_id)
                logging.info(f"为新会话 {new_session_id[:8]}... 生成了 Token")

                # 2. 设置 Set-Cookie 响应头
                response.set_cookie(
                    'auth_token',
                    value=new_token,
                    max_age=3600,  # 1 小时
                    httponly=True,  # 关键：防止 JS 访问
                    secure=False,  # 开发环境 False，生产环境应设为 True (需要 HTTPS)
                    samesite='Lax'  # 推荐的 SameSite 策略
                )
                logging.info(
                    f"已为新会话 {new_session_id[:8]}... 设置 auth_token Cookie")

            except Exception as token_err:
                # 记录 Token 生成或 Cookie 设置错误，但不阻止会话创建本身
                logging.error(
                    f"为新会话 {new_session_id[:8]} 生成 Token 或设置 Cookie 时出错: {token_err}", exc_info=True)
                # 可以选择在这里修改 response_data 添加警告信息
                # response_data['warning'] = "Token未能成功设置，后续操作可能需要重新登录"
                # response = make_response(jsonify(response_data)) # 如果需要更新响应内容

        # 返回带有 Set-Cookie (如果适用) 的响应对象
        return response

    @app.route('/auth/admin/all_sessions', methods=['GET'])
    def auth_admin_all_sessions():
        """管理员：获取所有活跃会话（上帝模式）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        if not session_id:
            return jsonify({"success": False, "message": "缺少当前会话ID"}), 401

        with web_sessions_lock:
            if session_id not in web_sessions:
                # 尝试从文件加载，如果连文件都没有，那肯定没登录
                state = load_session_state(session_id)
                if not state:
                    return jsonify({"success": False, "message": "当前会话无效或已过期"}), 401
                # 如果文件存在，创建一个临时实例来获取用户名
                api_instance = Api(args)
                restore_session_to_api_instance(api_instance, state)
                web_sessions[session_id] = api_instance  # 加载成功后放入内存
                logging.info(f"创建新会话时，按需恢复了发起请求的会话 {session_id[:8]}")
            else:
                api_instance = web_sessions[session_id]

        auth_username = getattr(api_instance, 'auth_username', None)
        is_guest = getattr(api_instance, 'is_guest', True)

        # 只对非游客用户验证 Token
        if not is_guest and auth_username:
            token_from_cookie = request.cookies.get('auth_token')
            if not token_from_cookie:
                logging.warning(
                    f"用户 {auth_username} 尝试创建会话但缺少 auth_token cookie")
                return jsonify({"success": False, "message": "缺少认证令牌(cookie)，请重新登录", "need_login": True}), 401

            is_valid, reason = token_manager.verify_token(
                auth_username, session_id, token_from_cookie)
            if not is_valid:
                logging.warning(f"用户 {auth_username} 尝试创建会话但令牌无效: {reason}")
                # 令牌无效，要求重新登录，并清除无效cookie
                response_data = {
                    "success": False, "message": f"令牌验证失败 ({reason})，请重新登录", "need_login": True}
                response = make_response(jsonify(response_data), 401)
                response.set_cookie('auth_token', '', max_age=0)
                return response
            else:
                # Token有效，刷新它
                token_manager.refresh_token(auth_username, session_id)
                logging.debug(
                    f"用户 {auth_username} (会话 {session_id[:8]}) Token 验证通过并已刷新")
        elif not is_guest and not auth_username:
            # 有会话实例但没有用户名，说明状态异常
            logging.error(f"会话 {session_id[:8]} 存在但缺少用户名，无法创建新会话")
            return jsonify({"success": False, "message": "当前会话状态异常，请重新登录", "need_login": True}), 401

        # 检查上帝模式权限
        if not auth_system.check_permission(auth_username, 'god_mode'):
            return jsonify({"success": False, "message": "需要上帝模式权限"}), 403

        # 获取所有会话信息（包括内存中的和磁盘上的）
        all_sessions = []
        session_ids_in_memory = set()

        # 1. 首先获取内存中的会话
        with web_sessions_lock:
            for sid, api in web_sessions.items():
                session_info = {
                    'session_id': sid,
                    'session_hash': hashlib.sha256(sid.encode()).hexdigest()[:16],
                    'auth_username': getattr(api, 'auth_username', None),
                    'auth_group': getattr(api, 'auth_group', 'guest'),
                    'is_authenticated': getattr(api, 'is_authenticated', False),
                    'is_guest': getattr(api, 'is_guest', False),
                    'created_at': getattr(api, '_session_created_at', 0),
                    'login_success': getattr(api, 'login_success', False),
                    'user_info': getattr(api, 'user_info', {}),
                    'is_current': sid == session_id,
                    # 添加username字段供前端使用
                    'username': getattr(api, 'auth_username', None)
                }
                all_sessions.append(session_info)
                session_ids_in_memory.add(sid)

        # 2. 然后扫描磁盘上的会话文件，添加未在内存中的会话
        try:
            if os.path.exists(SESSION_STORAGE_DIR):
                for filename in os.listdir(SESSION_STORAGE_DIR):
                    if filename == '_index.json' or not filename.endswith('.json'):
                        continue

                    session_file = os.path.join(SESSION_STORAGE_DIR, filename)
                    try:
                        with open(session_file, 'r', encoding='utf-8') as f:
                            state = json.load(f)

                        sid = state.get('session_id')
                        if not sid or sid in session_ids_in_memory:
                            continue  # 跳过已在内存中的会话

                        # 从文件中读取会话信息
                        session_info = {
                            'session_id': sid,
                            'session_hash': hashlib.sha256(sid.encode()).hexdigest()[:16],
                            'auth_username': state.get('auth_username', None),
                            'auth_group': state.get('auth_group', 'guest'),
                            'is_authenticated': state.get('is_authenticated', False),
                            'is_guest': state.get('is_guest', False),
                            'created_at': state.get('created_at', 0),
                            'login_success': state.get('login_success', False),
                            'user_info': state.get('user_info', {}),
                            'is_current': sid == session_id,
                            # 添加username字段供前端使用
                            'username': state.get('auth_username', None)
                        }
                        all_sessions.append(session_info)
                    except Exception as e:
                        logging.warning(f"读取会话文件 {filename} 失败: {e}")
                        continue
        except Exception as e:
            logging.error(f"扫描会话文件目录失败: {e}")

        return jsonify({
            "success": True,
            "sessions": all_sessions,
            "total_count": len(all_sessions)
        })

    @app.route('/auth/admin/destroy_session', methods=['POST'])
    def auth_admin_destroy_session():
        """管理员：强制销毁任意会话（上帝模式）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查上帝模式权限
        if not auth_system.check_permission(auth_username, 'god_mode'):
            return jsonify({"success": False, "message": "需要上帝模式权限"}), 403

        data = request.json
        target_session_id = data.get('session_id', '')

        if not target_session_id:
            return jsonify({"success": False, "message": "会话ID缺失"})

        # 不能销毁自己的会话
        if target_session_id == session_id:
            return jsonify({"success": False, "message": "不能销毁当前会话"})

        # 获取目标会话的用户信息（用于日志）
        target_username = 'unknown'
        with web_sessions_lock:
            if target_session_id in web_sessions:
                target_api = web_sessions[target_session_id]
                target_username = getattr(
                    target_api, 'auth_username', 'unknown')

        # 从用户数据中移除会话关联
        if target_username != 'unknown' and target_username != 'guest':
            auth_system.unlink_session_from_user(
                target_username, target_session_id)

        # 删除会话文件

        session_hash = hashlib.sha256(target_session_id.encode()).hexdigest()
        session_file = os.path.join(
            SESSION_STORAGE_DIR, f"{session_hash}.json")
        if os.path.exists(session_file):
            try:
                os.remove(session_file)
            except (FileNotFoundError, PermissionError) as e:
                logging.debug(f"[会话强制登出] 删除会话文件失败: {e}")

        # 从内存中移除
        with web_sessions_lock:
            if target_session_id in web_sessions:
                del web_sessions[target_session_id]

        # 记录审计日志
        ip_address = request.headers.get(
            'X-Forwarded-For', request.remote_addr)
        auth_system.log_audit(
            auth_username,
            'destroy_session',
            f'强制销毁用户 {target_username} 的会话: {target_session_id[:32]}...',
            ip_address,
            session_id
        )

        return jsonify({
            "success": True,
            "message": f"已销毁用户 {target_username} 的会话"
        })

    @app.route('/auth/admin/audit_logs', methods=['GET'])
    def auth_admin_audit_logs():
        """获取审计日志（管理员）"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401

        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')

        # 检查权限
        if not auth_system.check_permission(auth_username, 'view_audit_logs'):
            return jsonify({"success": False, "message": "权限不足"}), 403

        # 获取参数
        username = request.args.get('username', None)
        action = request.args.get('action', None)
        limit = int(request.args.get('limit', 100))

        logs = auth_system.get_audit_logs(username, action, limit)
        return jsonify({
            "success": True,
            "logs": logs
        })

    # ====================
    # 前端日志接收API
    # ====================

    @app.route('/api/log_frontend', methods=['POST'])
    def log_frontend():
        """接收前端日志并保存到后端日志文件"""
        try:
            data = request.get_json() or {}
            level = data.get('level', 'INFO').upper()
            message = data.get('message', '')
            timestamp = data.get('timestamp', '')
            source = data.get('source', 'unknown')

            if (data == None) or (not message):
                return jsonify({"success": False, "message": "无效的日志数据"}),
           

            # 获取 Session ID
            session_id = request.headers.get('X-Session-ID', 'UnknownSession')
            # session_id_short = session_id[:8] if session_id else None # 取前8位用于日志

            # 获取 IP 地址 (考虑代理)
            ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)

            # 获取用户名 (通过 Session ID)
            username = 'Guest/Unknown' # 默认值
            with web_sessions_lock:
                if session_id in web_sessions:
                    api_instance = web_sessions[session_id]
                    # 优先使用 auth_username，如果不存在则尝试 user_data.username
                    username_attr = getattr(api_instance, 'auth_username', None)
                    if not username_attr and hasattr(api_instance, 'user_data'):
                         username_attr = getattr(api_instance.user_data, 'username', None)

                    if username_attr: # 确保获取到的用户名非空
                         username = username_attr
                    elif getattr(api_instance, 'is_guest', False): # 明确是游客
                         username = 'Guest'
                    # 如果 session 存在但无法确定用户名，保留 'Guest/Unknown'

            # 构造新的日志消息，包含 IP、用户名和 Session ID
            log_message = f"[前端日志][IP:{ip_address}][前端时间:{timestamp}][用户:{username}][Session Id:{session_id}][{source}] {message}"

            # 根据级别记录日志
            if level == 'DEBUG':
                logging.debug(log_message)
            elif level == 'INFO':
                logging.info(log_message)
            elif level == 'WARNING' or level == 'WARN':
                logging.warning(log_message)
            elif level == 'ERROR':
                logging.error(log_message)
            elif level == 'CRITICAL':
                logging.critical(log_message)
            else:
                logging.info(log_message)

            return jsonify({"success": True})
        except Exception as e:
            # 在错误日志中也尝试包含 IP 和 Session ID
            session_id_err = request.headers.get('X-Session-ID', 'UnknownSession')
            ip_address_err = request.headers.get('X-Forwarded-For', request.remote_addr)
            logging.error(f"[前端日志处理错误][IP:{ip_address_err}][Sess:{session_id_err[:8]}] {e}", exc_info=True)
            return jsonify({"success": False, "message": str(e)}), 500
        
    # ====================
    # 应用主路由
    # ====================

    @app.route('/')
    def index():
        """首页：显示登录页面，等待用户认证后分配UUID"""
        # 不再自动分配UUID，直接返回HTML让前端处理认证
        # UUID将在用户完成认证（游客登录或系统账号登录）后由前端或后端API分配
        return render_template_string(html_content)

    @app.route('/uuid=<uuid>')
    def session_view(uuid):
        """会话页面：显示应用界面"""
        # --- 修复 UUID 验证 ---
        # 验证UUID格式 - 标准UUID v4格式
        uuid_pattern = re.compile(
            r'^[a-f0-9]{8}-[a-f0-9]{4}-4[a-f0-9]{3}-[89ab][a-f0-9]{3}-[a-f0-9]{12}$',
            re.IGNORECASE
        )
        if not uuid or not uuid_pattern.match(uuid):
            # 日志记录修改为记录实际收到的UUID（截断）和失败原因
            logging.warning(
                f"无效的UUID格式或不匹配标准格式: {uuid[:40] if uuid else 'None'}...")
            return redirect(url_for('index'))  # 验证失败，重定向到首页

        # 确保Api实例存在（从URL或文件恢复会话，不依赖Flask session）
        with web_sessions_lock:
            if uuid not in web_sessions:
                # 尝试从文件加载会话状态
                state = load_session_state(uuid)
                api_instance = Api(args)
                api_instance._session_created_at = time.time()
                api_instance._web_session_id = uuid  # 关键：保存session_id到实例，用于后台任务自动保存

                if state and state.get('login_success'):
                    # 恢复登录状态
                    api_instance.login_success = True
                    api_instance.user_info = state.get('user_info')
                    api_instance._session_created_at = state.get(
                        'created_at', time.time())

                    # 使用新的恢复函数恢复完整状态
                    restore_session_to_api_instance(api_instance, state)

                    logging.info(
                        f"从文件恢复已登录会话 : {uuid[:32]}... (用户: {state.get('user_info', {}).get('username', 'Unknown')}, 任务数: {len(api_instance.all_run_data)})")
                else:
                    logging.info(f"创建新会话 : {uuid[:32]}...")

                web_sessions[uuid] = api_instance
                # save_session_state(uuid, api_instance)
            else:
                # 确保已有会话也有_web_session_id属性
                api_instance = web_sessions[uuid]
                if not hasattr(api_instance, '_web_session_id'):
                    api_instance._web_session_id = uuid
                logging.debug(f"使用现有会话: {uuid[:32]}...")

        # 返回HTML内容
        return render_template_string(html_content)

    @app.route('/api/<path:method>', methods=['GET', 'POST'])
    def api_call(method):
        """API调用端点：将前端调用转发到Python后端"""
        # 从请求头获取session_id（前端会在每次调用时添加）
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id:
            return jsonify({"success": False, "message": "缺少会话ID"}), 401

        # ===== Token验证 (仅非游客会话) =====
        with web_sessions_lock:
            if session_id in web_sessions:
                api_instance = web_sessions[session_id]

                # 如果是已认证的非游客用户，验证token
                if hasattr(api_instance, 'is_authenticated') and api_instance.is_authenticated:
                    if hasattr(api_instance, 'is_guest') and not api_instance.is_guest:
                        username = getattr(api_instance, 'auth_username', None)

                        if username:
                            # 从cookie获取token
                            token = request.cookies.get('auth_token')

                            if not token:
                                return jsonify({"success": False, "message": "未找到认证令牌，请重新登录", "need_login": True}), 401

                            # 验证token
                            is_valid, reason = token_manager.verify_token(
                                username, session_id, token)

                            if not is_valid:
                                if reason == "token_expired":
                                    return jsonify({"success": False, "message": "令牌已过期，请重新登录", "need_login": True}), 401
                                elif reason == "token_mismatch":
                                    return jsonify({"success": False, "message": "令牌验证失败，可能账号在其他设备登录", "need_login": True, "logged_out_elsewhere": True}), 401
                                else:
                                    return jsonify({"success": False, "message": "令牌验证失败，请重新登录", "need_login": True}), 401

                            # Token有效，刷新过期时间
                            token_manager.refresh_token(username, session_id)

        # 更新会话活动时间
        update_session_activity(session_id)

        with web_sessions_lock:
            if session_id not in web_sessions:
                # 尝试恢复会话
                state = load_session_state(session_id)
                if state and state.get('login_success'):
                    api_instance = Api(args)
                    api_instance._session_created_at = state.get(
                        'created_at', time.time())
                    api_instance.login_success = True
                    api_instance.user_info = state.get('user_info')
                    restore_session_to_api_instance(api_instance, state)
                    web_sessions[session_id] = api_instance
                    logging.info(f"API调用时自动恢复会话: {session_id[:32]}...")
                else:
                    return jsonify({"success": False, "message": "会话已过期或无效"}), 401
            api_instance = web_sessions[session_id]

        # 获取请求参数
        if request.method == 'POST':
            params = request.get_json() or {}
        else:
            params = dict(request.args)

        # 调用对应的API方法
        try:
            # 权限检查：需要特定权限的方法
            permission_required_methods = {
                'mark_notification_read': 'mark_notifications_read',
                'mark_all_read': 'mark_notifications_read',
                'trigger_attendance': 'use_attendance',
                'multi_start_single_account': 'execute_multi_account',
                'multi_start_all_accounts': 'execute_multi_account',
            }

            if method in permission_required_methods:
                required_permission = permission_required_methods[method]
                if hasattr(api_instance, 'auth_username'):
                    if not auth_system.check_permission(api_instance.auth_username, required_permission):
                        return jsonify({"success": False, "message": f"权限不足：需要 {required_permission} 权限"}), 403
                else:
                    return jsonify({"success": False, "message": "请先登录认证"}), 401

            if hasattr(api_instance, method):
                func = getattr(api_instance, method)
                # 将参数展开调用
                if params:
                    result = func(**params) if isinstance(params,
                                                          dict) else func(*params)
                else:
                    result = func()

                # 关键改进：对于会改变会话状态的API调用，保存会话状态
                # 扩展自动保存的方法列表，包括所有可能改变状态的操作
                auto_save_methods = [
                    'login', 'logout', 'load_tasks', 'select_task',
                    'start_single_run', 'start_all_runs', 'stop_current_run',
                    'import_offline_file', 'export_offline_file',
                    'record_path', 'auto_generate_path', 'process_path', 'clear_path',
                    'update_param', 'generate_new_ua',
                    'enter_multi_account_mode', 'exit_multi_account_mode', 'enter_single_account_mode',  # 模式切换
                    'multi_add_account', 'multi_remove_account'  # 多账号操作
                ]
                if method in auto_save_methods:
                    save_session_state(session_id, api_instance)
                    logging.debug(f"API '{method}' 调用后自动保存会话状态")

                # 返回结果时刷新cookie（如果是非游客）
                response = jsonify(
                    result if result is not None else {"success": True})

                # 刷新cookie过期时间（仅非游客）
                if hasattr(api_instance, 'is_authenticated') and api_instance.is_authenticated:
                    if hasattr(api_instance, 'is_guest') and not api_instance.is_guest:
                        token = request.cookies.get('auth_token')
                        if token:
                            response.set_cookie(
                                'auth_token',
                                value=token,
                                max_age=3600,
                                httponly=True,
                                secure=False,
                                samesite='Lax'
                            )

                return response
            else:
                return jsonify({"success": False, "message": f"未知的API方法: {method}"}), 404
        except Exception as e:
            logging.error(f"API调用失败 {method}: {e}", exc_info=True)
            # 不暴露详细错误信息给前端，只记录到日志
            return jsonify({"success": False, "message": "服务器内部错误"}), 500

    @app.route('/execute_js', methods=['POST'])
    def execute_js():
        """在服务器端Chrome中执行JavaScript代码"""
        # 从请求头获取session_id
        session_id = request.headers.get('X-Session-ID', '')

        if not session_id:
            return jsonify({"success": False, "message": "缺少会话ID"}), 401
        data = request.get_json() or {}
        script = data.get('script', '')
        args_list = data.get('args', [])

        if not script:
            return jsonify({"success": False, "message": "缺少script参数"}), 400

        try:
            result = chrome_pool.execute_js(session_id, script, *args_list)
            return jsonify({"success": True, "result": result})
        except Exception as e:
            logging.error(f"执行JS失败: {e}")
            return jsonify({"success": False, "message": "JS执行失败"}), 500

    @app.route('/api/background_task/start', methods=['POST'])
    def start_background_task():
        """启动后台任务执行"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "会话无效或未登录"}), 401
        
        data = request.get_json() or {}
        task_indices = data.get('task_indices', [])
        auto_generate = data.get('auto_generate', False)
        
        if not task_indices:
            return jsonify({"success": False, "message": "未指定任务"}), 400
        
        api_instance = web_sessions[session_id]
        result = background_task_manager.start_background_task(
            session_id, api_instance, task_indices, auto_generate
        )
        return jsonify(result)

    @app.route('/api/background_task/status', methods=['GET'])
    def get_background_task_status():
        """获取后台任务状态"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id:
            return jsonify({"success": False, "message": "缺少会话ID"}), 401
        
        task_status = background_task_manager.get_task_status(session_id)
        if task_status:
            return jsonify({"success": True, "task_status": task_status})
        else:
            return jsonify({"success": False, "message": "未找到后台任务"})

    @app.route('/api/background_task/stop', methods=['POST'])
    def stop_background_task():
        """停止后台任务"""
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "会话无效或未登录"}), 401
        
        result = background_task_manager.stop_task(session_id)
        return jsonify(result)

    # ========== 留言板API ========== #
    
    @app.route('/api/messages/list', methods=['GET'])
    def get_messages():
        """获取留言列表"""
        # 验证会话
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401
        
        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')
        
        # 检查权限
        if not auth_system.check_permission(auth_username, 'view_messages'):
            return jsonify({"success": False, "message": "无权查看留言"}), 403
        
        # 读取留言文件
        messages_file = 'messages.json'
        messages = []
        
        if os.path.exists(messages_file):
            try:
                with open(messages_file, 'r', encoding='utf-8') as f:
                    messages = json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                logging.error(f"[留言板] 读取留言失败: {e}")
                messages = []
        
        # 按时间倒序排序（最新的在前）
        messages.sort(key=lambda x: x.get('timestamp', 0), reverse=True)
        
        return jsonify({"success": True, "messages": messages})
    
    @app.route('/api/messages/post', methods=['POST'])
    def post_message():
        """发表留言"""
        # 验证会话
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401
        
        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')
        
        # 检查权限
        if not auth_system.check_permission(auth_username, 'post_messages'):
            return jsonify({"success": False, "message": "无权发表留言"}), 403
        
        data = request.json
        content = data.get('content', '').strip()
        email = data.get('email', '').strip()
        nickname = data.get('nickname', '').strip()
        
        # 验证内容
        if not content:
            return jsonify({"success": False, "message": "留言内容不能为空"})
        
        if len(content) > 1000:
            return jsonify({"success": False, "message": "留言内容不能超过1000字"})
        
        # 游客必须填写邮箱和昵称
        is_guest = (auth_username == 'guest' or not auth_username)
        if is_guest:
            if not email:
                return jsonify({"success": False, "message": "游客必须填写邮箱"})
            if not nickname:
                return jsonify({"success": False, "message": "游客必须设置昵称"})
            # 简单的邮箱格式验证
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                return jsonify({"success": False, "message": "邮箱格式不正确"})
        
        # 构建留言对象
        import time
        import uuid
        message = {
            "id": str(uuid.uuid4()),
            "content": content,
            "username": auth_username if not is_guest else None,
            "nickname": nickname if is_guest else None,
            "email": email if is_guest else None,
            "is_guest": is_guest,
            "timestamp": time.time(),
            "ip": request.remote_addr
        }
        
        # 读取现有留言
        messages_file = 'messages.json'
        messages = []
        
        if os.path.exists(messages_file):
            try:
                with open(messages_file, 'r', encoding='utf-8') as f:
                    messages = json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                logging.error(f"[留言板] 读取留言失败: {e}")
                messages = []
        
        # 添加新留言
        messages.append(message)
        
        # 保存留言
        try:
            with open(messages_file, 'w', encoding='utf-8') as f:
                json.dump(messages, f, indent=2, ensure_ascii=False)
            
            logging.info(f"[留言板] 新留言 --> 用户: {auth_username}, 昵称: {nickname}, 内容长度: {len(content)}字")
            return jsonify({"success": True, "message": "留言发表成功", "message_id": message["id"]})
        except OSError as e:
            logging.error(f"[留言板] 保存留言失败: {e}")
            return jsonify({"success": False, "message": "保存留言失败"}), 500
    
    @app.route('/api/messages/delete', methods=['POST'])
    def delete_message():
        """删除留言"""
        # 验证会话
        session_id = request.headers.get('X-Session-ID', '')
        if not session_id or session_id not in web_sessions:
            return jsonify({"success": False, "message": "未登录"}), 401
        
        api_instance = web_sessions[session_id]
        auth_username = getattr(api_instance, 'auth_username', '')
        
        data = request.json
        message_id = data.get('message_id', '').strip()
        
        if not message_id:
            return jsonify({"success": False, "message": "留言ID不能为空"})
        
        # 读取留言
        messages_file = 'messages.json'
        messages = []
        
        if os.path.exists(messages_file):
            try:
                with open(messages_file, 'r', encoding='utf-8') as f:
                    messages = json.load(f)
            except (json.JSONDecodeError, OSError) as e:
                logging.error(f"[留言板] 读取留言失败: {e}")
                return jsonify({"success": False, "message": "读取留言失败"}), 500
        
        # 查找要删除的留言
        message_to_delete = None
        for msg in messages:
            if msg.get('id') == message_id:
                message_to_delete = msg
                break
        
        if not message_to_delete:
            return jsonify({"success": False, "message": "留言不存在"})
        
        # 检查权限
        can_delete_any = auth_system.check_permission(auth_username, 'delete_any_messages')
        can_delete_own = auth_system.check_permission(auth_username, 'delete_own_messages')
        is_own_message = (message_to_delete.get('username') == auth_username and not message_to_delete.get('is_guest'))
        
        if can_delete_any:
            # 管理员可以删除任何留言
            pass
        elif can_delete_own and is_own_message:
            # 普通用户可以删除自己的留言
            pass
        else:
            return jsonify({"success": False, "message": "无权删除此留言"}), 403
        
        # 删除留言
        messages = [msg for msg in messages if msg.get('id') != message_id]
        
        # 保存留言
        try:
            with open(messages_file, 'w', encoding='utf-8') as f:
                json.dump(messages, f, indent=2, ensure_ascii=False)
            
            logging.info(f"[留言板] 删除留言 --> 操作用户: {auth_username}, 留言ID: {message_id}")
            return jsonify({"success": True, "message": "留言已删除"})
        except OSError as e:
            logging.error(f"[留言板] 保存留言失败: {e}")
            return jsonify({"success": False, "message": "保存留言失败"}), 500

    @app.route('/health')
    def health():
        """健康检查端点"""
        return jsonify({
            "status": "ok",
            "sessions": len(web_sessions),
            "chrome_contexts": len(chrome_pool.contexts) if chrome_pool else 0
        })

    @socketio.on('connect')
    def handle_connect():
        # 前端 JS 连接时会触发
        # 从请求头或 cookie 中获取 session_id (这取决于前端如何发送)
        # 假设前端通过 cookie 或查询参数传递 session_id
        session_id = request.args.get('session_id') or request.cookies.get(
            'session_id_cookie')  # 示例，需要前端配合
        if not session_id:
            # 或者从 Flask session 获取，如果使用了 Flask-Session
            session_id = session.get('session_id')

        # 更可靠的方式: 让前端在连接后立即发送一个带有 session_id 的 'join' 事件
        logging.info(f"WebSocket client connected: {request.sid}")
        # 注意：此时还不知道是哪个 session_id，等待前端发送 'join' 事件

    @socketio.on('join')
    def handle_join(data):
        session_id = data.get('session_id')
        if session_id:
            # 将当前 WebSocket 连接加入以 session_id 命名的房间
            join_room(session_id)
            logging.info(
                f"WebSocket client {request.sid} joined room: {session_id[:8]}...")
            # 可以选择性地在这里发送一条欢迎消息
            # emit('log_message', {'msg': 'WebSocket connected successfully.'}, room=session_id)
        else:
            logging.warning(
                f"WebSocket client {request.sid} failed to join room: session_id missing.")

    @socketio.on('disconnect')
    def handle_disconnect():
        logging.info(f"WebSocket client disconnected: {request.sid}")

    # 定期清理过期会话

    def cleanup_sessions():
        """定期清理超过24小时无活动的会话"""
        while True:
            time.sleep(3600)  # 每小时检查一次
            with web_sessions_lock:
                # 这里可以添加更复杂的会话过期逻辑
                # 清理Chrome上下文
                if chrome_pool:
                    for session_id in list(web_sessions.keys()):
                        # 可以添加超时检查逻辑
                        pass

    cleanup_thread = threading.Thread(target=cleanup_sessions, daemon=True)
    cleanup_thread.start()

    # 加载持久化会话
    logging.info("正在加载持久化会话...")
    load_all_sessions(args)

    # 启动会话监控线程
    logging.info("正在启动会话监控...")
    start_session_monitor()

    # 扫描 .ini 文件，为已配置的账号启动服务
    # 这必须在加载会话之后，但在启动Web服务器（阻塞）之前
    try:
        start_background_auto_attendance(args)
    except Exception as e:
        logging.error(f"启动后台自动签到服务失败: {e}", exc_info=True)

    cleaned_sessions_count = 0
    with web_sessions_lock:
        for session_id, api_instance in web_sessions.items():
            try:
                # 强制设置所有停止标志为 True (停止状态)
                if hasattr(api_instance, 'stop_run_flag') and isinstance(api_instance.stop_run_flag, threading.Event):
                    api_instance.stop_run_flag.set()
                if hasattr(api_instance, 'multi_run_stop_flag') and isinstance(api_instance.multi_run_stop_flag, threading.Event):
                    api_instance.multi_run_stop_flag.set()
                if hasattr(api_instance, 'stop_auto_refresh') and isinstance(api_instance.stop_auto_refresh, threading.Event):
                    api_instance.stop_auto_refresh.set()
                if hasattr(api_instance, 'stop_multi_auto_refresh') and isinstance(api_instance.stop_multi_auto_refresh, threading.Event):
                    api_instance.stop_multi_auto_refresh.set()

                # 如果是多账号模式，也停止所有子账号的事件
                if getattr(api_instance, 'is_multi_account_mode', False) and hasattr(api_instance, 'accounts'):
                    for acc in api_instance.accounts.values():
                        if hasattr(acc, 'stop_event') and isinstance(acc.stop_event, threading.Event):
                            acc.stop_event.set()
                        # 清理可能存在的 worker_thread 引用（线程已不存在）
                        acc.worker_thread = None
                        # 重置状态文本为待命（如果不是错误状态）
                        if not api_instance._should_preserve_status(acc.status_text, "待命"):
                             acc.status_text = "待命"

                # 清理 Api 实例可能残留的 worker thread 引用
                if hasattr(api_instance, 'auto_refresh_thread'):
                    api_instance.auto_refresh_thread = None
                if hasattr(api_instance, 'multi_auto_refresh_thread'):
                    api_instance.multi_auto_refresh_thread = None

                # 重置单账号进度
                api_instance.current_run_idx = -1
                api_instance._first_center_done = False

                cleaned_sessions_count += 1
            except Exception as e:
                logging.error(f"重置会话 {session_id[:8]}... 状态时出错: {e}")

    logging.info(f"已重置 {cleaned_sessions_count} 个会话的运行状态标志。")

    logging.info("正在检查并修正持久化的后台任务状态...")
    interrupted_task_files = 0
    if background_task_manager and os.path.exists(background_task_manager.task_storage_dir):
        try:
            for filename in os.listdir(background_task_manager.task_storage_dir):
                # 只处理该管理器的JSON状态文件
                if filename.endswith('.json'):
                    filepath = os.path.join(background_task_manager.task_storage_dir, filename)
                    try:
                        # 读取-修改-写回，需要独占访问（使用文件锁或确保单线程启动）
                        # 这里假设启动阶段是单线程的，直接读写
                        task_state = None
                        with open(filepath, 'r', encoding='utf-8') as f:
                            task_state = json.load(f)

                        # 如果状态是 'running'，则修改为 'stopped' 并添加原因
                        if task_state and task_state.get('status') == 'running':
                            task_state['status'] = 'stopped' # 或者 'interrupted' 更准确
                            task_state['error'] = '程序意外重启导致任务中断。'
                            task_state['last_update'] = time.time() # 更新时间戳

                            # 写回文件
                            with open(filepath, 'w', encoding='utf-8') as f:
                                json.dump(task_state, f, indent=2, ensure_ascii=False)
                            interrupted_task_files += 1
                            logging.debug(f"已将持久化的后台任务 {filename} 状态修正为 stopped。")

                    except (IOError, json.JSONDecodeError) as e:
                        logging.warning(f"处理后台任务状态文件 {filename} 时出错: {e}，跳过此文件。")
                    except Exception as e: # 捕获其他可能的错误
                         logging.error(f"更新后台任务状态文件 {filename} 时发生意外错误: {e}", exc_info=True)

            if interrupted_task_files > 0:
                logging.info(f"已修正 {interrupted_task_files} 个持久化的 'running' 后台任务状态为 stopped。")
            else:
                logging.info("未发现需要修正状态的持久化后台任务文件。")
        except Exception as e:
            logging.error(f"检查持久化后台任务状态时发生错误: {e}", exc_info=True)
    else:
        logging.info("后台任务管理器未初始化或存储目录不存在，跳过持久化状态检查。")


    # # 清理后台任务管理器的内存状态（但不删除持久化文件）
    if background_task_manager:
        with background_task_manager.lock:
            initial_task_count = len(background_task_manager.tasks)
            background_task_manager.tasks.clear()
            logging.info(f"已清空后台任务管理器的内存状态（清理了 {initial_task_count} 个任务记录）。")

    # 启动服务器
    print(f"\n{'='*60}")
    print(f"  跑步助手 Web 模式已启动（服务器端Chrome渲染，2048位UUID）")
    print(f"  访问地址: http://{args.host}:{args.port}")
    print(f"  首次访问将自动分配2048位UUID并重定向")
    print(f"  会话持久化已启用（服务器重启后保留登录状态）")
    print(f"  JS计算在服务器端Chrome中执行，提升安全性")
    print(f"  会话监控已启用（5分钟不活跃自动清理）")
    print(f"{'='*60}\n")

    try:
        logging.info(
            f"正在启动带有 WebSocket 支持的 Web 服务器于 http://{args.host}:{args.port}")
        socketio.run(app, host=args.host, port=args.port, debug=False)
    except OSError as e:
        if "WinError 10013" in str(e) or "permission" in str(e).lower() or "访问权限" in str(e):
            print(f"\n{'='*60}")
            print(f"错误: 端口 {args.port} 被占用或无访问权限")
            print(f"")
            print(f"可能的原因：")
            print(f"  1. 端口已被其他程序使用")
            print(f"  2. Windows系统保留了该端口")
            print(f"  3. 需要管理员权限")
            print(f"")
            print(f"解决方法：")
            print(f"  方法1: 使用其他端口")
            print(f"    python main.py --port 8080")
            print(f"    python main.py --port 3000")
            print(f"")
            print(f"  方法2: 以管理员身份运行")
            print(f"    右键点击 PowerShell → 以管理员身份运行")
            print(f"")
            print(f"  方法3 (Windows): 检查并关闭占用端口的程序")
            print(f"    netstat -ano | findstr :{args.port}")
            print(f"    taskkill /PID <进程ID> /F")
            print(f"{'='*60}\n")
            logging.error(f"端口绑定失败: {e}")
        else:
            print(f"\n错误: 启动服务器失败 - {e}\n")
            logging.error(f"服务器启动失败: {e}", exc_info=True)
        sys.exit(1)
    except Exception as e:
        print(f"\n错误: 服务器运行时发生异常 - {e}\n")
        logging.error(f"服务器异常: {e}", exc_info=True)
        sys.exit(1)
    finally:
        # # 清理Chrome资源
        # if chrome_pool:
        #     logging.info("正在清理Chrome浏览器资源...")
        #     chrome_pool.cleanup()
        pass


def main():
    """主函数，启动Web服务器模式（已弃用桌面模式）"""

    parser = argparse.ArgumentParser(description='跑步助手 - Web服务器模式')
    parser.add_argument("--port", type=int, default=5000,
                        help="Web服务器端口（默认5000）")
    parser.add_argument("--host", type=str,
                        default="127.0.0.1", help="Web服务器地址（默认127.0.0.1）")
    parser.add_argument("--headless", action="store_true",
                        default=True, help="使用无头Chrome模式（默认启用）")
    parser.add_argument("--log-level", choices=['debug', 'info', 'warning',
                        'error', 'critical'], default='debug', help="设置日志级别（默认 debug）")
    parser.add_argument("--debug", action="store_true",
                        help="启用调试日志（兼容旧参数，等同于 --log-level debug）")
    args = parser.parse_args()

    # 配置详细的中文日志输出（确保UTF-8编码）
    selected_level_name = 'debug' if args.debug else args.log_level
    log_level = getattr(logging, selected_level_name.upper(), logging.DEBUG)

    log_format = "%(asctime)s [%(levelname)s] [%(filename)s:%(lineno)d] %(message)s"

    # 创建UTF-8编码的StreamHandler
    handler = logging.StreamHandler(sys.stdout)
    handler.setFormatter(logging.Formatter(log_format))

    logging.basicConfig(
        level=log_level,
        format=log_format,
        handlers=[handler]
    )

    logging.info("="*60)
    logging.info("跑步助手 Web 模式启动中...")
    level_name_map = {
        logging.DEBUG: "调试",
        logging.INFO: "信息",
        logging.WARNING: "警告",
        logging.ERROR: "错误",
        logging.CRITICAL: "严重"
    }
    logging.info(
        f"日志级别: {level_name_map.get(log_level, selected_level_name.upper())} ({selected_level_name.upper()})")
    logging.info(f"服务器地址: {args.host}:{args.port}")
    logging.info("="*60)

    check_and_import_dependencies()

    # 检查Playwright是否可用
    if not playwright_available:
        print("\n" + "="*60)
        print("错误: 需要安装 Playwright 以在服务器端运行Chrome")
        print("请运行以下命令:")
        print("  pip install playwright")
        print("  python -m playwright install chromium")
        print("="*60 + "\n")
        sys.exit(1)

    initial_port = args.port
    if not check_port_available(args.host, args.port):
        logging.warning(f"指定的端口 {args.port} 不可用或已被占用，尝试自动查找可用端口...")
        found_port = None
        # 尝试常用备选端口
        alternative_ports = [8080, 8000, 3000, 5001, 8888, 9000, 5005, 5050]
        for port in alternative_ports:
            logging.info(f"尝试端口 {port}...")
            if check_port_available(args.host, port):
                found_port = port
                logging.info(f"找到可用端口: {found_port}")
                break

        # 如果常用端口都不可用，尝试随机端口 (例如 10000 到 65535 之间)
        if not found_port:
            logging.info("常用备选端口均不可用，尝试在 10000-65535 范围内查找随机可用端口...")
            max_random_tries = 20  # 限制尝试次数
            for i in range(max_random_tries):
                random_port = random.randint(10000, 65535)
                # logging.debug(f"尝试随机端口 {random_port} ({i+1}/{max_random_tries})...") # Debug日志
                if check_port_available(args.host, random_port):
                    found_port = random_port
                    logging.info(f"找到可用随机端口: {found_port}")
                    break
                # 短暂等待避免CPU占用过高
                time.sleep(0.01)

        # 如果最终仍未找到可用端口
        if not found_port:
            logging.error(f"自动查找端口失败。初始端口 {initial_port} 及所有尝试的备选/随机端口均不可用。")
            print(f"\n{'='*60}")
            print(f"错误: 无法自动找到可用的网络端口。")
            print(f"请检查端口 {initial_port} 或其他常用端口是否被占用，或手动指定一个可用端口:")
            print(f"  python main.py --port <可用端口号>")
            print(f"{'='*60}\n")
            sys.exit(1)
        else:
            args.port = found_port  # 更新 args 中的端口号为找到的可用端口

    # 启动Web服务器模式
    logging.info("启动Web服务器模式（使用服务器端Chrome渲染）...")
    start_web_server(args)


if __name__ == "__main__":
    main()
